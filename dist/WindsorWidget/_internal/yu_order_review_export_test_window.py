
from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import sys
import shutil
import tempfile
import traceback
import zipfile
from collections import defaultdict
from copy import copy, deepcopy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

try:
    import pyodbc
except Exception:
    pyodbc = None

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter

from PySide6.QtCore import Qt, QSettings, QSize, QUrl, Signal
from PySide6.QtGui import QColor, QBrush, QDesktopServices
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QStatusBar,
    QTableWidget,
    QTableWidgetItem,
    QTextBrowser,
    QVBoxLayout,
    QWidget,
)


TABLE_FONT_SIZE_OPTIONS = (8, 9, 10, 11, 12, 14, 16, 18, 20)
TABLE_FONT_SETTINGS_PREFIX = "table_font_sizes"

DEFAULT_PREFIX = "yu_test"
DEFAULT_ORDER_CSV = ""
DEFAULT_TEMPLATE = ""
DEFAULT_OUTPUT_DIR = "yu_exports"

MYOB_PO_HEADERS = [
    "Co./Last Name",
    "First Name",
    "Addr 1 - Line 1",
    "Addr 1 - Line 2",
    "Addr 1 - Line 3",
    "Addr 1 - Line 4",
    "Inclusive",
    "Purchase No.",
    "Date",
    "Supplier Invoice No.",
    "Ship Via",
    "Delivery Status",
    "Item Number",
    "Quantity",
    "Description",
    "Price",
    "Discount",
    "Total",
    "Job",
    "Comment",
    "Journal Memo",
    "Shipping Date",
    "Tax Code",
    "Tax Amount",
    "Freight Amount",
    "Freight Tax Code",
    "Freight Tax Amount",
    "Purchase Status",
    "Terms - Payment is Due",
    "           - Discount Days",
    "           - Balance Due Days",
    "           - % Discount",
    "Amount Paid",
    "Category",
    "Order",
    "Received",
    "Billed",
    "Location ID",
    "Card ID",
    "Record ID",
]
MYOB_YU_SUPPLIER = "Yuchang Textile Factory"
MYOB_YU_SHIP_VIA = "SEA"
MYOB_YU_DELIVERY_STATUS = "P"
MYOB_YU_TAX_CODE = "OSP"
MYOB_YU_PURCHASE_STATUS = "O"
MYOB_YU_LOCATION_ID = "Location1"


def _decimal_value(value: Any, field_name: str = "value") -> Decimal:
    if isinstance(value, Decimal):
        result = value
    else:
        text = str(value if value is not None else "").strip()
        text = text.replace("$", "").replace(",", "")
        if not text:
            raise ValueError(f"{field_name} is blank")
        try:
            result = Decimal(text)
        except InvalidOperation as exc:
            raise ValueError(f"Invalid {field_name}: {value!r}") from exc
    if not result.is_finite():
        raise ValueError(f"Invalid {field_name}: {value!r}")
    return result


def _format_myob_number(value: Any, max_places: int = 6) -> str:
    number = _decimal_value(value)
    quantum = Decimal(1).scaleb(-max_places)
    number = number.quantize(quantum)
    text = format(number, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _format_myob_quantity(value: Any) -> str:
    number = _decimal_value(value, "quantity")
    return f"{number.quantize(Decimal('0.001')):.3f}"


def _normalise_myob_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        raise ValueError("Order date is blank")
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    raise ValueError(f"Invalid order date: {value!r}")


def _single_line_description(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def write_myob_po_import_txt(
    output_path: str | Path,
    order_date: str,
    order_number: str,
    lines: Iterable[dict[str, Any]],
) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    normalised_date = _normalise_myob_date(order_date)
    order_number = str(order_number or "").strip()
    if not order_number:
        raise ValueError("Purchase order number is blank")

    prepared_rows = []
    for line in lines:
        item_number = str(line.get("item_number") or "").strip()
        description = _single_line_description(line.get("description"))
        quantity = _decimal_value(line.get("quantity"), f"quantity for {item_number}")
        price = _decimal_value(line.get("price"), f"price for {item_number}")
        if not item_number:
            raise ValueError("A MYOB export line has a blank item number")
        if not description:
            raise ValueError(f"{item_number} has no description")
        if quantity <= 0:
            raise ValueError(f"{item_number} has a non-positive quantity")
        if price <= 0:
            raise ValueError(f"{item_number} has a non-positive YU cost")
        qty_text = _format_myob_quantity(quantity)
        # AccountRight validates the transaction-level total even when importing
        # an Order.  Omitting Total while supplying a non-zero Price can produce
        # error -4317 ("An unbalanced transaction may not be recorded").
        # Match the native AccountRight export format and round the line total
        # to currency precision.
        line_total = (quantity * price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        prepared_rows.append({
            "Co./Last Name": MYOB_YU_SUPPLIER,
            "First Name": "",
            "Addr 1 - Line 1": "",
            "Addr 1 - Line 2": "",
            "Addr 1 - Line 3": "",
            "Addr 1 - Line 4": "",
            "Inclusive": "",
            "Purchase No.": order_number,
            "Date": normalised_date,
            # The Widget order number belongs in Purchase No. only.
            # Supplier Invoice No. is the supplier's own invoice/reference and
            # must remain blank when creating a new purchase order.
            "Supplier Invoice No.": "",
            "Ship Via": MYOB_YU_SHIP_VIA,
            "Delivery Status": MYOB_YU_DELIVERY_STATUS,
            "Item Number": item_number,
            "Quantity": qty_text,
            "Description": description,
            "Price": _format_myob_number(price, 6),
            "Discount": "0%",
            "Total": f"{line_total:.2f}",
            "Job": "",
            "Comment": "",
            "Journal Memo": f"Purchase; {MYOB_YU_SUPPLIER}",
            "Shipping Date": "",
            "Tax Code": MYOB_YU_TAX_CODE,
            "Tax Amount": "0.00",
            "Freight Amount": "0.00",
            "Freight Tax Code": MYOB_YU_TAX_CODE,
            "Freight Tax Amount": "0.00",
            "Purchase Status": MYOB_YU_PURCHASE_STATUS,
            "Terms - Payment is Due": "",
            "           - Discount Days": "",
            "           - Balance Due Days": "",
            "           - % Discount": "",
            "Amount Paid": "0.00",
            "Category": "",
            "Order": qty_text,
            "Received": "0.000",
            "Billed": "0.000",
            "Location ID": MYOB_YU_LOCATION_ID,
            "Card ID": "",
            "Record ID": "",
        })

    if not prepared_rows:
        raise ValueError("There are no MYOB purchase-order lines to export")

    # AccountRight recognises the first-line {} marker as the modern format.
    # cp1252 is safer than UTF-8 with BOM for the legacy Import/Export Assistant.
    with output_path.open("w", encoding="cp1252", errors="replace", newline="") as handle:
        handle.write("{}\r\n")
        writer = csv.DictWriter(
            handle,
            fieldnames=MYOB_PO_HEADERS,
            delimiter=",",
            lineterminator="\r\n",
            quoting=csv.QUOTE_MINIMAL,
            extrasaction="raise",
        )
        writer.writeheader()
        writer.writerows(prepared_rows)
        handle.write("\r\n")
    return str(output_path)


# ------------------------------------------------------------
# Table font helpers copied to match main app behaviour
# ------------------------------------------------------------
def _safe_table_settings_token(table, fallback_scope="table"):
    token = str(table.property("_table_font_settings_token") or "").strip()
    if token:
        return token

    parts = []
    current = table
    while current is not None:
        name = ""
        try:
            name = str(current.objectName() or "").strip()
        except Exception:
            name = ""
        if name and not name.startswith("qt_"):
            parts.append(name)
        try:
            current = current.parentWidget()
        except Exception:
            current = None
    parts.reverse()

    token = "/".join(parts) if parts else fallback_scope
    token = re.sub(r"[^A-Za-z0-9_./-]+", "_", token)
    return token or fallback_scope


def table_font_settings_key(scope_key, table):
    token = _safe_table_settings_token(table, scope_key)
    return f"{TABLE_FONT_SETTINGS_PREFIX}/{scope_key}/{token}"


def stored_table_font_size(settings, scope_key, table):
    if settings is None or table is None:
        return None
    try:
        value = settings.value(table_font_settings_key(scope_key, table), None)
    except Exception:
        return None
    if value in (None, ""):
        return None
    try:
        size = int(value)
    except Exception:
        return None
    return size if size > 0 else None


def current_table_font_size(table):
    if table is None:
        return 10
    try:
        size = int(round(float(table.font().pointSizeF())))
        if size > 0:
            return size
    except Exception:
        pass
    try:
        size = int(table.font().pointSize())
        if size > 0:
            return size
    except Exception:
        pass
    return 10


def sync_table_item_font_sizes(table, size):
    if table is None or not isinstance(table, QTableWidget):
        return
    for row in range(table.rowCount()):
        for column in range(table.columnCount()):
            item = table.item(row, column)
            if item is None:
                continue
            item_font = item.font()
            item_font.setPointSize(size)
            item.setFont(item_font)


def apply_table_font_size(table, size, settings=None, scope_key=None, persist=True):
    if table is None:
        return
    try:
        size = int(size)
    except Exception:
        return
    if size < 6:
        size = 6

    font = table.font()
    font.setPointSize(size)
    table.setFont(font)
    sync_table_item_font_sizes(table, size)

    vertical_header = getattr(table, "verticalHeader", lambda: None)()
    if vertical_header is not None:
        try:
            vertical_header.setDefaultSectionSize(max(24, size * 2 + 8))
        except Exception:
            pass
    try:
        table.resizeRowsToContents()
    except Exception:
        pass
    try:
        table.viewport().update()
    except Exception:
        pass

    if persist and settings is not None and scope_key:
        try:
            settings.setValue(table_font_settings_key(scope_key, table), size)
        except Exception:
            pass


def reset_table_font_size(table, settings=None, scope_key=None):
    if table is None:
        return
    default_size = None
    try:
        default_size = int(table.property("_table_font_default_size") or 0)
    except Exception:
        default_size = None
    if not default_size:
        default_size = 10

    if settings is not None and scope_key:
        try:
            settings.remove(table_font_settings_key(scope_key, table))
        except Exception:
            pass
    apply_table_font_size(table, default_size, settings=settings, scope_key=scope_key, persist=False)


def install_table_font_context_menu(owner, table, settings, scope_key, settings_token=None):
    if table is None:
        return
    if table.property("_table_font_menu_installed"):
        saved_size = stored_table_font_size(settings, scope_key, table)
        if saved_size:
            apply_table_font_size(table, saved_size, settings=settings, scope_key=scope_key, persist=False)
        return

    if settings_token:
        table.setProperty("_table_font_settings_token", settings_token)
    elif not str(table.objectName() or "").strip():
        table.setObjectName(_safe_table_settings_token(table, scope_key))

    table.setProperty("_table_font_menu_installed", True)
    table.setProperty("_table_font_scope_key", scope_key)
    table.setProperty("_table_font_default_size", current_table_font_size(table))
    table.setContextMenuPolicy(Qt.CustomContextMenu)
    table.customContextMenuRequested.connect(
        lambda pos, tbl=table, own=owner, s=settings, scope=scope_key: show_table_font_context_menu(own, tbl, pos, s, scope)
    )

    saved_size = stored_table_font_size(settings, scope_key, table)
    if saved_size:
        apply_table_font_size(table, saved_size, settings=settings, scope_key=scope_key, persist=False)


def show_table_font_context_menu(owner, table, pos, settings, scope_key):
    if table is None:
        return

    menu = QMenu(table)
    font_menu = menu.addMenu("Font Size")
    current_size = current_table_font_size(table)

    for size in TABLE_FONT_SIZE_OPTIONS:
        action = font_menu.addAction(str(size))
        action.setCheckable(True)
        action.setChecked(size == current_size)
        action.triggered.connect(
            lambda _checked=False, tbl=table, selected_size=size, s=settings, scope=scope_key: apply_table_font_size(
                tbl, selected_size, settings=s, scope_key=scope, persist=True
            )
        )

    font_menu.addSeparator()
    reset_action = font_menu.addAction("Reset")
    reset_action.triggered.connect(
        lambda _checked=False, tbl=table, s=settings, scope=scope_key: reset_table_font_size(tbl, settings=s, scope_key=scope)
    )

    viewport = getattr(table, "viewport", lambda: None)()
    if viewport is not None:
        global_pos = viewport.mapToGlobal(pos)
    else:
        global_pos = table.mapToGlobal(pos)
    menu.exec(global_pos)


# ------------------------------------------------------------
# DB
# ------------------------------------------------------------
def get_database_config_candidate_paths(base_dir: Path | None = None) -> list[Path]:
    base_dir = Path(base_dir or Path.cwd())
    app_dir = Path(__file__).resolve().parent

    candidate_paths: list[Path] = []
    env_path = os.environ.get("WINDSOR_WIDGET_CONFIG", "").strip()
    if env_path:
        candidate_paths.append(Path(env_path))

    candidate_paths.extend([
        app_dir / "client_config.json",
        app_dir / "data" / "client_config.json",
        Path(os.environ.get("ProgramData", r"C:\ProgramData")) / "WindsorWidget" / "client_config.json",
        base_dir / "client_config.json",
        base_dir / "data" / "client_config.json",
    ])

    deduped: list[Path] = []
    seen: set[str] = set()
    for path in candidate_paths:
        try:
            resolved = str(path.resolve())
        except Exception:
            resolved = str(path)
        if resolved in seen:
            continue
        seen.add(resolved)
        deduped.append(path)
    return deduped


def load_database_config(base_dir: Path | None = None) -> dict:
    last_error = None
    checked_paths = []
    for path in get_database_config_candidate_paths(base_dir=base_dir):
        checked_paths.append(str(path))
        if not path.exists():
            continue
        for encoding_name in ("utf-8-sig", "utf-8"):
            try:
                data = json.loads(path.read_text(encoding=encoding_name))
                if isinstance(data, dict):
                    data["_loaded_from_path"] = str(path)
                    return data
            except Exception as exc:
                last_error = f"{path}: {exc}"
                continue
    return {"_checked_paths": checked_paths, "_last_error": last_error}


class SQLHelper:
    def __init__(self, base_dir: Path | None = None) -> None:
        if pyodbc is None:
            raise RuntimeError("pyodbc is required for SQL Server mode.")

        db_config = load_database_config(base_dir=base_dir)
        provider = str(db_config.get("provider", "")).strip().lower()
        server = str(db_config.get("server", "")).strip()
        database = str(db_config.get("database", "")).strip()

        if provider != "sqlserver" or not server or not database:
            checked_paths = db_config.get("_checked_paths", [])
            last_error = db_config.get("_last_error", "")
            checked_text = "\n".join(checked_paths) if checked_paths else "(none)"
            extra = f"\n\nLast config parse error:\n{last_error}" if last_error else ""
            raise RuntimeError(
                "A valid SQL Server client_config.json was not found "
                "or does not declare provider='sqlserver'.\n\nChecked paths:\n"
                f"{checked_text}{extra}"
            )

        driver = db_config.get("driver") or "ODBC Driver 18 for SQL Server"
        port = str(db_config.get("port", "")).strip()
        username = str(db_config.get("username", "")).strip()
        password = str(db_config.get("password", "")).strip()
        trusted = bool(db_config.get("trusted_connection", False))
        encrypt = str(db_config.get("encrypt", "no")).strip().lower()
        trust_cert = str(db_config.get("trust_server_certificate", "yes")).strip().lower()
        timeout = int(db_config.get("timeout", 5) or 5)
        server_spec = f"{server},{port}" if port else server

        parts = [
            f"DRIVER={{{driver}}}",
            f"SERVER={server_spec}",
            f"DATABASE={database}",
            f"Encrypt={'yes' if encrypt in {'1', 'true', 'yes'} else 'no'}",
            f"TrustServerCertificate={'yes' if trust_cert in {'1', 'true', 'yes'} else 'no'}",
            f"Connection Timeout={timeout}",
        ]
        if trusted:
            parts.append("Trusted_Connection=yes")
        else:
            parts.append(f"UID={username}")
            parts.append(f"PWD={password}")

        self.loaded_from_path = db_config.get("_loaded_from_path", "")
        self.conn = pyodbc.connect(";".join(parts))

    def close(self):
        self.conn.close()

    def all(self, sql: str, params: tuple | list = ()) -> list[dict[str, Any]]:
        cur = self.conn.cursor()
        cur.execute(sql, params)
        columns = [desc[0] for desc in cur.description] if cur.description else []
        rows = []
        for row in cur.fetchall():
            rows.append(dict(zip(columns, row)))
        return rows

    def one(self, sql: str, params: tuple | list = ()) -> dict[str, Any] | None:
        cur = self.conn.cursor()
        cur.execute(sql, params)
        row = cur.fetchone()
        if row is None:
            return None
        columns = [desc[0] for desc in cur.description] if cur.description else []
        return dict(zip(columns, row))

    def execute(self, sql: str, params: tuple | list = ()) -> None:
        cur = self.conn.cursor()
        cur.execute(sql, params)
        self.conn.commit()

    def scalar(self, sql: str, params: tuple | list = ()) -> Any:
        cur = self.conn.cursor()
        cur.execute(sql, params)
        row = cur.fetchone()
        return row[0] if row else None


class ExistingDBAdapter:
    """
    Wrap the main app's existing database backend so the YU review window can
    reuse the live connection instead of bootstrapping a second SQLHelper.
    This avoids installer / config path differences across PCs.
    """
    def __init__(self, backend) -> None:
        if backend is None:
            raise RuntimeError("ExistingDBAdapter requires a live database backend.")
        self.backend = backend
        self.loaded_from_path = "reused-main-app-connection"

    def close(self):
        # Deliberately do nothing. The main app owns this connection.
        return None

    def all(self, sql: str, params: tuple | list = ()) -> list[dict[str, Any]]:
        cur = self.backend.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
        return [dict(row) if not isinstance(row, dict) else row for row in rows]

    def one(self, sql: str, params: tuple | list = ()) -> dict[str, Any] | None:
        cur = self.backend.cursor()
        cur.execute(sql, params)
        row = cur.fetchone()
        if row is None:
            return None
        return dict(row) if not isinstance(row, dict) else row

    def execute(self, sql: str, params: tuple | list = ()) -> None:
        cur = self.backend.cursor()
        cur.execute(sql, params)
        self.backend.commit()

    def scalar(self, sql: str, params: tuple | list = ()) -> Any:
        cur = self.backend.cursor()
        cur.execute(sql, params)
        row = cur.fetchone()
        if row is None:
            return None
        try:
            return row[0]
        except Exception:
            values = list(dict(row).values()) if isinstance(row, dict) else list(row)
            return values[0] if values else None




# ------------------------------------------------------------
# Item-number resolver helpers
# ------------------------------------------------------------
def clean_item_key(value: str) -> str:
    """Return the space-insensitive item key used for non-FASTENERS lookups."""
    return re.sub(r"[\s\u00A0]+", "", str(value or "").strip()).upper()


def clean_item_sql_expr(expression: str) -> str:
    """SQL Server expression matching clean_item_key()."""
    return (
        "UPPER(REPLACE(REPLACE(REPLACE("
        f"LTRIM(RTRIM(ISNULL({expression}, ''))), "
        "' ', ''), CHAR(9), ''), CHAR(160), ''))"
    )


def _normalise_workbook_text(value: Any) -> str:
    """Return a stable comparison value for supplier-line workbook fields."""
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isfinite(value):
            return format(value, ".12g").upper()
        return str(value).upper()
    return re.sub(r"[\s\u00A0]+", " ", str(value).strip()).upper()


def _workbook_row_signature_from_values(values: Iterable[Any]) -> tuple[str, ...]:
    return tuple(_normalise_workbook_text(value) for value in values)


def _workbook_row_is_detail_values(values: list[Any] | tuple[Any, ...]) -> bool:
    """Mirror the YU template's detail-row rule using columns B:H."""
    if not values:
        return False
    description = values[0] if len(values) > 0 else None
    if description in (None, ""):
        return False
    return any(value not in (None, "") for value in values[1:7])


def detect_yuchang_footer_rows(
    worksheet,
    *,
    header_end_row: int = 14,
    export_min_col: int = 1,
    export_max_col: int = 14,
) -> tuple[int, int]:
    """Locate the movable YU footer from its labels instead of fixed row numbers."""
    transport_row: int | None = None
    total_row: int | None = None

    for row in range(max(header_end_row + 1, 1), int(worksheet.max_row or 0) + 1):
        values = [
            _normalise_workbook_text(worksheet.cell(row, col).value)
            for col in range(export_min_col, export_max_col + 1)
        ]
        if total_row is None and any("ORDER TOTAL" == value or value.startswith("ORDER TOTAL ") for value in values):
            total_row = row
        if transport_row is None and any("METHOD OF TRANSPORT" in value for value in values):
            transport_row = row

    if total_row is None:
        total_row = int(worksheet.max_row or header_end_row)

    if transport_row is None or transport_row > total_row:
        transport_row = max(header_end_row + 1, total_row - 5)

    footer_start = transport_row
    previous_row = transport_row - 1
    if previous_row > header_end_row:
        previous_values = [
            worksheet.cell(previous_row, col).value
            for col in range(export_min_col, export_max_col + 1)
        ]
        if all(value in (None, "") for value in previous_values):
            footer_start = previous_row

    return int(footer_start), int(total_row)


def scan_yuchang_workbook(
    template_path: str,
    *,
    sheet_name: str = "Sheet1",
    item_col: str = "A",
    header_end_row: int = 14,
    export_min_col: str = "A",
    export_max_col: str = "N",
) -> dict[str, Any]:
    """Build a live workbook index in one worksheet pass.

    Column A is the permanent item identifier. Row numbers are deliberately
    treated as current workbook addresses only.
    """
    workbook_path = Path(str(template_path or "").strip())
    if not workbook_path.exists():
        raise FileNotFoundError(f"YU workbook was not found: {workbook_path}")

    wb = load_workbook(str(workbook_path), read_only=True, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        ws = wb[sheet_name]

        item_idx = column_index_from_string(item_col)
        export_min_idx = column_index_from_string(export_min_col)
        export_max_idx = column_index_from_string(export_max_col)
        read_max_col = max(export_max_idx, item_idx, 8)

        # Read the worksheet only once. Calling ReadOnlyWorksheet.cell() for every
        # coordinate reparses the XML repeatedly and is extremely slow.
        row_cache: list[tuple[int, tuple[Any, ...]]] = []
        transport_row: int | None = None
        total_row: int | None = None

        for row_number, row_values in enumerate(
            ws.iter_rows(
                min_row=header_end_row + 1,
                max_row=int(ws.max_row or 0),
                min_col=1,
                max_col=read_max_col,
                values_only=True,
            ),
            start=header_end_row + 1,
        ):
            values = tuple(row_values)
            row_cache.append((row_number, values))

            export_values = [
                _normalise_workbook_text(values[col - 1] if col - 1 < len(values) else None)
                for col in range(export_min_idx, export_max_idx + 1)
            ]
            if total_row is None and any(
                value == "ORDER TOTAL" or value.startswith("ORDER TOTAL ")
                for value in export_values
            ):
                total_row = row_number
            if transport_row is None and any("METHOD OF TRANSPORT" in value for value in export_values):
                transport_row = row_number

        if total_row is None:
            total_row = int(ws.max_row or header_end_row)
        if transport_row is None or transport_row > total_row:
            transport_row = max(header_end_row + 1, total_row - 5)

        footer_start = int(transport_row)
        previous_row_number = footer_start - 1
        previous_values = next(
            (values for row_number, values in row_cache if row_number == previous_row_number),
            (),
        )
        if previous_values:
            previous_export_values = [
                previous_values[col - 1] if col - 1 < len(previous_values) else None
                for col in range(export_min_idx, export_max_idx + 1)
            ]
            if all(value in (None, "") for value in previous_export_values):
                footer_start = previous_row_number

        exact_rows: dict[str, list[int]] = defaultdict(list)
        clean_rows: dict[str, list[int]] = defaultdict(list)
        row_item_numbers: dict[int, str] = {}
        detail_rows: set[int] = set()
        signature_rows: dict[tuple[str, ...], list[int]] = defaultdict(list)
        loose_signature_rows: dict[tuple[str, ...], list[int]] = defaultdict(list)

        for row_number, values_tuple in row_cache:
            if row_number >= footer_start:
                break
            values = list(values_tuple)
            item_value = values[item_idx - 1] if item_idx - 1 < len(values) else None
            item_number = str(item_value or "").strip()

            supplier_values = values[1:8]
            if _workbook_row_is_detail_values(supplier_values):
                detail_rows.add(row_number)
                # Exact signature: B:G. Loose signature omits F (unit size),
                # which can be represented differently by Excel and SQL Server.
                exact_signature = _workbook_row_signature_from_values(supplier_values[:6])
                loose_signature = _workbook_row_signature_from_values(
                    [supplier_values[0], supplier_values[1], supplier_values[2], supplier_values[3], supplier_values[5]]
                )
                signature_rows[exact_signature].append(row_number)
                loose_signature_rows[loose_signature].append(row_number)

            if not item_number:
                continue

            row_item_numbers[row_number] = item_number
            exact_rows[item_number.upper()].append(row_number)
            clean_rows[clean_item_key(item_number)].append(row_number)

        return {
            "path": str(workbook_path),
            "mtime_ns": workbook_path.stat().st_mtime_ns,
            "sheet_name": sheet_name,
            "header_end_row": int(header_end_row),
            "footer_start_row": int(footer_start),
            "footer_end_row": int(total_row),
            "exact_rows": dict(exact_rows),
            "clean_rows": dict(clean_rows),
            "row_item_numbers": row_item_numbers,
            "detail_rows": detail_rows,
            "signature_rows": dict(signature_rows),
            "loose_signature_rows": dict(loose_signature_rows),
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def workbook_rows_for_item_from_scan(scan: dict[str, Any], item_number: str, *, exact: bool = False) -> list[int]:
    item_number = str(item_number or "").strip()
    if not item_number:
        return []
    if exact:
        return list((scan.get("exact_rows") or {}).get(item_number.upper(), []))
    return list((scan.get("clean_rows") or {}).get(clean_item_key(item_number), []))


def resolve_yuchang_items_to_current_rows(
    template_path: str,
    item_numbers_with_qty: Iterable[tuple[str, float | int]],
    *,
    sheet_name: str = "Sheet1",
    exact_item_numbers: Iterable[str] | None = None,
) -> tuple[list[tuple[int, float]], dict[str, Any]]:
    """Resolve order items against the current Column A positions."""
    scan = scan_yuchang_workbook(template_path, sheet_name=sheet_name)
    exact_keys = {str(item or "").strip().upper() for item in (exact_item_numbers or [])}
    rows_with_qty: list[tuple[int, float]] = []
    missing: list[str] = []
    duplicates: dict[str, list[int]] = {}

    for item_number, qty in item_numbers_with_qty:
        item_number = str(item_number or "").strip()
        rows = workbook_rows_for_item_from_scan(
            scan,
            item_number,
            exact=item_number.upper() in exact_keys,
        )
        if not rows:
            missing.append(item_number)
            continue
        if len(rows) != 1:
            duplicates[item_number] = sorted(set(int(row) for row in rows))
            continue
        rows_with_qty.append((int(rows[0]), float(qty)))

    if missing or duplicates:
        parts: list[str] = []
        if missing:
            parts.append("Missing from current workbook Column A: " + ", ".join(sorted(set(missing))))
        if duplicates:
            duplicate_text = "; ".join(
                f"{item}: rows {', '.join(str(row) for row in rows)}"
                for item, rows in sorted(duplicates.items())
            )
            parts.append("Duplicate item numbers in current workbook Column A: " + duplicate_text)
        raise ValueError("\n".join(parts))

    return rows_with_qty, scan

# ------------------------------------------------------------
# Order CSV + resolution + export
# ------------------------------------------------------------
@dataclass
class ReviewHit:
    source_row: int
    match_type: str
    confidence: float | None
    note: str


@dataclass
class OrderResolveResult:
    item_number: str
    quantity: float
    date: str
    order_number: str
    status: str
    source_row: int | None
    source: str
    note: str
    review_hits: list[ReviewHit]


def read_order_csv(csv_path: str) -> list[dict]:
    rows: list[dict] = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        expected = {"Date", "Order Number", "Item Number", "QTY"}
        missing = expected.difference(reader.fieldnames or [])
        if missing:
            raise ValueError(f"CSV is missing required columns: {sorted(missing)}")
        for line_no, row in enumerate(reader, start=2):
            item_number = str(row.get("Item Number", "")).strip()
            if not item_number:
                continue
            try:
                qty = float(row.get("QTY", 0) or 0)
            except Exception as exc:
                raise ValueError(f"Invalid QTY on CSV line {line_no}: {row.get('QTY')!r}") from exc
            rows.append(
                {
                    "Date": str(row.get("Date", "")).strip(),
                    "Order Number": str(row.get("Order Number", "")).strip(),
                    "Item Number": item_number,
                    "QTY": qty,
                }
            )
    return rows


def group_order_rows(rows: Iterable[dict]) -> list[dict]:
    grouped: dict[tuple[str, str, str], float] = defaultdict(float)
    for row in rows:
        key = (row["Date"], row["Order Number"], row["Item Number"])
        grouped[key] += float(row["QTY"])
    out: list[dict] = []
    for (date_text, order_no, item_number), qty in sorted(grouped.items(), key=lambda x: (x[0][1], x[0][0], x[0][2])):
        out.append({
            "Date": date_text,
            "Order Number": order_no,
            "Item Number": item_number,
            "QTY": qty,
        })
    return out


def write_audit_csv(audit_path: str, results: list[OrderResolveResult]) -> None:
    out_path = Path(audit_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "date",
                "order_number",
                "item_number",
                "quantity",
                "status",
                "resolved_workbook_row",
                "resolution_source",
                "note",
                "review_hit_1",
                "review_hit_2",
                "review_hit_3",
            ]
        )
        for result in results:
            hit_text = []
            for hit in result.review_hits[:3]:
                conf = "" if hit.confidence is None else f" {hit.confidence:.1%}"
                hit_text.append(f"row {hit.source_row} | {hit.match_type}{conf} | {hit.note}")
            while len(hit_text) < 3:
                hit_text.append("")
            writer.writerow(
                [
                    result.date,
                    result.order_number,
                    result.item_number,
                    result.quantity,
                    result.status,
                    result.source_row or "",
                    result.source,
                    result.note,
                    *hit_text,
                ]
            )



SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
XLSX_IGNORABLE_NAMESPACES = {
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
    "xr6": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6",
    "xr10": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10",
}
ET_NS = {"s": SHEET_NS, "r": OFFICE_REL_NS, "rel": PACKAGE_REL_NS}


def _register_xlsx_namespaces() -> None:
    try:
        import xml.etree.ElementTree as ET
        ET.register_namespace("", SHEET_NS)
        ET.register_namespace("r", OFFICE_REL_NS)
        ET.register_namespace("mc", MC_NS)
        for prefix, namespace_uri in XLSX_IGNORABLE_NAMESPACES.items():
            ET.register_namespace(prefix, namespace_uri)
    except Exception:
        pass


def _xml_namespace_is_used(root, namespace_uri: str) -> bool:
    namespace_marker = f"{{{namespace_uri}}}"
    for node in root.iter():
        if str(node.tag).startswith(namespace_marker):
            return True
        for attr_name in node.attrib.keys():
            if str(attr_name).startswith(namespace_marker):
                return True
    return False


def _ensure_mc_ignorable_namespace_declarations(root) -> None:
    ignorable_attr = f"{{{MC_NS}}}Ignorable"
    ignorable_text = str(root.attrib.get(ignorable_attr, "") or "").strip()
    if not ignorable_text:
        return

    for prefix in ignorable_text.split():
        namespace_uri = XLSX_IGNORABLE_NAMESPACES.get(prefix)
        if not namespace_uri:
            continue
        # ElementTree drops namespace declarations that are only mentioned inside
        # mc:Ignorable. Excel treats that as a damaged worksheet part, so add
        # explicit xmlns declarations for ignorable prefixes that are otherwise unused.
        if not _xml_namespace_is_used(root, namespace_uri):
            root.set(f"xmlns:{prefix}", namespace_uri)


def _xlsx_xml_tostring(root) -> bytes:
    import xml.etree.ElementTree as ET

    _ensure_mc_ignorable_namespace_declarations(root)
    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    if data.startswith(b"<?xml version='1.0' encoding='utf-8'?>"):
        data = data.replace(
            b"<?xml version='1.0' encoding='utf-8'?>",
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            1,
        )
    return data


def _xlsx_part_for_sheet(zip_file: zipfile.ZipFile, sheet_name: str) -> str:
    import xml.etree.ElementTree as ET

    workbook_root = ET.fromstring(zip_file.read("xl/workbook.xml"))
    rels_root = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {rel.attrib.get("Id"): rel.attrib.get("Target", "") for rel in rels_root}

    sheets_node = workbook_root.find("s:sheets", ET_NS)
    if sheets_node is None:
        raise ValueError("Workbook contains no sheets list.")

    for sheet in sheets_node.findall("s:sheet", ET_NS):
        if sheet.attrib.get("name") != sheet_name:
            continue
        rel_id = sheet.attrib.get(f"{{{OFFICE_REL_NS}}}id")
        target = rel_targets.get(rel_id, "")
        if not target:
            break
        if target.startswith("/"):
            return target.lstrip("/")
        return str(PurePosixPath("xl") / target)

    raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")


def _xlsx_shared_strings(zip_file: zipfile.ZipFile) -> list[str]:
    import xml.etree.ElementTree as ET

    if "xl/sharedStrings.xml" not in zip_file.namelist():
        return []
    root = ET.fromstring(zip_file.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for si in root.findall("s:si", ET_NS):
        parts = []
        for node in si.iter(f"{{{SHEET_NS}}}t"):
            parts.append(node.text or "")
        strings.append("".join(parts))
    return strings


def _cell_text(cell, shared_strings: list[str]) -> str:
    if cell is None:
        return ""
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        parts = []
        for node in cell.iter(f"{{{SHEET_NS}}}t"):
            parts.append(node.text or "")
        return "".join(parts)
    value_node = cell.find("s:v", ET_NS)
    if value_node is None or value_node.text is None:
        return ""
    value = value_node.text
    if cell_type == "s":
        try:
            return shared_strings[int(value)]
        except Exception:
            return value
    return value


def _excel_int(value: Any) -> int | None:
    try:
        if value is None:
            return None
        if isinstance(value, int):
            return int(value)
        text = str(value).strip()
        if not text:
            return None
        return int(float(text))
    except Exception:
        return None


def _split_cell_ref(cell_ref: str) -> tuple[str, int]:
    match = re.match(r"^([A-Z]+)(\d+)$", str(cell_ref or "").upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    return match.group(1), int(match.group(2))


def _cell_sort_key(cell) -> int:
    ref = cell.attrib.get("r", "A1")
    col_text, _row = _split_cell_ref(ref)
    return column_index_from_string(col_text)


def _sheet_rows_by_number(root) -> dict[int, Any]:
    sheet_data = root.find("s:sheetData", ET_NS)
    if sheet_data is None:
        raise ValueError("Worksheet has no sheetData node.")
    rows: dict[int, Any] = {}
    for row in sheet_data.findall("s:row", ET_NS):
        try:
            rows[int(row.attrib.get("r", "0"))] = row
        except Exception:
            pass
    return rows


def _ensure_xml_row(root, row_number: int):
    import xml.etree.ElementTree as ET

    sheet_data = root.find("s:sheetData", ET_NS)
    if sheet_data is None:
        sheet_data = ET.SubElement(root, f"{{{SHEET_NS}}}sheetData")
    rows = _sheet_rows_by_number(root)
    row = rows.get(row_number)
    if row is not None:
        return row
    row = ET.Element(f"{{{SHEET_NS}}}row", {"r": str(row_number)})
    inserted = False
    for index, existing in enumerate(list(sheet_data)):
        try:
            if int(existing.attrib.get("r", "0")) > row_number:
                sheet_data.insert(index, row)
                inserted = True
                break
        except Exception:
            continue
    if not inserted:
        sheet_data.append(row)
    return row


def _ensure_xml_cell(root, cell_ref: str):
    import xml.etree.ElementTree as ET

    _col_text, row_number = _split_cell_ref(cell_ref)
    row = _ensure_xml_row(root, row_number)
    for cell in row.findall("s:c", ET_NS):
        if cell.attrib.get("r") == cell_ref:
            return cell
    cell = ET.Element(f"{{{SHEET_NS}}}c", {"r": cell_ref})
    children = list(row)
    target_key = _cell_sort_key(cell)
    inserted = False
    for index, existing in enumerate(children):
        try:
            if _cell_sort_key(existing) > target_key:
                row.insert(index, cell)
                inserted = True
                break
        except Exception:
            continue
    if not inserted:
        row.append(cell)
    return cell


def _set_xml_cell_text(root, cell_ref: str, value: str | None) -> None:
    import xml.etree.ElementTree as ET

    cell = _ensure_xml_cell(root, cell_ref)
    for child in list(cell):
        tag = child.tag
        if tag in {f"{{{SHEET_NS}}}v", f"{{{SHEET_NS}}}is", f"{{{SHEET_NS}}}f"}:
            cell.remove(child)
    if value is None or str(value).strip() == "":
        cell.attrib.pop("t", None)
        return
    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, f"{{{SHEET_NS}}}is")
    text_node = ET.SubElement(inline, f"{{{SHEET_NS}}}t")
    text = str(value).strip()
    text_node.text = text
    if text.startswith(" ") or text.endswith(" "):
        text_node.attrib["{http://www.w3.org/XML/1998/namespace}space"] = "preserve"


def _find_header_column_in_xml(root, shared_strings: list[str], header_text: str, header_row: int = 4, default_col: str = "A") -> int:
    wanted = str(header_text or "").strip().lower()
    row = _sheet_rows_by_number(root).get(header_row)
    if row is not None:
        for cell in row.findall("s:c", ET_NS):
            if _cell_text(cell, shared_strings).strip().lower() == wanted:
                col_text, _row_number = _split_cell_ref(cell.attrib.get("r", "A1"))
                return column_index_from_string(col_text)
    return column_index_from_string(default_col)


def _review_row_map_by_source(root, shared_strings: list[str], source_col: int, header_row: int = 4) -> dict[int, int]:
    rows_by_number = _sheet_rows_by_number(root)
    source_letter = get_column_letter(source_col)
    out: dict[int, int] = {}
    for row_number, row in rows_by_number.items():
        if row_number <= header_row:
            continue
        source_cell = None
        source_ref = f"{source_letter}{row_number}"
        for cell in row.findall("s:c", ET_NS):
            if cell.attrib.get("r") == source_ref:
                source_cell = cell
                break
        source_row = _excel_int(_cell_text(source_cell, shared_strings))
        if source_row is not None:
            out[source_row] = row_number
    return out


def save_yuchang_workbook_matches(
    template_path: str,
    row_matches: dict[int, str | None],
    *,
    order_sheet_name: str = "Sheet1",
    match_col: str = "A",
    review_sheet_name: str = "Match_Review",
    review_header_row: int = 4,
    update_review_sheet: bool = False,
) -> dict:
    import xml.etree.ElementTree as ET

    if not row_matches:
        return {"updated_rows": [], "path": str(template_path or "")}

    workbook_path = Path(str(template_path or "").strip())
    if not workbook_path.exists():
        raise FileNotFoundError(f"YU workbook was not found: {workbook_path}")

    _register_xlsx_namespaces()
    updated_parts: dict[str, bytes] = {}

    with zipfile.ZipFile(str(workbook_path), "r") as zin:
        shared_strings = _xlsx_shared_strings(zin)
        order_part = _xlsx_part_for_sheet(zin, order_sheet_name)
        order_root = ET.fromstring(zin.read(order_part))
        match_letter = get_column_letter(column_index_from_string(match_col))
        updated_rows: list[int] = []

        for source_row, item_number in row_matches.items():
            row_number = int(source_row)
            if row_number <= 0:
                continue
            value = str(item_number or "").strip() or None
            _set_xml_cell_text(order_root, f"{match_letter}{row_number}", value)
            updated_rows.append(row_number)

        updated_parts[order_part] = _xlsx_xml_tostring(order_root)

        # Match_Review stores historical source row numbers. Once users insert or
        # move rows, writing back to it by row number can attach a part number to
        # the wrong supplier line. Column A on Sheet1 is now the permanent mapping,
        # so the review sheet is left untouched unless an older caller explicitly
        # opts in.
        if update_review_sheet:
            try:
                review_part = _xlsx_part_for_sheet(zin, review_sheet_name)
            except Exception:
                review_part = ""
            if review_part:
                review_root = ET.fromstring(zin.read(review_part))
                source_col = _find_header_column_in_xml(review_root, shared_strings, "Source Row", review_header_row, "A")
                final_col = _find_header_column_in_xml(review_root, shared_strings, "Final Selection", review_header_row, "B")
                final_letter = get_column_letter(final_col)
                review_row_by_source = _review_row_map_by_source(review_root, shared_strings, source_col, review_header_row)

                for source_row, item_number in row_matches.items():
                    review_row = review_row_by_source.get(int(source_row))
                    if review_row is None:
                        continue
                    value = str(item_number or "").strip() or None
                    _set_xml_cell_text(review_root, f"{final_letter}{review_row}", value)

                updated_parts[review_part] = _xlsx_xml_tostring(review_root)

        fd, tmp_name = tempfile.mkstemp(prefix=f".{workbook_path.stem}_", suffix=workbook_path.suffix, dir=str(workbook_path.parent))
        os.close(fd)
        tmp_path = Path(tmp_name)
        try:
            with zipfile.ZipFile(str(tmp_path), "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = updated_parts.get(item.filename)
                    if data is None:
                        data = zin.read(item.filename)
                    zout.writestr(item, data)
            shutil.move(str(tmp_path), str(workbook_path))
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception:
                    pass

    return {"updated_rows": sorted(set(updated_rows)), "path": str(workbook_path)}


def _make_compact_export_row_plan(
    kept_rows: list[int],
    row_has_item_fn,
    *,
    header_end_row: int,
    footer_start_row: int,
    separator_source_gap: int = 50,
) -> list[int | None]:
    """
    Build the output row plan for compact exports.

    None entries are deliberately inserted blank rows. This keeps separated order
    sections readable after the export removes hundreds of unmatched supplier rows.
    """
    row_plan: list[int | None] = []
    previous_detail_row: int | None = None

    for old_row in kept_rows:
        is_detail_row = header_end_row < old_row < footer_start_row and row_has_item_fn(old_row)
        if (
            is_detail_row
            and previous_detail_row is not None
            and old_row - previous_detail_row > separator_source_gap
        ):
            # Do not insert an extra blank if a header/note row was just kept immediately
            # before this item. Header/note rows already act as the visual separator.
            previous_kept = row_plan[-1] if row_plan else None
            previous_kept_is_header_or_note = (
                isinstance(previous_kept, int)
                and header_end_row < previous_kept < footer_start_row
                and not row_has_item_fn(previous_kept)
            )
            if not previous_kept_is_header_or_note:
                row_plan.append(None)

        row_plan.append(old_row)

        if is_detail_row:
            previous_detail_row = old_row

    return row_plan


def _collect_related_header_note_rows_above(
    source_row: int,
    *,
    header_end_row: int,
    matched_rows: set[int],
    row_has_item_fn,
    row_has_export_content_fn,
    max_blank_scan: int = 3,
    max_unmatched_item_scan: int = 25,
) -> set[int]:
    """
    Find instruction/header rows that belong to a selected source row.

    The old compact-export scan stopped as soon as it hit an item row above the
    selected line. That misses headers for cases like the CAN mattress-tape block
    where the first item under the note is not ordered, but a later item is.

    This scan walks past nearby unmatched item rows so it can still pick up the
    note/header immediately above the block, while refusing to cross a blank gap
    after it has already skipped item rows.
    """
    related_rows: set[int] = set()
    scan = int(source_row) - 1
    blank_count = 0
    skipped_unmatched_items = 0
    found_header_or_note = False

    while scan > header_end_row:
        if row_has_item_fn(scan):
            if scan in matched_rows:
                break
            skipped_unmatched_items += 1
            if skipped_unmatched_items > max_unmatched_item_scan:
                break
            blank_count = 0
            scan -= 1
            continue

        if row_has_export_content_fn(scan):
            related_rows.add(scan)
            found_header_or_note = True
            blank_count = 0
            scan -= 1
            continue

        blank_count += 1
        if found_header_or_note or skipped_unmatched_items > 0 or blank_count > max_blank_scan:
            break
        scan -= 1

    return related_rows


def _cell_text_for_autofit(cell) -> str:
    value = cell.value
    if value in (None, ""):
        return ""
    number_format = str(getattr(cell, "number_format", "") or "")
    if isinstance(value, str):
        # Formula text is not the displayed value and can make columns wildly wide.
        # Estimate simple amount formulas where possible, otherwise ignore formula text.
        if value.startswith("="):
            formula_match = re.fullmatch(r"=\s*([A-Z]+)(\d+)\s*\*\s*([A-Z]+)(\d+)\s*", value.strip(), flags=re.IGNORECASE)
            if formula_match:
                try:
                    ws = cell.parent
                    left = ws[f"{formula_match.group(1).upper()}{formula_match.group(2)}"].value
                    right = ws[f"{formula_match.group(3).upper()}{formula_match.group(4)}"].value
                    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
                        value = float(left) * float(right)
                    else:
                        return ""
                except Exception:
                    return ""
            else:
                return ""
        else:
            text = value
            return max(str(part) for part in text.splitlines() or [""])

    if isinstance(value, (int, float)):
        if "$" in number_format:
            decimals = 4 if "0000" in number_format else 2
            text = f"${value:,.{decimals}f}"
        elif "," in number_format or "#,##" in number_format:
            text = f"{value:,.0f}" if float(value).is_integer() else f"{value:,.2f}"
        else:
            text = str(value)
    else:
        text = str(value)
    return max(str(part) for part in text.splitlines() or [""])


def _merged_cell_coordinates_to_skip_for_autofit(ws) -> set[str]:
    skip: set[str] = set()
    try:
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_col == max_col and min_row == max_row:
                continue
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    skip.add(f"{get_column_letter(col)}{row}")
    except Exception:
        return set()
    return skip


def auto_size_columns_from_row(
    ws,
    *,
    start_row: int = 12,
    min_width: float = 8.0,
    max_width: float = 45.0,
    padding: float = 2.0,
) -> None:
    """Auto-size columns using worksheet content from start_row down only."""
    if ws.max_row < start_row:
        return

    skip_merged = _merged_cell_coordinates_to_skip_for_autofit(ws)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(start_row, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.coordinate in skip_merged:
                continue
            text = _cell_text_for_autofit(cell)
            if text:
                max_len = max(max_len, len(text))
        if max_len <= 0:
            continue
        ws.column_dimensions[letter].width = min(max_width, max(min_width, max_len + padding))
        ws.column_dimensions[letter].bestFit = True


def export_yuchang_po_compact_by_rows(
    template_path: str,
    output_path: str,
    order_date: str,
    order_number: str,
    source_rows_with_qty: list[tuple[int, float | int]],
    *,
    sheet_name: str = "Sheet1",
    qty_col: str = "L",
    date_cell: str = "C10",
    order_no_cell: str = "H10",
    export_min_col: str = "A",
    export_max_col: str = "N",
    header_start_row: int = 1,
    header_end_row: int = 14,
    footer_start_row: int | None = None,
    footer_end_row: int | None = None,
    max_blank_scan: int = 3,
    separator_source_gap: int = 50,
    auto_size_start_row: int = 12,
) -> dict:
    template_path = str(template_path)
    output_path = str(output_path)

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

    src = wb[sheet_name]

    qty_idx = column_index_from_string(qty_col)
    export_min_idx = column_index_from_string(export_min_col)
    export_max_idx = column_index_from_string(export_max_col)

    if export_min_idx > export_max_idx:
        raise ValueError("export_min_col cannot be after export_max_col")

    detected_footer_start, detected_footer_end = detect_yuchang_footer_rows(
        src,
        header_end_row=header_end_row,
        export_min_col=export_min_idx,
        export_max_col=export_max_idx,
    )
    if footer_start_row is None:
        footer_start_row = detected_footer_start
    if footer_end_row is None:
        footer_end_row = detected_footer_end
    footer_start_row = int(footer_start_row)
    footer_end_row = int(footer_end_row)
    if footer_start_row <= header_end_row or footer_end_row < footer_start_row:
        raise ValueError(
            f"Invalid YU footer range detected: {footer_start_row}-{footer_end_row}"
        )

    src[date_cell] = order_date
    src[order_no_cell] = order_number

    qty_by_row: dict[int, float] = defaultdict(float)
    for source_row, qty in source_rows_with_qty:
        qty_by_row[int(source_row)] += float(qty)

    matched_rows: set[int] = set(qty_by_row.keys())
    for row, qty in qty_by_row.items():
        src.cell(row, qty_idx).value = qty

    def row_has_item(r: int) -> bool:
        if src.cell(r, 2).value in (None, ""):
            return False
        detail_cols = (3, 4, 5, 6, 7, 8)
        return any(src.cell(r, c).value not in (None, "") for c in detail_cols)

    def row_has_export_content(r: int) -> bool:
        for c in range(export_min_idx, export_max_idx + 1):
            if src.cell(r, c).value not in (None, ""):
                return True
        return False

    rows_to_keep: set[int] = set(range(header_start_row, header_end_row + 1))
    rows_to_keep.update(range(footer_start_row, footer_end_row + 1))
    rows_to_keep.update(matched_rows)

    for row in sorted(matched_rows):
        rows_to_keep.update(
            _collect_related_header_note_rows_above(
                row,
                header_end_row=header_end_row,
                matched_rows=matched_rows,
                row_has_item_fn=row_has_item,
                row_has_export_content_fn=row_has_export_content,
                max_blank_scan=max_blank_scan,
            )
        )

    kept_rows = sorted(rows_to_keep)
    row_plan = _make_compact_export_row_plan(
        kept_rows,
        row_has_item,
        header_end_row=header_end_row,
        footer_start_row=footer_start_row,
        separator_source_gap=separator_source_gap,
    )
    row_map: dict[int, int] = {}
    output_row = 1
    for old_row in row_plan:
        if old_row is None:
            output_row += 1
            continue
        row_map[old_row] = output_row
        output_row += 1

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name

    out_ws.sheet_view.showGridLines = src.sheet_view.showGridLines
    out_ws.sheet_properties = copy(src.sheet_properties)
    out_ws.page_margins = copy(src.page_margins)
    out_ws.page_setup = copy(src.page_setup)
    out_ws.print_options = copy(src.print_options)
    out_ws.sheet_format = copy(src.sheet_format)

    if src.freeze_panes:
        try:
            fp_col = src.freeze_panes.column
            fp_row = src.freeze_panes.row
            if fp_row in row_map and export_min_idx <= fp_col <= export_max_idx:
                new_fp = f"{get_column_letter(fp_col - export_min_idx + 1)}{row_map[fp_row]}"
                out_ws.freeze_panes = new_fp
        except Exception:
            pass

    for src_col in range(export_min_idx, export_max_idx + 1):
        out_col = src_col - export_min_idx + 1
        src_letter = get_column_letter(src_col)
        out_letter = get_column_letter(out_col)
        src_dim = src.column_dimensions[src_letter]
        out_dim = out_ws.column_dimensions[out_letter]
        out_dim.width = src_dim.width
        out_dim.hidden = src_dim.hidden
        out_dim.bestFit = src_dim.bestFit
        out_dim.collapsed = src_dim.collapsed
        out_dim.outlineLevel = src_dim.outlineLevel

    for old_row in row_plan:
        if old_row is None:
            continue
        new_row = row_map[old_row]
        src_row_dim = src.row_dimensions[old_row]
        out_row_dim = out_ws.row_dimensions[new_row]
        out_row_dim.height = src_row_dim.height
        out_row_dim.hidden = src_row_dim.hidden
        out_row_dim.outlineLevel = src_row_dim.outlineLevel
        out_row_dim.collapsed = src_row_dim.collapsed

        for src_col in range(export_min_idx, export_max_idx + 1):
            out_col = src_col - export_min_idx + 1
            s = src.cell(old_row, src_col)
            d = out_ws.cell(new_row, out_col)

            is_instruction_row = header_end_row < old_row < footer_start_row and not row_has_item(old_row)
            if out_col == 1 and is_instruction_row:
                # Supplier instruction/header rows can contain helper formulas in column A.
                # Those formulas are not part of the purchase order and become #REF! after
                # compact export, so leave column A blank for those rows.
                d.value = None
            elif isinstance(s.value, str) and s.value.startswith("="):
                old_coord = f"{get_column_letter(src_col)}{old_row}"
                new_coord = f"{get_column_letter(out_col)}{new_row}"
                try:
                    d.value = Translator(s.value, origin=old_coord).translate_formula(new_coord)
                except Exception:
                    d.value = s.value
            else:
                d.value = s.value

            if s.has_style:
                d.font = copy(s.font)
                d.fill = copy(s.fill)
                d.border = copy(s.border)
                d.alignment = copy(s.alignment)
                d.number_format = s.number_format
                d.protection = copy(s.protection)
            if s.hyperlink:
                d._hyperlink = copy(s.hyperlink)
            if s.comment:
                d.comment = copy(s.comment)

    for merged_range in src.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col < export_min_idx or max_col > export_max_idx:
            continue
        if not all(r in row_map for r in range(min_row, max_row + 1)):
            continue
        out_ws.merge_cells(
            start_row=row_map[min_row],
            start_column=min_col - export_min_idx + 1,
            end_row=row_map[max_row],
            end_column=max_col - export_min_idx + 1,
        )

    for img in getattr(src, "_images", []):
        try:
            anc = img.anchor
            if not hasattr(anc, "_from"):
                continue

            old_from_row = anc._from.row + 1
            old_from_col = anc._from.col + 1

            # Keep header images like the logo even when they sit outside the export column window.
            if old_from_row not in row_map:
                if old_from_row <= header_end_row:
                    target_row = old_from_row
                else:
                    continue
            else:
                target_row = row_map[old_from_row]

            img_bytes = img._data()
            new_img = XLImage(BytesIO(img_bytes))
            new_anchor = deepcopy(anc)
            new_anchor._from.row = max(0, target_row - 1)

            if old_from_col < export_min_idx:
                new_anchor._from.col = 0
            elif old_from_col > export_max_idx:
                new_anchor._from.col = export_max_idx - export_min_idx
            else:
                new_anchor._from.col = old_from_col - export_min_idx

            if hasattr(new_anchor, "to") and hasattr(new_anchor.to, "row") and hasattr(new_anchor.to, "col"):
                old_to_row = anc.to.row + 1
                old_to_col = anc.to.col + 1

                if old_to_row not in row_map:
                    if old_to_row <= header_end_row:
                        target_to_row = old_to_row
                    else:
                        target_to_row = target_row
                else:
                    target_to_row = row_map[old_to_row]

                new_anchor.to.row = max(0, target_to_row - 1)

                if old_to_col < export_min_idx:
                    new_anchor.to.col = 0
                elif old_to_col > export_max_idx:
                    new_anchor.to.col = export_max_idx - export_min_idx
                else:
                    new_anchor.to.col = old_to_col - export_min_idx

            new_img.anchor = new_anchor
            out_ws.add_image(new_img)
        except Exception:
            pass

    amount_src_col = column_index_from_string("M")
    if export_min_idx <= amount_src_col <= export_max_idx and footer_end_row in row_map:
        amount_out_col = get_column_letter(amount_src_col - export_min_idx + 1)
        total_row = row_map[footer_end_row]
        detail_rows = [row_map[r] for r in kept_rows if header_end_row < r < footer_start_row]
        if detail_rows:
            first_detail = min(detail_rows)
            last_detail = max(detail_rows)
            out_ws[f"{amount_out_col}{total_row}"] = f"=SUM({amount_out_col}{first_detail}:{amount_out_col}{last_detail})"

    auto_size_columns_from_row(out_ws, start_row=auto_size_start_row)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    return {
        "output_path": output_path,
        "matched_rows": sorted(matched_rows),
        "kept_source_rows": kept_rows,
        "inserted_blank_rows": sum(1 for row in row_plan if row is None),
    }


def export_yuchang_po_compact_by_items(
    template_path: str,
    output_path: str,
    order_date: str,
    order_number: str,
    item_numbers_with_qty: Iterable[tuple[str, float | int]],
    *,
    sheet_name: str = "Sheet1",
    exact_item_numbers: Iterable[str] | None = None,
    **kwargs,
) -> dict:
    """Export by permanent item number, resolving current row positions at runtime."""
    current_rows, scan = resolve_yuchang_items_to_current_rows(
        template_path,
        item_numbers_with_qty,
        sheet_name=sheet_name,
        exact_item_numbers=exact_item_numbers,
    )
    result = export_yuchang_po_compact_by_rows(
        template_path=template_path,
        output_path=output_path,
        order_date=order_date,
        order_number=order_number,
        source_rows_with_qty=current_rows,
        sheet_name=sheet_name,
        footer_start_row=int(scan["footer_start_row"]),
        footer_end_row=int(scan["footer_end_row"]),
        **kwargs,
    )
    result["resolved_item_rows"] = [
        {"source_row": int(row), "quantity": float(qty)}
        for row, qty in current_rows
    ]
    result["footer_start_row"] = int(scan["footer_start_row"])
    result["footer_end_row"] = int(scan["footer_end_row"])
    return result


# ------------------------------------------------------------
# Window
# ------------------------------------------------------------
class YUOrderReviewWindow(QMainWindow):
    orders_exported = Signal(object)

    def __init__(
        self,
        db: Any,
        prefix: str,
        template_path: str,
        order_csv_path: str,
        output_dir: str,
    ):
        super().__init__()
        self.db = db
        self.prefix = prefix
        self.tables = {
            "supplier_lines": f"{prefix}_supplier_lines",
            "match_review": f"{prefix}_match_review",
            "match_candidates": f"{prefix}_match_candidates",
            "import_runs": f"{prefix}_import_runs",
        }
        self.template_path = str(template_path or "").strip()
        self.order_csv_path = str(order_csv_path or "").strip()
        self.output_dir = str(output_dir or "").strip()
        self.settings = QSettings("Windsor", "WidgetApp")

        self.raw_order_rows: list[dict] = []
        self.grouped_order_rows: list[dict] = []
        self.current_rows: list[dict] = []
        self.current_selected_detail: dict[str, Any] | None = None

        # Live Sheet1 index. Column A is the permanent item mapping; source rows
        # are only current workbook addresses used during validation.
        self._workbook_scan: dict[str, Any] = {}
        self._workbook_scan_error = ""

        self.setObjectName("YUOrderReviewWindow")
        self.setWindowTitle("YU Order Review")
        self.resize(1480, 920)
        self.setMinimumSize(QSize(1280, 760))

        self.build_ui()
        self.apply_theme_from_settings()
        self.check_required_tables()
        self.load_order_csv()
        self.refresh_all()

    # ---------------- UI
    def build_ui(self):
        central = QWidget(self)
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(8)

        self.top_frame = QFrame(central)
        self.top_frame.setObjectName("top_frame")
        self.top_frame.setMinimumHeight(58)
        top_layout = QHBoxLayout(self.top_frame)

        self.last_import_browser = QTextBrowser(self.top_frame)
        self.last_import_browser.setObjectName("last_import_browser")
        self.last_import_browser.setMaximumWidth(180)
        self.last_import_browser.setMaximumHeight(42)
        self.last_import_browser.setReadOnly(True)
        top_layout.addWidget(self.last_import_browser)

        self.title_label = QLabel("YU Order Review", self.top_frame)
        self.title_label.setObjectName("title_label")
        top_layout.addWidget(self.title_label)

        self.refresh_button = QPushButton("Refresh", self.top_frame)
        self.refresh_button.clicked.connect(self.refresh_all)
        top_layout.addWidget(self.refresh_button)

        self.open_workbook_button = QPushButton("Open Workbook", self.top_frame)
        self.open_workbook_button.clicked.connect(self.open_workbook)
        self.open_workbook_button.setEnabled(bool(self.template_path))
        top_layout.addWidget(self.open_workbook_button)

        self.export_visible_button = QPushButton("Export Visible", self.top_frame)
        self.export_visible_button.clicked.connect(self.export_visible_orders)
        top_layout.addWidget(self.export_visible_button)

        self.export_myob_po_button = QPushButton("Export MYOB PO", self.top_frame)
        self.export_myob_po_button.setToolTip(
            "Create an AccountRight Item Purchases TXT import using the visible, resolved order lines, "
            "the Widget item descriptions, and the latest imported YU costs."
        )
        self.export_myob_po_button.clicked.connect(self.export_myob_po)
        top_layout.addWidget(self.export_myob_po_button)

        self.open_output_button = QPushButton("Open Output Folder", self.top_frame)
        self.open_output_button.clicked.connect(self.open_output_dir)
        top_layout.addWidget(self.open_output_button)

        top_layout.addStretch(1)

        top_layout.addWidget(QLabel("Order", self.top_frame))
        self.order_filter_combo = QComboBox(self.top_frame)
        self.order_filter_combo.currentTextChanged.connect(self.load_main_table)
        top_layout.addWidget(self.order_filter_combo)

        top_layout.addWidget(QLabel("Filter", self.top_frame))
        self.filter_combo = QComboBox(self.top_frame)
        self.filter_combo.addItems(["All", "Resolved", "Needs Review", "Unmatched", "Error"])
        self.filter_combo.currentTextChanged.connect(self.load_main_table)
        top_layout.addWidget(self.filter_combo)

        self.search_edit = QLineEdit(self.top_frame)
        self.search_edit.setPlaceholderText("Search item or description...")
        self.search_edit.returnPressed.connect(self.load_main_table)
        self.search_edit.textChanged.connect(self.on_search_text_changed)
        self.search_edit.setMinimumWidth(240)
        top_layout.addWidget(self.search_edit)

        self.search_button = QPushButton("Show Data", self.top_frame)
        self.search_button.clicked.connect(self.load_main_table)
        top_layout.addWidget(self.search_button)

        root.addWidget(self.top_frame)

        stats_frame = QFrame(central)
        stats_frame.setObjectName("stats_frame")
        stats_layout = QHBoxLayout(stats_frame)
        stats_layout.setSpacing(8)

        self.total_box = self.make_stat_box("Total Lines")
        self.resolved_box = self.make_stat_box("Resolved")
        self.needs_review_box = self.make_stat_box("Needs Review")
        self.unmatched_box = self.make_stat_box("Unmatched / Error")

        stats_layout.addWidget(self.total_box["frame"])
        stats_layout.addWidget(self.resolved_box["frame"])
        stats_layout.addWidget(self.needs_review_box["frame"])
        stats_layout.addWidget(self.unmatched_box["frame"])
        stats_layout.addStretch(1)

        root.addWidget(stats_frame)

        splitter = QSplitter(Qt.Horizontal, central)

        left_panel = QWidget(splitter)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)

        self.main_table = QTableWidget(left_panel)
        self.main_table.setObjectName("main_table")
        self.main_table.setColumnCount(8)
        self.main_table.setHorizontalHeaderLabels([
            "Date", "Order No", "Item Number", "Qty", "Status", "Workbook Row", "Source", "Best Hit"
        ])
        self.main_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.main_table.setSelectionMode(QTableWidget.SingleSelection)
        self.main_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.main_table.verticalHeader().setVisible(False)
        self.main_table.itemSelectionChanged.connect(self.on_main_selection_changed)
        left_layout.addWidget(self.main_table)

        right_panel = QWidget(splitter)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(8)

        self.detail_frame = QFrame(right_panel)
        self.detail_frame.setObjectName("detail_frame")
        detail_layout = QVBoxLayout(self.detail_frame)
        detail_layout.setSpacing(6)

        self.date_box = self.make_value_row(detail_layout, "Date")
        self.order_no_box = self.make_value_row(detail_layout, "Order No")
        self.item_box = self.make_value_row(detail_layout, "Item Number")
        self.qty_box = self.make_value_row(detail_layout, "Qty")
        self.status_box = self.make_value_row(detail_layout, "Status")
        self.resolved_row_box = self.make_value_row(detail_layout, "Current Workbook Row")

        right_layout.addWidget(self.detail_frame)

        self.preview_frame = QFrame(right_panel)
        self.preview_frame.setObjectName("preview_frame")
        preview_layout = QVBoxLayout(self.preview_frame)
        preview_layout.addWidget(QLabel("Supplier Preview / Notes", self.preview_frame))
        self.preview_browser = QTextBrowser(self.preview_frame)
        self.preview_browser.setObjectName("preview_browser")
        preview_layout.addWidget(self.preview_browser)
        right_layout.addWidget(self.preview_frame, 1)

        self.candidate_frame = QFrame(right_panel)
        self.candidate_frame.setObjectName("candidate_frame")
        candidate_layout = QVBoxLayout(self.candidate_frame)
        candidate_layout.addWidget(QLabel("Candidate Supplier Rows", self.candidate_frame))

        self.candidate_table = QTableWidget(self.candidate_frame)
        self.candidate_table.setObjectName("candidate_table")
        self.candidate_table.setColumnCount(7)
        self.candidate_table.setHorizontalHeaderLabels([
            "Workbook Row", "Hit Type", "Conf %", "Item", "Description", "Size", "Labelled As"
        ])
        self.candidate_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.candidate_table.setSelectionMode(QTableWidget.SingleSelection)
        self.candidate_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.candidate_table.verticalHeader().setVisible(False)
        self.candidate_table.itemSelectionChanged.connect(self.on_candidate_selection_changed)
        self.candidate_table.itemDoubleClicked.connect(lambda *_args: self.confirm_selected_candidate())
        candidate_header = self.candidate_table.horizontalHeader()
        candidate_header.setStretchLastSection(False)
        candidate_header.setSectionsClickable(True)
        # Keep the candidate columns manually resizable.
        # resizeColumnToContents() is still called after data loads for a good starting width.
        for column_index in range(self.candidate_table.columnCount()):
            candidate_header.setSectionResizeMode(column_index, QHeaderView.Interactive)
        candidate_layout.addWidget(self.candidate_table)

        action_bar = QHBoxLayout()
        self.confirm_button = QPushButton("Confirm Candidate", self.candidate_frame)
        self.confirm_button.clicked.connect(self.confirm_selected_candidate)
        action_bar.addWidget(self.confirm_button)

        self.clear_button = QPushButton("Clear Confirm", self.candidate_frame)
        self.clear_button.clicked.connect(self.clear_confirmation)
        action_bar.addWidget(self.clear_button)

        self.manual_row_edit = QLineEdit(self.candidate_frame)
        self.manual_row_edit.setPlaceholderText("Type current workbook row...")
        self.manual_row_edit.returnPressed.connect(self.confirm_manual_row)
        action_bar.addWidget(self.manual_row_edit)

        self.manual_row_button = QPushButton("Use Workbook Row", self.candidate_frame)
        self.manual_row_button.clicked.connect(self.confirm_manual_row)
        action_bar.addWidget(self.manual_row_button)

        candidate_layout.addLayout(action_bar)
        right_layout.addWidget(self.candidate_frame, 1)

        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([920, 520])
        root.addWidget(splitter, 1)

        status_bar = QStatusBar(self)
        self.setStatusBar(status_bar)

        install_table_font_context_menu(self, self.main_table, self.settings, "yu_order_review", "yu_order_review/main_table")
        install_table_font_context_menu(self, self.candidate_table, self.settings, "yu_order_review", "yu_order_review/candidate_table")

    def make_stat_box(self, label_text: str) -> dict[str, QWidget]:
        frame = QFrame(self)
        frame.setObjectName("stat_box_frame")
        frame.setMinimumSize(QSize(160, 78))
        frame.setMaximumHeight(88)
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(4)
        label = QLabel(label_text, frame)
        label.setObjectName("stat_label")
        value = QLabel("", frame)
        value.setObjectName("stat_value")
        value.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        layout.addWidget(value)
        return {"frame": frame, "label": label, "value": value}

    def make_value_row(self, parent_layout: QVBoxLayout, label_text: str) -> QLabel:
        row = QHBoxLayout()
        label = QLabel(label_text, self.detail_frame)
        label.setObjectName("field_label")
        label.setMinimumWidth(120)
        value = QLabel("", self.detail_frame)
        value.setObjectName("field_value")
        value.setMinimumHeight(34)
        value.setAlignment(Qt.AlignCenter)
        row.addWidget(label)
        row.addWidget(value, 1)
        parent_layout.addLayout(row)
        return value

    def apply_theme_from_settings(self):
        theme_name = str(self.settings.value("theme", "dark"))
        app = QApplication.instance()
        if app is None:
            return
        if theme_name == "light":
            app.setStyleSheet("""
                QWidget { background: #e3e8ee; color: #1b1f23; }
                QMainWindow, QSplitter, QScrollArea, QScrollArea > QWidget > QWidget {
                    background: #e3e8ee; color: #1b1f23;
                }
                QFrame#top_frame, QFrame#stats_frame, QFrame#detail_frame,
                QFrame#preview_frame, QFrame#candidate_frame, QFrame#stat_box_frame {
                    background: #eef2f5;
                    border: 1px solid #c3cbd3;
                    border-radius: 8px;
                }
                QLabel#title_label {
                    font-size: 16px;
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_label, QLabel#stat_label {
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_value, QLabel#stat_value {
                    background: #f8fafb;
                    color: #111111;
                    border: 1px solid #aab4be;
                    border-radius: 4px;
                    padding: 6px;
                    font-weight: 700;
                }
                QLineEdit, QComboBox, QTableWidget, QTextBrowser {
                    background: #f8fafb;
                    color: #111111;
                    border: 1px solid #aab4be;
                    border-radius: 4px;
                    selection-background-color: #c9def5;
                    selection-color: #111111;
                }
                QHeaderView::section {
                    background: #d9e0e7;
                    color: #1b1f23;
                    border: 1px solid #bac4ce;
                    padding: 4px;
                    font-weight: 600;
                }
                QPushButton {
                    background: #d6dde5;
                    color: #1b1f23;
                    border: 1px solid #aab4be;
                    border-radius: 6px;
                    padding: 7px 12px;
                    min-height: 28px;
                    font-weight: 600;
                }
                QPushButton:hover {
                    background: #c9d3dc;
                }
                QPushButton:disabled {
                    background: #e3e8ee;
                    color: #808890;
                }
            """)
        elif theme_name == "high":
            app.setStyleSheet("""
                QWidget { background: #000000; color: #ffff00; }
                QMainWindow, QSplitter { background: #000000; color: #ffff00; }
                QFrame#top_frame, QFrame#stats_frame, QFrame#detail_frame,
                QFrame#preview_frame, QFrame#candidate_frame, QFrame#stat_box_frame {
                    background: #000000;
                    border: 2px solid #ffff00;
                    border-radius: 8px;
                }
                QLabel#title_label {
                    font-size: 16px;
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_label, QLabel#stat_label {
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_value, QLabel#stat_value {
                    background: #000000;
                    color: #ffff00;
                    border: 2px solid #ffff00;
                    border-radius: 4px;
                    padding: 6px;
                    font-weight: 700;
                }
                QLineEdit, QComboBox, QTableWidget, QTextBrowser {
                    background: #000000;
                    color: #ffff00;
                    border: 2px solid #ffff00;
                    border-radius: 4px;
                    selection-background-color: #ffff00;
                    selection-color: #000000;
                }
                QHeaderView::section {
                    background: #000000;
                    color: #ffff00;
                    border: 2px solid #ffff00;
                    padding: 4px;
                    font-weight: 700;
                }
                QPushButton {
                    background: #000000;
                    color: #ffff00;
                    border: 2px solid #ffff00;
                    border-radius: 6px;
                    padding: 7px 12px;
                    min-height: 28px;
                    font-weight: 700;
                }
                QPushButton:hover {
                    background: #333300;
                }
            """)
        else:
            app.setStyleSheet("""
                QWidget { background: #23272e; color: #e8eaed; }
                QMainWindow, QSplitter, QScrollArea, QScrollArea > QWidget > QWidget {
                    background: #23272e; color: #e8eaed;
                }
                QFrame#top_frame, QFrame#stats_frame, QFrame#detail_frame,
                QFrame#preview_frame, QFrame#candidate_frame, QFrame#stat_box_frame {
                    background: #2d333b;
                    border: 1px solid #48515b;
                    border-radius: 8px;
                }
                QLabel#title_label {
                    font-size: 16px;
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_label, QLabel#stat_label {
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_value, QLabel#stat_value {
                    background: #3b424d;
                    color: #f1f3f4;
                    border: 1px solid #59626d;
                    border-radius: 4px;
                    padding: 6px;
                    font-weight: 700;
                }
                QLineEdit, QComboBox, QTableWidget, QTextBrowser {
                    background: #3b424d;
                    color: #f1f3f4;
                    border: 1px solid #59626d;
                    border-radius: 4px;
                    selection-background-color: #5b7394;
                    selection-color: #ffffff;
                }
                QHeaderView::section {
                    background: #2f353d;
                    color: #e8eaed;
                    border: 1px solid #59626d;
                    padding: 4px;
                    font-weight: 600;
                }
                QPushButton {
                    background: #3c444f;
                    color: #e8eaed;
                    border: 1px solid #59626d;
                    border-radius: 6px;
                    padding: 7px 12px;
                    min-height: 28px;
                    font-weight: 600;
                }
                QPushButton:hover {
                    background: #4a5461;
                }
                QPushButton:disabled {
                    background: #323840;
                    color: #7b838c;
                }
            """)

    # ---------------- data load
    def check_required_tables(self):
        missing = []
        for table_name in self.tables.values():
            exists = self.db.scalar(
                "SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE' AND TABLE_NAME = ?",
                (table_name,),
            )
            if not exists:
                missing.append(table_name)
        if missing:
            raise RuntimeError(
                "Required test tables were not found.\n\n"
                "Run the importer first, for example:\n"
                "python yu_sqlserver_test_import_v3.py rebuild --workbook \"yuchang_order_form_matched_third_pass.xlsx\"\n\n"
                f"Missing tables: {', '.join(missing)}"
            )

    def load_order_csv(self):
        if not self.order_csv_path:
            raise RuntimeError("An order CSV path is required.")
        self.raw_order_rows = read_order_csv(self.order_csv_path)
        self.grouped_order_rows = group_order_rows(self.raw_order_rows)
        order_numbers = sorted({str(row["Order Number"]) for row in self.grouped_order_rows}, key=str.lower)

        self.order_filter_combo.blockSignals(True)
        self.order_filter_combo.clear()
        self.order_filter_combo.addItem("All")
        for order_no in order_numbers:
            self.order_filter_combo.addItem(order_no)
        self.order_filter_combo.blockSignals(False)


    # ---------------- Item-number resolving
    def item_is_fasteners(self, item_number: str) -> bool:
        """FASTENERS must be exact-match only; they are not space-normalised."""
        item_number = str(item_number or "").strip()
        if not item_number:
            return False
        try:
            row = self.db.one(
                """
                SELECT TOP 1 item_number, [Custom List 1] AS custom_list_1
                FROM dbo.items
                WHERE LTRIM(RTRIM(item_number)) = LTRIM(RTRIM(?))
                """,
                (item_number,),
            )
        except Exception:
            row = None
        if row is None:
            return False
        group = str(row.get("custom_list_1") or "").strip().upper()
        return "FASTENERS" in group

    def item_match_condition(self, column_exprs: list[str], item_number: str) -> tuple[str, tuple]:
        """Build an item-number comparison for supplier/review tables.

        FASTENERS keep exact item numbers. All other groups are matched using
        the no-whitespace canonical key so ABC 123 and ABC123 resolve together.
        """
        item_number = str(item_number or "").strip()
        if self.item_is_fasteners(item_number):
            parts = [f"ISNULL({expr}, '') = ?" for expr in column_exprs]
            return "(" + " OR ".join(parts) + ")", tuple(item_number for _ in column_exprs)
        clean = clean_item_key(item_number)
        parts = [f"{clean_item_sql_expr(expr)} = ?" for expr in column_exprs]
        return "(" + " OR ".join(parts) + ")", tuple(clean for _ in column_exprs)

    def refresh_workbook_index(self, force: bool = False) -> bool:
        """Refresh the live Sheet1 mapping used for resolution and export."""
        if not self.template_path:
            self._workbook_scan = {}
            self._workbook_scan_error = "No YU workbook template path is set."
            return False

        try:
            workbook_path = Path(self.template_path)
            current_mtime = workbook_path.stat().st_mtime_ns
            if (
                not force
                and self._workbook_scan
                and int(self._workbook_scan.get("mtime_ns") or -1) == int(current_mtime)
            ):
                return True

            self._workbook_scan = scan_yuchang_workbook(self.template_path)
            self._workbook_scan_error = ""
            return True
        except Exception as exc:
            self._workbook_scan = {}
            self._workbook_scan_error = f"{type(exc).__name__}: {exc}"
            return False

    def workbook_rows_for_item(self, item_number: str, *, force_refresh: bool = False) -> list[int]:
        if force_refresh or not self._workbook_scan:
            self.refresh_workbook_index(force=force_refresh)
        if not self._workbook_scan:
            return []

        exact = self.item_is_fasteners(item_number)
        return workbook_rows_for_item_from_scan(
            self._workbook_scan,
            item_number,
            exact=exact,
        )

    @staticmethod
    def supplier_record_signatures(record: dict[str, Any]) -> tuple[tuple[str, ...], tuple[str, ...]]:
        exact_signature = _workbook_row_signature_from_values(
            [
                record.get("description"),
                record.get("size_text"),
                record.get("colour"),
                record.get("pack_type"),
                record.get("unit_size"),
                record.get("labelled_as"),
            ]
        )
        loose_signature = _workbook_row_signature_from_values(
            [
                record.get("description"),
                record.get("size_text"),
                record.get("colour"),
                record.get("pack_type"),
                record.get("labelled_as"),
            ]
        )
        return exact_signature, loose_signature

    def locate_current_workbook_rows_for_supplier_record(
        self,
        record: dict[str, Any],
        preferred_row: int | None = None,
    ) -> list[int]:
        """Relocate a database snapshot row after workbook insertions or moves."""
        if not self._workbook_scan:
            self.refresh_workbook_index(force=False)
        scan = self._workbook_scan
        if not scan:
            return []

        exact_signature, loose_signature = self.supplier_record_signatures(record)
        exact_rows = list((scan.get("signature_rows") or {}).get(exact_signature, []))
        loose_rows = list((scan.get("loose_signature_rows") or {}).get(loose_signature, []))

        candidates = exact_rows or loose_rows
        candidates = sorted(set(int(row) for row in candidates))
        if not candidates:
            return []

        if preferred_row is not None:
            preferred_row = int(preferred_row)
            if preferred_row in candidates:
                return [preferred_row]
            # A row insertion normally moves a line only a short distance. If
            # otherwise-identical supplier lines exist, choose only when one
            # candidate is uniquely closest to the old snapshot row.
            distances = sorted((abs(row - preferred_row), row) for row in candidates)
            if len(distances) == 1 or distances[0][0] < distances[1][0]:
                return [distances[0][1]]

        return candidates

    def workbook_row_is_detail(self, row_number: int) -> bool:
        if not self._workbook_scan:
            self.refresh_workbook_index(force=False)
        return int(row_number) in set(self._workbook_scan.get("detail_rows") or set())

    def refresh_all(self):
        self.refresh_workbook_index(force=True)
        self.load_last_import_date()
        self.load_counts()
        self.load_main_table()

    def load_last_import_date(self):
        row = self.db.one(
            f"SELECT TOP 1 imported_at, workbook_path FROM dbo.{self.tables['import_runs']} ORDER BY imported_at DESC"
        )
        if row is None:
            self.last_import_browser.setPlainText("")
            return
        text = str(row.get("imported_at") or "")
        self.last_import_browser.setPlainText(text)

    def resolve_item(self, item_number: str, quantity: float, date_text: str, order_number: str) -> OrderResolveResult:
        item_number = str(item_number or "").strip()

        if not self._workbook_scan and not self.refresh_workbook_index(force=False):
            return OrderResolveResult(
                item_number=item_number,
                quantity=quantity,
                date=date_text,
                order_number=order_number,
                status="error",
                source_row=None,
                source="workbook_unavailable",
                note=f"The current YU workbook could not be scanned: {self._workbook_scan_error}",
                review_hits=[],
            )

        # Permanent resolution comes only from the current workbook's Column A.
        # The database source_row is a historical validation snapshot and is not
        # trusted as an export destination.
        workbook_rows = self.workbook_rows_for_item(item_number)
        if len(workbook_rows) == 1:
            current_row = int(workbook_rows[0])
            return OrderResolveResult(
                item_number=item_number,
                quantity=quantity,
                date=date_text,
                order_number=order_number,
                status="resolved",
                source_row=current_row,
                source="workbook_column_a",
                note=f"Resolved from current Sheet1 Column A at row {current_row}.",
                review_hits=[],
            )
        if len(workbook_rows) > 1:
            return OrderResolveResult(
                item_number=item_number,
                quantity=quantity,
                date=date_text,
                order_number=order_number,
                status="error",
                source_row=None,
                source="workbook_column_a_duplicate",
                note=f"Item number appears more than once in current Sheet1 Column A: {workbook_rows}.",
                review_hits=[],
            )

        hits: list[ReviewHit] = []

        direct_condition, direct_params = self.item_match_condition(
            ["current_item_number", "literal_item_number"], item_number
        )
        direct_rows_raw = self.db.all(
            f"""
            SELECT DISTINCT source_row
            FROM dbo.{self.tables["supplier_lines"]}
            WHERE row_kind = 'detail'
              AND {direct_condition}
            ORDER BY source_row
            """,
            direct_params,
        )
        for row in direct_rows_raw:
            hits.append(
                ReviewHit(
                    source_row=int(row["source_row"]),
                    match_type="supplier_snapshot",
                    confidence=None,
                    note="Item exists in the imported supplier snapshot but is missing from current workbook Column A.",
                )
            )

        final_condition, final_params = self.item_match_condition(["final_selection"], item_number)
        final_rows_raw = self.db.all(
            f"""
            SELECT DISTINCT source_row
            FROM dbo.{self.tables["match_review"]}
            WHERE {final_condition}
            ORDER BY source_row
            """,
            final_params,
        )
        for row in final_rows_raw:
            hits.append(
                ReviewHit(
                    source_row=int(row["source_row"]),
                    match_type="database_confirmation",
                    confidence=None,
                    note="Database confirmation exists, but Column A is the permanent mapping and currently has no unique match.",
                )
            )

        suggested_condition, suggested_params = self.item_match_condition(["suggested_match"], item_number)
        suggested_hits = self.db.all(
            f"""
            SELECT source_row, confidence_pct, review_row
            FROM dbo.{self.tables["match_review"]}
            WHERE {suggested_condition}
            ORDER BY source_row
            """,
            suggested_params,
        )
        for row in suggested_hits:
            conf = row.get("confidence_pct")
            hits.append(
                ReviewHit(
                    source_row=int(row["source_row"]),
                    match_type="suggested_match",
                    confidence=float(conf) if conf is not None else None,
                    note=f"suggested in Match_Review row {row.get('review_row')}",
                )
            )

        candidate_condition, candidate_params = self.item_match_condition(["candidate_item_number"], item_number)
        candidate_hits = self.db.all(
            f"""
            SELECT source_row, confidence_pct, review_row, candidate_rank
            FROM dbo.{self.tables["match_candidates"]}
            WHERE {candidate_condition}
            ORDER BY source_row, candidate_rank
            """,
            candidate_params,
        )
        for row in candidate_hits:
            conf = row.get("confidence_pct")
            hits.append(
                ReviewHit(
                    source_row=int(row["source_row"]),
                    match_type=f"candidate_{row.get('candidate_rank')}",
                    confidence=float(conf) if conf is not None else None,
                    note=f"candidate in Match_Review row {row.get('review_row')}",
                )
            )

        if hits:
            hits_sorted = sorted(
                hits,
                key=lambda x: (
                    x.confidence if x.confidence is not None else -1,
                    -int(x.source_row),
                ),
                reverse=True,
            )
            best = hits_sorted[0]
            confidence_txt = "" if best.confidence is None else f" Best confidence: {best.confidence:.1%}."
            note = (
                f"No unique current Column A mapping exists. Best validation hit is snapshot row "
                f"{best.source_row} via {best.match_type}.{confidence_txt} "
                "Confirm the current workbook row to write the item number into Column A."
            )
            return OrderResolveResult(
                item_number=item_number,
                quantity=quantity,
                date=date_text,
                order_number=order_number,
                status="needs_review",
                source_row=None,
                source="validation_required",
                note=note,
                review_hits=hits_sorted,
            )

        return OrderResolveResult(
            item_number=item_number,
            quantity=quantity,
            date=date_text,
            order_number=order_number,
            status="unmatched",
            source_row=None,
            source="not_found",
            note="Item number is not in current workbook Column A and no validation candidates were found.",
            review_hits=[],
        )

    def load_counts(self):
        resolved = 0
        needs_review = 0
        unmatched_error = 0

        for row in self.grouped_order_rows:
            result = self.resolve_item(
                item_number=row["Item Number"],
                quantity=float(row["QTY"]),
                date_text=row["Date"],
                order_number=row["Order Number"],
            )
            if result.status == "resolved":
                resolved += 1
            elif result.status == "needs_review":
                needs_review += 1
            else:
                unmatched_error += 1

        self.total_box["value"].setText(str(len(self.grouped_order_rows)))
        self.resolved_box["value"].setText(str(resolved))
        self.needs_review_box["value"].setText(str(needs_review))
        self.unmatched_box["value"].setText(str(unmatched_error))

    def passes_filters(self, row: dict, result: OrderResolveResult) -> bool:
        order_filter = (self.order_filter_combo.currentText() or "").strip()
        if order_filter and order_filter.lower() != "all":
            if str(row["Order Number"]) != order_filter:
                return False

        status_filter = (self.filter_combo.currentText() or "").strip().lower()
        if status_filter == "resolved" and result.status != "resolved":
            return False
        if status_filter == "needs review" and result.status != "needs_review":
            return False
        if status_filter == "unmatched" and result.status != "unmatched":
            return False
        if status_filter == "error" and result.status != "error":
            return False

        search = (self.search_edit.text() or "").strip().lower()
        if search:
            haystack = " ".join([
                str(row["Date"]),
                str(row["Order Number"]),
                str(row["Item Number"]),
                str(result.note),
                str(result.source),
            ]).lower()
            if search not in haystack:
                return False

        return True

    def load_main_table(self):
        self.current_rows = []
        table = self.main_table
        table.setRowCount(0)

        for row in self.grouped_order_rows:
            result = self.resolve_item(
                item_number=row["Item Number"],
                quantity=float(row["QTY"]),
                date_text=row["Date"],
                order_number=row["Order Number"],
            )
            if not self.passes_filters(row, result):
                continue

            detail = dict(row)
            detail["resolve_result"] = result
            self.current_rows.append(detail)

            row_index = table.rowCount()
            table.insertRow(row_index)
            values = [
                row["Date"],
                row["Order Number"],
                row["Item Number"],
                self.format_qty(row["QTY"]),
                result.status,
                "" if result.source_row is None else str(result.source_row),
                result.source,
                self.best_hit_text(result),
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem("" if value is None else str(value))
                if col in {3, 5}:
                    item.setTextAlignment(Qt.AlignCenter)
                if col == 2:
                    item.setData(Qt.UserRole, row["Item Number"])
                table.setItem(row_index, col, item)

            self.apply_main_row_styles(row_index, result)

        self.auto_size_all_columns(table)
        table.resizeRowsToContents()

        if table.rowCount() > 0:
            table.selectRow(0)
            self.on_main_selection_changed()
        else:
            self.current_selected_detail = None
            self.clear_detail_panel()

        self.statusBar().showMessage(f"Loaded {table.rowCount()} order lines.", 4000)

    def apply_main_row_styles(self, row_index: int, result: OrderResolveResult):
        if result.status == "resolved":
            bg = QColor(0, 255, 0)
            fg = QColor(Qt.black)
        elif result.status == "needs_review":
            bg = QColor(255, 199, 206)
            fg = QColor(156, 0, 6)
        elif result.status == "unmatched":
            bg = QColor(214, 214, 214) if row_index % 2 == 0 else QColor(190, 190, 190)
            fg = QColor(Qt.black)
        else:
            bg = QColor(255, 235, 156)
            fg = QColor(156, 101, 0)

        for col in range(self.main_table.columnCount()):
            item = self.main_table.item(row_index, col)
            if item is not None:
                item.setBackground(QBrush(bg))
                item.setForeground(QBrush(fg))

    def best_hit_text(self, result: OrderResolveResult) -> str:
        if result.status == "resolved":
            return ""
        if result.review_hits:
            hit = result.review_hits[0]
            conf = "" if hit.confidence is None else f" {hit.confidence:.1%}"
            return f"snapshot row {hit.source_row} {hit.match_type}{conf}"
        return ""

    def selected_row_detail(self) -> dict[str, Any] | None:
        row_index = self.main_table.currentRow()
        if row_index < 0 or row_index >= len(self.current_rows):
            return None
        return self.current_rows[row_index]

    def supplier_candidate_values(self, row: dict) -> dict:
        return {
            "sheet_col_b": str(row.get("sheet_col_b") or ""),
            "sheet_col_c": str(row.get("sheet_col_c") or ""),
            "sheet_col_d": str(row.get("sheet_col_d") or ""),
            "sheet_col_g": str(row.get("sheet_col_g") or ""),
            "description": row.get("description"),
            "size_text": row.get("size_text"),
            "colour": row.get("colour"),
            "pack_type": row.get("pack_type"),
            "unit_size": row.get("unit_size"),
            "labelled_as": row.get("labelled_as"),
        }

    def candidate_supplier_select_sql(self, prefix: str = "") -> str:
        p = prefix
        return (
            f"COALESCE(NULLIF({p}current_item_number, ''), {p}literal_item_number, '') AS sheet_col_b, "
            f"{p}description AS sheet_col_c, "
            f"{p}size_text AS sheet_col_d, "
            f"{p}labelled_as AS sheet_col_g, "
            f"{p}description AS description, "
            f"{p}size_text AS size_text, "
            f"{p}colour AS colour, "
            f"{p}pack_type AS pack_type, "
            f"{p}unit_size AS unit_size, "
            f"{p}labelled_as AS labelled_as"
        )

    def decorate_candidate_row(
        self,
        row: dict,
        *,
        hit_type: str,
        confidence: float | None,
    ) -> dict:
        db_source_row = int(row["source_row"])
        values = self.supplier_candidate_values(row)
        current_rows = self.locate_current_workbook_rows_for_supplier_record(
            values,
            preferred_row=db_source_row,
        )
        workbook_row = int(current_rows[0]) if len(current_rows) == 1 else None

        if workbook_row is None:
            row_state = "ambiguous" if current_rows else "not found in current workbook"
        elif workbook_row == db_source_row:
            row_state = "current"
        else:
            row_state = f"moved from snapshot row {db_source_row}"

        return {
            "source_row": workbook_row,
            "workbook_row": workbook_row,
            "db_source_row": db_source_row,
            "row_state": row_state,
            "candidate_workbook_rows": current_rows,
            "hit_type": hit_type,
            "confidence": confidence,
            **values,
        }

    def load_candidates_for_item(self, item_number: str, current_result: OrderResolveResult) -> list[dict]:
        rows: list[dict] = []
        supplier_cols = self.candidate_supplier_select_sql("s.")

        direct_condition, direct_params = self.item_match_condition(
            ["s.current_item_number", "s.literal_item_number"], item_number
        )
        direct_rows = self.db.all(
            f"""
            SELECT s.source_row, {supplier_cols}
            FROM dbo.{self.tables["supplier_lines"]} s
            WHERE s.row_kind = 'detail'
              AND {direct_condition}
            ORDER BY s.source_row
            """,
            direct_params,
        )
        for row in direct_rows:
            rows.append(self.decorate_candidate_row(row, hit_type="direct", confidence=None))

        final_condition, final_params = self.item_match_condition(["r.final_selection"], item_number)
        final_rows = self.db.all(
            f"""
            SELECT r.source_row, {supplier_cols}
            FROM dbo.{self.tables["match_review"]} r
            LEFT JOIN dbo.{self.tables["supplier_lines"]} s ON s.source_row = r.source_row
            WHERE {final_condition}
            ORDER BY r.source_row
            """,
            final_params,
        )
        for row in final_rows:
            rows.append(self.decorate_candidate_row(row, hit_type="database_confirmation", confidence=None))

        suggested_condition, suggested_params = self.item_match_condition(["r.suggested_match"], item_number)
        suggested_rows = self.db.all(
            f"""
            SELECT r.source_row, r.confidence_pct, {supplier_cols}
            FROM dbo.{self.tables["match_review"]} r
            LEFT JOIN dbo.{self.tables["supplier_lines"]} s ON s.source_row = r.source_row
            WHERE {suggested_condition}
            ORDER BY r.source_row
            """,
            suggested_params,
        )
        for row in suggested_rows:
            rows.append(
                self.decorate_candidate_row(
                    row,
                    hit_type="suggested_match",
                    confidence=float(row["confidence_pct"]) if row.get("confidence_pct") is not None else None,
                )
            )

        candidate_condition, candidate_params = self.item_match_condition(["c.candidate_item_number"], item_number)
        candidate_rows = self.db.all(
            f"""
            SELECT c.source_row, c.confidence_pct, c.candidate_rank, {supplier_cols}
            FROM dbo.{self.tables["match_candidates"]} c
            LEFT JOIN dbo.{self.tables["supplier_lines"]} s ON s.source_row = c.source_row
            WHERE {candidate_condition}
            ORDER BY c.source_row, c.candidate_rank
            """,
            candidate_params,
        )
        for row in candidate_rows:
            rows.append(
                self.decorate_candidate_row(
                    row,
                    hit_type=f"candidate_{row.get('candidate_rank')}",
                    confidence=float(row["confidence_pct"]) if row.get("confidence_pct") is not None else None,
                )
            )

        # When Column A already contains the item, always show its current row,
        # even if no matching database snapshot row is available.
        if current_result.source_row is not None:
            current_row = int(current_result.source_row)
            if not any(r.get("workbook_row") == current_row for r in rows):
                rows.insert(
                    0,
                    {
                        "source_row": current_row,
                        "workbook_row": current_row,
                        "db_source_row": None,
                        "row_state": "current Column A mapping",
                        "candidate_workbook_rows": [current_row],
                        "hit_type": "current_workbook",
                        "confidence": None,
                        "sheet_col_b": item_number,
                        "sheet_col_c": "",
                        "sheet_col_d": "",
                        "sheet_col_g": "",
                        "description": None,
                        "size_text": None,
                        "colour": None,
                        "pack_type": None,
                        "unit_size": None,
                        "labelled_as": None,
                    },
                )

        deduped: list[dict] = []
        seen: set[tuple[Any, str]] = set()
        for row in rows:
            key = (row.get("db_source_row"), str(row.get("hit_type")))
            if key in seen:
                continue
            seen.add(key)
            deduped.append(row)
        return deduped

    def on_main_selection_changed(self):
        detail = self.selected_row_detail()
        self.current_selected_detail = detail
        if detail is None:
            self.clear_detail_panel()
            return

        result: OrderResolveResult = detail["resolve_result"]

        self.date_box.setText(str(detail["Date"]))
        self.order_no_box.setText(str(detail["Order Number"]))
        self.item_box.setText(str(detail["Item Number"]))
        self.qty_box.setText(self.format_qty(detail["QTY"]))
        self.status_box.setText(result.status)
        self.resolved_row_box.setText("" if result.source_row is None else str(result.source_row))

        preview_lines = [
            f"<b>Resolution Source:</b> {self.html_text(result.source)}",
            f"<b>Note:</b> {self.html_text(result.note)}",
        ]
        self.preview_browser.setHtml("<br>".join(preview_lines))

        candidates = self.load_candidates_for_item(str(detail["Item Number"]), result)
        self.candidate_table.setRowCount(0)
        for cand in candidates:
            row_index = self.candidate_table.rowCount()
            self.candidate_table.insertRow(row_index)

            workbook_row = cand.get("workbook_row")
            row_text = str(workbook_row) if workbook_row is not None else "?"
            hit_type = str(cand.get("hit_type") or "")
            row_state = str(cand.get("row_state") or "")
            if row_state and row_state not in {"current", "current Column A mapping"}:
                hit_type = f"{hit_type} ({row_state})"

            values = [
                row_text,
                hit_type,
                self.format_confidence(cand["confidence"]),
                cand.get("sheet_col_b", ""),
                cand.get("sheet_col_c", ""),
                cand.get("sheet_col_d", ""),
                cand.get("sheet_col_g", ""),
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem("" if value is None else str(value))
                if col in {0, 2}:
                    item.setTextAlignment(Qt.AlignCenter)
                if col == 0:
                    item.setData(Qt.UserRole, dict(cand))
                    if workbook_row is None:
                        item.setToolTip(
                            "This database snapshot row could not be located uniquely in the current workbook. "
                            "Use the manual current workbook row field."
                        )
                self.candidate_table.setItem(row_index, col, item)

        self.auto_size_all_columns(self.candidate_table)
        self.candidate_table.resizeRowsToContents()
        if self.candidate_table.rowCount() > 0:
            self.candidate_table.selectRow(0)
            self.on_candidate_selection_changed()

    def on_candidate_selection_changed(self):
        row_index = self.candidate_table.currentRow()
        if row_index < 0:
            return
        item = self.candidate_table.item(row_index, 0)
        if item is None:
            return

        candidate = item.data(Qt.UserRole)
        if not isinstance(candidate, dict):
            return

        workbook_row = candidate.get("workbook_row")
        db_source_row = candidate.get("db_source_row")
        row_state = str(candidate.get("row_state") or "")
        possible_rows = candidate.get("candidate_workbook_rows") or []

        if db_source_row is None:
            lines = [
                f"<b>Current Workbook Row:</b> {self.html_text(workbook_row)}",
                f"<b>Row State:</b> {self.html_text(row_state)}",
                "<b>Mapping Source:</b> Current Sheet1 Column A",
            ]
            self.preview_browser.setHtml("<br>".join(lines))
            return

        preview = self.db.one(
            f"""
            SELECT
                s.source_row,
                s.current_item_number,
                s.literal_item_number,
                s.description,
                s.size_text,
                s.colour,
                s.pack_type,
                s.unit_size,
                s.labelled_as,
                r.final_selection,
                r.suggested_match,
                r.review_reasons
            FROM dbo.{self.tables["supplier_lines"]} s
            LEFT JOIN dbo.{self.tables["match_review"]} r ON r.source_row = s.source_row
            WHERE s.source_row = ?
            """,
            (int(db_source_row),),
        )
        if preview is None:
            return

        current_row_text = (
            str(workbook_row)
            if workbook_row is not None
            else ("Ambiguous: " + ", ".join(str(row) for row in possible_rows) if possible_rows else "Not found")
        )
        lines = [
            f"<b>Current Workbook Row:</b> {self.html_text(current_row_text)}",
            f"<b>Database Snapshot Row:</b> {self.html_text(preview.get('source_row'))}",
            f"<b>Row State:</b> {self.html_text(row_state)}",
            f"<b>Current Item:</b> {self.html_text(preview.get('current_item_number') or preview.get('literal_item_number'))}",
            f"<b>Description:</b> {self.html_text(preview.get('description'))}",
            f"<b>Size:</b> {self.html_text(preview.get('size_text'))}",
            f"<b>Colour:</b> {self.html_text(preview.get('colour'))}",
            f"<b>Pack:</b> {self.html_text(preview.get('pack_type'))}",
            f"<b>Unit Size:</b> {self.html_text(preview.get('unit_size'))}",
            f"<b>Labelled As:</b> {self.html_text(preview.get('labelled_as'))}",
            f"<b>Final Selection:</b> {self.html_text(preview.get('final_selection'))}",
            f"<b>Suggested Match:</b> {self.html_text(preview.get('suggested_match'))}",
            "",
            f"<b>Review Reasons:</b><br>{self.html_text(preview.get('review_reasons'))}",
        ]
        self.preview_browser.setHtml("<br>".join(lines))

    def clear_detail_panel(self):
        for box in (self.date_box, self.order_no_box, self.item_box, self.qty_box, self.status_box, self.resolved_row_box):
            box.setText("")
        self.preview_browser.clear()
        self.candidate_table.setRowCount(0)
        self.manual_row_edit.clear()

    # ---------------- actions
    def selected_candidate_mapping(self) -> dict[str, Any] | None:
        row_index = self.candidate_table.currentRow()
        if row_index < 0:
            return None
        item = self.candidate_table.item(row_index, 0)
        if item is None:
            return None
        value = item.data(Qt.UserRole)
        return dict(value) if isinstance(value, dict) else None

    def next_review_row(self) -> int:
        value = self.db.scalar(f"SELECT ISNULL(MAX(review_row), 4) + 1 FROM dbo.{self.tables['match_review']}")
        return int(value or 5)

    def ensure_match_review_row(self, source_row: int) -> None:
        exists = self.db.scalar(
            f"SELECT COUNT(*) FROM dbo.{self.tables['match_review']} WHERE source_row = ?",
            (source_row,),
        )
        if exists:
            return
        src = self.db.one(
            f"""
            SELECT source_row, current_item_number, size_text, colour, pack_type, labelled_as
            FROM dbo.{self.tables['supplier_lines']}
            WHERE source_row = ?
            """,
            (source_row,),
        ) or {}
        self.db.execute(
            f"""
            INSERT INTO dbo.{self.tables["match_review"]} (
                source_row,
                review_row,
                final_selection,
                suggested_match,
                confidence_pct,
                gap_score,
                review_reasons,
                duplicate_state,
                source_item,
                size_text,
                colour,
                pack_type,
                unit_size,
                labelled_as,
                review_status
            )
            VALUES (?, ?, NULL, NULL, NULL, NULL, NULL, NULL, ?, ?, ?, ?, NULL, ?, 'unmatched')
            """,
            (
                source_row,
                self.next_review_row(),
                src.get("current_item_number"),
                src.get("size_text"),
                src.get("colour"),
                src.get("pack_type"),
                src.get("labelled_as"),
            ),
        )

    def reset_row_status_for_item(self, item_number: str, exclude_source_row: int | None = None):
        final_condition, final_params = self.item_match_condition(["r.final_selection"], item_number)
        rows = self.db.all(
            f"""
            SELECT r.source_row,
                   CASE
                     WHEN ISNULL(r.suggested_match, '') <> ''
                          OR EXISTS (
                              SELECT 1 FROM dbo.{self.tables["match_candidates"]} c
                              WHERE c.source_row = r.source_row
                          )
                     THEN 'needs_review'
                     ELSE 'unmatched'
                   END AS new_status
            FROM dbo.{self.tables["match_review"]} r
            WHERE {final_condition}
            """,
            final_params,
        )
        for row in rows:
            source_row = int(row["source_row"])
            if exclude_source_row is not None and source_row == int(exclude_source_row):
                continue
            self.db.execute(
                f"""
                UPDATE dbo.{self.tables["match_review"]}
                SET final_selection = NULL, review_status = ?
                WHERE source_row = ?
                """,
                (str(row["new_status"]), source_row),
            )

    def source_rows_for_final_selection(self, item_number: str) -> list[int]:
        final_condition, final_params = self.item_match_condition(["final_selection"], item_number)
        rows = self.db.all(
            f"""
            SELECT source_row
            FROM dbo.{self.tables["match_review"]}
            WHERE {final_condition}
            ORDER BY source_row
            """,
            final_params,
        )
        out: list[int] = []
        for row in rows:
            try:
                out.append(int(row["source_row"]))
            except Exception:
                pass
        return out

    def save_workbook_matches_or_warn(self, row_matches: dict[int, str | None], action_title: str = "Save YU Match") -> bool:
        if not row_matches:
            return True
        try:
            result = save_yuchang_workbook_matches(self.template_path, row_matches)
        except PermissionError as exc:
            QMessageBox.warning(
                self,
                action_title,
                "The match could not be saved to the YU workbook.\n\n"
                "Close the workbook in Excel and try again.\n\n"
                f"{exc}",
            )
            return False
        except Exception as exc:
            QMessageBox.warning(
                self,
                action_title,
                "The match could not be saved to the YU workbook, so the local review table was not changed.\n\n"
                f"{exc}",
            )
            return False
        updated_rows = result.get("updated_rows") or []
        if updated_rows:
            self.statusBar().showMessage(
                f"Saved YU workbook match row(s): {', '.join(str(r) for r in updated_rows)}.",
                5000,
            )
        return True

    def sync_confirmed_matches_to_workbook(self) -> bool:
        """Legacy compatibility hook.

        Confirmed mappings are no longer replayed from database source rows.
        Column A is authoritative and is scanned in its current position.
        """
        if self.refresh_workbook_index(force=True):
            return True
        QMessageBox.warning(
            self,
            "YU Order Review",
            "The current YU workbook could not be scanned.\n\n"
            f"{self._workbook_scan_error}",
        )
        return False

    def db_source_row_for_current_workbook_row(self, workbook_row: int) -> int | None:
        """Best-effort link from a current workbook row to the old DB snapshot row."""
        records = self.db.all(
            f"""
            SELECT
                source_row,
                description,
                size_text,
                colour,
                pack_type,
                unit_size,
                labelled_as
            FROM dbo.{self.tables["supplier_lines"]}
            WHERE row_kind = 'detail'
            ORDER BY source_row
            """
        )
        matches: list[int] = []
        for record in records:
            db_source_row = int(record["source_row"])
            current_rows = self.locate_current_workbook_rows_for_supplier_record(
                record,
                preferred_row=db_source_row,
            )
            if len(current_rows) == 1 and int(current_rows[0]) == int(workbook_row):
                matches.append(db_source_row)
        return matches[0] if len(matches) == 1 else None

    def confirm_item_to_source_row(
        self,
        item_number: str,
        source_row: int,
        *,
        db_source_row: int | None = None,
    ):
        """Write the part number to the current workbook row.

        source_row is intentionally a current Sheet1 address, not a permanent key.
        """
        item_number = str(item_number or "").strip()
        source_row = int(source_row)

        if not self.refresh_workbook_index(force=True):
            QMessageBox.warning(
                self,
                "Confirm YU Match",
                "The current YU workbook could not be scanned.\n\n"
                f"{self._workbook_scan_error}",
            )
            return

        if not self.workbook_row_is_detail(source_row):
            QMessageBox.warning(
                self,
                "Confirm YU Match",
                f"Current workbook row {source_row} is not recognised as a YU item detail row.",
            )
            return

        existing_at_target = str(
            (self._workbook_scan.get("row_item_numbers") or {}).get(source_row) or ""
        ).strip()
        if existing_at_target:
            same_item = (
                existing_at_target.upper() == item_number.upper()
                if self.item_is_fasteners(item_number)
                else clean_item_key(existing_at_target) == clean_item_key(item_number)
            )
            if not same_item:
                QMessageBox.warning(
                    self,
                    "Confirm YU Match",
                    f"Current workbook row {source_row} already contains item "
                    f"{existing_at_target} in Column A.\n\n"
                    "Clear or correct that mapping before assigning another item.",
                )
                return

        # Enforce one Column A location per item. Any previous occurrence is
        # cleared from its current row, regardless of how the template moved.
        row_matches: dict[int, str | None] = {source_row: item_number}
        for existing_row in self.workbook_rows_for_item(item_number):
            if int(existing_row) != source_row:
                row_matches[int(existing_row)] = None

        if not self.save_workbook_matches_or_warn(row_matches, "Confirm YU Match"):
            return

        self.refresh_workbook_index(force=True)

        if db_source_row is None:
            db_source_row = self.db_source_row_for_current_workbook_row(source_row)

        if db_source_row is not None:
            db_source_row = int(db_source_row)
            self.ensure_match_review_row(db_source_row)
            self.reset_row_status_for_item(item_number, exclude_source_row=db_source_row)
            self.db.execute(
                f"""
                UPDATE dbo.{self.tables["match_review"]}
                SET final_selection = ?, review_status = 'confirmed'
                WHERE source_row = ?
                """,
                (item_number, db_source_row),
            )
            self.db.execute(
                f"""
                UPDATE dbo.{self.tables["supplier_lines"]}
                SET current_item_number = ?
                WHERE source_row = ?
                """,
                (item_number, db_source_row),
            )

        snapshot_text = (
            f" Database snapshot row {db_source_row} was updated."
            if db_source_row is not None
            else " No unique database snapshot row was required."
        )
        self.statusBar().showMessage(
            f"Confirmed {item_number} to current workbook row {source_row}.{snapshot_text}",
            6000,
        )
        self.refresh_all()
        self.reselect_item(item_number)

    def confirm_selected_candidate(self):
        detail = self.current_selected_detail
        if detail is None:
            QMessageBox.warning(self, "YU Order Review", "Select an order line first.")
            return

        candidate = self.selected_candidate_mapping()
        if candidate is None:
            QMessageBox.warning(self, "YU Order Review", "Select a candidate supplier row first.")
            return

        workbook_row = candidate.get("workbook_row")
        if workbook_row is None:
            possible_rows = candidate.get("candidate_workbook_rows") or []
            possible_text = (
                "\n\nPossible current workbook rows: "
                + ", ".join(str(row) for row in possible_rows)
                if possible_rows
                else ""
            )
            QMessageBox.warning(
                self,
                "YU Order Review",
                "The selected database candidate could not be located uniquely in the current workbook."
                f"{possible_text}\n\nUse the manual current workbook row field.",
            )
            return

        self.confirm_item_to_source_row(
            str(detail["Item Number"]),
            int(workbook_row),
            db_source_row=candidate.get("db_source_row"),
        )

    def confirm_manual_row(self):
        detail = self.current_selected_detail
        if detail is None:
            QMessageBox.warning(self, "YU Order Review", "Select an order line first.")
            return
        text = (self.manual_row_edit.text() or "").strip()
        if not text:
            QMessageBox.warning(self, "YU Order Review", "Type a current workbook row first.")
            self.manual_row_edit.setFocus()
            return
        try:
            workbook_row = int(text)
        except Exception:
            QMessageBox.warning(self, "YU Order Review", "Workbook row must be a whole number.")
            return

        if not self.refresh_workbook_index(force=True):
            QMessageBox.warning(
                self,
                "YU Order Review",
                "The current YU workbook could not be scanned.\n\n"
                f"{self._workbook_scan_error}",
            )
            return
        if not self.workbook_row_is_detail(workbook_row):
            QMessageBox.warning(
                self,
                "YU Order Review",
                f"Current workbook row {workbook_row} is not an item detail row.",
            )
            return

        self.confirm_item_to_source_row(str(detail["Item Number"]), workbook_row)

    def clear_confirmation(self):
        detail = self.current_selected_detail
        if detail is None:
            QMessageBox.warning(self, "YU Order Review", "Select an order line first.")
            return
        item_number = str(detail["Item Number"])

        if not self.refresh_workbook_index(force=True):
            QMessageBox.warning(
                self,
                "Clear YU Match",
                "The current YU workbook could not be scanned.\n\n"
                f"{self._workbook_scan_error}",
            )
            return

        workbook_rows = self.workbook_rows_for_item(item_number)
        final_condition, final_params = self.item_match_condition(["r.final_selection"], item_number)
        rows = self.db.all(
            f"""
            SELECT r.source_row,
                   CASE
                     WHEN ISNULL(r.suggested_match, '') <> ''
                          OR EXISTS (
                              SELECT 1 FROM dbo.{self.tables["match_candidates"]} c
                              WHERE c.source_row = r.source_row
                          )
                     THEN 'needs_review'
                     ELSE 'unmatched'
                   END AS new_status
            FROM dbo.{self.tables["match_review"]} r
            WHERE {final_condition}
            """,
            final_params,
        )

        if not workbook_rows and not rows:
            QMessageBox.information(self, "YU Order Review", f"There is no confirmed mapping for {item_number}.")
            return

        if workbook_rows:
            row_matches = {int(row): None for row in workbook_rows}
            if not self.save_workbook_matches_or_warn(row_matches, "Clear YU Match"):
                return

        for row in rows:
            self.db.execute(
                f"""
                UPDATE dbo.{self.tables["match_review"]}
                SET final_selection = NULL, review_status = ?
                WHERE source_row = ?
                """,
                (str(row["new_status"]), int(row["source_row"])),
            )

        self.refresh_workbook_index(force=True)
        self.statusBar().showMessage(
            f"Cleared the current Column A mapping for {item_number}.",
            5000,
        )
        self.refresh_all()
        self.reselect_item(item_number)

    def reselect_item(self, item_number: str):
        for row_index in range(self.main_table.rowCount()):
            item = self.main_table.item(row_index, 2)
            if item and (item.text() or "") == item_number:
                self.main_table.selectRow(row_index)
                self.main_table.scrollToItem(item)
                self.on_main_selection_changed()
                break

    def _load_myob_item_master(self) -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
        try:
            rows = self.db.all(
                """
                SELECT
                    TRIM(COALESCE(item_number, '')) AS item_number,
                    COALESCE(
                        NULLIF(TRIM(COALESCE(item_name, '')), ''),
                        NULLIF(TRIM(COALESCE(description, '')), ''),
                        ''
                    ) AS myob_description,
                    yu_last_cost,
                    yu_last_cost_date
                FROM items
                WHERE TRIM(COALESCE(item_number, '')) <> ''
                """
            )
        except Exception as exc:
            raise RuntimeError(
                "Could not read item descriptions and YU costs from the items table. "
                "Run the YU cost database update first and confirm the items table contains "
                "yu_last_cost and yu_last_cost_date."
            ) from exc

        by_key: dict[str, dict[str, Any]] = {}
        collisions: dict[str, list[str]] = defaultdict(list)
        for raw_row in rows:
            row = dict(raw_row)
            item_number = str(row.get("item_number") or "").strip()
            key = clean_item_key(item_number)
            if not key:
                continue
            collisions[key].append(item_number)
            if key not in by_key:
                by_key[key] = row

        collisions = {
            key: sorted(set(values))
            for key, values in collisions.items()
            if len(set(values)) > 1
        }
        return by_key, collisions

    def export_myob_po(self):
        try:
            self._export_myob_po_impl()
        except Exception as exc:
            self.show_myob_export_error(exc)

    def _export_myob_po_impl(self):
        order_filter = str(self.order_filter_combo.currentText() or "").strip()
        selected_order_rows = [
            row
            for row in self.grouped_order_rows
            if not order_filter
            or order_filter.lower() == "all"
            or str(row.get("Order Number") or "").strip() == order_filter
        ]
        if not selected_order_rows:
            QMessageBox.warning(self, "Export MYOB PO", "There are no order rows available to export.")
            return

        # Deliberately ignore the Status and Search display filters here. Accounting
        # exports must contain the complete selected order, not an accidentally
        # filtered subset of the table.
        visible_rows: list[dict[str, Any]] = []
        for source_row in selected_order_rows:
            result = self.resolve_item(
                item_number=source_row["Item Number"],
                quantity=float(source_row["QTY"]),
                date_text=source_row["Date"],
                order_number=source_row["Order Number"],
            )
            detail = dict(source_row)
            detail["resolve_result"] = result
            visible_rows.append(detail)

        unresolved = [row for row in visible_rows if row["resolve_result"].status != "resolved"]
        if unresolved:
            item_list = ", ".join(sorted({str(row.get("Item Number") or "") for row in unresolved}))
            QMessageBox.warning(
                self,
                "Export MYOB PO",
                "MYOB export stopped because some visible rows are unresolved.\n\n"
                f"Items: {item_list}",
            )
            return

        item_master, collisions = self._load_myob_item_master()
        problems: list[str] = []
        grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
        total_value = Decimal("0")
        cost_dates: list[str] = []

        for row in visible_rows:
            requested_item = str(row.get("Item Number") or "").strip()
            key = clean_item_key(requested_item)
            if key in collisions:
                problems.append(
                    f"{requested_item}: item-key collision in the Widget database ({', '.join(collisions[key])})"
                )
                continue

            master = item_master.get(key)
            if not master:
                problems.append(f"{requested_item}: not found in the Widget items table")
                continue

            actual_item = str(master.get("item_number") or requested_item).strip()
            description = _single_line_description(master.get("myob_description"))
            if not description:
                problems.append(f"{actual_item}: description is blank in the Widget items table")

            try:
                price = _decimal_value(master.get("yu_last_cost"), f"YU cost for {actual_item}")
                if price <= 0:
                    raise ValueError("cost is zero or negative")
            except Exception:
                problems.append(f"{actual_item}: latest YU cost is missing or zero")
                continue

            try:
                quantity = _decimal_value(row.get("QTY"), f"quantity for {actual_item}")
                if quantity <= 0:
                    raise ValueError("quantity is zero or negative")
            except Exception:
                problems.append(f"{actual_item}: quantity is invalid or zero")
                continue

            cost_date = str(master.get("yu_last_cost_date") or "").strip()
            if cost_date:
                cost_dates.append(cost_date)

            order_date = _normalise_myob_date(row.get("Date"))
            order_number = str(row.get("Order Number") or "").strip()
            if not order_number:
                problems.append(f"{actual_item}: order number is blank")
                continue

            grouped[(order_date, order_number)].append({
                "item_number": actual_item,
                "description": description,
                "quantity": quantity,
                "price": price,
                "cost_date": cost_date,
            })
            total_value += quantity * price

        if problems:
            QMessageBox.warning(
                self,
                "Export MYOB PO",
                "MYOB export stopped. Every line must have an item description and a positive YU cost.\n\n"
                + "\n".join(problems[:30])
                + ("\n\nMore problems were omitted." if len(problems) > 30 else ""),
            )
            return

        if not grouped:
            QMessageBox.warning(self, "Export MYOB PO", "There are no valid MYOB order lines to export.")
            return

        cost_date_text = "not recorded"
        parsed_cost_dates = []
        for value in cost_dates:
            text = str(value).split(" ")[0]
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
                try:
                    parsed_cost_dates.append(datetime.strptime(text, fmt).date())
                    break
                except ValueError:
                    continue
        if parsed_cost_dates:
            oldest = min(parsed_cost_dates).strftime("%d/%m/%Y")
            newest = max(parsed_cost_dates).strftime("%d/%m/%Y")
            cost_date_text = oldest if oldest == newest else f"{oldest} to {newest}"

        answer = QMessageBox.question(
            self,
            "Export MYOB PO",
            "Create AccountRight purchase-order import file(s)?\n\n"
            f"Orders: {len(grouped)}\n"
            f"Lines: {sum(len(lines) for lines in grouped.values())}\n"
            f"Order value from latest YU costs: ${total_value:,.2f}\n"
            f"Cost dates: {cost_date_text}\n\n"
            "The import creates Orders, not Bills. Review the MYOB total before saving the order.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if answer != QMessageBox.Yes:
            return

        output_dir = QFileDialog.getExistingDirectory(
            self,
            "Choose MYOB PO export folder",
            self.output_dir or str(Path.cwd() / DEFAULT_OUTPUT_DIR),
        )
        if not output_dir:
            return

        planned_paths = [
            Path(output_dir) / f"myob_po_{order_number}.txt"
            for (_date_text, order_number) in sorted(grouped, key=lambda value: (value[1], value[0]))
        ]
        existing = [path for path in planned_paths if path.exists()]
        if existing:
            overwrite = QMessageBox.question(
                self,
                "Overwrite MYOB PO file",
                "The following file(s) already exist:\n\n"
                + "\n".join(str(path) for path in existing)
                + "\n\nOverwrite them?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if overwrite != QMessageBox.Yes:
                return

        exports: list[str] = []
        for (date_text, order_number), lines in sorted(grouped.items(), key=lambda value: (value[0][1], value[0][0])):
            output_path = Path(output_dir) / f"myob_po_{order_number}.txt"
            exports.append(
                write_myob_po_import_txt(
                    output_path=output_path,
                    order_date=date_text,
                    order_number=order_number,
                    lines=lines,
                )
            )

        self.output_dir = output_dir
        self.statusBar().showMessage(f"Exported {len(exports)} MYOB PO import file(s).", 6000)
        QMessageBox.information(
            self,
            "Export MYOB PO",
            "MYOB purchase-order import export complete.\n\n"
            + "\n".join(exports)
            + "\n\nImport in AccountRight using:\n"
            "File > Import/Export Assistant > Import data > Purchases > Item Purchases\n"
            "Format: Commas; First record: Headers or Labels",
        )

    def show_myob_export_error(self, exc: Exception):
        error_path = None
        try:
            target_dir = Path(self.output_dir or tempfile.gettempdir())
            target_dir.mkdir(parents=True, exist_ok=True)
            error_path = target_dir / "myob_po_export_error.log"
            error_path.write_text(traceback.format_exc(), encoding="utf-8")
        except Exception:
            error_path = None

        message = (
            "The MYOB PO export failed.\n\n"
            f"{type(exc).__name__}: {exc}\n\n"
            "No MYOB import file was completed. Fix the issue and try again."
        )
        if error_path is not None:
            message += f"\n\nTechnical error log:\n{error_path}"
        QMessageBox.critical(self, "Export MYOB PO", message)
        try:
            self.statusBar().showMessage("MYOB PO export failed.", 8000)
        except Exception:
            pass

    def export_visible_orders(self):
        """Export visible YU order rows with user-facing error reporting.

        Qt can otherwise swallow exceptions raised inside a clicked slot, leaving
        the button appearing to do nothing. Keep this wrapper thin and put the
        actual export work in _export_visible_orders_impl().
        """
        try:
            self._export_visible_orders_impl()
        except Exception as exc:
            self.show_export_error(exc)

    def _export_visible_orders_impl(self):
        if not self.template_path:
            QMessageBox.warning(self, "YU Order Review", "No workbook template path is set.")
            return

        if not Path(str(self.template_path)).exists():
            QMessageBox.warning(
                self,
                "YU Order Review",
                "The YU workbook template could not be found.\n\n"
                f"{self.template_path}",
            )
            return

        if not self.current_rows:
            QMessageBox.warning(self, "YU Order Review", "There are no visible order rows to export.")
            return

        if not self.refresh_workbook_index(force=True):
            QMessageBox.warning(
                self,
                "YU Order Review",
                "The current YU workbook could not be scanned.\n\n"
                f"{self._workbook_scan_error}",
            )
            return

        # Re-resolve immediately before export. This deliberately ignores stored
        # database source rows and follows the item number's current Column A row.
        visible_rows: list[dict] = []
        for source in self.current_rows:
            row = dict(source)
            row["resolve_result"] = self.resolve_item(
                item_number=str(row.get("Item Number") or ""),
                quantity=float(row.get("QTY") or 0),
                date_text=str(row.get("Date") or ""),
                order_number=str(row.get("Order Number") or ""),
            )
            visible_rows.append(row)

        output_dir = QFileDialog.getExistingDirectory(
            self,
            "Choose export output folder",
            self.output_dir or str(Path.cwd() / DEFAULT_OUTPUT_DIR),
        )
        if not output_dir:
            return

        unresolved = [row for row in visible_rows if row["resolve_result"].status != "resolved"]
        audit_path = str(Path(output_dir) / "yu_order_audit.csv")
        write_audit_csv(audit_path, [row["resolve_result"] for row in visible_rows])

        if unresolved:
            item_list = ", ".join(sorted({row["Item Number"] for row in unresolved}))
            QMessageBox.warning(
                self,
                "YU Order Review",
                "Export stopped because some visible items do not have one unique current "
                "Column A mapping in the YU workbook.\n\n"
                f"Items: {item_list}\n\nAudit written to:\n{audit_path}",
            )
            return

        invalid_rows: list[str] = []
        for row in visible_rows:
            item_number = str(row.get("Item Number") or "").strip()
            try:
                qty = float(row.get("QTY") or 0)
                if qty <= 0:
                    raise ValueError("quantity must be greater than zero")
            except Exception:
                invalid_rows.append(f"{item_number}: invalid quantity {row.get('QTY')!r}")

        if invalid_rows:
            QMessageBox.warning(
                self,
                "YU Order Review",
                "Export stopped because one or more rows contain invalid quantities.\n\n"
                + "\n".join(invalid_rows[:20])
                + ("\n\nMore rows were omitted from this message." if len(invalid_rows) > 20 else "")
                + f"\n\nAudit written to:\n{audit_path}",
            )
            return

        grouped: dict[tuple[str, str], list[tuple[str, float]]] = defaultdict(list)
        grouped_items: dict[tuple[str, str], list[str]] = defaultdict(list)
        for row in visible_rows:
            order_key = (str(row["Date"]), str(row["Order Number"]))
            item_number = str(row["Item Number"]).strip()
            qty = float(row["QTY"])
            grouped[order_key].append((item_number, qty))
            current_row = row["resolve_result"].source_row
            grouped_items[order_key].append(
                f"{item_number} current row {current_row} qty {self.format_qty(qty)}"
            )

        exports = []
        for (date_text, order_number), item_lines in sorted(grouped.items(), key=lambda x: (x[0][1], x[0][0])):
            filename = f"yuchang_order_{order_number}.xlsx"
            output_path = str(Path(output_dir) / filename)
            try:
                exact_item_numbers = {
                    item_number
                    for item_number, _qty in item_lines
                    if self.item_is_fasteners(item_number)
                }
                result = export_yuchang_po_compact_by_items(
                    template_path=self.template_path,
                    output_path=output_path,
                    order_date=date_text,
                    order_number=order_number,
                    item_numbers_with_qty=item_lines,
                    exact_item_numbers=exact_item_numbers,
                )
            except PermissionError as exc:
                raise PermissionError(
                    "Could not write the exported YU workbook.\n\n"
                    "Close the output workbook if it is open in Excel, then try again.\n\n"
                    f"Order: {order_number}\n"
                    f"Output file: {output_path}"
                ) from exc
            except Exception as exc:
                item_details = "\n".join(grouped_items.get((date_text, order_number), [])[:30])
                raise RuntimeError(
                    "Export failed while resolving current workbook item positions.\n\n"
                    f"Order: {order_number}\n"
                    f"Date: {date_text}\n"
                    f"Output file: {output_path}\n\n"
                    "Items in this export:\n"
                    f"{item_details}\n\n"
                    f"Reason: {exc}"
                ) from exc
            exports.append(output_path)

            # Keep the audit display aligned with the rows actually used by the
            # final runtime scan.
            current_row_map = {
                str(item).strip(): int(row)
                for (item, _qty), row_info in zip(
                    item_lines,
                    result.get("resolved_item_rows") or [],
                )
                for row in [row_info.get("source_row")]
                if row is not None
            }
            for row in visible_rows:
                if (
                    str(row.get("Date")) == date_text
                    and str(row.get("Order Number")) == order_number
                    and str(row.get("Item Number")).strip() in current_row_map
                ):
                    row["resolve_result"].source_row = current_row_map[str(row.get("Item Number")).strip()]

        # Re-write the audit after the final runtime scan so it records the
        # exact workbook rows used for the completed exports.
        write_audit_csv(audit_path, [row["resolve_result"] for row in visible_rows])

        self.output_dir = output_dir
        exported_payload = [
            {
                "order_number": str(row.get("Order Number") or "").strip(),
                "item_number": str(row.get("Item Number") or "").strip(),
                "qty": float(row.get("QTY") or 0),
            }
            for row in visible_rows
        ]
        try:
            self.orders_exported.emit(exported_payload)
        except Exception:
            pass
        self.statusBar().showMessage(f"Exported {len(exports)} workbook(s) by current Column A item mapping.", 6000)
        QMessageBox.information(
            self,
            "YU Order Review",
            "Export complete.\n\n"
            + "\n".join(exports)
            + f"\n\nAudit:\n{audit_path}\n\n"
            "Item locations were resolved from the current workbook Column A. "
            "Inserted or moved rows do not require database row-number changes.",
        )

    def show_export_error(self, exc: Exception):
        error_path = None
        try:
            target_dir = Path(self.output_dir or tempfile.gettempdir())
            target_dir.mkdir(parents=True, exist_ok=True)
            error_path = target_dir / "yu_order_export_error.log"
            error_path.write_text(traceback.format_exc(), encoding="utf-8")
        except Exception:
            error_path = None

        message = (
            "The YU export failed.\n\n"
            f"{type(exc).__name__}: {exc}\n\n"
            "The order was not exported. Fix the issue and try again."
        )
        if error_path is not None:
            message += f"\n\nA technical error log was written to:\n{error_path}"

        try:
            self.statusBar().showMessage("YU export failed.", 8000)
        except Exception:
            pass

        QMessageBox.critical(self, "YU Export Failed", message)

    def open_workbook(self):
        if not self.template_path:
            QMessageBox.information(self, "YU Order Review", "No workbook path is set.")
            return
        try:
            resolved = str(Path(self.template_path).resolve())
            if sys.platform.startswith("win"):
                os.startfile(resolved)
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(resolved))
        except Exception as exc:
            QMessageBox.warning(self, "YU Order Review", f"Could not open workbook:\n{exc}")

    def open_output_dir(self):
        target = Path(self.output_dir or (Path.cwd() / DEFAULT_OUTPUT_DIR))
        target.mkdir(parents=True, exist_ok=True)
        try:
            resolved = str(target.resolve())
            if sys.platform.startswith("win"):
                os.startfile(resolved)
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(resolved))
        except Exception as exc:
            QMessageBox.warning(self, "YU Order Review", f"Could not open output folder:\n{exc}")

    def on_search_text_changed(self, text: str):
        if not text.strip():
            self.load_main_table()

    # helpers
    def format_confidence(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        try:
            return f"{float(value):.1%}"
        except Exception:
            return str(value)

    def format_qty(self, value: Any) -> str:
        try:
            number = float(value)
        except Exception:
            return str(value)
        if number.is_integer():
            return f"{int(number)}"
        return f"{number:,.2f}"

    def html_text(self, value: Any) -> str:
        text = "" if value is None else str(value)
        return (
            text.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace("\n", "<br>")
        )

    def auto_size_all_columns(self, table: QTableWidget):
        for column_index in range(table.columnCount()):
            try:
                table.resizeColumnToContents(column_index)
            except Exception:
                pass


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Standalone YU order review/export test window.")
    parser.add_argument("--prefix", default=DEFAULT_PREFIX, help=f"Table prefix. Default: {DEFAULT_PREFIX}")
    parser.add_argument("--template", default=DEFAULT_TEMPLATE, help="Path to the matched workbook/template")
    parser.add_argument("--order-csv", default=DEFAULT_ORDER_CSV, help="Path to the order CSV")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help=f"Default output folder. Default: {DEFAULT_OUTPUT_DIR}")
    parser.add_argument("--base-dir", default=None, help="Optional base directory to search for client_config.json")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    app = QApplication(sys.argv)
    db = None
    try:
        db = SQLHelper(base_dir=Path(args.base_dir).resolve() if args.base_dir else None)
        window = YUOrderReviewWindow(
            db=db,
            prefix=args.prefix,
            template_path=args.template,
            order_csv_path=args.order_csv,
            output_dir=args.output_dir,
        )
        window.show()
        return app.exec()
    except Exception as exc:
        QMessageBox.critical(None, "YU Order Review", str(exc))
        return 1
    finally:
        if db is not None:
            try:
                db.close()
            except Exception:
                pass


if __name__ == "__main__":
    raise SystemExit(main())
