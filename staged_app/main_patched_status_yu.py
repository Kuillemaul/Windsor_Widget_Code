import sys
import re
import math
try:
    import pyodbc
except Exception:
    pyodbc = None
import os
import tempfile
import json
import subprocess
from pathlib import Path
from datetime import datetime, date, timedelta
from collections import Counter

from PySide6.QtCore import Qt, QDate, QSettings, QMargins, QUrl, QEvent, QSignalBlocker
from PySide6.QtGui import (
    QColor,
    QFont,
    QPen,
    QStandardItem,
    QStandardItemModel,
    QPixmap,
    QPainter,
    QDesktopServices,
    QTextCursor,
    QDoubleValidator,
    QBrush,
)
from PySide6.QtWidgets import (QFrame, QTextEdit, 
    QApplication,
    QMainWindow,
    QMessageBox,
    QCompleter,
    QTableWidgetItem,
    QVBoxLayout,
    QLabel,
    QAbstractItemView,
    QAbstractSpinBox,
    QHeaderView,
    QDialog,
    QPushButton,
    QHBoxLayout,
    QTableWidget,
    QTabWidget,
    QTextEdit,
    QWidget,
    QSizePolicy,
    QLineEdit,
    QDateEdit,
    QFileDialog,
    QInputDialog,
    QProgressDialog,
    QCheckBox,
    QTableView,
    QFormLayout,
    QComboBox,
    QDialogButtonBox,
    QStyledItemDelegate,
    QMenu,
)
from PySide6.QtCharts import (
    QChart,
    QChartView,
    QLineSeries,
    QBarCategoryAxis,
    QValueAxis,
)

from ui_mainwindow import Ui_MainWindow
from shipments_window_ui import Ui_MainWindow as Ui_ShipmentsWindow
from yu_order_workflow import YUOrderEntryDialog, load_yu_review_module

TABLE_FONT_SIZE_OPTIONS = (8, 9, 10, 11, 12, 14, 16, 18, 20)
TABLE_FONT_SETTINGS_PREFIX = "table_font_sizes"
APP_VERSION = "1.0c"
APP_DESIGNER = "Bradley Mayze"


def _safe_table_settings_token(table, fallback_scope="table"):
    token = str(table.property("_table_font_settings_token") or "").strip()
    if token:
        return token

    parts = []
    current = table
    while current is not None:
        name = ""
        try:
            name = str(current.objectName() or "").strip()
        except Exception:
            name = ""
        if name and not name.startswith("qt_"):
            parts.append(name)
        try:
            current = current.parentWidget()
        except Exception:
            current = None
    parts.reverse()

    token = "/".join(parts) if parts else fallback_scope
    token = re.sub(r"[^A-Za-z0-9_./-]+", "_", token)
    return token or fallback_scope


def table_font_settings_key(scope_key, table):
    token = _safe_table_settings_token(table, scope_key)
    return f"{TABLE_FONT_SETTINGS_PREFIX}/{scope_key}/{token}"


def stored_table_font_size(settings, scope_key, table):
    if settings is None or table is None:
        return None
    try:
        value = settings.value(table_font_settings_key(scope_key, table), None)
    except Exception:
        return None
    if value in (None, ""):
        return None
    try:
        size = int(value)
    except Exception:
        return None
    return size if size > 0 else None


def current_table_font_size(table):
    if table is None:
        return 10
    try:
        size = int(round(float(table.font().pointSizeF())))
        if size > 0:
            return size
    except Exception:
        pass
    try:
        size = int(table.font().pointSize())
        if size > 0:
            return size
    except Exception:
        pass
    return 10


def sync_table_item_font_sizes(table, size):
    if table is None:
        return
    if not isinstance(table, QTableWidget):
        return
    blocker = QSignalBlocker(table)
    try:
        for row in range(table.rowCount()):
            for column in range(table.columnCount()):
                item = table.item(row, column)
                if item is None:
                    continue
                item_font = item.font()
                item_font.setPointSize(size)
                item.setFont(item_font)
    finally:
        del blocker


def apply_table_font_size(table, size, settings=None, scope_key=None, persist=True):
    if table is None:
        return
    try:
        size = int(size)
    except Exception:
        return
    if size < 6:
        size = 6

    font = table.font()
    font.setPointSize(size)
    table.setFont(font)
    sync_table_item_font_sizes(table, size)

    vertical_header = getattr(table, "verticalHeader", lambda: None)()
    if vertical_header is not None:
        try:
            vertical_header.setDefaultSectionSize(max(24, size * 2 + 8))
        except Exception:
            pass
    try:
        table.resizeRowsToContents()
    except Exception:
        pass
    try:
        table.viewport().update()
    except Exception:
        pass

    if persist and settings is not None and scope_key:
        try:
            settings.setValue(table_font_settings_key(scope_key, table), size)
        except Exception:
            pass


def reset_table_font_size(table, settings=None, scope_key=None):
    if table is None:
        return
    default_size = None
    try:
        default_size = int(table.property("_table_font_default_size") or 0)
    except Exception:
        default_size = None
    if not default_size:
        default_size = 10

    if settings is not None and scope_key:
        try:
            settings.remove(table_font_settings_key(scope_key, table))
        except Exception:
            pass
    apply_table_font_size(table, default_size, settings=settings, scope_key=scope_key, persist=False)


def install_table_font_context_menu(owner, table, settings, scope_key, settings_token=None):
    if table is None:
        return
    if table.property("_table_font_menu_installed"):
        saved_size = stored_table_font_size(settings, scope_key, table)
        if saved_size:
            apply_table_font_size(table, saved_size, settings=settings, scope_key=scope_key, persist=False)
        return

    if settings_token:
        table.setProperty("_table_font_settings_token", settings_token)
    elif not str(table.objectName() or "").strip():
        table.setObjectName(_safe_table_settings_token(table, scope_key))

    table.setProperty("_table_font_menu_installed", True)
    table.setProperty("_table_font_scope_key", scope_key)
    table.setProperty("_table_font_default_size", current_table_font_size(table))
    table.setContextMenuPolicy(Qt.CustomContextMenu)
    table.customContextMenuRequested.connect(
        lambda pos, tbl=table, own=owner, s=settings, scope=scope_key: show_table_font_context_menu(own, tbl, pos, s, scope)
    )

    saved_size = stored_table_font_size(settings, scope_key, table)
    if saved_size:
        apply_table_font_size(table, saved_size, settings=settings, scope_key=scope_key, persist=False)


def show_table_font_context_menu(owner, table, pos, settings, scope_key):
    if table is None:
        return

    menu = QMenu(table)
    font_menu = menu.addMenu("Font Size")
    current_size = current_table_font_size(table)

    for size in TABLE_FONT_SIZE_OPTIONS:
        action = font_menu.addAction(str(size))
        action.setCheckable(True)
        action.setChecked(size == current_size)
        action.triggered.connect(
            lambda _checked=False, tbl=table, selected_size=size, s=settings, scope=scope_key: apply_table_font_size(
                tbl, selected_size, settings=s, scope_key=scope, persist=True
            )
        )

    font_menu.addSeparator()
    reset_action = font_menu.addAction("Reset")
    reset_action.triggered.connect(
        lambda _checked=False, tbl=table, s=settings, scope=scope_key: reset_table_font_size(tbl, settings=s, scope_key=scope)
    )

    viewport = getattr(table, "viewport", lambda: None)()
    if viewport is not None:
        global_pos = viewport.mapToGlobal(pos)
    else:
        global_pos = table.mapToGlobal(pos)
    menu.exec(global_pos)


EMBEDDED_SUPPLIER_MASTER = [
    "Aerosolve Pty Ltd",
    "AICA ADTEK",
    "Airquip Pipetool - SA",
    "Alloy Wire - UK",
    "Ann Chain",
    "Aplix Hong Kong Ltd",
    "Australian Surface Treatments",
    "AZI ( Australian Zipper PTY LTD )",
    "Baoding Yongwei Changsheng Metal - SXJ",
    "Baolinda Thread Co",
    "BCSM Enterprise - Supplier",
    "Bunnings Warehouse",
    "C & N Building Supplies",
    "Cansew Inc",
    "Changrui Interlining Manufacture (Shenzhen) Co Ltd",
    "Charles Parsons & Co. Pty Ltd",
    "CIP - Xuzhou - Coil Nails",
    "CJ Humphries P/L",
    "CNISOO - Dongguan Spraying Acc",
    "Danfield Ltd.",
    "Datum Limited ( HK )",
    "Dave's Industrial Sewing Machines and Spares",
    "Direct Freight",
    "Dongguan Caihua Thread Products Co Ltd",
    "Durak Tekstil",
    "Elizabeth Sewing Machines",
    "Embroidery Source P/L",
    "Emtex Inc",
    "Fedelor Fasteners",
    "Festo Pty Ltd",
    "Fil-Tec",
    "Foshan Nanhai Nonwoven",
    "Giolite Lumian",
    "GNG Sales",
    "Gran Brothers Co",
    "Groz-Beckert",
    "Gumitex",
    "Hangzhou Yueda Hardware",
    "Happy Embroidery Machines",
    "HJ Corp",
    "Holdfast Components P/L",
    "Hoz Fasteners",
    "Industrial Air Tools",
    "Irrigation Warehouse Group P/L",
    "Jack Stock",
    "Jackson WM Pty Ltd",
    "Jiangsu Aidefu Latex",
    "Jiaxing Media Technology",
    "Jinshuai - Ningbo Good Shine Import and Export Co",
    "Kalamir J. P/L",
    "KGL - Taiwan",
    "Kingswell Industries",
    "KL Hai",
    "KS Textiles",
    "KYA Fasteners",
    "Leggett & Platt",
    "Liberty Fabrics",
    "Linkron Australia Pty Ltd",
    "Linyi Shuangfeng",
    "LM FASTENERS",
    "M Recht Accessories",
    "Marfar P/L",
    "Master Q P/L",
    "Melbourne Office Supplies",
    "Merello Ingenieros",
    "MPE Plastic Extrusions",
    "Nanjing Caiqing Hardware Co., Ltd (Yue Chang)",
    "Nanmu Yarns",
    "Nantong Qingluo Group",
    "NINGBO SCREW FASTENERS",
    "Ningbo Xinkai ( RGN TOOLS )",
    "Norjet Distributors",
    "Phillro Industries",
    "Plain Pallets",
    "Poly Global",
    "Rayson Non Wovens",
    "Rise Time Industrial",
    "Ross Hanna Aust",
    "Ryson Fasteners",
    "Saba Pacific",
    "Saba USA North America LLC",
    "San Esu",
    "Schappe",
    "Sew Fix",
    "Shandong Huifeng Supply Chain Mgt Co Ltd",
    "Shandong Laifen",
    "Shann Australia P/L",
    "Sheng Hung",
    "Shenzhen Taiqiang Investment Holding Co Ltd",
    "SMC Australia",
    "Somac UK",
    "South West Timber",
    "STOCK",
    "Sundry Creditor",
    "Testo Industry",
    "Thanh Xuan",
    "Threads ( India ) Ltd",
    "Tianjin Coways Meter",
    "Tianjin Hweschun Fasteners",
    "TLNT - Taiwan",
    "Toolkwip Pumps",
    "Tower Fastening Systems",
    "United Fasteners Aust P/L",
    "Venus",
    "Vietnam - TLNT",
    "VIM International - Taiwan",
    "Wenzhou LKK Zipper",
    "Windsor Textiles",
    "WIP",
    "Yama Ribbons",
    "Yuchang Textile Factory",
    "Zhejiang Kenking ( Xinchang )",
    "ZZZZZZZhejiang Hongming Weaving"
]

SHIPMENT_TYPE_OPTIONS = ["Melbourne", "Sydney", "SABA"]

SHIPMENT_FIELD_MAP = [
    ("Date", "entry_date"),
    ("Shipment Type", "shipment_type"),
    ("O/No", "order_no"),
    ("Supplier", "supplier_name"),
    ("Container No", "container_no"),
    ("Product", "product"),
    ("Qty", "qty"),
    ("Ready Date", "ready_date"),
    ("Shipment Date", "shipment_date"),
    ("Due Date", "due_date"),
    ("Status", "status"),
    ("Vessel", "vessel"),
    ("Notes", "notes"),
]

SHIPMENT_REMOVE_HEADER = "Remove"
SHIPMENT_HEADERS = [header for header, _field in SHIPMENT_FIELD_MAP] + [SHIPMENT_REMOVE_HEADER]
SHIPMENT_QTY_OPTIONS = ["", "LCL", "20' FCL", "40' FCL"]
SHIPMENT_STATUS_OPTIONS = ["", "BOOKED", "SHIPPED", "ARRIVED"]
SHIPMENT_TABLE_NAME = "shipments"
SHIPMENT_BUTTON_MAP = {
    "Melbourne": "newMelbourne_pushButton",
    "Sydney": "newSydney_pushButton",
    "SABA": "newSaba_pushButton",
}



try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.numbers import is_date_format
    from openpyxl.styles.colors import COLOR_INDEX
    from openpyxl.utils.datetime import from_excel
except Exception:
    load_workbook = None
    Workbook = None
    get_column_letter = None
    Font = None
    PatternFill = None
    Alignment = None
    is_date_format = None
    COLOR_INDEX = None
    from_excel = None


class DBRow(dict):
    def __init__(self, columns, values):
        super().__init__(zip(columns, values))
        self._columns = list(columns)
        self._values = list(values)

    def __getitem__(self, key):
        if isinstance(key, int):
            return self._values[key]
        return super().__getitem__(key)


class CursorAdapter:
    def __init__(self, cursor, backend):
        self._cursor = cursor
        self._backend = backend

    def execute(self, sql, params=()):
        translated = self._backend.translate_sql(sql)
        self._cursor.execute(translated, params)
        return self

    def executemany(self, sql, seq_of_params):
        translated = self._backend.translate_sql(sql)
        self._cursor.executemany(translated, seq_of_params)
        return self

    def fetchone(self):
        row = self._cursor.fetchone()
        return self._backend.wrap_row(self._cursor, row)

    def fetchall(self):
        rows = self._cursor.fetchall()
        return [self._backend.wrap_row(self._cursor, row) for row in rows]

    def __iter__(self):
        for row in self._cursor:
            yield self._backend.wrap_row(self._cursor, row)

    def __getattr__(self, name):
        return getattr(self._cursor, name)


class SQLServerBackend:
    def __init__(self, connection):
        self._connection = connection

    def cursor(self):
        return CursorAdapter(self._connection.cursor(), self)

    def commit(self):
        self._connection.commit()

    def close(self):
        self._connection.close()

    def wrap_row(self, cursor, row):
        if row is None:
            return None
        columns = [desc[0] for desc in cursor.description] if cursor.description else []
        values = list(row)
        return DBRow(columns, values)

    def _replace_sql_function_calls(self, sql, func_name, replacer):
        upper_sql = sql.upper()
        target = f"{func_name.upper()}("
        result = []
        index = 0
        while True:
            start = upper_sql.find(target, index)
            if start == -1:
                result.append(sql[index:])
                break
            result.append(sql[index:start])
            open_paren = start + len(func_name)
            pos = open_paren + 1
            depth = 1
            while pos < len(sql) and depth > 0:
                char = sql[pos]
                if char == '(':
                    depth += 1
                elif char == ')':
                    depth -= 1
                pos += 1
            if depth != 0:
                result.append(sql[start:])
                break
            inner = sql[open_paren + 1:pos - 1]
            result.append(replacer(inner))
            index = pos
        return ''.join(result)

    def _translate_date_calls(self, sql):
        return self._replace_sql_function_calls(sql, 'DATE', lambda inner: f"CAST({inner} AS date)")

    def translate_sql(self, sql):
        translated = sql
        translated = re.sub(
            r"SELECT\s+name\s+FROM\s+sqlite_master\s+WHERE\s+type='table'\s+AND\s+name\s*=\s*\?",
            "SELECT TABLE_NAME AS name FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE' AND TABLE_NAME = ?",
            translated, flags=re.IGNORECASE,
        )
        translated = re.sub(
            r"SELECT\s+name\s+FROM\s+sqlite_master\s+WHERE\s+type='table'\s+AND\s+name='([^']+)'",
            lambda m: f"SELECT TABLE_NAME AS name FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE' AND TABLE_NAME = '{m.group(1)}'",
            translated, flags=re.IGNORECASE,
        )

        def pragma_repl(match):
            table_name = match.group(1)
            return (
                "SELECT ORDINAL_POSITION - 1 AS cid, COLUMN_NAME AS name, DATA_TYPE AS type, "
                "CASE WHEN IS_NULLABLE = 'NO' THEN 1 ELSE 0 END AS notnull, "
                "COLUMN_DEFAULT AS dflt_value, 0 AS pk "
                f"FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '{table_name}' ORDER BY ORDINAL_POSITION"
            )

        translated = re.sub(r"PRAGMA\s+table_info\(([^)]+)\)", pragma_repl, translated, flags=re.IGNORECASE)
        translated = re.sub(r"\bCOLLATE\s+NOCASE\b", "", translated, flags=re.IGNORECASE)
        translated = re.sub(r"CAST\((.*?)\s+AS\s+TEXT\)", r"CAST(\1 AS NVARCHAR(MAX))", translated, flags=re.IGNORECASE)
        translated = re.sub(
            r"STRFTIME\(\s*'%Y-%m'\s*,\s*DATE\((.*?)\)\s*\)",
            r"CONVERT(char(7), CAST(\1 AS date), 120)",
            translated,
            flags=re.IGNORECASE | re.DOTALL,
        )
        translated = re.sub(
            r"STRFTIME\(\s*'%Y-%m'\s*,\s*CAST\((.*?)\s+AS\s+date\)\s*\)",
            r"CONVERT(char(7), CAST(\1 AS date), 120)",
            translated,
            flags=re.IGNORECASE | re.DOTALL,
        )
        translated = self._translate_date_calls(translated)

        if re.search(r"\bLIMIT\s+1\b", translated, flags=re.IGNORECASE):
            translated = re.sub(r"\bLIMIT\s+1\b", "", translated, flags=re.IGNORECASE)
            translated = re.sub(r"^\s*SELECT\b", lambda m: m.group(0) + " TOP 1", translated, count=1, flags=re.IGNORECASE)

        return translated




class NewShipmentDialog(QDialog):
    def __init__(self, supplier_names, parent=None):
        super().__init__(parent)
        self.setWindowTitle("New Shipment")
        self.resize(420, 160)

        cleaned_suppliers = []
        seen = set()
        for supplier_name in supplier_names or []:
            name = str(supplier_name or "").strip()
            if not name:
                continue
            key = name.casefold()
            if key in seen:
                continue
            seen.add(key)
            cleaned_suppliers.append(name)
        self._supplier_names = sorted(cleaned_suppliers, key=lambda value: value.casefold())

        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.order_number_edit = QLineEdit(self)
        self.order_number_edit.setPlaceholderText("Enter order number")
        form.addRow("Order Number", self.order_number_edit)

        self.supplier_combo = QComboBox(self)
        self.supplier_combo.setEditable(True)
        self.supplier_combo.addItems(self._supplier_names)
        self.supplier_combo.setCurrentIndex(-1)
        try:
            self.supplier_combo.setInsertPolicy(QComboBox.NoInsert)
            self.supplier_combo.setMaxVisibleItems(25)
        except Exception:
            pass
        supplier_completer = QCompleter(self._supplier_names, self)
        supplier_completer.setCaseSensitivity(Qt.CaseInsensitive)
        supplier_completer.setFilterMode(Qt.MatchStartsWith)
        supplier_completer.setCompletionMode(QCompleter.PopupCompletion)
        supplier_completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
        self.supplier_combo.setCompleter(supplier_completer)
        if self.supplier_combo.lineEdit() is not None:
            self.supplier_combo.lineEdit().setPlaceholderText("Choose supplier")
        form.addRow("Supplier", self.supplier_combo)

        layout.addLayout(form)

        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, parent=self)
        self.ok_button = self.button_box.button(QDialogButtonBox.Ok)
        if self.ok_button is not None:
            self.ok_button.setEnabled(False)
        self.button_box.accepted.connect(self.validate_and_accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

        self.order_number_edit.textChanged.connect(self.update_ok_button_state)
        supplier_line = self.supplier_combo.lineEdit()
        if supplier_line is not None:
            supplier_line.textChanged.connect(self.update_ok_button_state)

        self.order_number_edit.setFocus()

    def update_ok_button_state(self):
        has_order = bool(self.order_number().strip())
        has_supplier = bool((self.supplier_combo.currentText() or "").strip())
        if self.ok_button is not None:
            self.ok_button.setEnabled(has_order and has_supplier)

    def order_number(self):
        return (self.order_number_edit.text() or "").strip()

    def normalised_supplier_name(self):
        typed = (self.supplier_combo.currentText() or "").strip()
        if not typed:
            return ""

        typed_lower = typed.casefold()
        for supplier_name in self._supplier_names:
            if supplier_name.casefold() == typed_lower:
                return supplier_name

        prefix_matches = [
            supplier_name for supplier_name in self._supplier_names
            if supplier_name.casefold().startswith(typed_lower)
        ]
        if len(prefix_matches) == 1:
            return prefix_matches[0]
        return ""

    def validate_and_accept(self):
        order_number = self.order_number()
        supplier_name = self.normalised_supplier_name()

        if not order_number:
            QMessageBox.warning(self, "New Shipment", "Enter an order number.")
            self.order_number_edit.setFocus()
            return

        if not supplier_name:
            QMessageBox.warning(self, "New Shipment", "Choose a valid supplier from the list.")
            self.supplier_combo.setFocus()
            return

        self.supplier_combo.setCurrentText(supplier_name)
        self.accept()


class ShipmentComboBoxDelegate(QStyledItemDelegate):
    def __init__(self, options, parent=None):
        super().__init__(parent)
        self.options = list(options or [])

    def createEditor(self, parent, option, index):
        editor = QComboBox(parent)
        editor.addItems(self.options)
        editor.setEditable(False)
        return editor

    def setEditorData(self, editor, index):
        value = str(index.data(Qt.EditRole) or index.data(Qt.DisplayRole) or "")
        pos = editor.findText(value)
        editor.setCurrentIndex(pos if pos >= 0 else 0)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.EditRole)


class ShipmentDateDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.setInputMask("00/00/00;_")
        editor.setPlaceholderText("dd/mm/yy")
        return editor

    def setEditorData(self, editor, index):
        editor.setText(str(index.data(Qt.EditRole) or index.data(Qt.DisplayRole) or ""))

    def setModelData(self, editor, model, index):
        value = (editor.text() or "").replace("_", "").strip()
        if value in {"", "//", "/ /"}:
            value = ""
        model.setData(index, value, Qt.EditRole)


class ShipmentsWindow(QMainWindow):
    RECORD_ID_ROLE = Qt.UserRole + 500

    def __init__(self, main_window):
        super().__init__(main_window)
        self.main_window = main_window
        self.ui = Ui_ShipmentsWindow()
        self.ui.setupUi(self)
        self.setWindowTitle("Shipments")
        self.setAttribute(Qt.WA_DeleteOnClose, False)
        self._loading_table_data = False
        self._handling_table_change = False
        self.supplier_names = []
        self._qty_delegate = ShipmentComboBoxDelegate(SHIPMENT_QTY_OPTIONS, self)
        self._status_delegate = ShipmentComboBoxDelegate(SHIPMENT_STATUS_OPTIONS, self)
        self._type_delegate = ShipmentComboBoxDelegate(SHIPMENT_TYPE_OPTIONS, self)
        self._date_delegate = ShipmentDateDelegate(self)

        self.configure_ui()
        self.connect_signals()
        self.refresh_from_database()
        self.destroyed.connect(self._clear_owner_reference)

    def _clear_owner_reference(self, *_args):
        if getattr(self.main_window, "shipments_window", None) is self:
            self.main_window.shipments_window = None

    def table_widget(self):
        return getattr(self.ui, "mainshipping_table", None)

    def shipment_remove_column(self):
        return len(SHIPMENT_FIELD_MAP)

    def shipment_column_index(self, field_name):
        for index, (_header, candidate_field_name) in enumerate(SHIPMENT_FIELD_MAP):
            if candidate_field_name == field_name:
                return index
        return -1

    def format_display_date(self, value):
        parsed = self.main_window.parse_date_value(value)
        if parsed is None:
            return ""
        return parsed.strftime("%d/%m/%y")

    def parse_shipment_date(self, value):
        text = str(value or "").strip()
        if not text:
            return None
        return self.main_window.parse_date_value(text)

    def normalise_date_text(self, text):
        text = str(text or "").strip()
        if not text:
            return ""
        parsed = self.parse_shipment_date(text)
        if parsed is None:
            return text
        return parsed.strftime("%d/%m/%y")

    def normalise_qty_value(self, text):
        text = str(text or "").strip()
        if not text:
            return ""
        cleaned = (
            text.upper()
            .replace('”', "'")
            .replace('′', "'")
            .replace('FT', '')
            .replace('FEET', '')
            .replace(' ', '')
        )
        if cleaned == 'LCL':
            return 'LCL'
        if re.fullmatch(r"(?:1X)?20(?:'|)?(?:FCL|HQ|HC)?", cleaned):
            return "20' FCL"
        if re.fullmatch(r"(?:1X)?40(?:'|)?(?:FCL|HQ|HC)?", cleaned):
            return "40' FCL"
        return text

    def normalise_status_value(self, text):
        text = str(text or "").strip()
        if not text:
            return ""
        cleaned = text.upper()
        if cleaned in {"BOOKED", "SHIPPED", "ARRIVED"}:
            return cleaned
        return cleaned

    def normalise_shipment_type_value(self, text):
        text = str(text or "").strip()
        if not text:
            return ""
        lowered = text.casefold()
        for option in SHIPMENT_TYPE_OPTIONS:
            if lowered == option.casefold():
                return option
        if lowered == 'saba':
            return 'SABA'
        if lowered.startswith('melb'):
            return 'Melbourne'
        if lowered.startswith('syd'):
            return 'Sydney'
        return text

    def normalise_field_value(self, field_name, value):
        text = str(value or "").strip()
        if field_name in {"entry_date", "ready_date", "shipment_date", "due_date"}:
            return self.normalise_date_text(text)
        if field_name == "shipment_type":
            return self.normalise_shipment_type_value(text)
        if field_name == "qty":
            return self.normalise_qty_value(text)
        if field_name == "status":
            return self.normalise_status_value(text)
        return text

    def configure_ui(self):
        date_box = getattr(self.ui, "dateUpdated_textBrowser", None)
        if date_box is not None:
            date_box.setReadOnly(True)
            date_box.setOpenLinks(False)
            date_box.setPlaceholderText("Last Updated")

        label = getattr(self.ui, "newShipment_label", None)
        if label is not None:
            label.setText("Add New Shipment")

        extra_tables = [
            getattr(self.ui, "sydneyShipping_table", None),
            getattr(self.ui, "sabaShipping_table", None),
        ]
        for extra_table in extra_tables:
            if extra_table is not None:
                extra_table.hide()
                extra_table.setParent(None)

        top_layout = getattr(self.ui, "horizontalLayout", None)
        if top_layout is not None and not hasattr(self, "_shipment_filter_combo"):
            filter_label = QLabel("Filter Type", self)
            self._shipment_filter_label = filter_label
            self._shipment_filter_combo = QComboBox(self)
            self._shipment_filter_combo.addItem("All")
            for option in SHIPMENT_TYPE_OPTIONS:
                self._shipment_filter_combo.addItem(option)
            self._shipment_filter_combo.setCurrentText("All")
            top_layout.addSpacing(12)
            top_layout.addWidget(filter_label)
            top_layout.addWidget(self._shipment_filter_combo)

        table = self.table_widget()
        if table is None:
            return

        table.clear()
        table.setColumnCount(len(SHIPMENT_HEADERS))
        table.setHorizontalHeaderLabels(SHIPMENT_HEADERS)
        table.setRowCount(0)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.EditKeyPressed
            | QAbstractItemView.AnyKeyPressed
            | QAbstractItemView.SelectedClicked
        )
        table.verticalHeader().setVisible(False)
        table.setItemDelegateForColumn(self.shipment_column_index("shipment_type"), self._type_delegate)
        table.setItemDelegateForColumn(self.shipment_column_index("qty"), self._qty_delegate)
        table.setItemDelegateForColumn(self.shipment_column_index("status"), self._status_delegate)
        for field_name in ("entry_date", "ready_date", "shipment_date", "due_date"):
            column_index = self.shipment_column_index(field_name)
            if column_index >= 0:
                table.setItemDelegateForColumn(column_index, self._date_delegate)

        header = table.horizontalHeader()
        header.setStretchLastSection(False)
        for column_index in range(len(SHIPMENT_HEADERS)):
            header.setSectionResizeMode(column_index, QHeaderView.Interactive)
        try:
            header.sectionHandleDoubleClicked.connect(lambda logical_index, tbl=table: self.auto_size_shipment_column(tbl, logical_index))
        except Exception:
            pass
        header.sectionDoubleClicked.connect(lambda logical_index, tbl=table: self.auto_size_shipment_column(tbl, logical_index))
        self.install_table_font_menus()

    def install_table_font_menus(self):
        table = self.table_widget()
        if table is None:
            return

        if not str(table.objectName() or "").strip():
            table.setObjectName("mainshipping_table")
        table.setProperty("_table_font_settings_token", "shipments_window/mainshipping_table")

        if not table.property("_shipments_context_menu_installed"):
            table.setProperty("_shipments_context_menu_installed", True)
            table.setProperty("_table_font_scope_key", "shipments_window")
            table.setProperty("_table_font_default_size", current_table_font_size(table))
            table.setContextMenuPolicy(Qt.CustomContextMenu)
            table.customContextMenuRequested.connect(self.show_shipments_table_context_menu)

        saved_size = stored_table_font_size(self.main_window.settings, "shipments_window", table)
        if saved_size:
            apply_table_font_size(
                table,
                saved_size,
                settings=self.main_window.settings,
                scope_key="shipments_window",
                persist=False,
            )

    def shipment_row_text(self, table, row_index, field_name):
        if table is None or row_index is None or row_index < 0:
            return ""
        column_index = self.shipment_column_index(field_name)
        if column_index < 0:
            return ""
        item = table.item(row_index, column_index)
        return (item.text() if item is not None else "").strip()

    def show_shipments_table_context_menu(self, pos):
        table = self.table_widget()
        if table is None:
            return

        clicked_item = table.itemAt(pos)
        row_index = clicked_item.row() if clicked_item is not None else table.currentRow()
        has_row = row_index is not None and row_index >= 0
        vessel_name = self.shipment_row_text(table, row_index, "vessel") if has_row else ""
        container_no = self.shipment_row_text(table, row_index, "container_no") if has_row else ""

        menu = QMenu(table)

        open_vessel_action = None
        open_container_action = None
        if has_row:
            open_vessel_action = menu.addAction("Open Vessel in MarineTraffic")
            open_vessel_action.setEnabled(bool(vessel_name))
            open_container_action = menu.addAction("Open Container in findTEU")
            open_container_action.setEnabled(bool(container_no))
            menu.addSeparator()

        font_menu = menu.addMenu("Font Size")
        current_size = current_table_font_size(table)
        for size in TABLE_FONT_SIZE_OPTIONS:
            action = font_menu.addAction(str(size))
            action.setCheckable(True)
            action.setChecked(size == current_size)
            action.triggered.connect(
                lambda _checked=False, tbl=table, selected_size=size: apply_table_font_size(
                    tbl,
                    selected_size,
                    settings=self.main_window.settings,
                    scope_key="shipments_window",
                    persist=True,
                )
            )

        font_menu.addSeparator()
        reset_action = font_menu.addAction("Reset")
        reset_action.triggered.connect(
            lambda _checked=False, tbl=table: reset_table_font_size(
                tbl,
                settings=self.main_window.settings,
                scope_key="shipments_window",
            )
        )

        viewport = getattr(table, "viewport", lambda: None)()
        global_pos = viewport.mapToGlobal(pos) if viewport is not None else table.mapToGlobal(pos)
        chosen_action = menu.exec(global_pos)
        if chosen_action is None:
            return

        if chosen_action == open_vessel_action:
            self.open_vessel_in_marinetraffic(vessel_name)
        elif chosen_action == open_container_action:
            self.open_container_in_findteu(container_no)

    def open_vessel_in_marinetraffic(self, vessel_name):
        vessel_name = str(vessel_name or "").strip()
        if not vessel_name:
            return
        try:
            QApplication.clipboard().setText(vessel_name)
        except Exception:
            pass
        QDesktopServices.openUrl(QUrl("https://www.marinetraffic.com/"))
        try:
            self.statusBar().showMessage("Vessel copied to clipboard. MarineTraffic opened.", 4000)
        except Exception:
            pass

    def open_container_in_findteu(self, container_no):
        container_no = str(container_no or "").strip()
        if not container_no:
            return
        try:
            QApplication.clipboard().setText(container_no)
        except Exception:
            pass
        QDesktopServices.openUrl(QUrl("https://www.findteu.com/"))
        try:
            self.statusBar().showMessage("Container copied to clipboard. findTEU opened.", 4000)
        except Exception:
            pass

    def connect_signals(self):
        for shipment_type, button_name in SHIPMENT_BUTTON_MAP.items():
            button = getattr(self.ui, button_name, None)
            if button is not None:
                button.clicked.connect(lambda _checked=False, shipment_type=shipment_type: self.create_new_shipment(shipment_type))

        filter_combo = getattr(self, "_shipment_filter_combo", None)
        if filter_combo is not None:
            filter_combo.currentTextChanged.connect(lambda _text: self.load_table())

        table = self.table_widget()
        if table is not None:
            table.itemChanged.connect(self.handle_table_item_changed)
            table.cellDoubleClicked.connect(self.handle_table_cell_double_clicked)

    def auto_size_shipment_column(self, table, logical_index):
        if table is None or logical_index < 0:
            return
        try:
            table.resizeColumnToContents(logical_index)
        except Exception:
            pass

    def auto_size_all_shipment_columns(self, table):
        if table is None:
            return
        for column_index in range(table.columnCount()):
            self.auto_size_shipment_column(table, column_index)

    def refresh_supplier_names(self):
        self.main_window.load_reference_lists()
        self.supplier_names = sorted(
            {(name or "").strip() for name in self.main_window.supplier_names if (name or "").strip()},
            key=lambda value: value.casefold(),
        )

    def refresh_from_database(self):
        self.refresh_supplier_names()
        self.load_table()
        self.refresh_last_updated_display()

    def create_new_shipment(self, shipment_type):
        self.refresh_supplier_names()
        dialog = NewShipmentDialog(self.supplier_names, self)
        if dialog.exec() != QDialog.Accepted:
            return

        shipment_type = self.normalise_shipment_type_value(shipment_type)
        today_display = date.today().strftime("%d/%m/%y")
        today_iso = date.today().isoformat()
        order_number = dialog.order_number()
        supplier_name = dialog.normalised_supplier_name()

        cur = self.main_window.db_conn.cursor()
        cur.execute(
            f"""
            INSERT INTO {SHIPMENT_TABLE_NAME} (
                entry_date, shipment_type, order_no, supplier_name, container_no, product, qty,
                ready_date, shipment_date, due_date, status, vessel, notes, updated_on
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (today_display, shipment_type, order_number, supplier_name, "", "", "", "", "", "", "", "", "", today_iso),
        )
        self.main_window.db_conn.commit()
        self.mark_shipments_changed(today_iso)
        self.load_table()
        self.refresh_last_updated_display()

    def format_cell_text(self, field_name, value):
        if value is None:
            return ""
        if field_name in {"entry_date", "ready_date", "shipment_date", "due_date"}:
            return self.format_display_date(value)
        if field_name == "shipment_type":
            return self.normalise_shipment_type_value(value)
        if field_name == "qty":
            return self.normalise_qty_value(value)
        if field_name == "status":
            return self.normalise_status_value(value)
        return str(value)

    def current_shipment_type_filter(self):
        filter_combo = getattr(self, "_shipment_filter_combo", None)
        if filter_combo is None:
            return ""
        selected = str(filter_combo.currentText() or "").strip()
        if not selected or selected.casefold() == "all":
            return ""
        return self.normalise_shipment_type_value(selected)

    def shipment_sort_key(self, row_data):
        due_date = self.parse_shipment_date(row_data.get("due_date", ""))
        shipment_date = self.parse_shipment_date(row_data.get("shipment_date", ""))
        entry_date = self.parse_shipment_date(row_data.get("entry_date", ""))
        return (
            1 if due_date is not None else 0,
            due_date or date.min,
            shipment_date or date.min,
            entry_date or date.min,
            int(row_data.get("id", 0) or 0),
        )

    def make_remove_item(self):
        item = QTableWidgetItem(SHIPMENT_REMOVE_HEADER)
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags((item.flags() | Qt.ItemIsSelectable | Qt.ItemIsEnabled) & ~Qt.ItemIsEditable)
        font = item.font()
        font.setBold(True)
        item.setFont(font)
        return item

    def load_table(self):
        table = self.table_widget()
        if table is None:
            return

        rows = self.main_window.db_all(
            f"""
            SELECT id, entry_date, shipment_type, order_no, supplier_name, container_no, product, qty,
                   ready_date, shipment_date, due_date, status, vessel, notes
            FROM {SHIPMENT_TABLE_NAME}
            """
        )

        shipment_type_filter = self.current_shipment_type_filter()
        if shipment_type_filter:
            rows = [
                row_data for row_data in rows
                if self.normalise_shipment_type_value(row_data.get("shipment_type", "")) == shipment_type_filter
            ]

        rows = sorted(rows, key=self.shipment_sort_key, reverse=True)

        self._loading_table_data = True
        blocker = QSignalBlocker(table)
        try:
            table.setRowCount(0)
            for row_data in rows:
                row_index = table.rowCount()
                table.insertRow(row_index)
                for column_index, (_header, field_name) in enumerate(SHIPMENT_FIELD_MAP):
                    item = QTableWidgetItem(self.format_cell_text(field_name, row_data[field_name]))
                    if column_index == 0:
                        item.setData(self.RECORD_ID_ROLE, row_data["id"])
                    if field_name in {"entry_date", "shipment_type", "qty", "ready_date", "shipment_date", "due_date", "status"}:
                        item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row_index, column_index, item)
                table.setItem(row_index, self.shipment_remove_column(), self.make_remove_item())
                self.apply_row_styles(table, row_index)
        finally:
            del blocker
            self._loading_table_data = False

        table.resizeRowsToContents()
        saved_size = stored_table_font_size(self.main_window.settings, "shipments_window", table)
        if saved_size:
            apply_table_font_size(table, saved_size, settings=self.main_window.settings, scope_key="shipments_window", persist=False)
        self.auto_size_all_shipment_columns(table)

    def record_id_for_row(self, table, row_index):
        for column_index in range(min(table.columnCount(), len(SHIPMENT_FIELD_MAP))):
            item = table.item(row_index, column_index)
            if item is None:
                continue
            record_id = item.data(self.RECORD_ID_ROLE)
            if record_id not in (None, ""):
                return int(record_id)
        return None

    def row_values(self, table, row_index):
        values = {}
        for column_index, (_header, field_name) in enumerate(SHIPMENT_FIELD_MAP):
            item = table.item(row_index, column_index)
            values[field_name] = (item.text() if item is not None else "").strip()
        return values

    def normalise_row_items(self, table, row_index):
        values = {}
        blocker = QSignalBlocker(table)
        try:
            for column_index, (_header, field_name) in enumerate(SHIPMENT_FIELD_MAP):
                item = table.item(row_index, column_index)
                if item is None:
                    item = QTableWidgetItem("")
                    table.setItem(row_index, column_index, item)
                normalised_value = self.normalise_field_value(field_name, item.text())
                values[field_name] = normalised_value
                if item.text() != normalised_value:
                    item.setText(normalised_value)
                if field_name in {"entry_date", "shipment_type", "qty", "ready_date", "shipment_date", "due_date", "status"}:
                    item.setTextAlignment(Qt.AlignCenter)
        finally:
            del blocker
        return values

    def handle_table_item_changed(self, item):
        if self._loading_table_data or self._handling_table_change or item is None:
            return

        table = self.table_widget()
        if table is None:
            return

        row_index = item.row()
        if item.column() == self.shipment_remove_column():
            return

        record_id = self.record_id_for_row(table, row_index)
        if record_id is None:
            return

        self._handling_table_change = True
        try:
            values = self.normalise_row_items(table, row_index)
            updated_on = date.today().isoformat()

            cur = self.main_window.db_conn.cursor()
            cur.execute(
                f"""
                UPDATE {SHIPMENT_TABLE_NAME}
                SET entry_date = ?,
                    shipment_type = ?,
                    order_no = ?,
                    supplier_name = ?,
                    container_no = ?,
                    product = ?,
                    qty = ?,
                    ready_date = ?,
                    shipment_date = ?,
                    due_date = ?,
                    status = ?,
                    vessel = ?,
                    notes = ?,
                    updated_on = ?
                WHERE id = ?
                """,
                (
                    values["entry_date"],
                    values["shipment_type"],
                    values["order_no"],
                    values["supplier_name"],
                    values["container_no"],
                    values["product"],
                    values["qty"],
                    values["ready_date"],
                    values["shipment_date"],
                    values["due_date"],
                    values["status"],
                    values["vessel"],
                    values["notes"],
                    updated_on,
                    record_id,
                ),
            )
            self.main_window.db_conn.commit()
            self.mark_shipments_changed(updated_on)
            self.apply_row_styles(table, row_index)
            self.refresh_last_updated_display()
        finally:
            self._handling_table_change = False

    def handle_table_cell_double_clicked(self, row_index, column_index):
        table = self.table_widget()
        if table is None:
            return
        if column_index != self.shipment_remove_column():
            return
        record_id = self.record_id_for_row(table, row_index)
        if record_id is None:
            return
        answer = QMessageBox.question(
            self,
            "Remove Shipment",
            "Remove this shipment row?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return

        cur = self.main_window.db_conn.cursor()
        cur.execute(f"DELETE FROM {SHIPMENT_TABLE_NAME} WHERE id = ?", (record_id,))
        self.main_window.db_conn.commit()
        self.mark_shipments_changed(date.today().isoformat())
        self.load_table()
        self.refresh_last_updated_display()

    def set_row_item_colors(self, item, background=None, foreground=None):
        if item is None:
            return
        if background is None:
            item.setBackground(QBrush())
        else:
            item.setBackground(QBrush(background))
        if foreground is None:
            item.setForeground(QBrush())
        else:
            item.setForeground(QBrush(foreground))

    def apply_row_styles(self, table, row_index):
        values = self.row_values(table, row_index)
        status = self.normalise_status_value(values.get("status", ""))
        shipment_date = self.parse_shipment_date(values.get("shipment_date", ""))
        due_date = self.parse_shipment_date(values.get("due_date", ""))
        today = date.today()

        shipped_row_background = QColor(0, 255, 0)
        booked_row_background = QColor(160, 200, 235)
        blank_row_background_even = QColor(214, 214, 214)
        blank_row_background_odd = QColor(190, 190, 190)
        overdue_background = QColor(255, 199, 206)
        overdue_foreground = QColor(156, 0, 6)

        for column_index in range(table.columnCount()):
            self.set_row_item_colors(table.item(row_index, column_index), None, None)

        if status in {"SHIPPED", "ARRIVED"}:
            for column_index in range(table.columnCount()):
                self.set_row_item_colors(table.item(row_index, column_index), shipped_row_background, QColor(Qt.black))
        elif status == "BOOKED":
            for column_index in range(table.columnCount()):
                self.set_row_item_colors(table.item(row_index, column_index), booked_row_background, QColor(Qt.black))
        elif not status:
            blank_row_background = blank_row_background_even if row_index % 2 == 0 else blank_row_background_odd
            for column_index in range(table.columnCount()):
                self.set_row_item_colors(table.item(row_index, column_index), blank_row_background, QColor(Qt.black))

        shipment_column = self.shipment_column_index("shipment_date")
        if shipment_column >= 0 and shipment_date is not None and shipment_date < today and status not in {"SHIPPED", "ARRIVED"}:
            self.set_row_item_colors(table.item(row_index, shipment_column), overdue_background, overdue_foreground)

        due_column = self.shipment_column_index("due_date")
        if due_column >= 0:
            due_item = table.item(row_index, due_column)
            if due_item is not None:
                due_font = due_item.font()
                due_font.setBold(True)
                due_item.setFont(due_font)
        if due_column >= 0 and due_date is not None and due_date < today and status != "ARRIVED":
            self.set_row_item_colors(table.item(row_index, due_column), overdue_background, overdue_foreground)

    def mark_shipments_changed(self, updated_on=None):
        updated_on = updated_on or date.today().isoformat()
        cur = self.main_window.db_conn.cursor()
        cur.execute("DELETE FROM shipments_meta WHERE meta_key = ?", ("last_updated",))
        cur.execute(
            "INSERT INTO shipments_meta (meta_key, meta_value) VALUES (?, ?)",
            ("last_updated", updated_on),
        )
        self.main_window.db_conn.commit()

    def refresh_last_updated_display(self):
        latest_date = None
        row = self.main_window.db_one(
            "SELECT meta_value FROM shipments_meta WHERE meta_key = ?",
            ("last_updated",),
        )
        if row is not None:
            latest_date = self.main_window.parse_date_value(row["meta_value"])

        if latest_date is None:
            fallback_row = self.main_window.db_one(f"SELECT MAX(updated_on) AS updated_on FROM {SHIPMENT_TABLE_NAME}")
            if fallback_row is not None:
                latest_date = self.main_window.parse_date_value(fallback_row["updated_on"])

        text_value = latest_date.strftime("%d/%m/%y") if latest_date else ""
        date_box = getattr(self.ui, "dateUpdated_textBrowser", None)
        if date_box is not None:
            date_box.setPlainText(text_value)



class CustomerFileViewerDialog(QDialog):
    MAX_ROWS = 3000
    MAX_COLS = 200
    MAX_SHEETS = 20

    def __init__(self, path, parent=None, is_customer_file_context=False):
        super().__init__(parent)
        self.path = Path(path)
        self.is_customer_file_context = bool(is_customer_file_context)
        self.file_type = self.detect_file_type()
        self.edit_mode = False
        self.is_dirty = False
        self.workbook = None
        self.sheet_tables = {}
        self.changed_excel_cells = {}
        self.csv_rows = []
        self.csv_table = None
        self.text_editor = None
        self.title_label = None
        self.path_label = None

        self.setWindowTitle(f"Customer File - {self.path.name}")
        self.resize(1280, 820)
        self.apply_viewer_style()
        self.build_ui()
        self.load_file()
        self.update_buttons()

    def table_font_settings(self):
        parent = self.parent()
        settings = getattr(parent, "settings", None) if parent is not None else None
        return settings or QSettings("Windsor", "WidgetApp")

    def install_table_font_menu(self, table, settings_token):
        install_table_font_context_menu(
            self,
            table,
            self.table_font_settings(),
            "customer_file_viewer",
            settings_token=settings_token,
        )

    def apply_viewer_style(self):
        self.setStyleSheet("""
            QDialog {
                background: #d9dee4;
                color: #111111;
            }
            QLabel {
                color: #111111;
                background: transparent;
            }
            QTabWidget::pane {
                border: 1px solid #a4adb7;
                background: #dfe4e9;
                top: -1px;
            }
            QTabBar::tab {
                background: #c7cfd8;
                color: #111111;
                border: 1px solid #9aa5b1;
                padding: 6px 12px;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background: #eef1f4;
                font-weight: 700;
            }
            QTableWidget, QTextEdit {
                background: #eceff3;
                color: #111111;
                border: 1px solid #98a2ad;
                gridline-color: #bcc4cd;
                selection-background-color: #b9d3f0;
                selection-color: #111111;
            }
            QTableCornerButton::section,
            QHeaderView::section {
                background: #cfd6de;
                color: #111111;
                border: 1px solid #a2acb7;
                padding: 4px;
                font-weight: 600;
            }
            QPushButton {
                background: #cdd4db;
                color: #111111;
                border: 1px solid #8f99a4;
                border-radius: 4px;
                padding: 6px 10px;
            }
            QPushButton:hover {
                background: #c1c9d1;
            }
            QPushButton:disabled {
                background: #dfe3e8;
                color: #6b7280;
                border: 1px solid #b6bec7;
            }
        """)

    def detect_file_type(self):
        suffix = self.path.suffix.lower()
        if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm'} and load_workbook is not None:
            return 'excel'
        if suffix == '.csv':
            return 'csv'
        if suffix in {'.txt', '.log', '.md'}:
            return 'text'
        return 'unsupported'

    def build_ui(self):
        layout = QVBoxLayout(self)

        self.title_label = QLabel(self.path.name)
        title_font = self.title_label.font()
        title_font.setPointSize(title_font.pointSize() + 1)
        title_font.setBold(True)
        self.title_label.setFont(title_font)
        layout.addWidget(self.title_label)

        self.path_label = QLabel(str(self.path))
        self.path_label.setWordWrap(True)
        self.path_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self.path_label)

        self.status_label = QLabel('Read only')
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)

        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.content_widget, 1)

        button_row = QHBoxLayout()
        button_row.addStretch(1)

        self.select_button = QPushButton('Select File')
        self.select_button.clicked.connect(self.select_new_file)
        button_row.addWidget(self.select_button)

        self.edit_button = QPushButton('Enable Editing')
        self.edit_button.clicked.connect(self.toggle_edit_mode)
        button_row.addWidget(self.edit_button)

        self.open_button = QPushButton('Open in Excel')
        self.open_button.clicked.connect(self.open_in_excel)
        button_row.addWidget(self.open_button)

        self.save_close_button = QPushButton('Save and Close')
        self.save_close_button.clicked.connect(self.save_and_close)
        button_row.addWidget(self.save_close_button)

        self.close_button = QPushButton('Close')
        self.close_button.clicked.connect(self.reject)
        button_row.addWidget(self.close_button)

        layout.addLayout(button_row)

    def refresh_header(self):
        if self.title_label is not None:
            self.title_label.setText(self.path.name)
        if self.path_label is not None:
            self.path_label.setText(str(self.path))
        self.setWindowTitle(f"Customer File - {self.path.name}{' *' if self.is_dirty else ''}")

    def load_new_path(self, new_path):
        self.path = Path(new_path)
        self.file_type = self.detect_file_type()
        self.edit_mode = False
        self.workbook = None
        self.refresh_header()
        self.load_file()

    def select_new_file(self):
        parent_window = self.parent()
        if parent_window is None or not hasattr(parent_window, "select_customer_file_for_current_customer"):
            QMessageBox.warning(self, "Customer File", "File selection is not available from this dialog.")
            return

        if self.is_dirty:
            answer = QMessageBox.question(
                self,
                'Discard changes?',
                'There are unsaved changes in this viewer. Select a different file without saving?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                return

        parent_window.select_customer_file_for_current_customer(self)

    def clear_content(self):
        while self.content_layout.count():
            item = self.content_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

    def load_file(self):
        self.clear_content()
        self.sheet_tables = {}
        self.changed_excel_cells = {}
        self.csv_rows = []
        self.csv_table = None
        self.text_editor = None
        self.is_dirty = False

        if self.file_type == 'excel':
            self.load_excel_file()
        elif self.file_type == 'csv':
            self.load_csv_file()
        elif self.file_type == 'text':
            self.load_text_file()
        else:
            self.load_unsupported_file()

        self.update_title()
        self.update_buttons()

    def load_excel_file(self):
        try:
            self.workbook = load_workbook(filename=str(self.path), data_only=False)
        except Exception as exc:
            self.load_message(f"Could not open workbook.\n\n{exc}")
            return

        tabs = QTabWidget()
        tabs.setDocumentMode(True)

        sheet_names = self.workbook.sheetnames[:self.MAX_SHEETS]
        for sheet_name in sheet_names:
            sheet = self.workbook[sheet_name]
            row_count = max(1, min(sheet.max_row or 1, self.MAX_ROWS))
            col_count = max(1, min(sheet.max_column or 1, self.MAX_COLS))

            table = QTableWidget(row_count, col_count)
            table.setObjectName(f"preview_table_{re.sub(r'[^A-Za-z0-9_]+', '_', sheet_name).strip('_') or 'sheet'}")
            table.setAlternatingRowColors(False)
            table.setSelectionBehavior(QAbstractItemView.SelectItems)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

            if get_column_letter is not None:
                table.setHorizontalHeaderLabels([get_column_letter(i + 1) for i in range(col_count)])

            default_col_width = table.horizontalHeader().defaultSectionSize()
            for col_idx in range(1, col_count + 1):
                col_letter = get_column_letter(col_idx) if get_column_letter is not None else None
                col_dim = sheet.column_dimensions.get(col_letter) if col_letter else None
                if col_dim is not None and getattr(col_dim, 'hidden', False):
                    table.setColumnHidden(col_idx - 1, True)
                    continue
                width = getattr(col_dim, 'width', None) if col_dim else None
                if width is not None:
                    table.setColumnWidth(col_idx - 1, self.excel_column_width_to_pixels(width))
                else:
                    table.setColumnWidth(col_idx - 1, default_col_width)

            default_row_height = table.verticalHeader().defaultSectionSize()
            for row_idx in range(1, row_count + 1):
                row_dim = sheet.row_dimensions.get(row_idx)
                if row_dim is not None and getattr(row_dim, 'hidden', False):
                    table.setRowHidden(row_idx - 1, True)
                    continue
                height = getattr(row_dim, 'height', None) if row_dim else None
                if height is not None:
                    table.setRowHeight(row_idx - 1, self.excel_row_height_to_pixels(height))
                else:
                    table.setRowHeight(row_idx - 1, default_row_height)

                for col_idx in range(1, col_count + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    item = QTableWidgetItem(
                        self.cell_display_text(
                            cell.value,
                            col_idx,
                            cell,
                            row_index=row_idx,
                            context=sheet,
                        )
                    )
                    self.apply_excel_cell_style(item, cell)
                    table.setItem(row_idx - 1, col_idx - 1, item)

            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                if min_row > row_count or min_col > col_count:
                    continue
                span_rows = min(max_row, row_count) - min_row + 1
                span_cols = min(max_col, col_count) - min_col + 1
                if span_rows > 1 or span_cols > 1:
                    table.setSpan(min_row - 1, min_col - 1, span_rows, span_cols)

            table.itemChanged.connect(lambda item, s=sheet_name: self.on_excel_item_changed(s, item))

            if sheet.max_row > row_count or sheet.max_column > col_count:
                info_text = f"Showing {row_count:,} of {sheet.max_row:,} rows and {col_count:,} of {sheet.max_column:,} columns."
            else:
                info_text = f"Showing full sheet: {sheet.max_row:,} rows x {sheet.max_column:,} columns."
            if sheet.freeze_panes:
                info_text += f"  Freeze panes: {sheet.freeze_panes}"

            info = QLabel(info_text)
            info.setWordWrap(True)

            page = QWidget()
            page_layout = QVBoxLayout(page)
            page_layout.addWidget(info)
            page_layout.addWidget(table, 1)
            tabs.addTab(page, sheet_name)
            self.sheet_tables[sheet_name] = table
            self.install_table_font_menu(table, f"customer_file_viewer/excel/{table.objectName()}")

        self.content_layout.addWidget(tabs, 1)

    def load_csv_file(self):
        import csv

        try:
            with open(self.path, 'r', newline='', encoding='utf-8-sig', errors='replace') as handle:
                self.csv_rows = [list(row) for row in csv.reader(handle)]
        except Exception as exc:
            self.load_message(f"Could not open CSV file.\n\n{exc}")
            return

        row_count = max(1, len(self.csv_rows))
        col_count = max(1, max((len(row) for row in self.csv_rows), default=0))
        table = QTableWidget(row_count, col_count)
        table.setObjectName("preview_csv_table")
        table.setAlternatingRowColors(False)
        table.setSelectionBehavior(QAbstractItemView.SelectItems)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        for row_idx, row in enumerate(self.csv_rows):
            for col_idx, value in enumerate(row):
                table.setItem(
                    row_idx,
                    col_idx,
                    QTableWidgetItem(
                        self.cell_display_text(
                            value,
                            col_idx + 1,
                            row_index=row_idx + 1,
                            context=self.csv_rows,
                        )
                    ),
                )

        table.itemChanged.connect(self.on_generic_content_changed)
        table.resizeColumnsToContents()
        table.horizontalHeader().setStretchLastSection(True)
        self.csv_table = table
        self.install_table_font_menu(table, "customer_file_viewer/csv/preview_csv_table")

        info = QLabel(f"Showing full CSV: {row_count:,} rows x {col_count:,} columns.")
        info.setWordWrap(True)
        self.content_layout.addWidget(info)
        self.content_layout.addWidget(table, 1)

    def load_text_file(self):
        editor = QTextEdit()
        editor.setReadOnly(True)
        try:
            text = self.path.read_text(encoding='utf-8', errors='replace')
        except Exception as exc:
            self.load_message(f"Could not open text file.\n\n{exc}")
            return

        editor.setPlainText(text)
        editor.textChanged.connect(self.on_generic_content_changed)
        self.text_editor = editor
        self.content_layout.addWidget(editor, 1)

    def load_unsupported_file(self):
        self.load_message(
            'Preview and in-app editing are not available for this file type. '
            'Use "Open in Excel" to open the original file in its default application.'
        )

    def load_message(self, message):
        label = QLabel(message)
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        self.content_layout.addStretch(1)
        self.content_layout.addWidget(label)
        self.content_layout.addStretch(1)

    def open_in_excel(self):
        try:
            if sys.platform.startswith('win'):
                os.startfile(str(self.path))
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(str(self.path)))
        except Exception as exc:
            QMessageBox.warning(self, 'Open File', f'Could not open file:\n{self.path}\n\n{exc}')

    def update_title(self):
        self.refresh_header()

    def update_buttons(self):
        can_edit = self.file_type in {'excel', 'csv', 'text'}
        self.edit_button.setEnabled(can_edit)
        self.edit_button.setText('Disable Editing' if self.edit_mode else 'Enable Editing')
        self.save_close_button.setEnabled(can_edit)

        if not can_edit:
            self.status_label.setText('Read only. In-app editing is not available for this file type.')
        elif self.edit_mode:
            self.status_label.setText('Unsaved changes.' if self.is_dirty else 'Editing enabled. No unsaved changes yet.')
        else:
            self.status_label.setText('Read only. Use Enable Editing before making changes.')

    def toggle_edit_mode(self):
        if self.file_type not in {'excel', 'csv', 'text'}:
            return

        self.edit_mode = not self.edit_mode

        if self.file_type == 'excel':
            for table in self.sheet_tables.values():
                table.setEditTriggers(QAbstractItemView.AllEditTriggers if self.edit_mode else QAbstractItemView.NoEditTriggers)
        elif self.file_type == 'csv' and self.csv_table is not None:
            self.csv_table.setEditTriggers(QAbstractItemView.AllEditTriggers if self.edit_mode else QAbstractItemView.NoEditTriggers)
        elif self.file_type == 'text' and self.text_editor is not None:
            self.text_editor.setReadOnly(not self.edit_mode)

        self.update_buttons()

    def on_excel_item_changed(self, sheet_name, item):
        if item is None or not self.edit_mode:
            return
        row = item.row() + 1
        col = item.column() + 1
        self.changed_excel_cells.setdefault(sheet_name, {})[(row, col)] = item.text()
        self.is_dirty = True
        self.update_title()
        self.update_buttons()

    def on_generic_content_changed(self, *_args):
        if not self.edit_mode:
            return
        self.is_dirty = True
        self.update_title()
        self.update_buttons()

    def is_customer_file(self):
        return self.is_customer_file_context or any(part.lower() == "customerfiles" for part in self.path.parts)

    def normalize_header(self, header):
        return re.sub(r"[^a-z0-9]+", "", str(header or "").strip().lower())

    def excel_column_width_to_pixels(self, width):
        try:
            width_value = float(width)
        except Exception:
            return 100
        if width_value <= 0:
            return 0
        return max(18, min(1200, int(round(width_value * 7 + 5))))

    def excel_row_height_to_pixels(self, height):
        try:
            height_value = float(height)
        except Exception:
            return 24
        if height_value <= 0:
            return 0
        return max(16, min(600, int(round(height_value * 96 / 72))))

    def cell_has_currency_format(self, cell):
        number_format = str(getattr(cell, 'number_format', '') or '')
        return ('$' in number_format) or ('[$' in number_format) or ('accounting' in number_format.lower())

    def cell_is_excel_date(self, value, cell=None):
        if isinstance(value, (datetime, date)):
            return True
        if cell is None:
            return False
        try:
            if getattr(cell, 'is_date', False):
                return True
        except Exception:
            pass
        if is_date_format is not None:
            try:
                return bool(is_date_format(getattr(cell, 'number_format', '') or ''))
            except Exception:
                pass
        return False

    def excel_date_to_date(self, value, cell=None):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if value is None:
            return None
        if self.cell_is_excel_date(value, cell) and isinstance(value, (int, float)) and from_excel is not None:
            try:
                epoch = getattr(getattr(cell, 'parent', None), 'epoch', None) if cell is not None else None
                converted = from_excel(value, epoch=epoch) if epoch is not None else from_excel(value)
                if isinstance(converted, datetime):
                    return converted.date()
                if isinstance(converted, date):
                    return converted
            except Exception:
                pass
        return None

    def try_parse_date_value(self, value, cell=None):
        direct_date = self.excel_date_to_date(value, cell)
        if direct_date is not None:
            return direct_date
        if value is None:
            return None

        text = str(value).strip()
        if not text:
            return None

        for fmt in (
            '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',
            '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
            '%d/%m/%y', '%d-%m-%y',
            '%d %b %Y', '%d %B %Y',
        ):
            try:
                return datetime.strptime(text, fmt).date()
            except Exception:
                pass

        return None

    def try_parse_numeric_value(self, value):
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if value is None:
            return None

        text = str(value).strip()
        if not text:
            return None

        cleaned = text.replace('$', '').replace(',', '').strip()
        try:
            return float(cleaned)
        except Exception:
            return None

    def get_context_cell_value(self, context, row_index, column_index):
        if context is None or row_index is None or column_index is None:
            return None

        if hasattr(context, 'cell'):
            try:
                return context.cell(row=row_index, column=column_index).value
            except Exception:
                return None

        if isinstance(context, list):
            list_row_index = row_index - 1
            list_col_index = column_index - 1
            if 0 <= list_row_index < len(context):
                row = context[list_row_index] or []
                if 0 <= list_col_index < len(row):
                    return row[list_col_index]
        return None

    def is_blank_context_row(self, context, row_index):
        if context is None or row_index is None:
            return False

        if hasattr(context, 'cell'):
            try:
                max_columns = min(max(1, getattr(context, 'max_column', 1) or 1), self.MAX_COLS)
            except Exception:
                max_columns = self.MAX_COLS
            for column_index in range(1, max_columns + 1):
                value = self.get_context_cell_value(context, row_index, column_index)
                if str(value or '').strip():
                    return False
            return True

        if isinstance(context, list):
            list_row_index = row_index - 1
            if 0 <= list_row_index < len(context):
                return all(not str(value or '').strip() for value in (context[list_row_index] or []))
        return False

    def customer_price_header_applies(self, row_index, column_index, context):
        if not self.is_customer_file() or row_index is None:
            return False

        for scan_row in range(row_index, 0, -1):
            if scan_row < row_index and self.is_blank_context_row(context, scan_row):
                return False

            raw_value = self.get_context_cell_value(context, scan_row, column_index)
            text_value = str(raw_value or '').strip()
            if not text_value:
                continue

            normalized_header = self.normalize_header(text_value)
            if normalized_header in {'price', 'unitprice', 'sellprice'}:
                return True

            if self.try_parse_numeric_value(raw_value) is not None:
                continue

            return False

        return False

    def format_customer_file_value(self, value, column_index, cell=None, row_index=None, context=None):
        if not self.is_customer_file():
            return None

        number = self.try_parse_numeric_value(value)
        if number is not None:
            if self.customer_price_header_applies(row_index, column_index, context) or (cell is not None and self.cell_has_currency_format(cell)):
                return f'${number:,.2f}'

        dt = self.try_parse_date_value(value, cell)
        if dt is not None and (column_index == 7 or self.cell_is_excel_date(value, cell)):
            return dt.strftime('%d/%m/%Y')

        return None

    def cell_display_text(self, value, column_index=None, cell=None, row_index=None, context=None):
        formatted = (
            self.format_customer_file_value(value, column_index, cell, row_index=row_index, context=context)
            if column_index is not None else None
        )
        if formatted is not None:
            return formatted
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        return str(value)

    def qt_color_from_openpyxl(self, color_obj):
        if color_obj is None:
            return QColor()

        rgb = getattr(color_obj, 'rgb', None)
        if rgb:
            rgb = str(rgb).strip()
            if len(rgb) == 8:
                rgb = rgb[2:]
            if len(rgb) == 6:
                color = QColor(f'#{rgb}')
                if color.isValid():
                    return color

        indexed = getattr(color_obj, 'indexed', None)
        if indexed is not None and COLOR_INDEX is not None:
            try:
                indexed_value = COLOR_INDEX[indexed]
                if indexed_value:
                    indexed_value = str(indexed_value).strip()
                    if len(indexed_value) == 8:
                        indexed_value = indexed_value[2:]
                    if len(indexed_value) == 6:
                        color = QColor(f'#{indexed_value}')
                        if color.isValid():
                            return color
            except Exception:
                pass

        value = getattr(color_obj, 'value', None)
        if isinstance(value, str):
            value = value.strip()
            if len(value) == 8:
                value = value[2:]
            if len(value) == 6:
                color = QColor(f'#{value}')
                if color.isValid():
                    return color

        return QColor()

    def apply_excel_cell_style(self, item, cell):
        font = item.font()
        if cell.font is not None:
            if cell.font.name:
                font.setFamily(cell.font.name)
            if cell.font.sz:
                try:
                    font.setPointSizeF(float(cell.font.sz))
                except Exception:
                    pass
            font.setBold(bool(cell.font.bold))
            font.setItalic(bool(cell.font.italic))
            font.setUnderline(bool(cell.font.underline and cell.font.underline != 'none'))
            font.setStrikeOut(bool(cell.font.strike))
        item.setFont(font)

        foreground = self.qt_color_from_openpyxl(getattr(cell.font, 'color', None))
        if foreground.isValid():
            item.setForeground(QBrush(foreground))

        fill = getattr(cell, 'fill', None)
        fill_color = QColor()
        if fill is not None and getattr(fill, 'fill_type', None) not in (None, ''):
            for color_attr in ('fgColor', 'start_color', 'bgColor', 'end_color'):
                fill_color = self.qt_color_from_openpyxl(getattr(fill, color_attr, None))
                if fill_color.isValid():
                    break
        if fill_color.isValid():
            item.setBackground(QBrush(fill_color))

        if cell.alignment is not None:
            horizontal_map = {
                'left': Qt.AlignLeft,
                'center': Qt.AlignHCenter,
                'centerContinuous': Qt.AlignHCenter,
                'right': Qt.AlignRight,
                'justify': Qt.AlignJustify,
                'distributed': Qt.AlignJustify,
                'fill': Qt.AlignLeft,
                'general': Qt.AlignLeft,
            }
            vertical_map = {
                'top': Qt.AlignTop,
                'center': Qt.AlignVCenter,
                'bottom': Qt.AlignBottom,
                'justify': Qt.AlignVCenter,
                'distributed': Qt.AlignVCenter,
            }
            alignment = horizontal_map.get(cell.alignment.horizontal, Qt.AlignLeft) | vertical_map.get(cell.alignment.vertical, Qt.AlignVCenter)
            item.setTextAlignment(alignment)

        if self.try_parse_numeric_value(cell.value) is not None:
            current_alignment = item.textAlignment()
            if current_alignment & (Qt.AlignHCenter | Qt.AlignLeft | Qt.AlignRight) == 0:
                item.setTextAlignment((current_alignment & ~Qt.AlignLeft) | Qt.AlignRight | Qt.AlignVCenter)

    def convert_text_for_excel_save(self, text, original_value):
        stripped = text.strip()
        if stripped == '':
            return None
        if stripped.startswith('='):
            return stripped

        if isinstance(original_value, bool):
            return stripped.lower() in {'1', 'true', 'yes', 'y'}
        if isinstance(original_value, int) and not isinstance(original_value, bool):
            try:
                if '.' in stripped:
                    value = float(stripped.replace(',', '').replace('$', ''))
                    return int(value) if value.is_integer() else value
                return int(stripped.replace(',', '').replace('$', ''))
            except Exception:
                return text
        if isinstance(original_value, float):
            try:
                return float(stripped.replace(',', '').replace('$', ''))
            except Exception:
                return text
        if isinstance(original_value, datetime):
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M', '%d/%m/%Y %H:%M:%S'):
                try:
                    return datetime.strptime(stripped, fmt)
                except Exception:
                    pass
            return text
        if isinstance(original_value, date):
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    return datetime.strptime(stripped, fmt).date()
                except Exception:
                    pass
            return text
        return text

    def save_excel_file(self):
        if self.workbook is None:
            return

        for sheet_name, changed_cells in self.changed_excel_cells.items():
            sheet = self.workbook[sheet_name]
            for (row, col), text in changed_cells.items():
                cell = sheet.cell(row=row, column=col)
                cell.value = self.convert_text_for_excel_save(text, cell.value)

        self.workbook.save(str(self.path))

    def save_csv_file(self):
        import csv

        if self.csv_table is None:
            return

        row_count = self.csv_table.rowCount()
        col_count = self.csv_table.columnCount()

        while len(self.csv_rows) < row_count:
            self.csv_rows.append([])

        for row_idx in range(row_count):
            while len(self.csv_rows[row_idx]) < col_count:
                self.csv_rows[row_idx].append('')
            for col_idx in range(col_count):
                item = self.csv_table.item(row_idx, col_idx)
                self.csv_rows[row_idx][col_idx] = item.text() if item is not None else ''

        with open(self.path, 'w', newline='', encoding='utf-8-sig') as handle:
            csv.writer(handle).writerows(self.csv_rows)

    def save_text_file(self):
        if self.text_editor is None:
            return
        self.path.write_text(self.text_editor.toPlainText(), encoding='utf-8')

    def save_and_close(self):
        try:
            if self.file_type == 'excel' and self.is_dirty:
                self.save_excel_file()
            elif self.file_type == 'csv' and self.is_dirty:
                self.save_csv_file()
            elif self.file_type == 'text' and self.is_dirty:
                self.save_text_file()
        except Exception as exc:
            QMessageBox.warning(self, 'Save File', f'Could not save file:\n{self.path}\n\n{exc}')
            return

        self.is_dirty = False
        self.update_title()
        self.accept()

    def reject(self):
        if self.is_dirty:
            answer = QMessageBox.question(
                self,
                'Discard changes?',
                'There are unsaved changes in this viewer. Close without saving?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                return
        super().reject()


class FrozenColumnsHelper:
    def __init__(self, main_view, frozen_columns, click_handler=None, double_click_handler=None):
        self.main_view = main_view
        self.frozen_columns = max(0, int(frozen_columns or 0))
        self.click_handler = click_handler
        self.double_click_handler = double_click_handler
        self.wrapper = None
        self.frozen_view = None
        self._syncing_section_width = False
        self._syncing_row_height = False
        self._install()

    def _install(self):
        if self.main_view is None or self.frozen_columns <= 0:
            return

        parent = self.main_view.parentWidget()
        if parent is None:
            return
        layout = parent.layout()
        if layout is None:
            return

        insert_index = None
        for i in range(layout.count()):
            item = layout.itemAt(i)
            if item is not None and item.widget() is self.main_view:
                insert_index = i
                break
        if insert_index is None:
            return

        layout.removeWidget(self.main_view)

        self.wrapper = QWidget(parent)
        wrapper_layout = QHBoxLayout(self.wrapper)
        wrapper_layout.setContentsMargins(0, 0, 0, 0)
        wrapper_layout.setSpacing(0)

        self.frozen_view = QTableView(self.wrapper)
        self.frozen_view.setObjectName(f"{self.main_view.objectName()}_frozen")
        self.frozen_view.setFocusPolicy(Qt.NoFocus)
        self.frozen_view.setSelectionBehavior(self.main_view.selectionBehavior())
        self.frozen_view.setSelectionMode(self.main_view.selectionMode())
        self.frozen_view.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.frozen_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.frozen_view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.frozen_view.setHorizontalScrollMode(self.main_view.horizontalScrollMode())
        self.frozen_view.setVerticalScrollMode(self.main_view.verticalScrollMode())
        self.frozen_view.setWordWrap(self.main_view.wordWrap())
        self.frozen_view.setAlternatingRowColors(self.main_view.alternatingRowColors())
        self.frozen_view.setShowGrid(self.main_view.showGrid())
        self.frozen_view.setStyleSheet(self.main_view.styleSheet())
        self.frozen_view.setFont(self.main_view.font())
        self.frozen_view.setSortingEnabled(self.main_view.isSortingEnabled())
        self.copy_header_settings()

        self.main_view.setParent(self.wrapper)
        wrapper_layout.addWidget(self.frozen_view, 0)
        wrapper_layout.addWidget(self.main_view, 1)
        layout.insertWidget(insert_index, self.wrapper, 1)

        self.main_view.verticalScrollBar().valueChanged.connect(self.frozen_view.verticalScrollBar().setValue)
        self.frozen_view.verticalScrollBar().valueChanged.connect(self.main_view.verticalScrollBar().setValue)
        self.main_view.verticalHeader().sectionResized.connect(self.update_row_height_from_main)
        self.frozen_view.verticalHeader().sectionResized.connect(self.update_row_height_from_frozen)
        self.main_view.horizontalHeader().sectionResized.connect(self.update_section_width_from_main)
        self.frozen_view.horizontalHeader().sectionResized.connect(self.update_section_width_from_frozen)
        self.main_view.horizontalHeader().geometriesChanged.connect(self.refresh)
        self.frozen_view.horizontalHeader().geometriesChanged.connect(self.refresh)

        if self.click_handler is not None:
            self.frozen_view.clicked.connect(self._forward_click)
        if self.double_click_handler is not None:
            self.frozen_view.doubleClicked.connect(self._forward_double_click)

        self.rebind()

    def copy_header_settings(self):
        if self.frozen_view is None or self.main_view is None:
            return

        main_h = self.main_view.horizontalHeader()
        frozen_h = self.frozen_view.horizontalHeader()
        frozen_h.setVisible(main_h.isVisible())
        frozen_h.setDefaultAlignment(main_h.defaultAlignment())
        frozen_h.setHighlightSections(main_h.highlightSections())
        frozen_h.setSectionsClickable(main_h.sectionsClickable())
        frozen_h.setSortIndicatorShown(main_h.isSortIndicatorShown())
        frozen_h.setCascadingSectionResizes(main_h.cascadingSectionResizes())
        frozen_h.setMinimumSectionSize(main_h.minimumSectionSize())
        frozen_h.setDefaultSectionSize(main_h.defaultSectionSize())
        frozen_h.setStretchLastSection(False)
        frozen_h.setSectionsMovable(False)

        main_v = self.main_view.verticalHeader()
        frozen_v = self.frozen_view.verticalHeader()
        frozen_v.setVisible(main_v.isVisible())
        frozen_v.setDefaultAlignment(main_v.defaultAlignment())
        frozen_v.setHighlightSections(main_v.highlightSections())
        frozen_v.setSectionsClickable(main_v.sectionsClickable())
        frozen_v.setMinimumSectionSize(main_v.minimumSectionSize())
        frozen_v.setDefaultSectionSize(main_v.defaultSectionSize())
        frozen_v.setStretchLastSection(False)

    def _forward_click(self, index):
        if self.click_handler is not None and index.isValid():
            self.click_handler(index)

    def _forward_double_click(self, index):
        if self.double_click_handler is not None and index.isValid():
            self.double_click_handler(index)

    def rebind(self):
        if self.frozen_view is None or self.main_view is None:
            return
        model = self.main_view.model()
        self.frozen_view.setModel(model)
        if self.main_view.selectionModel() is not None:
            self.frozen_view.setSelectionModel(self.main_view.selectionModel())
        self.copy_header_settings()
        self.refresh()

    def refresh(self):
        if self.frozen_view is None or self.main_view is None:
            return
        model = self.main_view.model()
        if model is None:
            return

        self.copy_header_settings()

        column_count = model.columnCount()
        row_count = model.rowCount()
        main_header = self.main_view.horizontalHeader()
        frozen_header = self.frozen_view.horizontalHeader()

        for column in range(column_count):
            freeze_this = column < self.frozen_columns
            self.frozen_view.setColumnHidden(column, not freeze_this)
            self.main_view.setColumnHidden(column, freeze_this)
            try:
                frozen_header.setSectionResizeMode(column, main_header.sectionResizeMode(column))
            except Exception:
                pass
            if freeze_this:
                width = self.main_view.columnWidth(column)
                self.frozen_view.setColumnWidth(column, width)

        for row in range(row_count):
            self.frozen_view.setRowHidden(row, self.main_view.isRowHidden(row))
            self.frozen_view.setRowHeight(row, self.main_view.rowHeight(row))

        self.frozen_view.verticalHeader().setDefaultSectionSize(self.main_view.verticalHeader().defaultSectionSize())
        self.frozen_view.horizontalHeader().setDefaultSectionSize(self.main_view.horizontalHeader().defaultSectionSize())
        self.frozen_view.horizontalHeader().setStretchLastSection(False)
        self.frozen_view.horizontalHeader().viewport().update()
        self.frozen_view.viewport().update()

        frozen_width = 0
        for column in range(min(self.frozen_columns, column_count)):
            if not self.frozen_view.isColumnHidden(column):
                frozen_width += self.frozen_view.columnWidth(column)
        frozen_width += self.frozen_view.frameWidth() * 2
        if self.frozen_view.verticalHeader().isVisible():
            frozen_width += self.frozen_view.verticalHeader().width()
        self.frozen_view.setFixedWidth(max(0, frozen_width))

    def update_section_width_from_main(self, logical_index, _old_size, new_size):
        if self.frozen_view is None or logical_index >= self.frozen_columns or self._syncing_section_width:
            return
        self._syncing_section_width = True
        try:
            self.frozen_view.setColumnWidth(logical_index, new_size)
        finally:
            self._syncing_section_width = False
        self.refresh()

    def update_section_width_from_frozen(self, logical_index, _old_size, new_size):
        if self.main_view is None or logical_index >= self.frozen_columns or self._syncing_section_width:
            return
        self._syncing_section_width = True
        try:
            self.main_view.setColumnWidth(logical_index, new_size)
        finally:
            self._syncing_section_width = False
        self.refresh()

    def update_row_height_from_main(self, logical_index, _old_size, new_size):
        if self.frozen_view is None or self._syncing_row_height:
            return
        self._syncing_row_height = True
        try:
            self.frozen_view.setRowHeight(logical_index, new_size)
        finally:
            self._syncing_row_height = False

    def update_row_height_from_frozen(self, logical_index, _old_size, new_size):
        if self.main_view is None or self._syncing_row_height:
            return
        self._syncing_row_height = True
        try:
            self.main_view.setRowHeight(logical_index, new_size)
        finally:
            self._syncing_row_height = False


class MainWindow(QMainWindow):
    STATE_SUFFIX_RE = re.compile(
        r"\s*(?:-\s*)?(N\.S\.W|N\.S\.W\.|NSW|VIC|QLD|WA|SA|TAS|NT|ACT|NEW SOUTH WALES|VICTORIA|QUEENSLAND|WESTERN AUSTRALIA|SOUTH AUSTRALIA|TASMANIA|NORTHERN TERRITORY|AUSTRALIAN CAPITAL TERRITORY)\s*$",
        re.IGNORECASE,
    )

    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setMinimumSize(1600, 900)
        self.resize(1600, 900)
        self.setMaximumSize(16777215, 16777215)
        sp = self.sizePolicy()
        sp.setHorizontalPolicy(QSizePolicy.Expanding)
        sp.setVerticalPolicy(QSizePolicy.Expanding)
        self.setSizePolicy(sp)


        self.app_dir = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
        self.base_dir = Path(getattr(sys, "_MEIPASS", str(self.app_dir))).resolve() if getattr(sys, "frozen", False) else self.app_dir
        self.settings = QSettings("Windsor", "WidgetApp")
        self.db_conn = None
        self.db_engine = "sqlserver"
        self.db_config = {}
        self.detached_windows = []

        self.customer_names = []
        self.item_numbers = []
        self.supplier_names = []
        self.current_customer_months = []
        self.current_customer_pivot = {}
        self.current_customer_name = None
        self.current_customer_file_path = None
        self.current_item_number = None
        self._last_item_total_value = 0.0
        self.customer_file_preview_dialog = None
        self.sales_table_frozen_helper = None
        self.customer_purchase_frozen_helper = None

        self.customer_start_picker = getattr(self.ui, "startMonthPickerCustomer_2", None)
        self.customer_end_picker = getattr(self.ui, "endMonthPickerCustomer", None)
        self.item_start_picker = getattr(self.ui, "startMonthPickerItem", None)
        self.item_end_picker = getattr(self.ui, "endMonthPickerItem", None)
        self.lead_time_picker = getattr(self.ui, "leadTimePicker", None)
        self.customer_chart_view = None
        self.item_chart_view = None
        self._logo_label = None
        self.order_item_completer = None
        self.order_analysis_supplier_completer = None
        self.saba_customer_completer = None
        self.saba_show_all_checkbox = None
        self.combine_threads_checkbox = None
        self.current_saba_customer = None
        self.current_order_analysis_supplier = None
        self._updating_order_table = False
        self._updating_on_order_table = False
        self.order_on_order_column = 2
        self.order_qty_column = 3
        self.order_supplier_column = 4
        self.order_priority_column = 5
        self.order_status_column = self.order_priority_column
        self.order_remove_column = 6
        self._updating_container_table = False
        self._spin_arrow_icon_cache = {}
        self.container_item_completer = None
        self.current_container_ref = None
        self.current_container_notes = ""
        self.shipments_window = None
        self.shipments_button = None
        self.left_shipments_button = None
        self.on_order_button = None
        self.yu_order_entry_dialog = None
        self.yu_order_review_windows = []
        self.yu_order_review_db_helpers = []
        self.item_summary_editable_fields = {
            "rollSpool_box": {"column": "roll", "label": "Roll / Spool"},
            "mtUnit_box": {"column": "per_roll", "label": "Mt / Unit"},
            "box_box": {"column": "carton", "label": "Box"},
            "palletCarton_box": {"column": "pallet", "label": "Pallet / Carton"},
        }
        self.container_columns = {
            "order": 0,
            "item": 1,
            "description": 2,
            "qty": 3,
            "cartons": 4,
            "additional_cartons": 5,
            "urgent": 6,
            "additional": 7,
            "remove": 8,
        }
        self.container_sort_column = None
        self.container_sort_descending = False
        self.order_analysis_columns = {
            "item_number": 0,
            "item_name": 1,
            "sales_for_period": 2,
            "avg_monthly_sales": 3,
            "soh": 4,
            "stock_on_order": 5,
            "on_order_form": 6,
            "on_next_container": 7,
            "shipped_container": 8,
            "suggested_order": 9,
            "at_risk": 10,
        }

        self.setup_navigation()
        self.add_left_shipments_nav_button()
        self.setup_theme_controls()
        self.setup_chart_views()

        self.open_database()
        self.ensure_to_order_lines_table()
        self.ensure_on_order_lines_table()
        self.ensure_on_order_meta_table()
        self.ensure_supplier_master_table()
        self.ensure_shipment_tables()
        self.load_reference_lists()

        self.setup_customer_autocomplete()
        self.setup_item_autocomplete()
        self.setup_item_summary_combine_checkbox()
        self.setup_supplier_autocomplete()
        self.setup_date_ranges()
        self.setup_sales_table()
        self.setup_customer_info_table()
        self.setup_customer_purchase_table()
        self.setup_update_page()
        self.setup_order_table()
        self.setup_on_order_page()
        self.load_saved_order_lines()
        self.load_saved_on_order_lines()
        self.load_on_order_general_comments()
        self.setup_order_analysis_table()
        self.setup_saba_review_page()
        self.setup_container_table()
        self.setup_containers_list_table()
        self.add_shipments_button()
        self.setup_order_entry()
        self.setup_container_entry()
        self.setup_logo()
        self.update_version_display()
        self.clear_item_summary_fields()
        self.update_charge_freight_box(False)

        self.connect_signals()
        self.install_all_table_font_menus()
        self.restore_theme()

    def install_all_table_font_menus(self):
        tables = []
        seen_ids = set()
        for table_class in (QTableWidget, QTableView):
            for table in self.findChildren(table_class):
                if table is None or id(table) in seen_ids:
                    continue
                seen_ids.add(id(table))
                object_name = str(table.objectName() or "").strip()
                if not object_name or object_name.startswith("qt_") or object_name.endswith("_frozen"):
                    continue
                tables.append(table)

        for table in tables:
            install_table_font_context_menu(
                self,
                table,
                self.settings,
                "main_window",
                settings_token=f"main_window/{table.objectName()}",
            )

    # -----------------------------
    # Setup helpers
    # -----------------------------
    def setup_navigation(self):
        if not hasattr(self.ui, "stackedWidget"):
            return

        nav_map = [
            ("customerSummary_button", "customerSummary_page"),
            ("itemSummary_button", "itemSummary_page"),
            ("toOrderSheet_button", "toOrderSheet_page"),
            ("onOrder_button", "onOrder_page"),
            ("buildContainerSheet_button", "buildContainer_page"),
            ("orderAnalasys_button", "orderAnalysy_page"),
            ("sabaReview_button", "sabaReview_page"),
            ("updateData_button", "update_page"),
        ]

        self.nav_buttons = []
        self.nav_button_page_map = {}
        for button_name, page_name in nav_map:
            button = getattr(self.ui, button_name, None)
            page = getattr(self.ui, page_name, None)
            if button is not None and page is not None:
                self.nav_buttons.append((button, page))
                self.nav_button_page_map[button] = page_name
                button.clicked.connect(lambda _checked=False, p=page: self.ui.stackedWidget.setCurrentWidget(p))
                self.install_left_nav_context_menu(button, page_name, button.text())

        try:
            self.ui.stackedWidget.currentChanged.connect(self.handle_stacked_widget_changed)
        except Exception:
            pass
        self.handle_stacked_widget_changed()

    def update_version_display(self):
        version_widget = getattr(self.ui, "textEdit", None)
        if version_widget is None:
            return
        try:
            version_widget.setReadOnly(True)
        except Exception:
            pass
        html = (
            "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" "
            "\"http://www.w3.org/TR/REC-html40/strict.dtd\">"
            "<html><head><meta charset=\"utf-8\" />"
            "<style type=\"text/css\">p, li { white-space: pre-wrap; }</style>"
            "</head><body style=\" font-family:'Segoe UI'; font-size:9pt; font-weight:400; font-style:normal;\">"
            f"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; "
            f"-qt-block-indent:0; text-indent:0px;\">Widget Version {APP_VERSION}<br/>"
            f"Written and Designed by {APP_DESIGNER}</p></body></html>"
        )
        try:
            version_widget.setHtml(html)
        except Exception:
            try:
                version_widget.setPlainText(f"Widget Version {APP_VERSION}\nWritten and Designed by {APP_DESIGNER}")
            except Exception:
                pass

    def install_left_nav_context_menu(self, button, page_name, button_label=""):
        if button is None:
            return
        button.setContextMenuPolicy(Qt.CustomContextMenu)
        button.customContextMenuRequested.connect(
            lambda pos, b=button, p=page_name, t=button_label: self.show_left_nav_context_menu(b, pos, p, t)
        )

    def show_left_nav_context_menu(self, button, pos, page_name, button_label=""):
        if button is None:
            return
        menu = QMenu(button)
        label_text = (button_label or page_name or "page").strip()
        open_action = menu.addAction(f"Open {label_text} in new window")
        chosen = menu.exec(button.mapToGlobal(pos))
        if chosen == open_action:
            self.open_page_in_new_window(page_name, label_text)

    def register_detached_window(self, window):
        if window is None:
            return
        self.detached_windows.append(window)
        try:
            window.destroyed.connect(lambda *_args, w=window: self._forget_detached_window(w))
        except Exception:
            pass

    def _forget_detached_window(self, window):
        self.detached_windows = [w for w in getattr(self, "detached_windows", []) if w is not window]

    def open_page_in_new_window(self, page_name, page_label=""):
        detached = MainWindow()
        self.register_detached_window(detached)
        detached.show()
        target_page = getattr(detached.ui, page_name, None)
        stacked_widget = getattr(detached.ui, "stackedWidget", None)
        if stacked_widget is not None and target_page is not None:
            stacked_widget.setCurrentWidget(target_page)
            try:
                detached.handle_stacked_widget_changed()
            except Exception:
                pass
        label_text = (page_label or "").strip()
        if label_text:
            detached.setWindowTitle(f"Windsor Widget - {label_text}")
        return detached

    def open_shipments_window_detached(self):
        detached = ShipmentsWindow(self)
        self.register_detached_window(detached)
        detached.show()
        try:
            detached.raise_()
            detached.activateWindow()
        except Exception:
            pass
        return detached

    def navigation_selected_style(self):
        if hasattr(self.ui, "radioHighContrast") and self.ui.radioHighContrast.isChecked():
            return "background: #005a9e; color: white; border: 2px solid cyan; font-weight: 700;"
        if hasattr(self.ui, "radioLight") and self.ui.radioLight.isChecked():
            return "background: #6d6d6d; color: white; border: 1px solid #4f4f4f; font-weight: 700;"
        return "background: #b0b4b9; color: #111111; border: 1px solid #8a8d91; font-weight: 700;"

    def update_navigation_button_highlight(self, *_args):
        stacked_widget = getattr(self.ui, "stackedWidget", None)
        current_page = stacked_widget.currentWidget() if stacked_widget is not None else None
        for button, page in getattr(self, "nav_buttons", []):
            if current_page is not None and page is current_page:
                button.setStyleSheet(self.navigation_selected_style())
            else:
                button.setStyleSheet("")

    def handle_stacked_widget_changed(self, *_args):
        self.update_navigation_button_highlight()
        self.refresh_page_data_on_navigation()

    def refresh_page_data_on_navigation(self):
        stacked_widget = getattr(self.ui, "stackedWidget", None)
        current_page = stacked_widget.currentWidget() if stacked_widget is not None else None
        if current_page is None or self.db_conn is None:
            return

        to_order_page = getattr(self.ui, "toOrderSheet_page", None)
        build_container_page = getattr(self.ui, "buildContainer_page", None)

        if current_page in {to_order_page, build_container_page}:
            self.load_reference_lists()
            self.setup_customer_autocomplete()
            self.setup_item_autocomplete()
            self.setup_supplier_autocomplete()
            self.setup_order_entry()
            self.setup_container_entry()

        if current_page is to_order_page:
            self.load_saved_order_lines()
            self.refresh_item_summary_context_boxes()
            return

        if current_page is build_container_page:
            self.refresh_containers_list()
            if not self.reload_current_container_state():
                self.refresh_container_note_rows()
                self.refresh_container_totals()
                self.refresh_item_summary_context_boxes()

    def add_left_shipments_nav_button(self):
        if getattr(self, "left_shipments_button", None) is not None:
            return

        layout = getattr(self.ui, "verticalLayout_4", None)
        order_button = getattr(self.ui, "orderAnalasys_button", None)
        saba_button = getattr(self.ui, "sabaReview_button", None)
        frame = getattr(self.ui, "frame_7", None)
        if layout is None:
            return

        button = QPushButton("Shipments", frame or self)
        button.setObjectName("leftShipments_button")
        if order_button is not None:
            try:
                button.setFont(order_button.font())
                button.setSizePolicy(order_button.sizePolicy())
                button.setMinimumSize(order_button.minimumSize())
                button.setMaximumSize(order_button.maximumSize())
            except Exception:
                pass
        button.clicked.connect(self.open_shipments_window)

        insert_index = -1
        if order_button is not None:
            insert_index = layout.indexOf(order_button)
        elif saba_button is not None:
            insert_index = layout.indexOf(saba_button) - 1

        if insert_index >= 0:
            layout.insertWidget(insert_index + 1, button)
        else:
            layout.addWidget(button)

        self.left_shipments_button = button

    def add_shipments_button(self):
        if getattr(self, "shipments_button", None) is not None:
            return

        layout = getattr(self.ui, "horizontalLayout_18", None)
        table = getattr(self.ui, "containers_tableWidget", None)
        frame = getattr(self.ui, "frame_46", None)
        if layout is None or table is None:
            return

        button = QPushButton("Shipments", frame or self)
        button.setObjectName("shipments_pushButton")
        export_button = getattr(self.ui, "exportEmail_pushButton", None)
        if export_button is not None:
            button.setFont(export_button.font())
        button.setMinimumWidth(120)
        button.clicked.connect(self.open_shipments_window)

        insert_index = layout.indexOf(table)
        if insert_index >= 0:
            layout.insertWidget(insert_index + 1, button)
        else:
            layout.addWidget(button)

        self.shipments_button = button

    def show_shipments_nav_context_menu(self, button, pos):
        if button is None:
            return
        menu = QMenu(button)
        open_action = menu.addAction("Open Shipments in new window")
        chosen = menu.exec(button.mapToGlobal(pos))
        if chosen == open_action:
            self.open_shipments_window_detached()

    def open_shipments_window(self):
        needs_new_window = self.shipments_window is None
        if not needs_new_window:
            try:
                self.shipments_window.isVisible()
                table = self.shipments_window.table_widget()
                if table is not None:
                    table.rowCount()
            except RuntimeError:
                self.shipments_window = None
                needs_new_window = True

        if needs_new_window:
            self.shipments_window = ShipmentsWindow(self)
        else:
            self.shipments_window.refresh_from_database()

        self.shipments_window.show()
        if self.shipments_window.isMinimized():
            self.shipments_window.showNormal()
        self.shipments_window.raise_()
        self.shipments_window.activateWindow()

    def setup_theme_controls(self):
        if hasattr(self.ui, "radioLight"):
            self.ui.radioLight.toggled.connect(lambda checked: checked and self.apply_theme("light"))
        if hasattr(self.ui, "radioDark"):
            self.ui.radioDark.toggled.connect(lambda checked: checked and self.apply_theme("dark"))
        if hasattr(self.ui, "radioHighContrast"):
            self.ui.radioHighContrast.toggled.connect(lambda checked: checked and self.apply_theme("high"))

        if self.lead_time_picker is not None:
            self.lead_time_picker.setRange(1, 18)
            self.lead_time_picker.setSingleStep(1)
            self.lead_time_picker.setValue(14)
            self.lead_time_picker.setAlignment(Qt.AlignCenter)
            self.lead_time_picker.setToolTip("Lead time in weeks")

        additional_spinner = getattr(self.ui, 'additional_spinner', None)
        if additional_spinner is not None:
            additional_spinner.setRange(0, 9999)
            additional_spinner.setSingleStep(1)
            additional_spinner.setAlignment(Qt.AlignCenter)

        for spin in (self.lead_time_picker, additional_spinner):
            if spin is not None:
                spin.setButtonSymbols(QAbstractSpinBox.UpDownArrows)
                spin.setAccelerated(True)
                spin.update()

    def setup_chart_views(self):
        self.customer_chart_view = self.create_chart_view(getattr(self.ui, "sales_graph", None))
        self.item_chart_view = self.create_chart_view(getattr(self.ui, "itemSales_graph", None))

    def create_chart_view(self, host_widget):
        if host_widget is None:
            return None

        layout = host_widget.layout()
        if layout is None:
            layout = QVBoxLayout(host_widget)
            layout.setContentsMargins(0, 0, 0, 0)
        else:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()

        chart_view = QChartView(host_widget)
        chart_view.setRenderHint(QPainter.Antialiasing)
        layout.addWidget(chart_view)
        return chart_view

    def connect_signals(self):
        if hasattr(self.ui, "searchButton"):
            self.ui.searchButton.clicked.connect(self.search_customer_sales)
        if hasattr(self.ui, "customerEdit"):
            self.ui.customerEdit.returnPressed.connect(self.search_customer_sales)
        if hasattr(self.ui, "combineStateAccountsCheck"):
            self.ui.combineStateAccountsCheck.stateChanged.connect(self.rerun_search_if_ready)
        if hasattr(self.ui, "salesTable"):
            self.ui.salesTable.cellClicked.connect(self.show_customer_row_chart)
            self.ui.salesTable.cellDoubleClicked.connect(self.handle_customer_table_double_click)

        if hasattr(self.ui, "loadItem"):
            self.ui.loadItem.clicked.connect(self.load_item_summary)
        if self.combine_threads_checkbox is not None:
            self.combine_threads_checkbox.stateChanged.connect(self.rerun_item_if_ready)

        supplier_edit = self.get_order_analysis_supplier_edit()
        if supplier_edit is not None:
            supplier_edit.returnPressed.connect(self.load_order_analysis)
            completer = supplier_edit.completer()
            if completer is not None:
                completer.activated.connect(lambda *_args: self.load_order_analysis())

        order_analysis_table = self.get_order_analysis_table()
        if order_analysis_table is not None:
            order_analysis_table.cellDoubleClicked.connect(self.handle_order_analysis_table_double_click)

        saba_customer_edit = self.get_saba_customer_edit()
        if saba_customer_edit is not None:
            saba_customer_edit.returnPressed.connect(self.load_saba_review)
            completer = saba_customer_edit.completer()
            if completer is not None:
                completer.activated.connect(lambda *_args: self.load_saba_review())
        if self.saba_show_all_checkbox is not None:
            self.saba_show_all_checkbox.stateChanged.connect(self.handle_saba_all_customers_changed)

        if hasattr(self.ui, "enterItem"):
            self.ui.enterItem.returnPressed.connect(self.load_item_summary)
            if self.ui.enterItem.completer() is not None:
                self.ui.enterItem.completer().activated.connect(lambda *_args: self.load_item_summary())

        if self.customer_start_picker:
            self.customer_start_picker.dateChanged.connect(self.rerun_search_if_ready)
        if self.customer_end_picker:
            self.customer_end_picker.dateChanged.connect(self.rerun_search_if_ready)
        if self.item_start_picker:
            self.item_start_picker.dateChanged.connect(self.rerun_item_if_ready)
        if self.item_end_picker:
            self.item_end_picker.dateChanged.connect(self.rerun_item_if_ready)
        if self.lead_time_picker:
            self.lead_time_picker.valueChanged.connect(self.rerun_item_if_ready)
        if hasattr(self.ui, "loadFile"):
            self.ui.loadFile.clicked.connect(self.open_customer_file)
        customer_purchase_table = getattr(self.ui, "customerPurchase_table", None)
        if customer_purchase_table is not None:
            customer_purchase_table.doubleClicked.connect(self.handle_customer_purchase_double_click)

        container_table = getattr(self.ui, "container_table", None)
        if container_table is not None:
            container_table.itemChanged.connect(self.handle_container_table_item_changed)
            container_table.cellDoubleClicked.connect(self.handle_container_table_double_click)

        item_number_container = getattr(self.ui, "itemNumberContainer_line", None)
        if item_number_container is not None:
            item_number_container.textChanged.connect(self.refresh_item_summary_context_boxes)
            item_number_container.returnPressed.connect(self.resolve_container_item_and_focus_qty)

        qty_container = getattr(self.ui, "qtyContainder_line", None)
        if qty_container is not None:
            qty_container.textChanged.connect(self.refresh_item_summary_context_boxes)
            qty_container.returnPressed.connect(self.add_container_line_from_inputs)

        order_number_container = getattr(self.ui, "orderNumberContainer_line", None)
        if order_number_container is not None:
            order_number_container.returnPressed.connect(self.focus_container_item_input)

        eta_date_edit = getattr(self.ui, "eta_dateEdit", None)
        if eta_date_edit is not None:
            eta_date_edit.dateChanged.connect(self.handle_container_header_changed)
            eta_date_edit.dateChanged.connect(self.refresh_item_summary_context_boxes)

        additional_spinner = getattr(self.ui, "additional_spinner", None)
        if additional_spinner is not None:
            additional_spinner.valueChanged.connect(self.handle_container_header_changed)

        dog_leads_check = getattr(self.ui, "checkBox", None)
        if dog_leads_check is not None:
            dog_leads_check.stateChanged.connect(self.handle_container_header_changed)

        notes_button = getattr(self.ui, "pushButton_8", None)
        if notes_button is not None:
            notes_button.clicked.connect(self.edit_container_notes)

        export_button = getattr(self.ui, "exportEmail_pushButton", None)
        if export_button is not None:
            export_button.clicked.connect(self.export_container_to_excel_and_email)

        containers_list = getattr(self.ui, "containers_tableWidget", None)
        if containers_list is not None:
            containers_list.cellDoubleClicked.connect(self.handle_saved_container_double_click)

        create_yu_order_button = getattr(self.ui, "createYUOrder_pushButton", None)
        if create_yu_order_button is not None:
            create_yu_order_button.clicked.connect(self.open_yu_order_entry_dialog)

        update_orders_button = getattr(self.ui, "updateOrders_button", None) or getattr(self.ui, "updatOrders_pushButton_3", None)
        if update_orders_button is not None:
            update_orders_button.clicked.connect(self.import_orders_from_dialog)

        update_sales_button = getattr(self.ui, "updateSales_pushButton", None)
        if update_sales_button is not None:
            update_sales_button.clicked.connect(self.import_sales_from_dialog)

        update_stock_button = getattr(self.ui, "updateStock_pushButton", None)
        if update_stock_button is not None:
            update_stock_button.clicked.connect(self.import_stock_from_dialog)

        self.setup_item_summary_editing()

    def setup_item_summary_editing(self):
        for object_name, field_meta in self.item_summary_editable_fields.items():
            widget = getattr(self.ui, object_name, None)
            if widget is None:
                continue
            widget.installEventFilter(self)
            widget.setCursor(Qt.PointingHandCursor)
            label = field_meta.get("label", object_name)
            widget.setToolTip(f"Double-click to edit {label}.")

        for object_name in ("stockOnOrder_label", "stockOnOrder_box"):
            widget = getattr(self.ui, object_name, None)
            if widget is None:
                continue
            widget.installEventFilter(self)
            widget.setCursor(Qt.PointingHandCursor)
            try:
                widget.setToolTip("Double-click to add the current item to the To Order Sheet.")
            except Exception:
                pass

    # -----------------------------
    # Database helpers
    # -----------------------------
    def find_data_file(self, *relative_names):
        for name in relative_names:
            p1 = self.base_dir / "data" / name
            if p1.exists():
                return p1
            p2 = self.base_dir / name
            if p2.exists():
                return p2
        return self.base_dir / "data" / relative_names[0]

    def get_database_config_candidate_paths(self):
        candidate_paths = []
        env_path = os.environ.get("WINDSOR_WIDGET_CONFIG", "").strip()
        if env_path:
            candidate_paths.append(Path(env_path))

        candidate_paths.extend([
            self.app_dir / "client_config.json",
            self.app_dir / "data" / "client_config.json",
            Path(os.environ.get("ProgramData", r"C:\ProgramData")) / "WindsorWidget" / "client_config.json",
            self.base_dir / "client_config.json",
            self.base_dir / "data" / "client_config.json",
        ])

        deduped = []
        seen = set()
        for path in candidate_paths:
            try:
                resolved = str(path.resolve())
            except Exception:
                resolved = str(path)
            if resolved in seen:
                continue
            seen.add(resolved)
            deduped.append(path)
        return deduped

    def load_database_config(self):
        last_error = None
        checked_paths = []
        for path in self.get_database_config_candidate_paths():
            checked_paths.append(str(path))
            if not path.exists():
                continue
            for encoding_name in ("utf-8-sig", "utf-8"):
                try:
                    data = json.loads(path.read_text(encoding=encoding_name))
                    if isinstance(data, dict):
                        data["_loaded_from_path"] = str(path)
                        return data
                except Exception as exc:
                    last_error = f"{path}: {exc}"
                    continue
        return {"_checked_paths": checked_paths, "_last_error": last_error}

    def open_database(self):
        self.db_config = self.load_database_config()
        provider = str(self.db_config.get("provider", "")).strip().lower()
        server = str(self.db_config.get("server", "")).strip()
        database = str(self.db_config.get("database", "")).strip()

        if provider != "sqlserver" or not server or not database:
            checked_paths = self.db_config.get("_checked_paths", [])
            last_error = self.db_config.get("_last_error", "")
            checked_text = "\n".join(checked_paths) if checked_paths else "(none)"
            extra = f"\n\nLast config parse error:\n{last_error}" if last_error else ""
            raise RuntimeError(
                "This build is SQL Server only. A valid SQL Server client_config.json was not found "
                "or does not declare provider='sqlserver'.\n\nChecked paths:\n"
                f"{checked_text}{extra}"
            )

        if pyodbc is None:
            QMessageBox.critical(self, "Missing dependency", "pyodbc is required for SQL Server mode.")
            raise RuntimeError("pyodbc is not installed")

        driver = self.db_config.get("driver") or "ODBC Driver 18 for SQL Server"
        port = str(self.db_config.get("port", "")).strip()
        username = str(self.db_config.get("username", "")).strip()
        password = str(self.db_config.get("password", "")).strip()
        trusted = bool(self.db_config.get("trusted_connection", False))
        encrypt = str(self.db_config.get("encrypt", "no")).strip().lower()
        trust_cert = str(self.db_config.get("trust_server_certificate", "yes")).strip().lower()
        timeout = int(self.db_config.get("timeout", 5) or 5)
        server_spec = f"{server},{port}" if port else server

        parts = [
            f"DRIVER={{{driver}}}",
            f"SERVER={server_spec}",
            f"DATABASE={database}",
            f"Encrypt={'yes' if encrypt in {'1', 'true', 'yes'} else 'no'}",
            f"TrustServerCertificate={'yes' if trust_cert in {'1', 'true', 'yes'} else 'no'}",
            f"Connection Timeout={timeout}",
        ]
        if trusted:
            parts.append("Trusted_Connection=yes")
        else:
            parts.append(f"UID={username}")
            parts.append(f"PWD={password}")

        conn_str = ";".join(parts)
        try:
            raw_conn = pyodbc.connect(conn_str)
        except Exception as exc:
            loaded_from = self.db_config.get("_loaded_from_path", "(unknown)")
            raise RuntimeError(
                "SQL Server connection failed.\n\n"
                f"Loaded config: {loaded_from}\n"
                f"Server: {server_spec}\n"
                f"Database: {database}\n\n"
                f"{exc}"
            ) from exc

        self.db_conn = SQLServerBackend(raw_conn)
        self.db_engine = "sqlserver"

    def closeEvent(self, event):
        try:
            if self.db_conn is not None:
                self.db_conn.close()
        finally:
            super().closeEvent(event)

    def db_all(self, sql, params=()):
        cur = self.db_conn.cursor()
        cur.execute(sql, params)
        return cur.fetchall()

    def db_one(self, sql, params=()):
        cur = self.db_conn.cursor()
        cur.execute(sql, params)
        return cur.fetchone()

    def has_table(self, table_name):
        row = self.db_one(
            "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
            (table_name,),
        )
        return row is not None

    def get_table_columns(self, table_name):
        if self.db_conn is None or not self.has_table(table_name):
            return []
        cur = self.db_conn.cursor()
        return [row[1] for row in cur.execute(f"PRAGMA table_info({table_name})").fetchall()]


    def ensure_supplier_master_table(self):
        cur = self.db_conn.cursor()

        if not self.has_table("supplier_master"):
            if self.db_engine == "sqlserver":
                cur.execute(
                    """
                    CREATE TABLE supplier_master (
                        id INT IDENTITY(1,1) PRIMARY KEY,
                        supplier_name NVARCHAR(255) NOT NULL
                    )
                    """
                )
            else:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS supplier_master (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        supplier_name TEXT NOT NULL
                    )
                    """
                )
        else:
            existing_columns = {str(name).lower(): name for name in self.get_table_columns("supplier_master")}
            if "supplier_name" not in existing_columns:
                if self.db_engine == "sqlserver":
                    cur.execute("ALTER TABLE supplier_master ADD [supplier_name] NVARCHAR(255) NULL")
                else:
                    cur.execute("ALTER TABLE supplier_master ADD COLUMN supplier_name TEXT")

        if self.db_engine == "sqlserver":
            try:
                cur.execute(
                    """
                    IF NOT EXISTS (
                        SELECT 1
                        FROM sys.indexes
                        WHERE name = 'idx_supplier_master_name'
                          AND object_id = OBJECT_ID('supplier_master')
                    )
                    CREATE INDEX idx_supplier_master_name ON supplier_master(supplier_name)
                    """
                )
            except Exception:
                pass
        else:
            cur.execute("CREATE INDEX IF NOT EXISTS idx_supplier_master_name ON supplier_master(supplier_name)")

        self.db_conn.commit()

        row = self.db_one("SELECT COUNT(*) AS row_count FROM supplier_master")
        row_count = 0 if row is None else int(self.parse_float(row["row_count"]))
        if row_count <= 0 and EMBEDDED_SUPPLIER_MASTER:
            self.seed_supplier_master(EMBEDDED_SUPPLIER_MASTER)

    def seed_supplier_master(self, supplier_names):
        cleaned = []
        seen = set()
        for supplier_name in supplier_names or []:
            name = str(supplier_name or "").strip()
            if not name:
                continue
            key = name.casefold()
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(name)
        if not cleaned:
            return

        cur = self.db_conn.cursor()
        cur.execute("DELETE FROM supplier_master")
        cur.executemany(
            "INSERT INTO supplier_master (supplier_name) VALUES (?)",
            [(name,) for name in cleaned],
        )
        self.db_conn.commit()

    def import_supplier_master_file(self, file_path, replace_existing=True):
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(str(path))

        supplier_names = []
        if path.suffix.lower() in {'.csv', '.txt'}:
            import csv
            with path.open('r', encoding='utf-8-sig', newline='') as handle:
                rows = list(csv.reader(handle))
            start_index = 0
            if rows:
                first_value = (rows[0][0] if rows[0] else '') or ''
                normalized_first = self.normalize_header(first_value)
                if normalized_first in {'supplier', 'suppliername', 'colastname', 'companylastname', 'name'}:
                    start_index = 1
            for row in rows[start_index:]:
                if not row:
                    continue
                supplier_name = str(row[0] or '').strip()
                if supplier_name:
                    supplier_names.append(supplier_name)
        elif path.suffix.lower() in {'.xlsx', '.xlsm'} and load_workbook is not None:
            workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
            try:
                sheet = workbook.active
                values = list(sheet.iter_rows(values_only=True))
            finally:
                workbook.close()
            start_index = 0
            if values:
                first_value = (values[0][0] if values[0] else '') or ''
                normalized_first = self.normalize_header(first_value)
                if normalized_first in {'supplier', 'suppliername', 'colastname', 'companylastname', 'name'}:
                    start_index = 1
            for row in values[start_index:]:
                supplier_name = str((row[0] if row else '') or '').strip()
                if supplier_name:
                    supplier_names.append(supplier_name)
        else:
            raise ValueError('Unsupported supplier list file type. Use CSV or Excel.')

        if replace_existing:
            self.seed_supplier_master(supplier_names)
        else:
            for supplier_name in supplier_names:
                self.ensure_supplier_exists(supplier_name)
        self.load_reference_lists()
        self.setup_supplier_autocomplete()
        return len(supplier_names)

    def get_supplier_master_names(self):
        if self.db_conn is None or not self.has_table("supplier_master"):
            return []
        rows = self.db_all(
            """
            SELECT DISTINCT TRIM(supplier_name) AS supplier_name
            FROM supplier_master
            WHERE TRIM(COALESCE(supplier_name, '')) <> ''
            ORDER BY supplier_name COLLATE NOCASE
            """
        )
        return [row["supplier_name"] for row in rows if row["supplier_name"]]


    def ensure_to_order_lines_table(self):
        cur = self.db_conn.cursor()
        desired_columns = [
            ("id", "INTEGER"),
            ("line_no", "INTEGER"),
            ("item_number", "TEXT"),
            ("description", "TEXT"),
            ("qty", "REAL"),
            ("supplier_name", "TEXT"),
            ("urgent", "INTEGER"),
            ("status", "TEXT"),
        ]

        if not self.has_table("to_order_lines"):
            if self.db_engine == "sqlserver":
                cur.execute(
                    """
                    CREATE TABLE to_order_lines (
                        id INT IDENTITY(1,1) PRIMARY KEY,
                        line_no INT NOT NULL DEFAULT 0,
                        item_number NVARCHAR(100) NOT NULL,
                        description NVARCHAR(255) NULL,
                        qty FLOAT NOT NULL DEFAULT 0,
                        supplier_name NVARCHAR(255) NULL,
                        urgent BIT NOT NULL DEFAULT 0,
                        status NVARCHAR(50) NULL
                    )
                    """
                )
            else:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS to_order_lines (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        line_no INTEGER NOT NULL DEFAULT 0,
                        item_number TEXT NOT NULL,
                        description TEXT,
                        qty REAL NOT NULL DEFAULT 0,
                        supplier_name TEXT,
                        urgent INTEGER NOT NULL DEFAULT 0,
                        status TEXT
                    )
                    """
                )
        else:
            existing_columns = {str(name).lower(): name for name in self.get_table_columns("to_order_lines")}
            for column_name, ddl in desired_columns:
                if column_name.lower() in existing_columns:
                    continue
                if column_name == "id":
                    continue
                if self.db_engine == "sqlserver":
                    mapped = {
                        "line_no": "INT NOT NULL DEFAULT 0",
                        "item_number": "NVARCHAR(100) NULL",
                        "description": "NVARCHAR(255) NULL",
                        "qty": "FLOAT NOT NULL DEFAULT 0",
                        "supplier_name": "NVARCHAR(255) NULL",
                        "urgent": "BIT NOT NULL DEFAULT 0",
                        "status": "NVARCHAR(50) NULL",
                    }[column_name]
                    cur.execute(f"ALTER TABLE to_order_lines ADD [{column_name}] {mapped}")
                else:
                    mapped = {
                        "line_no": "INTEGER NOT NULL DEFAULT 0",
                        "item_number": "TEXT",
                        "description": "TEXT",
                        "qty": "REAL NOT NULL DEFAULT 0",
                        "supplier_name": "TEXT",
                        "urgent": "INTEGER NOT NULL DEFAULT 0",
                        "status": "TEXT",
                    }[column_name]
                    cur.execute(f"ALTER TABLE to_order_lines ADD COLUMN {column_name} {mapped}")

        if self.db_engine == "sqlserver":
            try:
                cur.execute(
                    """
                    IF NOT EXISTS (
                        SELECT 1
                        FROM sys.indexes
                        WHERE name = 'idx_to_order_lines_line_no'
                          AND object_id = OBJECT_ID('to_order_lines')
                    )
                    CREATE INDEX idx_to_order_lines_line_no ON to_order_lines(line_no, id)
                    """
                )
            except Exception:
                pass
        else:
            cur.execute("CREATE INDEX IF NOT EXISTS idx_to_order_lines_line_no ON to_order_lines(line_no, id)")

        self.db_conn.commit()


    def ensure_on_order_lines_table(self):
        cur = self.db_conn.cursor()
        desired_columns = [
            ("id", "INTEGER"),
            ("line_no", "INTEGER"),
            ("order_number", "TEXT"),
            ("item_number", "TEXT"),
            ("description", "TEXT"),
            ("qty", "REAL"),
            ("supplier_name", "TEXT"),
            ("ready_date", "TEXT"),
            ("comments", "TEXT"),
            ("status", "TEXT"),
        ]

        if not self.has_table("on_order_lines"):
            if self.db_engine == "sqlserver":
                cur.execute(
                    """
                    CREATE TABLE on_order_lines (
                        id INT IDENTITY(1,1) PRIMARY KEY,
                        line_no INT NOT NULL DEFAULT 0,
                        order_number NVARCHAR(100) NULL,
                        item_number NVARCHAR(100) NOT NULL,
                        description NVARCHAR(255) NULL,
                        qty FLOAT NOT NULL DEFAULT 0,
                        supplier_name NVARCHAR(255) NULL,
                        ready_date NVARCHAR(20) NULL,
                        comments NVARCHAR(500) NULL,
                        status NVARCHAR(50) NULL
                    )
                    """
                )
            else:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS on_order_lines (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        line_no INTEGER NOT NULL DEFAULT 0,
                        order_number TEXT,
                        item_number TEXT NOT NULL,
                        description TEXT,
                        qty REAL NOT NULL DEFAULT 0,
                        supplier_name TEXT,
                        ready_date TEXT,
                        comments TEXT,
                        status TEXT
                    )
                    """
                )
        else:
            existing_columns = {str(name).lower(): name for name in self.get_table_columns("on_order_lines")}
            for column_name, ddl in desired_columns:
                if column_name.lower() in existing_columns:
                    continue
                if column_name == "id":
                    continue
                if self.db_engine == "sqlserver":
                    mapped = {
                        "line_no": "INT NOT NULL DEFAULT 0",
                        "order_number": "NVARCHAR(100) NULL",
                        "item_number": "NVARCHAR(100) NULL",
                        "description": "NVARCHAR(255) NULL",
                        "qty": "FLOAT NOT NULL DEFAULT 0",
                        "supplier_name": "NVARCHAR(255) NULL",
                        "ready_date": "NVARCHAR(20) NULL",
                        "comments": "NVARCHAR(500) NULL",
                        "status": "NVARCHAR(50) NULL",
                    }[column_name]
                    cur.execute(f"ALTER TABLE on_order_lines ADD [{column_name}] {mapped}")
                else:
                    mapped = {
                        "line_no": "INTEGER NOT NULL DEFAULT 0",
                        "order_number": "TEXT",
                        "item_number": "TEXT",
                        "description": "TEXT",
                        "qty": "REAL NOT NULL DEFAULT 0",
                        "supplier_name": "TEXT",
                        "ready_date": "TEXT",
                        "comments": "TEXT",
                        "status": "TEXT",
                    }[column_name]
                    cur.execute(f"ALTER TABLE on_order_lines ADD COLUMN {column_name} {mapped}")

        if self.db_engine == "sqlserver":
            try:
                cur.execute(
                    """
                    IF NOT EXISTS (
                        SELECT 1
                        FROM sys.indexes
                        WHERE name = 'idx_on_order_lines_line_no'
                          AND object_id = OBJECT_ID('on_order_lines')
                    )
                    CREATE INDEX idx_on_order_lines_line_no ON on_order_lines(line_no, id)
                    """
                )
            except Exception:
                pass
        else:
            cur.execute("CREATE INDEX IF NOT EXISTS idx_on_order_lines_line_no ON on_order_lines(line_no, id)")

        self.db_conn.commit()

    def ensure_on_order_meta_table(self):
        cur = self.db_conn.cursor()
        if not self.has_table("on_order_meta"):
            if self.db_engine == "sqlserver":
                cur.execute(
                    """
                    CREATE TABLE on_order_meta (
                        meta_key NVARCHAR(100) PRIMARY KEY,
                        meta_value NVARCHAR(MAX) NULL
                    )
                    """
                )
            else:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS on_order_meta (
                        meta_key TEXT PRIMARY KEY,
                        meta_value TEXT
                    )
                    """
                )
        self.db_conn.commit()

    def setup_on_order_page(self):
        parent = getattr(self.ui, "frame_onOrder", None)
        layout = getattr(self.ui, "verticalLayout_onOrder_inner", None)
        if parent is None or layout is None:
            return
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        top_frame = QWidget(parent)
        top_layout = QHBoxLayout(top_frame)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(14)

        entry_frame = QFrame(top_frame)
        entry_frame.setObjectName("onOrderEntry_frame")
        entry_frame.setMinimumSize(760, 120)
        entry_frame.setMaximumHeight(140)
        entry_layout = QHBoxLayout(entry_frame)
        entry_layout.setContentsMargins(16, 12, 16, 12)
        entry_layout.setSpacing(18)

        primary_frame = QWidget(entry_frame)
        primary_layout = QFormLayout(primary_frame)
        primary_layout.setContentsMargins(0, 0, 0, 0)
        primary_layout.setHorizontalSpacing(12)
        primary_layout.setVerticalSpacing(10)

        secondary_frame = QWidget(entry_frame)
        secondary_layout = QFormLayout(secondary_frame)
        secondary_layout.setContentsMargins(0, 0, 0, 0)
        secondary_layout.setHorizontalSpacing(12)
        secondary_layout.setVerticalSpacing(10)

        self.onOrderOrderNumber_label = QLabel("Order No", primary_frame)
        self.onOrderItem_label = QLabel("Enter Item", primary_frame)
        self.onOrderQty_label = QLabel("Enter Qty", primary_frame)
        self.onOrderReadyDate_label = QLabel("Ready Date (optional)", secondary_frame)
        self.onOrderComments_label = QLabel("Comments (optional)", secondary_frame)

        self.onOrderOrderNumber_lineEdit = QLineEdit(primary_frame)
        self.onOrderOrderNumber_lineEdit.setObjectName("onOrderOrderNumber_lineEdit")
        self.onOrderItem_lineEdit = QLineEdit(primary_frame)
        self.onOrderItem_lineEdit.setObjectName("onOrderItem_lineEdit")
        self.onOrderQty_lineEdit = QLineEdit(primary_frame)
        self.onOrderQty_lineEdit.setObjectName("onOrderQty_lineEdit")

        self.onOrderReadyDate_dateEdit = QDateEdit(secondary_frame)
        self.onOrderReadyDate_dateEdit.setObjectName("onOrderReadyDate_dateEdit")
        self.onOrderReadyDate_dateEdit.setCalendarPopup(True)
        self.onOrderReadyDate_dateEdit.setDisplayFormat("dd/MM/yy")
        self.onOrderReadyDate_dateEdit.setDate(QDate.currentDate())
        self.onOrderComments_lineEdit = QLineEdit(secondary_frame)
        self.onOrderComments_lineEdit.setObjectName("onOrderComments_lineEdit")
        self.onOrderComments_lineEdit.setPlaceholderText("Optional notes")

        qty_validator = QDoubleValidator(0.0, 999999999.0, 3, self.onOrderQty_lineEdit)
        qty_validator.setNotation(QDoubleValidator.StandardNotation)
        self.onOrderQty_lineEdit.setValidator(qty_validator)

        label_widgets = [
            self.onOrderOrderNumber_label,
            self.onOrderItem_label,
            self.onOrderQty_label,
            self.onOrderReadyDate_label,
            self.onOrderComments_label,
        ]
        input_widgets = [
            self.onOrderOrderNumber_lineEdit,
            self.onOrderItem_lineEdit,
            self.onOrderQty_lineEdit,
            self.onOrderReadyDate_dateEdit,
            self.onOrderComments_lineEdit,
        ]

        for label in label_widgets:
            try:
                label_font = label.font()
                label_font.setPointSize(max(label_font.pointSize(), 11))
                label_font.setBold(True)
                label.setFont(label_font)
            except Exception:
                pass

        for widget in input_widgets:
            try:
                widget.setMinimumHeight(34)
            except Exception:
                pass

        self.onOrderOrderNumber_lineEdit.setMinimumWidth(210)
        self.onOrderItem_lineEdit.setMinimumWidth(210)
        self.onOrderQty_lineEdit.setMinimumWidth(210)
        self.onOrderReadyDate_dateEdit.setMinimumWidth(150)
        self.onOrderComments_lineEdit.setMinimumWidth(280)

        primary_layout.setWidget(0, QFormLayout.ItemRole.LabelRole, self.onOrderOrderNumber_label)
        primary_layout.setWidget(0, QFormLayout.ItemRole.FieldRole, self.onOrderOrderNumber_lineEdit)
        primary_layout.setWidget(1, QFormLayout.ItemRole.LabelRole, self.onOrderItem_label)
        primary_layout.setWidget(1, QFormLayout.ItemRole.FieldRole, self.onOrderItem_lineEdit)
        primary_layout.setWidget(2, QFormLayout.ItemRole.LabelRole, self.onOrderQty_label)
        primary_layout.setWidget(2, QFormLayout.ItemRole.FieldRole, self.onOrderQty_lineEdit)

        secondary_layout.setWidget(0, QFormLayout.ItemRole.LabelRole, self.onOrderReadyDate_label)
        secondary_layout.setWidget(0, QFormLayout.ItemRole.FieldRole, self.onOrderReadyDate_dateEdit)
        secondary_layout.setWidget(1, QFormLayout.ItemRole.LabelRole, self.onOrderComments_label)
        secondary_layout.setWidget(1, QFormLayout.ItemRole.FieldRole, self.onOrderComments_lineEdit)

        entry_layout.addWidget(primary_frame, 0)
        entry_layout.addWidget(secondary_frame, 0)

        self.onOrderAdd_button = QPushButton("Add Line", top_frame)
        self.onOrderAdd_button.setMinimumSize(160, 72)
        self.onOrderAdd_button.setMaximumSize(160, 72)

        top_layout.addWidget(entry_frame, 0)
        top_layout.addWidget(self.onOrderAdd_button, 0)
        top_layout.addStretch(1)
        layout.addWidget(top_frame)

        self.onOrder_table = QTableWidget(parent)
        self.onOrder_table.setObjectName("onOrder_table")
        self.onOrder_table.setMinimumSize(528, 176)
        layout.addWidget(self.onOrder_table, 1)

        comments_frame = QFrame(parent)
        comments_frame.setObjectName("onOrderGeneralComments_frame")
        comments_layout = QVBoxLayout(comments_frame)
        comments_layout.setContentsMargins(0, 0, 0, 0)
        comments_layout.setSpacing(6)
        self.onOrderGeneralComments_label = QLabel("General Comments / Sundries", comments_frame)
        try:
            _font = self.onOrderGeneralComments_label.font()
            _font.setPointSize(max(_font.pointSize(), 11))
            _font.setBold(True)
            self.onOrderGeneralComments_label.setFont(_font)
        except Exception:
            pass
        comments_layout.addWidget(self.onOrderGeneralComments_label)
        self.onOrderGeneralComments_textEdit = QTextEdit(comments_frame)
        self.onOrderGeneralComments_textEdit.setObjectName("onOrderGeneralComments_textEdit")
        self.onOrderGeneralComments_textEdit.setMinimumHeight(90)
        self.onOrderGeneralComments_textEdit.setPlaceholderText("Add sundries or general comments that are not MYOB item numbers...")
        comments_layout.addWidget(self.onOrderGeneralComments_textEdit)
        layout.addWidget(comments_frame)

        self.setup_on_order_autocomplete()
        self.setup_on_order_table()

        self.onOrderAdd_button.clicked.connect(self.add_on_order_line_from_inputs)
        self.onOrderOrderNumber_lineEdit.returnPressed.connect(self._focus_on_order_item_or_qty)
        self.onOrderItem_lineEdit.returnPressed.connect(self._focus_on_order_qty_or_add)
        self.onOrderQty_lineEdit.returnPressed.connect(self.add_on_order_line_from_inputs)
        self.onOrderComments_lineEdit.returnPressed.connect(self.add_on_order_line_from_inputs)
        self.onOrderGeneralComments_textEdit.textChanged.connect(self.save_on_order_general_comments)

    def _focus_on_order_item_or_qty(self):
        order_edit = getattr(self, "onOrderOrderNumber_lineEdit", None)
        item_edit = getattr(self, "onOrderItem_lineEdit", None)
        qty_edit = getattr(self, "onOrderQty_lineEdit", None)
        if order_edit is None or item_edit is None or qty_edit is None:
            return
        if not (order_edit.text() or "").strip():
            order_edit.setFocus()
            order_edit.selectAll()
            return
        if not (item_edit.text() or "").strip():
            item_edit.setFocus()
            item_edit.selectAll()
            return
        qty_edit.setFocus()
        qty_edit.selectAll()

    def _focus_on_order_qty_or_add(self):
        item_edit = getattr(self, "onOrderItem_lineEdit", None)
        qty_edit = getattr(self, "onOrderQty_lineEdit", None)
        if item_edit is None or qty_edit is None:
            return
        if not (item_edit.text() or "").strip():
            item_edit.setFocus()
            return
        qty_edit.setFocus()
        qty_edit.selectAll()

    def setup_on_order_autocomplete(self):
        item_edit = getattr(self, "onOrderItem_lineEdit", None)
        if item_edit is None:
            return
        completer = QCompleter(self.item_numbers, self)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchStartsWith)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
        item_edit.setCompleter(completer)

    def setup_on_order_table(self):
        table = getattr(self, "onOrder_table", None)
        if table is None:
            return

        headers = ["Order No", "Item Number", "Description", "Qty", "Supplier", "Ready Date", "Comments", "Status", "Add To Container", "Remove"]
        self.on_order_order_number_column = 0
        self.on_order_item_column = 1
        self.on_order_description_column = 2
        self.on_order_qty_column = 3
        self.on_order_supplier_column = 4
        self.on_order_ready_date_column = 5
        self.on_order_comments_column = 6
        self.on_order_status_column = 7
        self.on_order_add_column = 8
        self.on_order_remove_column = 9

        table.clear()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(0)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(self.on_order_order_number_column, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(self.on_order_item_column, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(self.on_order_description_column, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(self.on_order_comments_column, QHeaderView.Stretch)
        for col in (self.on_order_qty_column, self.on_order_supplier_column, self.on_order_ready_date_column, self.on_order_status_column, self.on_order_add_column, self.on_order_remove_column):
            table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeToContents)
        table.setProperty("_table_font_settings_token", "main_window/on_order_table")
        table.setContextMenuPolicy(Qt.CustomContextMenu)
        table.customContextMenuRequested.connect(self.show_on_order_table_context_menu)
        table.cellDoubleClicked.connect(self.handle_on_order_table_double_click)
        table.itemChanged.connect(self.handle_on_order_table_item_changed)


    def format_on_order_ready_date(self, value):
        parsed = self.parse_date_value(value)
        if parsed is None:
            return ""
        return parsed.strftime("%d/%m/%y")

    def on_order_ready_date_is_ready(self, value):
        parsed = self.parse_date_value(value)
        if parsed is None:
            return False
        return parsed <= date.today()

    def apply_on_order_row_styles(self, row):
        table = getattr(self, "onOrder_table", None)
        if table is None or row < 0 or row >= table.rowCount():
            return

        ready_item = table.item(row, self.on_order_ready_date_column)
        ready_text = ready_item.text().strip() if ready_item is not None and ready_item.text() else ""
        is_ready = self.on_order_ready_date_is_ready(ready_text)

        ready_bg = QColor("#d9f2d9")
        ready_fg = QColor("#006400")

        for col in range(table.columnCount()):
            item = table.item(row, col)
            if item is None:
                continue
            if col in {self.on_order_status_column, self.on_order_add_column, self.on_order_remove_column}:
                continue
            if is_ready:
                item.setBackground(QBrush(ready_bg))
                item.setForeground(QBrush(ready_fg if col == self.on_order_ready_date_column else QColor(Qt.black)))
            else:
                item.setBackground(QBrush())
                item.setForeground(QBrush())

    def build_on_order_status_item(self, status_text):
        status_text = (status_text or "").strip().upper()
        background = None
        foreground = None
        if status_text == "IN CONTAINER":
            background = QColor("#dbeeff")
            foreground = QColor("#003f7f")
        return self.make_order_table_item(
            status_text,
            editable=False,
            align=Qt.AlignCenter,
            background=background,
            foreground=foreground,
            bold=bool(status_text),
        )

    def build_on_order_add_item(self):
        return self.make_order_table_item(
            "Add To Container",
            background=QColor("#d9f2d9"),
            foreground=QColor("#006400"),
            bold=True,
            align=Qt.AlignCenter,
        )

    def build_on_order_remove_item(self):
        return self.make_order_table_item(
            "Remove",
            background=QColor("#ffd6d6"),
            foreground=QColor("#8b0000"),
            bold=True,
            align=Qt.AlignCenter,
        )

    def populate_on_order_table_from_rows(self, rows):
        table = getattr(self, "onOrder_table", None)
        if table is None:
            return
        self._updating_on_order_table = True
        table.setRowCount(0)
        for line in rows:
            row = table.rowCount()
            table.insertRow(row)
            order_number = line.get("order_number", "")
            item_number = line.get("item_number", "")
            table.setItem(row, self.on_order_order_number_column, self.make_order_table_item(order_number, editable=True))
            table.setItem(row, self.on_order_item_column, self.make_order_table_item(item_number))
            table.setItem(row, self.on_order_description_column, self.make_order_table_item(line.get("description", "")))

            qty_value = self.parse_float(line.get("qty", 0))
            qty_item = self.make_order_table_item(
                self.format_value(qty_value),
                editable=True,
                align=Qt.AlignRight | Qt.AlignVCenter,
            )
            qty_item.setData(Qt.UserRole, qty_value)
            table.setItem(row, self.on_order_qty_column, qty_item)

            supplier_value = (line.get("supplier_name", "") or "").strip()
            supplier_item = self.make_order_table_item(supplier_value, editable=False)
            supplier_item.setData(Qt.UserRole, supplier_value)
            supplier_item.setToolTip("Double-click to choose a supplier.")
            table.setItem(row, self.on_order_supplier_column, supplier_item)

            ready_date_text = self.format_on_order_ready_date(line.get("ready_date", ""))
            table.setItem(row, self.on_order_ready_date_column, self.make_order_table_item(ready_date_text, editable=True, align=Qt.AlignCenter))
            table.setItem(row, self.on_order_comments_column, self.make_order_table_item(line.get("comments", ""), editable=True))
            table.setItem(row, self.on_order_status_column, self.build_on_order_status_item(line.get("status", "")))
            table.setItem(row, self.on_order_add_column, self.build_on_order_add_item())
            table.setItem(row, self.on_order_remove_column, self.build_on_order_remove_item())
            self.apply_on_order_row_styles(row)

        self._updating_on_order_table = False
        table.resizeRowsToContents()

    def get_on_order_table_rows(self):
        table = getattr(self, "onOrder_table", None)
        if table is None:
            return []
        rows = []
        for row in range(table.rowCount()):
            order_number = table.item(row, self.on_order_order_number_column).text().strip() if table.item(row, self.on_order_order_number_column) is not None and table.item(row, self.on_order_order_number_column).text() else ""
            item_number = table.item(row, self.on_order_item_column).text().strip() if table.item(row, self.on_order_item_column) is not None and table.item(row, self.on_order_item_column).text() else ""
            description = table.item(row, self.on_order_description_column).text().strip() if table.item(row, self.on_order_description_column) is not None and table.item(row, self.on_order_description_column).text() else ""
            qty_item = table.item(row, self.on_order_qty_column)
            qty_value = qty_item.data(Qt.UserRole) if qty_item is not None else None
            if qty_value in (None, "") and qty_item is not None:
                qty_value = qty_item.text()
            supplier_item = table.item(row, self.on_order_supplier_column)
            supplier_value = ""
            if supplier_item is not None:
                supplier_value = str(supplier_item.data(Qt.UserRole) or supplier_item.text() or "").strip()
            ready_date_item = table.item(row, self.on_order_ready_date_column)
            ready_date_text = ready_date_item.text().strip() if ready_date_item is not None and ready_date_item.text() else ""
            ready_date_value = self.format_on_order_ready_date(ready_date_text)
            comments_item = table.item(row, self.on_order_comments_column)
            comments_text = comments_item.text().strip() if comments_item is not None and comments_item.text() else ""
            status_item = table.item(row, self.on_order_status_column)
            status_text = status_item.text().strip().upper() if status_item is not None and status_item.text() else ""
            if not item_number:
                continue
            rows.append({
                "order_number": order_number,
                "item_number": item_number,
                "description": description,
                "qty": self.parse_float(qty_value),
                "supplier_name": supplier_value,
                "ready_date": ready_date_value,
                "comments": comments_text,
                "status": status_text,
            })
        return rows

    def save_on_order_table_state(self):
        if self.db_conn is None or not self.has_table("on_order_lines"):
            return
        rows = self.get_on_order_table_rows()
        cur = self.db_conn.cursor()
        cur.execute("DELETE FROM on_order_lines")
        for line_no, row in enumerate(rows, start=1):
            cur.execute(
                """
                INSERT INTO on_order_lines (line_no, order_number, item_number, description, qty, supplier_name, ready_date, comments, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    line_no,
                    row.get("order_number", ""),
                    row.get("item_number", ""),
                    row.get("description", ""),
                    float(row.get("qty", 0) or 0),
                    row.get("supplier_name", ""),
                    row.get("ready_date", ""),
                    row.get("comments", ""),
                    row.get("status", ""),
                ),
            )
        self.db_conn.commit()

    def load_saved_on_order_lines(self):
        if self.db_conn is None or not self.has_table("on_order_lines"):
            return
        rows = self.db_all(
            """
            SELECT order_number, item_number, description, qty, supplier_name, ready_date, comments, status
            FROM on_order_lines
            ORDER BY line_no, id
            """
        )
        parsed_rows = [
            {
                "order_number": row.get("order_number") or "",
                "item_number": row["item_number"] or "",
                "description": row["description"] or "",
                "qty": self.parse_float(row["qty"]),
                "supplier_name": row["supplier_name"] or "",
                "ready_date": row.get("ready_date") or "",
                "comments": row.get("comments") or "",
                "status": (row.get("status") or "").strip().upper(),
            }
            for row in rows
        ]
        self.populate_on_order_table_from_rows(parsed_rows)
        self.refresh_on_order_statuses_from_container_data(persist=False)

    def load_on_order_general_comments(self):
        editor = getattr(self, "onOrderGeneralComments_textEdit", None)
        if editor is None or self.db_conn is None or not self.has_table("on_order_meta"):
            return
        row = self.db_one("SELECT meta_value FROM on_order_meta WHERE meta_key = ?", ("general_comments",))
        text_value = str((row or {}).get("meta_value") or "")
        blocker = QSignalBlocker(editor)
        try:
            editor.setPlainText(text_value)
        finally:
            del blocker

    def save_on_order_general_comments(self):
        editor = getattr(self, "onOrderGeneralComments_textEdit", None)
        if editor is None or self.db_conn is None or not self.has_table("on_order_meta"):
            return
        text_value = editor.toPlainText().strip()
        cur = self.db_conn.cursor()
        if self.db_engine == "sqlserver":
            cur.execute("DELETE FROM on_order_meta WHERE meta_key = ?", ("general_comments",))
            cur.execute(
                "INSERT INTO on_order_meta (meta_key, meta_value) VALUES (?, ?)",
                ("general_comments", text_value),
            )
        else:
            cur.execute(
                """
                INSERT INTO on_order_meta (meta_key, meta_value)
                VALUES (?, ?)
                ON CONFLICT(meta_key) DO UPDATE SET meta_value = excluded.meta_value
                """,
                ("general_comments", text_value),
            )
        self.db_conn.commit()

    def on_order_line_is_in_any_container(self, order_number, item_number):
        if self.db_conn is None or not self.has_table("container_lines"):
            return False
        order_number = str(order_number or "").strip()
        item_number = str(item_number or "").strip()
        if not item_number:
            return False
        if self.db_engine == "sqlserver":
            row = self.db_one(
                """
                SELECT TOP 1 1 AS found
                FROM container_lines
                WHERE UPPER(TRIM(item_number)) = UPPER(TRIM(?))
                  AND UPPER(TRIM(COALESCE(order_number, ''))) = UPPER(TRIM(?))
                """,
                (item_number, order_number),
            )
        else:
            row = self.db_one(
                """
                SELECT 1 AS found
                FROM container_lines
                WHERE UPPER(TRIM(item_number)) = UPPER(TRIM(?))
                  AND UPPER(TRIM(COALESCE(order_number, ''))) = UPPER(TRIM(?))
                LIMIT 1
                """,
                (item_number, order_number),
            )
        return row is not None

    def refresh_on_order_statuses_from_container_data(self, persist=True):
        table = getattr(self, "onOrder_table", None)
        if table is None:
            return
        self._updating_on_order_table = True
        changed = False
        try:
            for row in range(table.rowCount()):
                row_data = self.get_on_order_row_data(row)
                if row_data is None:
                    continue
                should_be = "IN CONTAINER" if self.on_order_line_is_in_any_container(
                    row_data.get("order_number", ""),
                    row_data.get("item_number", ""),
                ) else ""
                status_item = table.item(row, self.on_order_status_column)
                current_text = (status_item.text().strip().upper() if status_item is not None and status_item.text() else "")
                if current_text != should_be:
                    table.setItem(row, self.on_order_status_column, self.build_on_order_status_item(should_be))
                    changed = True
                self.apply_on_order_row_styles(row)
        finally:
            self._updating_on_order_table = False
        if changed and persist:
            self.save_on_order_table_state()

    def add_on_order_line_from_inputs(self):
        item_edit = getattr(self, "onOrderItem_lineEdit", None)
        order_widget = getattr(self, "onOrderOrderNumber_lineEdit", None)
        qty_widget = getattr(self, "onOrderQty_lineEdit", None)
        ready_widget = getattr(self, "onOrderReadyDate_dateEdit", None)
        comments_widget = getattr(self, "onOrderComments_lineEdit", None)
        table = getattr(self, "onOrder_table", None)
        if item_edit is None or order_widget is None or qty_widget is None or ready_widget is None or comments_widget is None or table is None:
            return

        typed_item = item_edit.text().strip()
        item_number = self.find_item_number(typed_item)
        if not item_number:
            QMessageBox.warning(self, "Invalid item", "Please choose a valid item number from the list.")
            item_edit.setFocus()
            return

        order_number = (order_widget.text() or "").strip()
        if not order_number:
            QMessageBox.warning(self, "Invalid order number", "Please enter an order number.")
            order_widget.setFocus()
            order_widget.selectAll()
            return

        qty_text = qty_widget.text().strip()
        qty_value = self.parse_float(qty_text)
        if qty_value <= 0:
            QMessageBox.warning(self, "Invalid quantity", "Please enter a quantity greater than 0.")
            qty_widget.setFocus()
            qty_widget.selectAll()
            return

        ready_date_text = self.format_on_order_ready_date(ready_widget.date())
        comments_text = (comments_widget.text() or "").strip()

        for row in range(table.rowCount()):
            existing_order = table.item(row, self.on_order_order_number_column)
            existing_item = table.item(row, self.on_order_item_column)
            if (
                existing_order is not None and (existing_order.text() or "").strip() == order_number
                and existing_item is not None and (existing_item.text() or "").strip() == item_number
            ):
                qty_item = table.item(row, self.on_order_qty_column)
                current_qty = self.parse_float(qty_item.data(Qt.UserRole) if qty_item is not None else 0)
                new_qty = current_qty + qty_value
                self._updating_on_order_table = True
                if qty_item is not None:
                    qty_item.setText(self.format_value(new_qty))
                    qty_item.setData(Qt.UserRole, new_qty)
                ready_item = table.item(row, self.on_order_ready_date_column)
                if ready_item is not None and ready_date_text:
                    ready_item.setText(ready_date_text)
                comments_item = table.item(row, self.on_order_comments_column)
                if comments_item is not None and comments_text:
                    comments_item.setText(comments_text)
                self.apply_on_order_row_styles(row)
                self._updating_on_order_table = False
                self.save_on_order_table_state()
                item_edit.clear()
                order_widget.clear()
                qty_widget.clear()
                comments_widget.clear()
                item_edit.setFocus()
                return

        item_row = self.get_item_master_row(item_number)
        description = self.get_first(item_row, "description", "item_name", "Item Name", "Description")
        supplier = self.get_first(item_row, "supplier_name", "supplier_code", "Column1", "Supplier")

        row = table.rowCount()
        self._updating_on_order_table = True
        table.insertRow(row)
        table.setItem(row, self.on_order_order_number_column, self.make_order_table_item(order_number, editable=True))
        table.setItem(row, self.on_order_item_column, self.make_order_table_item(item_number))
        table.setItem(row, self.on_order_description_column, self.make_order_table_item(description))
        qty_item = self.make_order_table_item(
            self.format_value(qty_value),
            editable=True,
            align=Qt.AlignRight | Qt.AlignVCenter,
        )
        qty_item.setData(Qt.UserRole, qty_value)
        table.setItem(row, self.on_order_qty_column, qty_item)

        supplier_item = self.make_order_table_item(supplier, editable=False)
        supplier_item.setData(Qt.UserRole, supplier)
        supplier_item.setToolTip("Double-click to choose a supplier.")
        table.setItem(row, self.on_order_supplier_column, supplier_item)
        table.setItem(row, self.on_order_ready_date_column, self.make_order_table_item(ready_date_text, editable=True, align=Qt.AlignCenter))
        table.setItem(row, self.on_order_comments_column, self.make_order_table_item(comments_text, editable=True))
        table.setItem(row, self.on_order_status_column, self.build_on_order_status_item(""))
        table.setItem(row, self.on_order_add_column, self.build_on_order_add_item())
        table.setItem(row, self.on_order_remove_column, self.build_on_order_remove_item())
        self.apply_on_order_row_styles(row)
        self._updating_on_order_table = False

        table.resizeRowsToContents()
        self.save_on_order_table_state()
        item_edit.clear()
        order_widget.clear()
        qty_widget.clear()
        comments_widget.clear()
        item_edit.setFocus()

    def add_or_update_on_order_line(self, order_number, item_number, qty_value, ready_date="", comments=""):
        table = getattr(self, "onOrder_table", None)
        if table is None:
            return False, "On Order table is not available."

        order_number = str(order_number or "").strip()
        typed_item = str(item_number or "").strip()
        item_number = self.find_item_number(typed_item)
        try:
            qty_value = float(qty_value)
        except Exception:
            qty_value = 0.0
        ready_date_text = self.format_on_order_ready_date(ready_date)
        comments_text = str(comments or "").strip()

        if not order_number:
            return False, "Order number is required."
        if not item_number:
            return False, "Please choose a valid item number."
        if qty_value <= 0:
            return False, "Quantity must be greater than 0."

        for row in range(table.rowCount()):
            existing_order = table.item(row, self.on_order_order_number_column)
            existing_item = table.item(row, self.on_order_item_column)
            if (
                existing_order is not None and (existing_order.text() or "").strip() == order_number
                and existing_item is not None and (existing_item.text() or "").strip() == item_number
            ):
                qty_item = table.item(row, self.on_order_qty_column)
                current_qty = self.parse_float(qty_item.data(Qt.UserRole) if qty_item is not None else 0)
                new_qty = current_qty + qty_value
                self._updating_on_order_table = True
                if qty_item is not None:
                    qty_item.setText(self.format_value(new_qty))
                    qty_item.setData(Qt.UserRole, new_qty)
                ready_item = table.item(row, self.on_order_ready_date_column)
                if ready_item is not None and ready_date_text:
                    ready_item.setText(ready_date_text)
                comments_item = table.item(row, self.on_order_comments_column)
                if comments_item is not None and comments_text:
                    comments_item.setText(comments_text)
                self.apply_on_order_row_styles(row)
                self._updating_on_order_table = False
                self.save_on_order_table_state()
                return True, f"Updated On Order line for {item_number}."

        item_row = self.get_item_master_row(item_number)
        description = self.get_first(item_row, "description", "item_name", "Item Name", "Description")
        supplier = self.get_first(item_row, "supplier_name", "supplier_code", "Column1", "Supplier")

        row = table.rowCount()
        self._updating_on_order_table = True
        table.insertRow(row)
        table.setItem(row, self.on_order_order_number_column, self.make_order_table_item(order_number, editable=True))
        table.setItem(row, self.on_order_item_column, self.make_order_table_item(item_number))
        table.setItem(row, self.on_order_description_column, self.make_order_table_item(description))
        qty_item = self.make_order_table_item(
            self.format_value(qty_value),
            editable=True,
            align=Qt.AlignRight | Qt.AlignVCenter,
        )
        qty_item.setData(Qt.UserRole, qty_value)
        table.setItem(row, self.on_order_qty_column, qty_item)

        supplier_item = self.make_order_table_item(supplier, editable=False)
        supplier_item.setData(Qt.UserRole, supplier)
        supplier_item.setToolTip("Double-click to choose a supplier.")
        table.setItem(row, self.on_order_supplier_column, supplier_item)
        table.setItem(row, self.on_order_ready_date_column, self.make_order_table_item(ready_date_text, editable=True, align=Qt.AlignCenter))
        table.setItem(row, self.on_order_comments_column, self.make_order_table_item(comments_text, editable=True))
        table.setItem(row, self.on_order_status_column, self.build_on_order_status_item(""))
        table.setItem(row, self.on_order_add_column, self.build_on_order_add_item())
        table.setItem(row, self.on_order_remove_column, self.build_on_order_remove_item())
        self.apply_on_order_row_styles(row)
        self._updating_on_order_table = False

        table.resizeRowsToContents()
        self.save_on_order_table_state()
        return True, f"Added {item_number} to On Order."

    def get_on_order_row_data(self, row):
        table = getattr(self, "onOrder_table", None)
        if table is None or row < 0 or row >= table.rowCount():
            return None
        order_number = table.item(row, self.on_order_order_number_column).text().strip() if table.item(row, self.on_order_order_number_column) is not None and table.item(row, self.on_order_order_number_column).text() else ""
        item_number = table.item(row, self.on_order_item_column).text().strip() if table.item(row, self.on_order_item_column) is not None and table.item(row, self.on_order_item_column).text() else ""
        description = table.item(row, self.on_order_description_column).text().strip() if table.item(row, self.on_order_description_column) is not None and table.item(row, self.on_order_description_column).text() else ""
        qty_item = table.item(row, self.on_order_qty_column)
        qty_value = qty_item.data(Qt.UserRole) if qty_item is not None else None
        if qty_value in (None, "") and qty_item is not None:
            qty_value = qty_item.text()
        supplier_item = table.item(row, self.on_order_supplier_column)
        supplier_value = str(supplier_item.data(Qt.UserRole) or supplier_item.text() or "").strip() if supplier_item is not None else ""
        ready_item = table.item(row, self.on_order_ready_date_column)
        ready_date_text = ready_item.text().strip() if ready_item is not None and ready_item.text() else ""
        comments_item = table.item(row, self.on_order_comments_column)
        comments_text = comments_item.text().strip() if comments_item is not None and comments_item.text() else ""
        status_item = table.item(row, self.on_order_status_column)
        status_text = status_item.text().strip().upper() if status_item is not None and status_item.text() else ""
        if not item_number:
            return None
        return {
            "order_number": order_number,
            "item_number": item_number,
            "description": description,
            "qty": self.parse_float(qty_value),
            "supplier_name": supplier_value,
            "ready_date": self.format_on_order_ready_date(ready_date_text),
            "comments": comments_text,
            "status": status_text,
        }

    def show_on_order_table_context_menu(self, pos):
        table = getattr(self, "onOrder_table", None)
        if table is None:
            return

        clicked_item = table.itemAt(pos)
        row = clicked_item.row() if clicked_item is not None else table.currentRow()
        has_row = row is not None and row >= 0
        if has_row:
            table.selectRow(row)

        menu = QMenu(table)
        add_to_container_action = menu.addAction("Add to Container")
        remove_action = menu.addAction("Remove")
        add_to_container_action.setEnabled(bool(has_row))
        remove_action.setEnabled(bool(has_row))
        menu.addSeparator()

        font_menu = menu.addMenu("Font Size")
        current_size = current_table_font_size(table)
        for size in TABLE_FONT_SIZE_OPTIONS:
            action = font_menu.addAction(str(size))
            action.setCheckable(True)
            action.setChecked(size == current_size)
            action.triggered.connect(
                lambda _checked=False, tbl=table, selected_size=size: apply_table_font_size(
                    tbl,
                    selected_size,
                    settings=self.settings,
                    scope_key="main_window",
                    persist=True,
                )
            )
        font_menu.addSeparator()
        reset_action = font_menu.addAction("Reset")
        reset_action.triggered.connect(
            lambda _checked=False, tbl=table: reset_table_font_size(
                tbl,
                settings=self.settings,
                scope_key="main_window",
            )
        )

        viewport = getattr(table, "viewport", lambda: None)()
        global_pos = viewport.mapToGlobal(pos) if viewport is not None else table.mapToGlobal(pos)
        chosen_action = menu.exec(global_pos)
        if chosen_action is None or not has_row:
            return

        if chosen_action == add_to_container_action:
            self.add_on_order_row_to_container(row)
            return

        if chosen_action == remove_action:
            self.remove_on_order_row(row)
            return

    def add_on_order_row_to_container(self, row):
        table = getattr(self, "onOrder_table", None)
        if table is None:
            return
        row_data = self.get_on_order_row_data(row)
        if row_data is None:
            return

        container_ref, order_number = self.prompt_for_container_and_order(row_data.get("order_number", ""))
        if not container_ref or not order_number:
            return
        try:
            self.append_line_to_container_ref(
                container_ref=container_ref,
                order_number=order_number,
                item_number=row_data.get("item_number", ""),
                qty_value=row_data.get("qty", 0),
                is_urgent=False,
            )
        except Exception as exc:
            QMessageBox.warning(self, "Add to Container", str(exc))
            return

        remove_result = QMessageBox.question(
            self,
            "Added to Container",
            f"Added {row_data.get('item_number', '')} to {container_ref} with order number {order_number}.\n\nRemove it from On Order?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if remove_result == QMessageBox.Yes:
            table.removeRow(row)
        else:
            self._updating_on_order_table = True
            table.setItem(row, self.on_order_status_column, self.build_on_order_status_item("IN CONTAINER"))
            self._updating_on_order_table = False
        self.save_on_order_table_state()
        self.refresh_on_order_statuses_from_container_data(persist=True)
        self.refresh_item_summary_context_boxes()
        self.rerun_order_analysis_if_ready()

    def remove_on_order_row(self, row):
        table = getattr(self, "onOrder_table", None)
        if table is None or row < 0:
            return
        item_number = table.item(row, self.on_order_item_column).text() if table.item(row, self.on_order_item_column) else "this line"
        result = QMessageBox.question(
            self,
            "Remove line",
            f"Remove {item_number} from the On Order reminder sheet?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if result == QMessageBox.Yes:
            table.removeRow(row)
            self.save_on_order_table_state()

    def handle_on_order_table_double_click(self, row, column):
        table = getattr(self, "onOrder_table", None)
        if table is None or row < 0:
            return

        if column == self.on_order_add_column:
            self.add_on_order_row_to_container(row)
            return

        if column == self.on_order_remove_column:
            self.remove_on_order_row(row)
            return

        if column == self.on_order_order_number_column:
            item = table.item(row, column)
            if item is not None:
                table.editItem(item)
            return

        if column == self.on_order_qty_column:
            item = table.item(row, column)
            if item is not None:
                table.editItem(item)
            return

        if column == self.on_order_ready_date_column:
            item = table.item(row, column)
            if item is not None:
                table.editItem(item)
            return

        if column == self.on_order_comments_column:
            item = table.item(row, column)
            if item is not None:
                table.editItem(item)
            return

        if column == self.on_order_supplier_column:
            item = table.item(row, column)
            current_supplier = ""
            if item is not None:
                current_supplier = str(item.data(Qt.UserRole) or item.text() or "").strip()
            resolved_supplier = self.prompt_for_supplier_name(current_supplier)
            if resolved_supplier is None:
                return
            if item is None:
                item = self.make_order_table_item("", editable=False)
                table.setItem(row, column, item)
            self._updating_on_order_table = True
            item.setText(resolved_supplier)
            item.setData(Qt.UserRole, resolved_supplier)
            item.setToolTip("Double-click to choose a supplier.")
            self._updating_on_order_table = False
            item_number_item = table.item(row, self.on_order_item_column)
            item_number = item_number_item.text().strip() if item_number_item is not None and item_number_item.text() else ""
            self.update_item_supplier_in_database(item_number, resolved_supplier)
            self.load_reference_lists()
            self.setup_supplier_autocomplete()
            self.setup_on_order_autocomplete()
            self.apply_on_order_row_styles(row)
            self.save_on_order_table_state()
            return

    def handle_on_order_table_item_changed(self, item):
        if self._updating_on_order_table or item is None:
            return

        if item.column() == self.on_order_qty_column:
            new_qty = self.parse_float(item.text())
            old_qty = self.parse_float(item.data(Qt.UserRole))
            if new_qty <= 0:
                QMessageBox.warning(self, "Invalid quantity", "Quantity must be greater than 0.")
                self._updating_on_order_table = True
                item.setText(self.format_value(old_qty))
                self._updating_on_order_table = False
                return

            self._updating_on_order_table = True
            item.setText(self.format_value(new_qty))
            item.setData(Qt.UserRole, new_qty)
            self.apply_on_order_row_styles(item.row())
            self._updating_on_order_table = False
            self.save_on_order_table_state()
            return

        if item.column() == self.on_order_ready_date_column:
            parsed = self.parse_date_value(item.text())
            if parsed is None and (item.text() or "").strip():
                QMessageBox.warning(self, "Invalid ready date", "Ready Date must be a valid date like 13/04/26 or 13/04/2026.")
                self._updating_on_order_table = True
                item.setText("")
                self._updating_on_order_table = False
                return

            self._updating_on_order_table = True
            item.setText(self.format_on_order_ready_date(parsed))
            self.apply_on_order_row_styles(item.row())
            self._updating_on_order_table = False
            self.save_on_order_table_state()
            return

        if item.column() in {self.on_order_order_number_column, self.on_order_comments_column}:
            self._updating_on_order_table = True
            item.setText((item.text() or "").strip())
            self.apply_on_order_row_styles(item.row())
            self._updating_on_order_table = False
            self.save_on_order_table_state()
            return

    def ensure_shipment_tables(self):
        cur = self.db_conn.cursor()
        desired_columns = [
            ("id", "INTEGER"),
            ("entry_date", "TEXT"),
            ("shipment_type", "TEXT"),
            ("order_no", "TEXT"),
            ("supplier_name", "TEXT"),
            ("container_no", "TEXT"),
            ("product", "TEXT"),
            ("qty", "TEXT"),
            ("ready_date", "TEXT"),
            ("shipment_date", "TEXT"),
            ("due_date", "TEXT"),
            ("status", "TEXT"),
            ("vessel", "TEXT"),
            ("notes", "TEXT"),
            ("updated_on", "TEXT"),
        ]

        if not self.has_table(SHIPMENT_TABLE_NAME):
            if self.db_engine == "sqlserver":
                cur.execute(
                    f"""
                    CREATE TABLE {SHIPMENT_TABLE_NAME} (
                        id INT IDENTITY(1,1) PRIMARY KEY,
                        entry_date NVARCHAR(20) NULL,
                        shipment_type NVARCHAR(50) NULL,
                        order_no NVARCHAR(100) NULL,
                        supplier_name NVARCHAR(255) NULL,
                        container_no NVARCHAR(100) NULL,
                        product NVARCHAR(255) NULL,
                        qty NVARCHAR(50) NULL,
                        ready_date NVARCHAR(20) NULL,
                        shipment_date NVARCHAR(20) NULL,
                        due_date NVARCHAR(20) NULL,
                        status NVARCHAR(100) NULL,
                        vessel NVARCHAR(255) NULL,
                        notes NVARCHAR(MAX) NULL,
                        updated_on NVARCHAR(10) NULL
                    )
                    """
                )
            else:
                cur.execute(
                    f"""
                    CREATE TABLE IF NOT EXISTS {SHIPMENT_TABLE_NAME} (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        entry_date TEXT,
                        shipment_type TEXT,
                        order_no TEXT,
                        supplier_name TEXT,
                        container_no TEXT,
                        product TEXT,
                        qty TEXT,
                        ready_date TEXT,
                        shipment_date TEXT,
                        due_date TEXT,
                        status TEXT,
                        vessel TEXT,
                        notes TEXT,
                        updated_on TEXT
                    )
                    """
                )
        else:
            existing_columns = {str(name).lower(): name for name in self.get_table_columns(SHIPMENT_TABLE_NAME)}
            for column_name, _ddl in desired_columns:
                if column_name.lower() in existing_columns or column_name == "id":
                    continue
                if self.db_engine == "sqlserver":
                    mapped = {
                        "entry_date": "NVARCHAR(20) NULL",
                        "shipment_type": "NVARCHAR(50) NULL",
                        "order_no": "NVARCHAR(100) NULL",
                        "supplier_name": "NVARCHAR(255) NULL",
                        "container_no": "NVARCHAR(100) NULL",
                        "product": "NVARCHAR(255) NULL",
                        "qty": "NVARCHAR(50) NULL",
                        "ready_date": "NVARCHAR(20) NULL",
                        "shipment_date": "NVARCHAR(20) NULL",
                        "due_date": "NVARCHAR(20) NULL",
                        "status": "NVARCHAR(100) NULL",
                        "vessel": "NVARCHAR(255) NULL",
                        "notes": "NVARCHAR(MAX) NULL",
                        "updated_on": "NVARCHAR(10) NULL",
                    }[column_name]
                    cur.execute(f"ALTER TABLE {SHIPMENT_TABLE_NAME} ADD [{column_name}] {mapped}")
                else:
                    cur.execute(f"ALTER TABLE {SHIPMENT_TABLE_NAME} ADD COLUMN {column_name} TEXT")

        if not self.has_table("shipments_meta"):
            if self.db_engine == "sqlserver":
                cur.execute(
                    """
                    CREATE TABLE shipments_meta (
                        meta_key NVARCHAR(100) PRIMARY KEY,
                        meta_value NVARCHAR(100) NULL
                    )
                    """
                )
            else:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS shipments_meta (
                        meta_key TEXT PRIMARY KEY,
                        meta_value TEXT
                    )
                    """
                )

        legacy_tables = {
            "Melbourne": "shipments_melbourne",
            "Sydney": "shipments_sydney",
            "SABA": "shipments_saba",
        }
        shipments_count_row = self.db_one(f"SELECT COUNT(*) AS row_count FROM {SHIPMENT_TABLE_NAME}")
        shipments_count = int(shipments_count_row["row_count"] or 0) if shipments_count_row is not None else 0
        if shipments_count == 0:
            for shipment_type, legacy_table in legacy_tables.items():
                if not self.has_table(legacy_table):
                    continue
                legacy_rows = self.db_all(
                    f"""
                    SELECT entry_date, order_no, supplier_name, container_no, product, qty,
                           ready_date, shipment_date, due_date, status, vessel, notes, updated_on
                    FROM {legacy_table}
                    ORDER BY id
                    """
                )
                for row in legacy_rows:
                    cur.execute(
                        f"""
                        INSERT INTO {SHIPMENT_TABLE_NAME} (
                            entry_date, shipment_type, order_no, supplier_name, container_no, product, qty,
                            ready_date, shipment_date, due_date, status, vessel, notes, updated_on
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            row["entry_date"],
                            shipment_type,
                            row["order_no"],
                            row["supplier_name"],
                            row["container_no"],
                            row["product"],
                            row["qty"],
                            row["ready_date"],
                            row["shipment_date"],
                            row["due_date"],
                            row["status"],
                            row["vessel"],
                            row["notes"],
                            row["updated_on"],
                        ),
                    )

        self.db_conn.commit()

    def ensure_sales_table(self):
        if self.db_engine == "sqlserver":
            return

        cur = self.db_conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS sales (
                sale_date TEXT,
                customer_name TEXT,
                item_number TEXT,
                description TEXT,
                month_key TEXT,
                quantity REAL,
                price REAL,
                extended REAL
            )
            """
        )

        desired_columns = [
            ("sale_date", "TEXT"),
            ("customer_name", "TEXT"),
            ("item_number", "TEXT"),
            ("description", "TEXT"),
            ("month_key", "TEXT"),
            ("quantity", "REAL"),
            ("price", "REAL"),
            ("extended", "REAL"),
        ]
        existing_columns = {row[1] for row in cur.execute("PRAGMA table_info(sales)").fetchall()}
        for column_name, ddl in desired_columns:
            if column_name not in existing_columns:
                cur.execute(f"ALTER TABLE sales ADD COLUMN {column_name} {ddl}")

        cur.execute("CREATE INDEX IF NOT EXISTS idx_sales_item_date ON sales(item_number, sale_date)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_sales_customer_date ON sales(customer_name, sale_date)")
        self.db_conn.commit()

    def ensure_stock_table(self):
        if self.db_engine == "sqlserver":
            return

        cur = self.db_conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS stock (
                item_number TEXT,
                on_hand REAL,
                committed REAL,
                on_order REAL,
                available REAL
            )
            """
        )

        desired_columns = [
            ("item_number", "TEXT"),
            ("on_hand", "REAL"),
            ("committed", "REAL"),
            ("on_order", "REAL"),
            ("available", "REAL"),
        ]
        existing_columns = {row[1] for row in cur.execute("PRAGMA table_info(stock)").fetchall()}
        for column_name, ddl in desired_columns:
            if column_name not in existing_columns:
                cur.execute(f"ALTER TABLE stock ADD COLUMN {column_name} {ddl}")

        cur.execute("CREATE INDEX IF NOT EXISTS idx_stock_item_number ON stock(item_number)")
        self.db_conn.commit()

    def ensure_container_tables(self):
        if self.db_engine == "sqlserver":
            return

        cur = self.db_conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS containers (
                container_ref TEXT PRIMARY KEY,
                updated_on TEXT,
                eta_date TEXT,
                additional_cartons INTEGER DEFAULT 0,
                dog_leads INTEGER DEFAULT 0,
                notes TEXT DEFAULT ''
            )
            """
        )

        desired_columns = [
            ("id", "INTEGER PRIMARY KEY AUTOINCREMENT"),
            ("container_ref", "TEXT NOT NULL"),
            ("line_no", "INTEGER NOT NULL"),
            ("order_number", "TEXT"),
            ("item_number", "TEXT"),
            ("description", "TEXT"),
            ("qty", "REAL DEFAULT 0"),
            ("cartons", "INTEGER DEFAULT 0"),
            ("additional_cartons", "INTEGER DEFAULT 0"),
            ("urgent", "INTEGER DEFAULT 0"),
            ("additional", "INTEGER DEFAULT 0"),
        ]

        def create_container_lines_table():
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS container_lines (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    container_ref TEXT NOT NULL,
                    line_no INTEGER NOT NULL,
                    order_number TEXT,
                    item_number TEXT,
                    description TEXT,
                    qty REAL DEFAULT 0,
                    cartons INTEGER DEFAULT 0,
                    additional_cartons INTEGER DEFAULT 0,
                    urgent INTEGER DEFAULT 0,
                    additional INTEGER DEFAULT 0,
                    FOREIGN KEY(container_ref) REFERENCES containers(container_ref) ON DELETE CASCADE
                )
                """
            )

        table_exists = cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='container_lines'"
        ).fetchone()

        if not table_exists:
            create_container_lines_table()
        else:
            existing_info = cur.execute("PRAGMA table_info(container_lines)").fetchall()
            existing_cols = {row[1] for row in existing_info}
            required_cols = {name for name, _ in desired_columns}

            if not {"container_ref", "line_no"}.issubset(existing_cols):
                backup_name = "container_lines_legacy"
                suffix = 1
                while cur.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (backup_name,)
                ).fetchone():
                    suffix += 1
                    backup_name = f"container_lines_legacy_{suffix}"
                cur.execute(f"ALTER TABLE container_lines RENAME TO {backup_name}")
                create_container_lines_table()
            else:
                missing = [name for name, ddl in desired_columns if name not in existing_cols and name != "id"]
                for name in missing:
                    ddl = dict(desired_columns)[name]
                    cur.execute(f"ALTER TABLE container_lines ADD COLUMN {name} {ddl}")

        cur.execute(
            "CREATE INDEX IF NOT EXISTS idx_container_lines_ref ON container_lines(container_ref, line_no)"
        )
        self.db_conn.commit()

    def ensure_orders_table(self):
        if self.db_engine == "sqlserver":
            return

        cur = self.db_conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS orders (
                item_number TEXT,
                quantity REAL,
                purchase_no TEXT,
                order_date TEXT
            )
            """
        )

        existing_columns = {
            row[1] for row in cur.execute("PRAGMA table_info(orders)").fetchall()
        }
        if "order_date" not in existing_columns:
            cur.execute("ALTER TABLE orders ADD COLUMN order_date TEXT")
            if "date" in existing_columns:
                cur.execute(
                    "UPDATE orders SET order_date = COALESCE(NULLIF(TRIM(date), ''), order_date)"
                )

        rows_to_fix = cur.execute(
            "SELECT rowid, order_date FROM orders WHERE order_date IS NOT NULL AND TRIM(order_date) <> ''"
        ).fetchall()
        updates = []
        for row in rows_to_fix:
            try:
                rowid = row[0]
                raw_value = row[1]
            except Exception:
                rowid = row["rowid"]
                raw_value = row["order_date"]
            try:
                normalized = self.normalize_order_import_date(raw_value)
            except ValueError:
                continue
            if normalized and normalized != str(raw_value).strip():
                updates.append((normalized, rowid))
        if updates:
            cur.executemany("UPDATE orders SET order_date = ? WHERE rowid = ?", updates)

        self.db_conn.commit()

    def ensure_app_meta_table(self):
        if self.db_engine == "sqlserver":
            return

        cur = self.db_conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS app_meta (
                meta_key TEXT PRIMARY KEY,
                meta_value TEXT
            )
            """
        )
        self.db_conn.commit()

    def get_meta_value(self, key, default=""):
        row = self.db_one("SELECT meta_value FROM app_meta WHERE meta_key = ?", (key,))
        if row is None:
            return default
        try:
            return row["meta_value"]
        except Exception:
            return default

    def set_meta_value(self, key, value):
        cur = self.db_conn.cursor()
        value_text = "" if value is None else str(value)
        if self.db_engine == "sqlserver":
            cur.execute("UPDATE app_meta SET meta_value = ? WHERE meta_key = ?", (value_text, key))
            if getattr(cur, 'rowcount', 0) == 0:
                cur.execute("INSERT INTO app_meta (meta_key, meta_value) VALUES (?, ?)", (key, value_text))
        else:
            cur.execute(
                """
                INSERT INTO app_meta (meta_key, meta_value)
                VALUES (?, ?)
                ON CONFLICT(meta_key) DO UPDATE SET meta_value = excluded.meta_value
                """,
                (key, value_text),
            )
        self.db_conn.commit()

    def row_to_dict(self, row):
        if row is None:
            return {}
        return dict(row)

    def get_first(self, row, *keys, default=""):
        if row is None:
            return default
        for key in keys:
            if key in row and row[key] not in (None, ""):
                return row[key]
        return default

    def load_reference_lists(self):
        if not self.has_table("sales"):
            QMessageBox.critical(self, "Missing table", "The database does not contain a sales table.")
            return

        if self.has_table("customers"):
            customer_rows = self.db_all(
                """
                SELECT DISTINCT TRIM(customer_name) AS customer_name
                FROM customers
                WHERE customer_name IS NOT NULL AND TRIM(customer_name) <> ''
                ORDER BY customer_name COLLATE NOCASE
                """
            )
        else:
            customer_rows = self.db_all(
                """
                SELECT DISTINCT TRIM(customer_name) AS customer_name
                FROM sales
                WHERE customer_name IS NOT NULL AND TRIM(customer_name) <> ''
                ORDER BY customer_name COLLATE NOCASE
                """
            )
        self.customer_names = [r["customer_name"] for r in customer_rows if r["customer_name"]]

        item_rows = self.db_all(
            """
            SELECT item_number
            FROM (
                SELECT DISTINCT TRIM(item_number) AS item_number
                FROM items
                WHERE item_number IS NOT NULL AND TRIM(item_number) <> ''
                UNION
                SELECT DISTINCT TRIM(item_number) AS item_number
                FROM sales
                WHERE item_number IS NOT NULL AND TRIM(item_number) <> ''
            ) AS item_union
            ORDER BY item_number COLLATE NOCASE
            """
        )
        self.item_numbers = [r["item_number"] for r in item_rows if r["item_number"]]

        supplier_names = []
        supplier_seen = set()

        supplier_columns = self.get_items_supplier_column_names()
        if supplier_columns:
            supplier_selects = []
            for column_name in supplier_columns:
                supplier_selects.append(
                    f"SELECT DISTINCT TRIM([{column_name}]) AS supplier_name FROM items WHERE TRIM(COALESCE([{column_name}], '')) <> ''"
                )
            supplier_rows = self.db_all(
                "\nUNION\n".join(supplier_selects) + "\nORDER BY supplier_name COLLATE NOCASE"
            )
            for row in supplier_rows:
                supplier_name = (row["supplier_name"] or "").strip()
                if not supplier_name:
                    continue
                key = supplier_name.casefold()
                if key in supplier_seen:
                    continue
                supplier_seen.add(key)
                supplier_names.append(supplier_name)

        supplier_master_names = self.get_supplier_master_names()
        if not supplier_master_names and EMBEDDED_SUPPLIER_MASTER:
            supplier_master_names = list(EMBEDDED_SUPPLIER_MASTER)

        for supplier_name in supplier_master_names:
            supplier_name = (supplier_name or "").strip()
            if not supplier_name:
                continue
            key = supplier_name.casefold()
            if key in supplier_seen:
                continue
            supplier_seen.add(key)
            supplier_names.append(supplier_name)

        self.supplier_names = sorted(supplier_names, key=lambda value: value.casefold())

    # -----------------------------
    # UI configuration
    # -----------------------------
    def setup_customer_autocomplete(self):
        if not hasattr(self.ui, "customerEdit"):
            return
        completer = QCompleter(self.customer_names, self)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchStartsWith)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
        self.ui.customerEdit.setCompleter(completer)

    def setup_item_autocomplete(self):
        if not hasattr(self.ui, "enterItem"):
            return
        completer = QCompleter(self.item_numbers, self)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchStartsWith)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
        self.ui.enterItem.setCompleter(completer)

    def setup_item_summary_combine_checkbox(self):
        checkbox = getattr(self.ui, "combineThreads_checkBox", None)
        if checkbox is not None:
            checkbox.setToolTip(
                "Combine sales for the India thread code and the Liberty code with trailing L, for example BN40 101 + BN40 101 L."
            )
            self.combine_threads_checkbox = checkbox
            return

        frame = getattr(self.ui, "frame_17", None)
        layout = frame.layout() if frame is not None else None
        if layout is None:
            self.combine_threads_checkbox = None
            return
        if self.combine_threads_checkbox is None:
            checkbox = QCheckBox("Combine Threads", frame)
            checkbox.setObjectName("combineThreadsCheck")
            checkbox.setToolTip(
                "Combine sales for the India thread code and the Liberty code with trailing L, for example BN40 101 + BN40 101 L."
            )
            layout.addWidget(checkbox)
            self.combine_threads_checkbox = checkbox

    def is_combine_threads_enabled(self):
        return bool(self.combine_threads_checkbox is not None and self.combine_threads_checkbox.isChecked())

    def get_order_analysis_supplier_edit(self):
        return getattr(self.ui, "customer_lineEdit", None)

    def get_order_analysis_table(self):
        return getattr(self.ui, "tableWidget", None)

    def get_saba_customer_edit(self):
        return getattr(self.ui, "customerSaba_lineEdit", None)

    def get_saba_table(self):
        return getattr(self.ui, "sabaSales_table", None)

    def build_items_group_expression(self, alias=False):
        candidate_names = ["item_group", "group_name", "group", "itemgroup", "Item Group", "Group"]
        available_columns = {name.lower(): name for name in self.get_table_columns("items")}
        parts = []
        for candidate in candidate_names:
            actual = available_columns.get(candidate.lower())
            if actual:
                parts.append(f"NULLIF(TRIM([{actual}]), '')")
        if not parts:
            return "''" if alias else "''"
        expr = f"COALESCE({', '.join(parts)}, '')"
        return f"{expr} AS item_group" if alias else expr

    def build_items_supplier_expression(self, alias=False):
        candidate_names = ["supplier_name", "supplier_code", "Column1", "Supplier"]
        available_columns = {name.lower(): name for name in self.get_table_columns("items")}
        parts = []
        for candidate in candidate_names:
            actual = available_columns.get(candidate.lower())
            if actual:
                parts.append(f"NULLIF(TRIM([{actual}]), '')")
        if not parts:
            return "''" if alias else "''"
        expr = f"COALESCE({', '.join(parts)}, '')"
        return f"{expr} AS supplier_name" if alias else expr

    def get_items_supplier_column_names(self):
        candidate_names = ["supplier_name", "supplier_code", "Column1", "Supplier"]
        available_columns = {name.lower(): name for name in self.get_table_columns("items")}
        resolved = []
        seen = set()
        for candidate in candidate_names:
            actual = available_columns.get(candidate.lower())
            if not actual:
                continue
            key = actual.lower()
            if key in seen:
                continue
            seen.add(key)
            resolved.append(actual)
        return resolved

    def get_items_supplier_column_name(self):
        columns = self.get_items_supplier_column_names()
        return columns[0] if columns else None

    def build_items_supplier_match_clause(self):
        columns = self.get_items_supplier_column_names()
        if not columns:
            return "", []
        conditions = [f"UPPER(TRIM(COALESCE([{column_name}], ''))) = UPPER(TRIM(?))" for column_name in columns]
        return " OR ".join(conditions), columns

    def update_item_supplier_in_database(self, item_number, supplier_name):
        item_number = (item_number or "").strip()
        supplier_name = (supplier_name or "").strip()
        if supplier_name:
            self.ensure_supplier_exists(supplier_name)
        if not item_number:
            return False
        column_name = self.get_items_supplier_column_name()
        if not column_name:
            return False
        cur = self.db_conn.cursor()
        cur.execute(
            f"UPDATE items SET [{column_name}] = ? WHERE UPPER(TRIM(item_number)) = UPPER(TRIM(?))",
            (supplier_name, item_number),
        )
        self.db_conn.commit()
        return True

    def ensure_supplier_exists(self, supplier_name):
        supplier_name = (supplier_name or "").strip()
        if not supplier_name or self.db_conn is None or not self.has_table("supplier_master"):
            return False
        row = self.db_one(
            "SELECT TOP 1 supplier_name FROM supplier_master WHERE UPPER(TRIM(supplier_name)) = UPPER(TRIM(?))",
            (supplier_name,),
        )
        if row is not None:
            return False
        cur = self.db_conn.cursor()
        cur.execute("INSERT INTO supplier_master (supplier_name) VALUES (?)", (supplier_name,))
        self.db_conn.commit()
        return True

    def setup_supplier_autocomplete(self):
        edit = self.get_order_analysis_supplier_edit()
        if edit is None:
            return
        supplier_values = sorted({(name or "").strip() for name in self.supplier_names if (name or "").strip()}, key=lambda value: value.casefold())
        completer = QCompleter(supplier_values, self)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchStartsWith)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
        try:
            completer.setMaxVisibleItems(25)
        except Exception:
            pass
        edit.setCompleter(completer)
        self.order_analysis_supplier_completer = completer

    def setup_date_ranges(self):
        row = self.db_one(
            """
            SELECT MIN(DATE(sale_date)) AS min_date,
                   MAX(DATE(sale_date)) AS max_date
            FROM sales
            WHERE sale_date IS NOT NULL AND TRIM(CAST(sale_date AS TEXT)) <> ''
            """
        )

        if not row or not row["min_date"] or not row["max_date"]:
            today = date.today()
            start_qdate = QDate(today.year, today.month, 1)
            end_qdate = start_qdate
        else:
            min_date = self.parse_date_value(row["min_date"])
            max_date = self.parse_date_value(row["max_date"])
            if min_date is None or max_date is None:
                today = date.today()
                start_qdate = QDate(today.year, today.month, 1)
                end_qdate = start_qdate
            else:
                start_qdate = QDate(min_date.year, min_date.month, 1)
                end_qdate = QDate(max_date.year, max_date.month, 1)

        for picker in (self.customer_start_picker, self.item_start_picker):
            if picker is not None:
                picker.setDate(start_qdate)
        for picker in (self.customer_end_picker, self.item_end_picker):
            if picker is not None:
                picker.setDate(end_qdate)

    def setup_sales_table(self):
        table = getattr(self.ui, "salesTable", None)
        if table is None:
            return
        table.setRowCount(0)
        table.setColumnCount(0)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        header = table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setSectionsClickable(True)

    def setup_customer_info_table(self):
        table = getattr(self.ui, "customer_Info", None)
        if table is None:
            return

        model = QStandardItemModel(0, 1, self)
        table.setModel(model)
        table.horizontalHeader().setVisible(False)
        table.verticalHeader().setVisible(False)
        table.setShowGrid(False)
        table.setSelectionMode(QAbstractItemView.NoSelection)
        table.setFocusPolicy(Qt.NoFocus)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

    def setup_customer_purchase_table(self):
        table = getattr(self.ui, "customerPurchase_table", None)
        if table is None:
            return

        model = QStandardItemModel(0, 3, self)
        model.setHorizontalHeaderLabels(["Customer", "Last Sale Date", "Last Price"])
        table.setModel(model)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setWordWrap(False)
        header = table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionsClickable(True)
        for col in range(model.columnCount()):
            header.setSectionResizeMode(col, QHeaderView.Interactive)
            table.resizeColumnToContents(col)
        table.verticalHeader().setVisible(False)

    def setup_update_page(self):
        orders_status = getattr(self.ui, "lastUpdateOrders_textBrowser_3", None)
        if orders_status is not None:
            orders_status.setReadOnly(True)
            orders_status.setOpenLinks(False)
            orders_status.setPlainText(self.get_meta_value("orders_last_import_display", "No orders import yet."))

        sales_status = getattr(self.ui, "lastUpdateSales_textBrowser", None)
        if sales_status is not None:
            sales_status.setReadOnly(True)
            sales_status.setOpenLinks(False)
            sales_status.setPlainText(self.get_meta_value("sales_last_import_display", "No sales import yet."))

        stock_status = getattr(self.ui, "lastUPdateStock_textBrowser_2", None)
        if stock_status is not None:
            stock_status.setReadOnly(True)
            stock_status.setOpenLinks(False)
            stock_status.setPlainText(self.get_meta_value("stock_last_import_display", "No stock import yet."))

        freight_box = getattr(self.ui, "chargeFreight_textBrowser", None)
        if freight_box is not None:
            freight_box.setReadOnly(True)
            freight_box.setOpenLinks(False)
            freight_box.installEventFilter(self)
            freight_box.setCursor(Qt.PointingHandCursor)
            try:
                freight_box.viewport().installEventFilter(self)
                freight_box.viewport().setCursor(Qt.PointingHandCursor)
            except Exception:
                pass
            freight_box.setToolTip("Double-click to toggle Charge Freight for the current customer.")

    def setup_order_table(self):
        table = getattr(self.ui, "order_table", None)
        if table is None:
            return

        table.clear()
        table.setColumnCount(7)
        table.setHorizontalHeaderLabels(["Item Number", "Description", "On Order", "Qty", "Supplier", "Status", "Remove"])
        table.setRowCount(0)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        table.setProperty("_table_font_menu_installed", True)
        table.setProperty("_table_font_scope_key", "main_window")
        table.setProperty("_table_font_default_size", current_table_font_size(table))
        table.setProperty("_table_font_settings_token", "main_window/order_table")
        table.setContextMenuPolicy(Qt.CustomContextMenu)
        table.customContextMenuRequested.connect(self.show_order_table_context_menu)
        table.cellDoubleClicked.connect(self.handle_order_table_double_click)
        table.itemChanged.connect(self.handle_order_table_item_changed)

    def get_saved_container_refs(self):
        if self.db_conn is None or not self.has_table("containers"):
            return []
        rows = self.db_all(
            """
            SELECT container_ref
            FROM containers
            WHERE TRIM(COALESCE(container_ref, '')) <> ''
            ORDER BY COALESCE(CAST(updated_on AS date), CAST('1900-01-01' AS date)) DESC, container_ref
            """
        )
        refs = []
        seen = set()
        for row in rows:
            ref = str(row["container_ref"] or "").strip()
            if not ref:
                continue
            key = ref.casefold()
            if key in seen:
                continue
            seen.add(key)
            refs.append(ref)
        return refs

    def prompt_for_container_and_order(self, current_order_number=""):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add to Container")
        dialog.resize(420, 150)
        layout = QVBoxLayout(dialog)
        form = QFormLayout()

        container_combo = QComboBox(dialog)
        container_combo.setEditable(True)
        container_combo.addItems(self.get_saved_container_refs())
        container_combo.setCurrentIndex(-1)
        try:
            container_combo.setInsertPolicy(QComboBox.NoInsert)
            container_combo.setMaxVisibleItems(20)
        except Exception:
            pass
        completer = QCompleter(self.get_saved_container_refs(), dialog)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchStartsWith)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
        container_combo.setCompleter(completer)
        if container_combo.lineEdit() is not None:
            container_combo.lineEdit().setPlaceholderText("Select or type a container")
        form.addRow("Container", container_combo)

        order_edit = QLineEdit(dialog)
        order_edit.setText(current_order_number or "")
        order_edit.setPlaceholderText("Enter order number")
        form.addRow("Order Number", order_edit)

        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, parent=dialog)
        ok_button = buttons.button(QDialogButtonBox.Ok)
        if ok_button is not None:
            ok_button.setEnabled(False)
        layout.addWidget(buttons)

        def update_ok_state(*_args):
            has_container = bool((container_combo.currentText() or "").strip())
            has_order = bool((order_edit.text() or "").strip())
            if ok_button is not None:
                ok_button.setEnabled(has_container and has_order)

        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        order_edit.textChanged.connect(update_ok_state)
        if container_combo.lineEdit() is not None:
            container_combo.lineEdit().textChanged.connect(update_ok_state)
        update_ok_state()
        if container_combo.lineEdit() is not None:
            container_combo.lineEdit().setFocus()

        if dialog.exec() != QDialog.Accepted:
            return None, None

        container_ref = (container_combo.currentText() or "").strip()
        order_number = (order_edit.text() or "").strip()
        if not container_ref or not order_number:
            return None, None
        return container_ref, order_number

    def get_container_lines_for_ref(self, container_ref):
        ref = (container_ref or "").strip()
        if not ref or self.db_conn is None or not self.has_table("container_lines"):
            return []
        rows = self.db_all(
            """
            SELECT order_number, item_number, description, qty, cartons, additional_cartons, urgent, additional
            FROM container_lines
            WHERE UPPER(TRIM(container_ref)) = UPPER(TRIM(?))
            ORDER BY line_no, id
            """,
            (ref,),
        )
        return [
            {
                "order_number": row["order_number"] or "",
                "item_number": row["item_number"] or "",
                "description": row["description"] or "",
                "qty": self.parse_float(row["qty"]),
                "cartons": int(round(self.parse_float(row["cartons"]))),
                "additional_cartons": int(round(self.parse_float(row["additional_cartons"]))),
                "urgent": bool(row["urgent"]),
                "additional": bool(row["additional"]),
            }
            for row in rows
            if str(row["item_number"] or "").strip()
        ]

    def append_line_to_container_ref(self, container_ref, order_number, item_number, qty_value, is_urgent=False):
        ref = (container_ref or "").strip()
        order_number = (order_number or "").strip()
        item_number = (item_number or "").strip()
        if not ref:
            raise ValueError("Container reference is required.")
        if not order_number:
            raise ValueError("Order number is required.")
        if not item_number:
            raise ValueError("Item number is required.")
        if self.db_conn is None:
            raise ValueError("Database is not available.")

        if self.current_container_ref:
            try:
                self.save_current_container_state()
            except Exception:
                pass

        item_row = self.get_item_master_row(item_number)
        description = self.get_first(item_row, "item_name", "Item Name", "description", "Description")
        carton_size = self.get_item_carton_size(item_row)
        requested_qty = self.parse_float(qty_value)
        if requested_qty <= 0:
            raise ValueError("Quantity must be greater than 0.")
        cartons = max(1, math.ceil(requested_qty / carton_size))
        rounded_qty = cartons * carton_size

        row = self.db_one("SELECT * FROM containers WHERE UPPER(TRIM(container_ref)) = UPPER(TRIM(?))", (ref,))
        cur = self.db_conn.cursor()
        today_iso = date.today().isoformat()
        if row is None:
            cur.execute(
                "INSERT INTO containers (container_ref, updated_on, eta_date, additional_cartons, dog_leads, notes) VALUES (?, ?, ?, ?, ?, ?)",
                (ref, today_iso, today_iso, 0, 0, ""),
            )
            dog_leads = False
            notes = ""
            additional_needed = 0
        else:
            dog_leads = bool(row["dog_leads"])
            notes = (row["notes"] or "").strip()
            additional_needed = row["additional_cartons"] or 0

        lines = self.get_container_lines_for_ref(ref)
        lines.append({
            "order_number": order_number,
            "item_number": item_number,
            "description": description,
            "qty": rounded_qty,
            "cartons": cartons,
            "additional_cartons": 0,
            "urgent": bool(is_urgent),
            "additional": False,
        })

        cur.execute(
            "UPDATE containers SET updated_on = ?, eta_date = COALESCE(NULLIF(eta_date, ''), ?), additional_cartons = ?, dog_leads = ?, notes = ? WHERE UPPER(TRIM(container_ref)) = UPPER(TRIM(?))",
            (today_iso, today_iso, additional_needed, 1 if dog_leads else 0, notes, ref),
        )
        cur.execute("DELETE FROM container_lines WHERE UPPER(TRIM(container_ref)) = UPPER(TRIM(?))", (ref,))
        for line_no, line in enumerate(lines, start=1):
            cur.execute(
                """
                INSERT INTO container_lines (
                    container_ref, line_no, order_number, item_number, description, qty, cartons,
                    additional_cartons, urgent, additional
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    ref,
                    line_no,
                    line.get("order_number", ""),
                    line.get("item_number", ""),
                    line.get("description", ""),
                    float(line.get("qty", 0) or 0),
                    int(line.get("cartons", 0) or 0),
                    int(line.get("additional_cartons", 0) or 0),
                    1 if line.get("urgent") else 0,
                    1 if line.get("additional") else 0,
                ),
            )
        self.db_conn.commit()
        self.refresh_containers_list()
        if (self.current_container_ref or "").strip().casefold() == ref.casefold():
            self.reload_current_container_state()
        self.rerun_order_analysis_if_ready()

    def get_order_row_data(self, row):
        table = getattr(self.ui, "order_table", None)
        if table is None or row < 0 or row >= table.rowCount():
            return None
        item_number = table.item(row, 0).text().strip() if table.item(row, 0) is not None and table.item(row, 0).text() else ""
        description = table.item(row, 1).text().strip() if table.item(row, 1) is not None and table.item(row, 1).text() else ""
        qty_item = table.item(row, self.order_qty_column)
        qty_value = qty_item.data(Qt.UserRole) if qty_item is not None else None
        if qty_value in (None, "") and qty_item is not None:
            qty_value = qty_item.text()
        supplier_item = table.item(row, self.order_supplier_column)
        supplier_value = str(supplier_item.data(Qt.UserRole) or supplier_item.text() or "").strip() if supplier_item is not None else ""
        priority_item = table.item(row, self.order_priority_column)
        priority_text = priority_item.text().strip().upper() if priority_item is not None and priority_item.text() else ""
        if not item_number:
            return None
        return {
            "item_number": item_number,
            "description": description,
            "qty": self.parse_float(qty_value),
            "supplier_name": supplier_value,
            "urgent": priority_text == "URGENT",
            "status": self.normalise_order_status(priority_text),
            "on_order": self.get_item_on_order_qty(item_number),
        }

    def get_item_on_order_qty(self, item_number):
        if self.db_conn is None or not item_number or not self.has_table("orders"):
            return 0.0
        total_qty, _order_date = self.get_orders_table_item_summary(item_number)
        return total_qty

    def show_order_table_context_menu(self, pos):
        table = getattr(self.ui, "order_table", None)
        if table is None:
            return

        clicked_item = table.itemAt(pos)
        row = clicked_item.row() if clicked_item is not None else table.currentRow()
        has_row = row is not None and row >= 0
        if has_row:
            table.selectRow(row)

        menu = QMenu(table)
        create_order_action = menu.addAction("Create Order")
        add_to_container_action = menu.addAction("Add to Container")
        create_order_action.setEnabled(bool(has_row))
        add_to_container_action.setEnabled(bool(has_row))
        menu.addSeparator()

        font_menu = menu.addMenu("Font Size")
        current_size = current_table_font_size(table)
        for size in TABLE_FONT_SIZE_OPTIONS:
            action = font_menu.addAction(str(size))
            action.setCheckable(True)
            action.setChecked(size == current_size)
            action.triggered.connect(
                lambda _checked=False, tbl=table, selected_size=size: apply_table_font_size(
                    tbl,
                    selected_size,
                    settings=self.settings,
                    scope_key="main_window",
                    persist=True,
                )
            )
        font_menu.addSeparator()
        reset_action = font_menu.addAction("Reset")
        reset_action.triggered.connect(
            lambda _checked=False, tbl=table: reset_table_font_size(
                tbl,
                settings=self.settings,
                scope_key="main_window",
            )
        )

        viewport = getattr(table, "viewport", lambda: None)()
        global_pos = viewport.mapToGlobal(pos) if viewport is not None else table.mapToGlobal(pos)
        chosen_action = menu.exec(global_pos)
        if chosen_action is None or not has_row:
            return

        row_data = self.get_order_row_data(row)
        if row_data is None:
            return

        if chosen_action == create_order_action:
            default_po = str(row_data.get("order_number", "") or "").strip()
            order_number, accepted = QInputDialog.getText(
                self,
                "Create YU Order",
                "Order Number:",
                QLineEdit.Normal,
                default_po,
            )
            order_number = (order_number or "").strip()
            if not accepted or not order_number:
                return
            self.open_yu_order_entry_dialog(
                initial_order_number=order_number,
                initial_lines=[{
                    "item_number": row_data.get("item_number", ""),
                    "qty": row_data.get("qty", 0),
                }],
            )
            return

        if chosen_action == add_to_container_action:
            container_ref, order_number = self.prompt_for_container_and_order(row_data.get("order_number", ""))
            if not container_ref or not order_number:
                return
            try:
                self.append_line_to_container_ref(
                    container_ref=container_ref,
                    order_number=order_number,
                    item_number=row_data.get("item_number", ""),
                    qty_value=row_data.get("qty", 0),
                    is_urgent=bool(row_data.get("urgent")),
                )
            except Exception as exc:
                QMessageBox.warning(self, "Add to Container", str(exc))
                return

            remove_result = QMessageBox.question(
                self,
                "Added to Container",
                f"Added {row_data.get('item_number', '')} to {container_ref} with order number {order_number}.\n\nRemove it from To Order?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if remove_result == QMessageBox.Yes:
                table.removeRow(row)
            else:
                self._updating_order_table = True
                table.setItem(row, self.order_priority_column, self.build_order_status_item("IN CONTAINER"))
                self._updating_order_table = False
            self.save_order_table_state()
            self.refresh_item_summary_context_boxes()
            self.rerun_order_analysis_if_ready()
            return

    def setup_order_analysis_table(self):
        table = self.get_order_analysis_table()
        if table is None:
            return

        headers = [
            "Item Number",
            "Item Name",
            "Sales for Period",
            "Avg Monthly Sales",
            "SOH",
            "Stock On Order",
            "On Order Form",
            "On Next Container",
            "Shipped Container",
            "Suggested Order",
            "At Risk",
        ]
        table.clear()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(0)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(self.on_order_order_number_column, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(self.on_order_item_column, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(self.on_order_description_column, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(self.on_order_comments_column, QHeaderView.Stretch)
        for col in (self.on_order_qty_column, self.on_order_supplier_column, self.on_order_ready_date_column, self.on_order_status_column, self.on_order_add_column, self.on_order_remove_column):
            table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeToContents)

        self.setup_order_analysis_clear_button()
        self.setup_order_analysis_export_button()

    def setup_order_analysis_clear_button(self):
        supplier_edit = self.get_order_analysis_supplier_edit()
        if supplier_edit is None:
            return
        parent = supplier_edit.parentWidget()
        layout = parent.layout() if parent is not None else None
        if layout is None:
            return
        existing = getattr(self.ui, "orderAnalysisClear_button", None)
        if existing is None:
            clear_button = QPushButton("Clear", parent)
            clear_button.setObjectName("orderAnalysisClear_button")
            clear_button.clicked.connect(self.clear_order_analysis_page)
            layout.addWidget(clear_button)
            self.ui.orderAnalysisClear_button = clear_button

    def setup_order_analysis_export_button(self):
        supplier_edit = self.get_order_analysis_supplier_edit()
        if supplier_edit is None:
            return
        parent = supplier_edit.parentWidget()
        layout = parent.layout() if parent is not None else None
        if layout is None:
            return
        existing = getattr(self.ui, "orderAnalysisExport_button", None)
        if existing is None:
            export_button = QPushButton("Export to Excel", parent)
            export_button.setObjectName("orderAnalysisExport_button")
            export_button.clicked.connect(self.export_order_analysis_to_excel)
            layout.addWidget(export_button)
            self.ui.orderAnalysisExport_button = export_button

    def clear_order_analysis_page(self):
        supplier_edit = self.get_order_analysis_supplier_edit()
        table = self.get_order_analysis_table()
        if supplier_edit is not None:
            supplier_edit.clear()
            supplier_edit.setFocus()
        if table is not None:
            table.setRowCount(0)
            table.clearSelection()
        self.current_order_analysis_supplier = None

    def export_order_analysis_to_excel(self):
        if Workbook is None:
            QMessageBox.warning(self, "Export Order Analysis", "openpyxl is not installed, so Excel export is not available.")
            return

        table = self.get_order_analysis_table()
        if table is None or table.rowCount() == 0 or table.columnCount() == 0:
            QMessageBox.warning(self, "Export Order Analysis", "There is no order analysis data to export.")
            return

        supplier_name = (self.current_order_analysis_supplier or "").strip()
        safe_supplier = re.sub(r"[^A-Za-z0-9._-]+", "_", supplier_name or "order_analysis").strip("_") or "order_analysis"
        default_filename = f"{safe_supplier}_order_analysis.xlsx"
        home_dir = str(Path.home() / default_filename)
        export_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Order Analysis",
            home_dir,
            "Excel Workbook (*.xlsx)",
        )
        if not export_path:
            return
        if not export_path.lower().endswith(".xlsx"):
            export_path += ".xlsx"

        start_picker = getattr(self, "item_start_picker", None) or getattr(self, "item_end_picker", None)
        end_picker = getattr(self, "item_end_picker", None) or getattr(self, "item_start_picker", None)
        period_from = self.month_start_from_picker(start_picker) if start_picker is not None else None
        period_to = self.month_end_from_picker(end_picker) if end_picker is not None else None
        months_in_period = 0
        if period_from is not None and period_to is not None and period_from <= period_to:
            months_in_period = max(1, len(self.month_list_between(period_from, period_to)))

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Order Analysis"

        headers = []
        for column in range(table.columnCount()):
            header_item = table.horizontalHeaderItem(column)
            headers.append(header_item.text() if header_item is not None and header_item.text() else f"Column {column + 1}")

        meta_rows = [
            ("Supplier", supplier_name or ""),
            ("Period From", period_from.strftime("%d/%m/%Y") if period_from is not None else ""),
            ("Period To", period_to.strftime("%d/%m/%Y") if period_to is not None else ""),
            ("Months in Period", months_in_period if months_in_period else ""),
        ]
        for row_index, (label, value) in enumerate(meta_rows, start=1):
            label_cell = worksheet.cell(row=row_index, column=1, value=label)
            value_cell = worksheet.cell(row=row_index, column=2, value=value)
            if Font is not None:
                label_cell.font = Font(bold=True)
            if Alignment is not None:
                label_cell.alignment = Alignment(horizontal="left", vertical="center")
                value_cell.alignment = Alignment(horizontal="left", vertical="center")

        header_row = len(meta_rows) + 2
        for column, header_text in enumerate(headers, start=1):
            cell = worksheet.cell(row=header_row, column=column, value=header_text)
            if Font is not None:
                cell.font = Font(bold=True)
            if Alignment is not None:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell_border is not None:
                cell.border = cell_border

        numeric_columns = {
            self.order_analysis_columns["sales_for_period"],
            self.order_analysis_columns["avg_monthly_sales"],
            self.order_analysis_columns["soh"],
            self.order_analysis_columns["stock_on_order"],
            self.order_analysis_columns["on_order_form"],
            self.order_analysis_columns["on_next_container"],
            self.order_analysis_columns["shipped_container"],
            self.order_analysis_columns["suggested_order"],
            self.order_analysis_columns["at_risk"],
        }

        first_data_row = header_row + 1
        for row in range(table.rowCount()):
            for column in range(table.columnCount()):
                item = table.item(row, column)
                cell = worksheet.cell(row=first_data_row + row, column=column + 1)
                text_value = item.text().strip() if item is not None and item.text() else ""
                if column in numeric_columns:
                    numeric_value = self.parse_float(text_value)
                    cell.value = int(numeric_value) if float(numeric_value).is_integer() else numeric_value
                    cell.number_format = '#,##0.###'
                else:
                    cell.value = text_value

        if table.rowCount() > 0 and table.columnCount() > 0:
            worksheet.auto_filter.ref = (
                f"A{header_row}:{get_column_letter(table.columnCount())}{first_data_row + table.rowCount() - 1}"
            )
        worksheet.freeze_panes = f"A{first_data_row}"

        for column in range(1, worksheet.max_column + 1):
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                value = worksheet.cell(row=row, column=column).value
                length = len(str(value)) if value is not None else 0
                if length > max_length:
                    max_length = length
            worksheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 12), 40)

        key_sheet = workbook.create_sheet("Key")
        key_sheet.cell(row=1, column=1, value="Field")
        key_sheet.cell(row=1, column=2, value="Meaning / Calculation")
        if Font is not None:
            key_sheet["A1"].font = Font(bold=True)
            key_sheet["B1"].font = Font(bold=True)

        key_rows = [
            ("Sales for Period", "Total units sold between Period From and Period To."),
            ("Avg Monthly Sales", "Sales for Period divided by Months in Period."),
            ("SOH", "Stock on hand from the stock table."),
            ("Stock On Order", "General on-order quantity from the stock table."),
            ("On Order Form", "Inbound quantity currently on the order form within the calculation horizon."),
            ("On Next Container", "Inbound quantity assigned to the next container within the calculation horizon."),
            ("Shipped Container", "Inbound quantity already shipped and due within the calculation horizon."),
            ("Suggested Order", "Max(0, lead-time demand - (SOH + hard inbound by horizon)), then rounded to carton/pallet rules."),
            ("At Risk", "Max(0, demand until the earliest inbound or cutoff - supply available before that cutoff), then rounded to carton/pallet rules."),
            ("Lead-time demand", "Avg Monthly Sales multiplied by lead days divided by 30.4375."),
            ("Hard inbound", "Inbound quantities included in calculations up to the horizon date."),
            ("Rounding", "Order quantities are rounded using the item carton size and pallet size rules."),
        ]
        for row_index, (field_name, description) in enumerate(key_rows, start=2):
            key_sheet.cell(row=row_index, column=1, value=field_name)
            key_sheet.cell(row=row_index, column=2, value=description)
            if Alignment is not None:
                key_sheet.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        key_sheet.column_dimensions["A"].width = 24
        key_sheet.column_dimensions["B"].width = 110

        try:
            workbook.save(export_path)
        except Exception as exc:
            QMessageBox.critical(self, "Export Order Analysis", f"Could not export order analysis to Excel.\n\n{exc}")
            return

        QMessageBox.information(self, "Export Order Analysis", f"Order analysis exported to:\n{export_path}")

    def setup_saba_review_page(self):
        table = self.get_saba_table()
        if table is not None:
            headers = [
                "Customer",
                "Item Number",
                "Item Name",
                "Pack",
                "Avg Weeks Between Purchases",
                "Weeks Since Last Purchase",
                "Last Purchase Date",
                "Purchase Dates",
                "Status",
            ]
            model = QStandardItemModel(0, len(headers), self)
            model.setHorizontalHeaderLabels(headers)
            table.setModel(model)
            table.setSelectionBehavior(QAbstractItemView.SelectRows)
            table.setSelectionMode(QAbstractItemView.SingleSelection)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            table.setAlternatingRowColors(False)
            table.verticalHeader().setVisible(False)
            header = table.horizontalHeader()
            header.setStretchLastSection(False)
            header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
            header.setSectionResizeMode(2, QHeaderView.Stretch)
            for col in range(3, len(headers)):
                header.setSectionResizeMode(col, QHeaderView.ResizeToContents)

        avg_label = getattr(self.ui, "weeksPurchase_label", None)
        if avg_label is not None:
            avg_label.setText("Average weeks between purchases")

        last_label = getattr(self.ui, "weeksSinceLast_label", None)
        if last_label is not None:
            last_label.setText("Weeks since last purchase")

        edit = self.get_saba_customer_edit()
        if edit is not None:
            edit.setPlaceholderText("Enter customer or type All")
            completer_values = ["All"] + list(self.customer_names)
            completer = QCompleter(completer_values, self)
            completer.setCaseSensitivity(Qt.CaseInsensitive)
            completer.setFilterMode(Qt.MatchStartsWith)
            completer.setCompletionMode(QCompleter.PopupCompletion)
            completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
            edit.setCompleter(completer)
            self.saba_customer_completer = completer

        layout = getattr(self.ui, "horizontalLayout_14", None)
        parent = getattr(self.ui, "frame_39", None)
        if layout is not None and parent is not None and self.saba_show_all_checkbox is None:
            checkbox = QCheckBox("All Customers", parent)
            checkbox.setObjectName("sabaAllCustomers_checkBox")
            layout.addWidget(checkbox)
            self.saba_show_all_checkbox = checkbox

        for object_name in (
            "bagInBox_textBrowser",
            "keg_textBrowser",
            "ibc_textBrowser",
            "bagInBoxWeeksLast_textBrowser",
            "kegWeeksLast_textBrowser",
            "ibcWeeksLast_textBrowser",
        ):
            widget = getattr(self.ui, object_name, None)
            if widget is not None:
                widget.setReadOnly(True)
                widget.setOpenLinks(False)
                widget.setPlainText("No matching rows")

    def setup_container_table(self):
        table = getattr(self.ui, "container_table", None)
        if table is None:
            return

        self._updating_container_table = True
        table.clearSpans()
        table.clear()
        table.setColumnCount(9)
        table.setHorizontalHeaderLabels([
            "Order Number", "Item Number", "Description", "Qty", "Cartons",
            "Additional Cartons", "Urgent", "Additional", "Remove"
        ])
        table.setRowCount(0)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.setSortingEnabled(False)
        table.horizontalHeader().setSectionsClickable(True)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        if not bool(table.property("_container_header_sort_connected")):
            table.horizontalHeader().sectionClicked.connect(self.handle_container_header_sort_clicked)
            table.setProperty("_container_header_sort_connected", True)
        self._updating_container_table = False
        self.refresh_container_note_rows()
        self.refresh_container_totals()

    def container_sort_key_for_line(self, line, column):
        cols = self.container_columns
        if column == cols["order"]:
            return (line.get("order_number", "") or "").strip().upper()
        if column == cols["item"]:
            return (line.get("item_number", "") or "").strip().upper()
        if column == cols["description"]:
            return (line.get("description", "") or "").strip().upper()
        if column == cols["qty"]:
            return self.parse_float(line.get("qty", 0))
        if column == cols["cartons"]:
            return self.parse_float(line.get("cartons", 0))
        if column == cols["additional_cartons"]:
            return self.parse_float(line.get("additional_cartons", 0))
        if column == cols["urgent"]:
            return 1 if bool(line.get("urgent")) else 0
        if column == cols["additional"]:
            return 1 if bool(line.get("additional")) else 0
        return ""

    def handle_container_header_sort_clicked(self, column):
        cols = self.container_columns
        if column == cols["remove"]:
            return
        if self.container_sort_column == column:
            self.container_sort_descending = not self.container_sort_descending
        else:
            self.container_sort_column = column
            self.container_sort_descending = False

        lines = self.get_container_line_dicts()
        lines.sort(
            key=lambda line: self.container_sort_key_for_line(line, column),
            reverse=self.container_sort_descending,
        )
        self.populate_container_table(lines)

    def apply_container_sort_if_needed(self, lines):
        if self.container_sort_column is None:
            return list(lines)
        sorted_lines = list(lines)
        sorted_lines.sort(
            key=lambda line: self.container_sort_key_for_line(line, self.container_sort_column),
            reverse=self.container_sort_descending,
        )
        return sorted_lines

    def setup_containers_list_table(self):
        table = getattr(self.ui, "containers_tableWidget", None)
        if table is None:
            return
        table.clear()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Container", "ETA"])
        table.setRowCount(0)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.refresh_containers_list()

    def setup_container_entry(self):
        ref_widget = getattr(self.ui, "nextContainer_box", None)
        updated_box = getattr(self.ui, "updated_box", None)
        order_widget = getattr(self.ui, "orderNumberContainer_line", None)
        item_widget = getattr(self.ui, "itemNumberContainer_line", None)
        qty_widget = getattr(self.ui, "qtyContainder_line", None)
        eta_widget = getattr(self.ui, "eta_dateEdit", None)

        if ref_widget is not None:
            ref_widget.installEventFilter(self)
            ref_widget.setAcceptRichText(False)
            ref_widget.setTabChangesFocus(True)

        if updated_box is not None:
            updated_box.setOpenLinks(False)
            updated_box.setReadOnly(True)

        if qty_widget is not None:
            validator = QDoubleValidator(0.0, 999999999.0, 3, qty_widget)
            validator.setNotation(QDoubleValidator.StandardNotation)
            qty_widget.setValidator(validator)

        if item_widget is not None:
            completer = QCompleter(self.item_numbers, self)
            completer.setCaseSensitivity(Qt.CaseInsensitive)
            completer.setFilterMode(Qt.MatchStartsWith)
            completer.setCompletionMode(QCompleter.PopupCompletion)
            completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
            item_widget.setCompleter(completer)
            completer.activated.connect(self.container_item_completion_selected)
            self.container_item_completer = completer

        if eta_widget is not None and not eta_widget.date().isValid():
            eta_widget.setDate(QDate.currentDate())
        if eta_widget is not None:
            eta_widget.setDisplayFormat("dd/MM/yyyy")

        if ref_widget is not None:
            ref_widget.setFocusPolicy(Qt.StrongFocus)

    def make_container_table_item(self, text, align=None, bold=False, background=None, foreground=None, data=None):
        item = QTableWidgetItem("" if text is None else str(text))
        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        if align is not None:
            item.setTextAlignment(align)
        if background is not None:
            item.setBackground(background)
        if foreground is not None:
            item.setForeground(foreground)
        if bold:
            f = item.font()
            f.setBold(True)
            item.setFont(f)
        if data is not None:
            item.setData(Qt.UserRole, data)
        return item

    def update_charge_freight_box(self, should_charge, raw_value=""):
        box = getattr(self.ui, "chargeFreight_textBrowser", None)
        if box is None:
            return

        text_value = (raw_value or "").strip()
        if should_charge:
            box.setHtml('<div style="text-align:center; margin-top:38px; font-size:22px; font-weight:700; color:#8b0000;">Charge Freight</div>')
            box.setStyleSheet(
                "QTextBrowser {"
                "background: #fff1f1;"
                "color: #8b0000;"
                "border: 4px solid #d62828;"
                "border-radius: 8px;"
                "padding: 10px;"
                "}"
            )
            box.setToolTip(f"Freight flag: {text_value or 'Yes'}")
        else:
            box.clear()
            box.setStyleSheet(
                "QTextBrowser {"
                "background: transparent;"
                "color: inherit;"
                "border: 1px solid #b7bcc2;"
                "border-radius: 8px;"
                "padding: 10px;"
                "}"
            )
            box.setToolTip(f"Freight flag: {text_value or 'No'}")

    def is_yes_like(self, value):
        text_value = str(value or "").strip().lower()
        return text_value in {"1", "true", "t", "y", "yes", "charge freight", "charged"}

    def get_customer_freight_value(self, row):
        return self.get_first(
            row,
            "freight",
            "Freight",
            "charge_freight",
            "Charge Freight",
            "freight_flag",
            "Freight Flag",
            default="",
        )


    def get_customer_freight_column_name(self):
        columns = {name.lower(): name for name in self.get_table_columns("customers")}
        for candidate in ("charge_freight", "freight", "freight_flag", "charge freight", "freight flag"):
            actual = columns.get(candidate.lower())
            if actual:
                return actual
        return None

    def toggle_charge_freight_for_current_customer(self):
        customer_name = (self.current_customer_name or "").strip()
        if not customer_name:
            return False

        column_name = self.get_customer_freight_column_name()
        if not column_name:
            QMessageBox.warning(
                self,
                "Charge Freight",
                "Could not find a freight flag column in the customers table.",
            )
            return True

        row = self.row_to_dict(
            self.db_one(
                "SELECT * FROM customers WHERE LOWER(TRIM(customer_name)) = LOWER(TRIM(?)) LIMIT 1",
                (customer_name,),
            )
        )
        current_value = self.get_customer_freight_value(row)
        new_value = "No" if self.is_yes_like(current_value) else "Yes"

        cur = self.db_conn.cursor()
        cur.execute(
            f"UPDATE customers SET [{column_name}] = ? WHERE LOWER(TRIM(customer_name)) = LOWER(TRIM(?))",
            (new_value, customer_name),
        )
        self.db_conn.commit()
        self.populate_customer_info(customer_name)
        return True

    def get_container_ref_text(self):
        widget = getattr(self.ui, "nextContainer_box", None)
        return widget.toPlainText().strip() if widget is not None else ""

    def set_container_ref_text(self, text):
        widget = getattr(self.ui, "nextContainer_box", None)
        if widget is not None:
            widget.setPlainText(text or "")

    def display_to_iso_date(self, text_value):
        text_value = (text_value or "").strip()
        if not text_value:
            return ""
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text_value, fmt).date().isoformat()
            except ValueError:
                pass
        return ""

    def iso_to_display_date(self, text_value):
        parsed = self.parse_date_value(text_value)
        if parsed is not None:
            return parsed.strftime("%d/%m/%Y")
        text_value = "" if text_value is None else str(text_value).strip()
        return text_value

    def set_updated_box_today(self):
        self.set_label_text("updated_box", date.today().strftime("%d/%m/%Y"))

    def set_container_meta_widgets(self, updated_display="", eta_iso="", additional_value=0, dog_leads=False):
        blockers = []
        eta_widget = getattr(self.ui, "eta_dateEdit", None)
        additional_spinner = getattr(self.ui, "additional_spinner", None)
        dog_leads_check = getattr(self.ui, "checkBox", None)
        for widget in (eta_widget, additional_spinner, dog_leads_check):
            if widget is not None:
                blockers.append(QSignalBlocker(widget))
        try:
            self.set_label_text("updated_box", updated_display)
            if eta_widget is not None:
                if eta_iso:
                    parsed = QDate.fromString(eta_iso, "yyyy-MM-dd")
                    if parsed.isValid():
                        eta_widget.setDate(parsed)
                else:
                    eta_widget.setDate(QDate.currentDate())
            if additional_spinner is not None:
                additional_spinner.setValue(int(self.parse_float(additional_value)))
            if dog_leads_check is not None:
                dog_leads_check.setChecked(bool(dog_leads))
        finally:
            del blockers

    def focus_container_item_input(self):
        item_widget = getattr(self.ui, "itemNumberContainer_line", None)
        if item_widget is not None:
            item_widget.setFocus()
            item_widget.selectAll()

    def focus_container_qty_input(self):
        qty_widget = getattr(self.ui, "qtyContainder_line", None)
        if qty_widget is not None:
            qty_widget.setFocus()
            qty_widget.selectAll()

    def container_item_completion_selected(self, completion_text):
        item_widget = getattr(self.ui, "itemNumberContainer_line", None)
        item_number = self.find_item_number(completion_text)
        if item_widget is None or not item_number:
            return
        item_widget.setText(item_number)
        item_widget.setCursorPosition(len(item_number))
        if self.container_item_completer is not None:
            self.container_item_completer.popup().hide()
        self.focus_container_qty_input()

    def resolve_container_item_and_focus_qty(self):
        item_widget = getattr(self.ui, "itemNumberContainer_line", None)
        typed = item_widget.text().strip() if item_widget is not None else ""
        item_number = self.find_item_number(typed)
        if not item_number:
            if typed:
                QMessageBox.warning(self, "Invalid item", "Please choose a valid item number from the list.")
            return False
        self.container_item_completion_selected(item_number)
        return True

    def get_item_carton_size(self, item_row):
        carton_size = self.parse_float(self.get_first(item_row, "carton", "CARTON", default=0))
        return carton_size if carton_size > 0 else 1.0

    def get_container_line_dicts(self):
        table = getattr(self.ui, "container_table", None)
        if table is None:
            return []
        lines = []
        cols = self.container_columns
        for row in range(table.rowCount()):
            first_item = table.item(row, 0)
            meta = first_item.data(Qt.UserRole) if first_item is not None else None
            if isinstance(meta, dict) and meta.get("note_row"):
                continue
            line = {
                "order_number": first_item.text().strip() if first_item is not None and first_item.text() else "",
                "item_number": table.item(row, cols["item"]).text().strip() if table.item(row, cols["item"]) is not None and table.item(row, cols["item"]).text() else "",
                "description": table.item(row, cols["description"]).text().strip() if table.item(row, cols["description"]) is not None and table.item(row, cols["description"]).text() else "",
                "qty": self.parse_float(table.item(row, cols["qty"]).text() if table.item(row, cols["qty"]) is not None else 0),
                "cartons": int(round(self.parse_float(table.item(row, cols["cartons"]).text() if table.item(row, cols["cartons"]) is not None else 0))),
                "additional_cartons": int(round(self.parse_float(table.item(row, cols["additional_cartons"]).text() if table.item(row, cols["additional_cartons"]) is not None else 0))),
                "urgent": bool(meta.get("urgent")) if isinstance(meta, dict) else False,
                "additional": bool(meta.get("additional")) if isinstance(meta, dict) else False,
            }
            if line["item_number"]:
                lines.append(line)
        return lines

    def populate_container_table(self, lines):
        table = getattr(self.ui, "container_table", None)
        if table is None:
            return
        lines = self.apply_container_sort_if_needed(lines)
        cols = self.container_columns
        self._updating_container_table = True
        table.clearSpans()
        table.setRowCount(0)
        for line in lines:
            row = table.rowCount()
            table.insertRow(row)
            meta = {
                "urgent": bool(line.get("urgent")),
                "additional": bool(line.get("additional")),
                "note_row": False,
            }
            table.setItem(row, cols["order"], self.make_container_table_item(line.get("order_number", ""), data=meta))
            table.setItem(row, cols["item"], self.make_container_table_item(line.get("item_number", "")))
            table.setItem(row, cols["description"], self.make_container_table_item(line.get("description", "")))
            table.setItem(row, cols["qty"], self.make_container_table_item(self.format_value(line.get("qty", 0)), align=Qt.AlignRight | Qt.AlignVCenter))
            table.setItem(row, cols["cartons"], self.make_container_table_item(self.format_value(line.get("cartons", 0)), align=Qt.AlignRight | Qt.AlignVCenter))
            table.setItem(row, cols["additional_cartons"], self.make_container_table_item(self.format_value(line.get("additional_cartons", 0)), align=Qt.AlignRight | Qt.AlignVCenter))
            table.setItem(row, cols["urgent"], self.make_container_table_item("Urgent", align=Qt.AlignCenter, bold=True))
            table.setItem(row, cols["additional"], self.make_container_table_item("Additional", align=Qt.AlignCenter, bold=True))
            table.setItem(row, cols["remove"], self.make_container_table_item("Remove", align=Qt.AlignCenter, bold=True))
            self.apply_container_row_style(row, bool(line.get("urgent")), bool(line.get("additional")))
        self.append_container_note_rows()
        self._updating_container_table = False
        table.resizeRowsToContents()
        self.refresh_container_totals()

    def append_container_note_rows(self):
        table = getattr(self.ui, "container_table", None)
        if table is None:
            return
        notes = []
        if getattr(self.ui, "checkBox", None) is not None and self.ui.checkBox.isChecked():
            notes.append("ADD ALL DOG LEADS")
        extra_notes = [line.strip() for line in (self.current_container_notes or "").splitlines() if line.strip()]
        notes.extend(extra_notes)
        if not notes:
            return
        note_bg = QColor("#f6f1c1")
        note_fg = QColor("#4f3f00")
        for note_text in notes:
            row = table.rowCount()
            table.insertRow(row)
            note_item = self.make_container_table_item(f"NOTE: {note_text}", bold=True, background=note_bg, foreground=note_fg, data={"note_row": True})
            table.setItem(row, 0, note_item)
            for col in range(1, table.columnCount()):
                table.setItem(row, col, self.make_container_table_item("", background=note_bg, foreground=note_fg, data={"note_row": True}))
            table.setSpan(row, 0, 1, table.columnCount())

    def refresh_container_note_rows(self):
        self.populate_container_table(self.get_container_line_dicts())

    def apply_container_row_style(self, row, urgent=False, additional=False):
        table = getattr(self.ui, "container_table", None)
        if table is None:
            return
        if urgent:
            bg = QColor("#ffc9c9")
            fg = QColor("#7a0000")
        elif additional:
            bg = QColor("#fff5b1")
            fg = QColor("#5f4b00")
        else:
            bg = None
            fg = None
        for col in range(table.columnCount()):
            item = table.item(row, col)
            if item is None:
                continue
            meta = item.data(Qt.UserRole) if col == 0 else None
            if bg is None:
                item.setBackground(QBrush())
                item.setForeground(QBrush())
            else:
                item.setBackground(bg)
                item.setForeground(fg)
            if col == self.container_columns["urgent"]:
                if urgent:
                    item.setBackground(QColor("#ff8d8d"))
                elif bg is not None:
                    item.setBackground(bg)
            if col == self.container_columns["additional"]:
                if additional:
                    item.setBackground(QColor("#ffe36b"))
                elif bg is not None:
                    item.setBackground(bg)
            if col == self.container_columns["remove"]:
                if urgent:
                    item.setBackground(QColor("#d9d9d9"))
                    item.setForeground(QColor("#666666"))
                else:
                    item.setBackground(QColor("#ffd6d6"))
                    item.setForeground(QColor("#8b0000"))
            if col == 0 and isinstance(meta, dict):
                item.setData(Qt.UserRole, {**meta, "urgent": urgent, "additional": additional, "note_row": False})

    def get_remaining_additional_cartons(self):
        spinner = getattr(self.ui, "additional_spinner", None)
        if spinner is None:
            return 0
        return max(0, int(self.parse_float(spinner.value())))

    def set_remaining_additional_cartons(self, value):
        spinner = getattr(self.ui, "additional_spinner", None)
        if spinner is None:
            return
        remaining = max(0, int(round(self.parse_float(value))))
        blocker = QSignalBlocker(spinner)
        try:
            spinner.setValue(remaining)
        finally:
            del blocker
        spinner.setToolTip(f"Additional cartons still needed: {remaining}")

    def adjust_remaining_additional_cartons(self, delta):
        self.set_remaining_additional_cartons(self.get_remaining_additional_cartons() + int(round(self.parse_float(delta))))

    def refresh_container_totals(self):
        lines = self.get_container_line_dicts()
        total_cartons = sum(int(line.get("cartons", 0) or 0) for line in lines)
        remaining_additional_needed = self.get_remaining_additional_cartons()
        self.set_label_text("totalCarton_box", self.format_value(total_cartons))
        self.set_label_text("targetCarton_box", self.format_value(total_cartons + remaining_additional_needed))
        spinner = getattr(self.ui, "additional_spinner", None)
        if spinner is not None:
            spinner.setToolTip(f"Additional cartons still needed: {remaining_additional_needed}")
        self.refresh_item_summary_context_boxes()

    def handle_container_reference_submit(self):
        ref = self.get_container_ref_text()
        if not ref:
            QMessageBox.warning(self, "Missing container", "Enter a container reference first.")
            return
        row = self.db_one("SELECT * FROM containers WHERE UPPER(TRIM(container_ref)) = UPPER(TRIM(?))", (ref,))
        if row is not None:
            self.load_container_into_page(row["container_ref"])
        else:
            self.current_container_ref = ref
            self.current_container_notes = ""
            self.set_updated_box_today()
            self.set_container_meta_widgets(
                updated_display=date.today().strftime("%d/%m/%Y"),
                eta_iso=date.today().isoformat(),
                additional_value=0,
                dog_leads=False,
            )
            self.populate_container_table([])
            order_widget = getattr(self.ui, "orderNumberContainer_line", None)
            item_widget = getattr(self.ui, "itemNumberContainer_line", None)
            qty_widget = getattr(self.ui, "qtyContainder_line", None)
            if order_widget is not None:
                order_widget.clear()
            if item_widget is not None:
                item_widget.clear()
            if qty_widget is not None:
                qty_widget.clear()
            self.save_current_container_state()
        order_widget = getattr(self.ui, "orderNumberContainer_line", None)
        if order_widget is not None:
            order_widget.setFocus()
            order_widget.selectAll()

    def add_container_line_from_inputs(self):
        ref = self.get_container_ref_text()
        order_widget = getattr(self.ui, "orderNumberContainer_line", None)
        item_widget = getattr(self.ui, "itemNumberContainer_line", None)
        qty_widget = getattr(self.ui, "qtyContainder_line", None)
        if not ref:
            QMessageBox.warning(self, "Missing container", "Enter the container reference and press Enter first.")
            return
        order_number = order_widget.text().strip() if order_widget is not None else ""
        if not order_number:
            QMessageBox.warning(self, "Missing order number", "Enter an order number first.")
            if order_widget is not None:
                order_widget.setFocus()
            return
        item_number = self.find_item_number(item_widget.text().strip() if item_widget is not None else "")
        if not item_number:
            QMessageBox.warning(self, "Invalid item", "Please choose a valid item number from the list.")
            if item_widget is not None:
                item_widget.setFocus()
            return
        requested_qty = self.parse_float(qty_widget.text() if qty_widget is not None else 0)
        if requested_qty <= 0:
            QMessageBox.warning(self, "Invalid quantity", "Quantity must be greater than 0.")
            if qty_widget is not None:
                qty_widget.setFocus()
                qty_widget.selectAll()
            return

        item_row = self.get_item_master_row(item_number)
        description = self.get_first(item_row, "item_name", "Item Name", "description", "Description")
        carton_size = self.get_item_carton_size(item_row)
        cartons = max(1, math.ceil(requested_qty / carton_size))
        rounded_qty = cartons * carton_size

        lines = self.get_container_line_dicts()
        lines.append({
            "order_number": order_number,
            "item_number": item_number,
            "description": description,
            "qty": rounded_qty,
            "cartons": cartons,
            "additional_cartons": 0,
            "urgent": False,
            "additional": False,
        })
        self.populate_container_table(lines)
        self.set_updated_box_today()
        self.save_current_container_state()
        if item_widget is not None:
            item_widget.clear()
            item_widget.setFocus()
        if qty_widget is not None:
            qty_widget.clear()

    def handle_container_table_double_click(self, row, column):
        table = getattr(self.ui, "container_table", None)
        if table is None or row < 0:
            return
        first_item = table.item(row, 0)
        meta = first_item.data(Qt.UserRole) if first_item is not None else None
        if isinstance(meta, dict) and meta.get("note_row"):
            return

        lines = self.get_container_line_dicts()
        if row >= len(lines):
            return
        line = dict(lines[row])
        old_additional_cartons = int(line.get("additional_cartons", 0) or 0) if bool(line.get("additional")) else 0

        if column == self.container_columns["item"]:
            item_number = (line.get("item_number") or "").strip()
            if item_number:
                self.open_item_summary_from_order_analysis(item_number)
            return

        if column == self.container_columns["qty"]:
            item_row = self.get_item_master_row(line.get("item_number", ""))
            carton_size = self.get_item_carton_size(item_row)
            current_qty = self.parse_float(line.get("qty", 0))
            new_qty, accepted = QInputDialog.getDouble(
                self,
                "Edit Quantity",
                f"Quantity for {line.get('item_number', '')}:",
                current_qty,
                0.0,
                999999999.0,
                3,
            )
            if not accepted:
                return
            if new_qty <= 0:
                QMessageBox.warning(self, "Invalid quantity", "Quantity must be greater than 0.")
                return
            cartons = max(1, math.ceil(new_qty / carton_size))
            rounded_qty = cartons * carton_size
            line["qty"] = rounded_qty
            line["cartons"] = cartons
            if bool(line.get("additional")):
                line["additional_cartons"] = cartons
                self.adjust_remaining_additional_cartons(old_additional_cartons - cartons)
            lines[row] = line
            self.populate_container_table(lines)
            self.set_updated_box_today()
            self.save_current_container_state()
            return

        if column == self.container_columns["urgent"]:
            line["urgent"] = not bool(line.get("urgent"))
            if line["urgent"] and bool(line.get("additional")):
                self.adjust_remaining_additional_cartons(old_additional_cartons)
                line["additional"] = False
                line["additional_cartons"] = 0
            lines[row] = line
            self.populate_container_table(lines)
            self.set_updated_box_today()
            self.save_current_container_state()
            return

        if column == self.container_columns["additional"]:
            line["additional"] = not bool(line.get("additional"))
            if line["additional"]:
                line["urgent"] = False
                new_additional_cartons = int(line.get("cartons", 0) or 0)
                line["additional_cartons"] = new_additional_cartons
                self.adjust_remaining_additional_cartons(-new_additional_cartons)
            else:
                self.adjust_remaining_additional_cartons(old_additional_cartons)
                line["additional_cartons"] = 0
            lines[row] = line
            self.populate_container_table(lines)
            self.set_updated_box_today()
            self.save_current_container_state()
            return

        if column == self.container_columns["remove"]:
            if bool(line.get("urgent")):
                QMessageBox.information(self, "Urgent line", "Urgent lines cannot be removed until Urgent is toggled off.")
                return
            result = QMessageBox.question(
                self,
                "Remove line",
                f"Remove {line.get('item_number', 'this line')} from the container?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if result == QMessageBox.Yes:
                if bool(line.get("additional")):
                    self.adjust_remaining_additional_cartons(old_additional_cartons)
                del lines[row]
                self.populate_container_table(lines)
                self.set_updated_box_today()
                self.save_current_container_state()

    def handle_container_header_changed(self, *_args):
        if not self.get_container_ref_text():
            return
        self.refresh_container_note_rows()
        self.set_updated_box_today()
        self.save_current_container_state()

    def edit_container_notes(self):
        if not self.get_container_ref_text():
            QMessageBox.warning(self, "Missing container", "Enter the container reference and press Enter first.")
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("Container Notes")
        dialog.resize(520, 340)
        layout = QVBoxLayout(dialog)
        editor = QTextEdit(dialog)
        editor.setPlainText(self.current_container_notes or "")
        layout.addWidget(editor)
        buttons = QHBoxLayout()
        buttons.addStretch(1)
        save_button = QPushButton("Save", dialog)
        cancel_button = QPushButton("Cancel", dialog)
        buttons.addWidget(save_button)
        buttons.addWidget(cancel_button)
        layout.addLayout(buttons)
        save_button.clicked.connect(dialog.accept)
        cancel_button.clicked.connect(dialog.reject)
        if dialog.exec() == QDialog.Accepted:
            self.current_container_notes = editor.toPlainText().strip()
            self.refresh_container_note_rows()
            self.set_updated_box_today()
            self.save_current_container_state()

    def save_current_container_state(self):
        ref = self.get_container_ref_text() or self.current_container_ref
        if not ref:
            return
        self.current_container_ref = ref
        updated_display = getattr(self.ui, "updated_box", None).toPlainText().strip() if getattr(self.ui, "updated_box", None) is not None else date.today().strftime("%d/%m/%Y")
        updated_iso = self.display_to_iso_date(updated_display) or date.today().isoformat()
        eta_widget = getattr(self.ui, "eta_dateEdit", None)
        eta_iso = eta_widget.date().toString("yyyy-MM-dd") if eta_widget is not None else ""
        additional_spinner = getattr(self.ui, "additional_spinner", None)
        extra = int(additional_spinner.value()) if additional_spinner is not None else 0
        dog_leads = 1 if getattr(self.ui, "checkBox", None) is not None and self.ui.checkBox.isChecked() else 0
        notes = (self.current_container_notes or "").strip()
        lines = self.get_container_line_dicts()

        cur = self.db_conn.cursor()
        if self.db_engine == "sqlserver":
            cur.execute(
                "UPDATE containers SET updated_on = ?, eta_date = ?, additional_cartons = ?, dog_leads = ?, notes = ? WHERE container_ref = ?",
                (updated_iso, eta_iso, extra, dog_leads, notes, ref),
            )
            if getattr(cur, 'rowcount', 0) == 0:
                cur.execute(
                    "INSERT INTO containers (container_ref, updated_on, eta_date, additional_cartons, dog_leads, notes) VALUES (?, ?, ?, ?, ?, ?)",
                    (ref, updated_iso, eta_iso, extra, dog_leads, notes),
                )
        else:
            cur.execute(
                """
                INSERT INTO containers (container_ref, updated_on, eta_date, additional_cartons, dog_leads, notes)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(container_ref) DO UPDATE SET
                    updated_on = excluded.updated_on,
                    eta_date = excluded.eta_date,
                    additional_cartons = excluded.additional_cartons,
                    dog_leads = excluded.dog_leads,
                    notes = excluded.notes
                """,
                (ref, updated_iso, eta_iso, extra, dog_leads, notes),
            )
        cur.execute("DELETE FROM container_lines WHERE container_ref = ?", (ref,))
        for idx, line in enumerate(lines, start=1):
            cur.execute(
                """
                INSERT INTO container_lines (
                    container_ref, line_no, order_number, item_number, description, qty, cartons,
                    additional_cartons, urgent, additional
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    ref, idx, line.get("order_number", ""), line.get("item_number", ""), line.get("description", ""),
                    float(line.get("qty", 0) or 0), int(line.get("cartons", 0) or 0),
                    int(line.get("additional_cartons", 0) or 0), 1 if line.get("urgent") else 0, 1 if line.get("additional") else 0,
                ),
            )
        self.db_conn.commit()
        self.refresh_container_totals()
        self.refresh_containers_list()
        self.refresh_on_order_statuses_from_container_data(persist=True)

    def refresh_containers_list(self):
        table = getattr(self.ui, "containers_tableWidget", None)
        if table is None or self.db_conn is None:
            return
        rows = self.db_all(
            """
            SELECT
                c.container_ref,
                c.eta_date,
                c.updated_on
            FROM containers c
            ORDER BY COALESCE(CAST(c.updated_on AS date), CAST('1900-01-01' AS date)) DESC, c.container_ref COLLATE NOCASE
            """
        )
        table.setRowCount(0)
        for row_data in rows:
            row = table.rowCount()
            table.insertRow(row)
            table.setItem(row, 0, self.make_container_table_item(row_data["container_ref"]))
            table.setItem(row, 1, self.make_container_table_item(self.iso_to_display_date(row_data["eta_date"]), align=Qt.AlignCenter))
        table.resizeRowsToContents()

    def reload_current_container_state(self):
        container_ref = (self.current_container_ref or self.get_container_ref_text() or "").strip()
        if not container_ref:
            return False

        row = self.db_one("SELECT * FROM containers WHERE container_ref = ?", (container_ref,))
        if row is None:
            return False

        self.current_container_ref = row["container_ref"]
        self.current_container_notes = (row["notes"] or "").strip()
        self.set_container_ref_text(row["container_ref"])
        self.set_container_meta_widgets(
            updated_display=self.iso_to_display_date(row["updated_on"]),
            eta_iso=row["eta_date"] or "",
            additional_value=row["additional_cartons"] or 0,
            dog_leads=bool(row["dog_leads"]),
        )
        lines = self.db_all(
            """
            SELECT order_number, item_number, description, qty, cartons, additional_cartons, urgent, additional
            FROM container_lines
            WHERE container_ref = ?
            ORDER BY line_no, id
            """,
            (container_ref,),
        )
        parsed_lines = [
            {
                "order_number": r["order_number"] or "",
                "item_number": r["item_number"] or "",
                "description": r["description"] or "",
                "qty": self.parse_float(r["qty"]),
                "cartons": int(self.parse_float(r["cartons"])),
                "additional_cartons": int(self.parse_float(r["additional_cartons"])),
                "urgent": bool(r["urgent"]),
                "additional": bool(r["additional"]),
            }
            for r in lines
        ]
        self.populate_container_table(parsed_lines)
        return True

    def load_container_into_page(self, container_ref):
        self.current_container_ref = (container_ref or "").strip()
        if not self.reload_current_container_state():
            return
        self.ui.stackedWidget.setCurrentWidget(self.ui.buildContainer_page)
        order_widget = getattr(self.ui, "orderNumberContainer_line", None)
        if order_widget is not None:
            order_widget.setFocus()
            order_widget.selectAll()

    def handle_saved_container_double_click(self, row, _column):
        table = getattr(self.ui, "containers_tableWidget", None)
        if table is None or row < 0:
            return
        ref_item = table.item(row, 0)
        container_ref = ref_item.text().strip() if ref_item is not None and ref_item.text() else ""
        if not container_ref:
            return
        box = QMessageBox(self)
        box.setWindowTitle("Saved Container")
        box.setText(f"What do you want to do with {container_ref}?")
        edit_button = box.addButton("Edit", QMessageBox.AcceptRole)
        delete_button = box.addButton("Delete", QMessageBox.DestructiveRole)
        cancel_button = box.addButton(QMessageBox.Cancel)
        box.exec()
        clicked = box.clickedButton()
        if clicked is edit_button:
            self.load_container_into_page(container_ref)
        elif clicked is delete_button:
            confirm = QMessageBox.question(
                self,
                "Delete container",
                f"Delete container {container_ref}? This cannot be undone.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if confirm == QMessageBox.Yes:
                cur = self.db_conn.cursor()
                cur.execute("DELETE FROM container_lines WHERE container_ref = ?", (container_ref,))
                cur.execute("DELETE FROM containers WHERE container_ref = ?", (container_ref,))
                self.db_conn.commit()
                if (self.current_container_ref or "") == container_ref:
                    self.current_container_ref = None
                    self.current_container_notes = ""
                    self.set_container_ref_text("")
                    self.set_container_meta_widgets(updated_display="", eta_iso=date.today().isoformat(), additional_value=0, dog_leads=False)
                    self.populate_container_table([])
                self.refresh_containers_list()
                self.refresh_on_order_statuses_from_container_data(persist=True)

    def get_order_item_edit(self):
        return getattr(self.ui, "enterItemOrder_lineEdit", None)

    def get_order_qty_widget(self):
        return getattr(self.ui, "enterItemOrder_lineEdit_2", None)

    def get_order_qty_editor(self):
        return self.get_order_qty_widget()

    def setup_order_entry(self):
        item_edit = self.get_order_item_edit()
        qty_widget = self.get_order_qty_widget()
        qty_editor = self.get_order_qty_editor()

        if item_edit is not None:
            item_edit.installEventFilter(self)
            completer = QCompleter(self.item_numbers, self)
            completer.setCaseSensitivity(Qt.CaseInsensitive)
            completer.setFilterMode(Qt.MatchStartsWith)
            completer.setCompletionMode(QCompleter.PopupCompletion)
            completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
            item_edit.setCompleter(completer)
            completer.activated.connect(self.order_item_completion_selected)
            self.order_item_completer = completer
            item_edit.textEdited.connect(lambda _text: self.update_order_item_completion())

        if qty_widget is not None:
            validator = QDoubleValidator(0.0, 999999999.0, 3, qty_widget)
            validator.setNotation(QDoubleValidator.StandardNotation)
            qty_widget.setValidator(validator)
        if qty_editor is not None:
            qty_editor.installEventFilter(self)

    def eventFilter(self, obj, event):
        item_edit = self.get_order_item_edit()
        qty_editor = self.get_order_qty_editor()
        qty_widget = self.get_order_qty_widget()
        container_ref_widget = getattr(self.ui, "nextContainer_box", None)
        freight_box = getattr(self.ui, "chargeFreight_textBrowser", None)
        freight_viewport = freight_box.viewport() if freight_box is not None else None

        if event.type() == QEvent.KeyPress:
            if obj is item_edit:
                return self.handle_order_item_keypress(event)
            if obj is qty_editor or obj is qty_widget:
                return self.handle_order_qty_keypress(event)
            if obj is container_ref_widget and event.key() in (Qt.Key_Return, Qt.Key_Enter):
                self.handle_container_reference_submit()
                return True

        if event.type() == QEvent.MouseButtonDblClick:
            if obj is freight_box or obj is freight_viewport:
                return self.toggle_charge_freight_for_current_customer()
            for object_name in self.item_summary_editable_fields:
                if obj is getattr(self.ui, object_name, None):
                    self.edit_item_summary_field(object_name)
                    return True
            if obj is getattr(self.ui, "stockOnOrder_label", None) or obj is getattr(self.ui, "stockOnOrder_box", None):
                self.prompt_add_current_item_to_order_sheet()
                return True

        return super().eventFilter(obj, event)

    def add_or_update_to_order_line(self, item_number, qty_value, supplier_name="", urgent=False, status=""):
        table = getattr(self.ui, "order_table", None)
        if table is None:
            return False, "To Order table is not available."

        typed_item = str(item_number or "").strip()
        item_number = self.find_item_number(typed_item)
        try:
            qty_value = float(qty_value)
        except Exception:
            qty_value = 0.0

        if not item_number:
            return False, "Please choose a valid item number."
        if qty_value <= 0:
            return False, "Quantity must be greater than 0."

        item_row = self.get_item_master_row(item_number)
        description = self.get_first(item_row, "description", "item_name", "Item Name", "Description")
        supplier = (supplier_name or "").strip() or self.get_first(item_row, "supplier_name", "supplier_code", "Column1", "Supplier")

        for row in range(table.rowCount()):
            existing_item = table.item(row, 0)
            if existing_item is None or (existing_item.text() or "").strip() != item_number:
                continue

            qty_item = table.item(row, self.order_qty_column)
            current_qty = self.parse_float(qty_item.data(Qt.UserRole) if qty_item is not None else 0)
            new_qty = current_qty + qty_value

            self._updating_order_table = True
            if qty_item is not None:
                qty_item.setText(self.format_value(new_qty))
                qty_item.setData(Qt.UserRole, new_qty)

            if supplier:
                supplier_item = table.item(row, self.order_supplier_column)
                if supplier_item is not None:
                    supplier_item.setText(supplier)
                    supplier_item.setData(Qt.UserRole, supplier)
            if urgent:
                table.setItem(row, self.order_priority_column, self.build_order_status_item("URGENT", True))
            self._updating_order_table = False

            self.save_order_table_state()
            if self.current_item_number:
                self.refresh_item_summary_context_boxes()
            return True, f"Updated To Order line for {item_number}."

        row = table.rowCount()
        self._updating_order_table = True
        table.insertRow(row)
        table.setItem(row, 0, self.make_order_table_item(item_number))
        table.setItem(row, 1, self.make_order_table_item(description))
        table.setItem(
            row,
            self.order_on_order_column,
            self.make_order_table_item(
                self.format_value(self.get_item_on_order_qty(item_number)),
                align=Qt.AlignRight | Qt.AlignVCenter,
            ),
        )

        qty_item = self.make_order_table_item(
            self.format_value(qty_value),
            editable=True,
            align=Qt.AlignRight | Qt.AlignVCenter,
        )
        qty_item.setData(Qt.UserRole, qty_value)
        table.setItem(row, self.order_qty_column, qty_item)

        supplier_item = self.make_order_table_item(supplier, editable=False)
        supplier_item.setData(Qt.UserRole, supplier)
        supplier_item.setToolTip("Double-click to choose a supplier.")
        table.setItem(row, self.order_supplier_column, supplier_item)

        table.setItem(row, self.order_priority_column, self.build_order_status_item("URGENT" if urgent else status, urgent))
        table.setItem(row, self.order_remove_column, self.build_order_remove_item())
        self._updating_order_table = False
        table.resizeRowsToContents()

        self.save_order_table_state()
        if self.current_item_number:
            self.refresh_item_summary_context_boxes()
        return True, f"Added {item_number} to To Order."

    def prompt_add_current_item_to_order_sheet(self):
        if not self.current_item_number:
            return False
        current_value = self.parse_float(getattr(self.ui, "stockOnOrder_box", None).text() if getattr(self.ui, "stockOnOrder_box", None) is not None and hasattr(getattr(self.ui, "stockOnOrder_box", None), "text") else 0)
        qty_value, accepted = QInputDialog.getDouble(
            self,
            "Add to To Order Sheet",
            f"Qty for {self.current_item_number}:",
            current_value if current_value > 0 else 0.0,
            0.0,
            999999999.0,
            3,
        )
        if not accepted or qty_value <= 0:
            return False
        ok, message = self.add_or_update_to_order_line(self.current_item_number, qty_value)
        if ok:
            QMessageBox.information(self, "To Order Sheet", message)
            return True
        QMessageBox.warning(self, "To Order Sheet", message)
        return False

    def edit_item_summary_field(self, object_name):
        if self.db_conn is None or not self.current_item_number:
            return

        field_meta = self.item_summary_editable_fields.get(object_name)
        widget = getattr(self.ui, object_name, None)
        if field_meta is None or widget is None:
            return

        current_value = self.parse_float(widget.text() if hasattr(widget, "text") else 0)
        label = field_meta.get("label", object_name)
        new_value, accepted = QInputDialog.getDouble(
            self,
            f"Edit {label}",
            f"{label}:",
            current_value,
            0.0,
            999999999.0,
            3,
        )
        if not accepted:
            return

        cur = self.db_conn.cursor()
        cur.execute(
            f"UPDATE items SET [{field_meta['column']}] = ? WHERE UPPER(TRIM(item_number)) = UPPER(TRIM(?))",
            (float(new_value), self.current_item_number),
        )
        self.db_conn.commit()

        self.load_item_summary()

    def update_order_item_completion(self):
        item_edit = self.get_order_item_edit()
        completer = self.order_item_completer
        if item_edit is None or completer is None:
            return

        prefix = item_edit.text().strip()
        if not prefix:
            completer.popup().hide()
            return

        completer.setCompletionPrefix(prefix)
        popup = completer.popup()
        rect = item_edit.cursorRect()
        rect.setWidth(max(item_edit.width(), popup.sizeHintForColumn(0) + popup.verticalScrollBar().sizeHint().width() + 24))
        completer.complete(rect)

    def order_item_completion_selected(self, completion_text):
        item_edit = self.get_order_item_edit()
        qty_widget = self.get_order_qty_widget()
        item_number = self.find_item_number(completion_text)
        if item_edit is None or not item_number:
            return

        item_edit.setText(item_number)
        item_edit.setCursorPosition(len(item_number))
        if self.order_item_completer is not None:
            self.order_item_completer.popup().hide()

        if qty_widget is not None:
            qty_widget.setFocus()
            if hasattr(qty_widget, "selectAll"):
                qty_widget.selectAll()

    def handle_order_item_keypress(self, event):
        key = event.key()
        popup = self.order_item_completer.popup() if self.order_item_completer is not None else None
        popup_visible = bool(popup is not None and popup.isVisible())

        if popup_visible and key in (Qt.Key_Down, Qt.Key_Up, Qt.Key_PageDown, Qt.Key_PageUp):
            QApplication.sendEvent(popup, event)
            return True

        if key in (Qt.Key_Return, Qt.Key_Enter, Qt.Key_Tab):
            if popup_visible and self.order_item_completer.currentCompletion():
                self.order_item_completion_selected(self.order_item_completer.currentCompletion())
                return True
            self.resolve_order_item_and_focus_qty(show_warning=True)
            return True

        if key == Qt.Key_Escape and popup_visible:
            popup.hide()
            return True

        return False

    def handle_order_qty_keypress(self, event):
        key = event.key()
        if key in (Qt.Key_Return, Qt.Key_Enter, Qt.Key_Tab):
            self.add_order_line_from_inputs()
            return True
        return False

    def resolve_order_item_and_focus_qty(self, show_warning=False):
        item_edit = self.get_order_item_edit()
        typed = item_edit.text().strip() if item_edit is not None else ""
        valid_item = self.find_item_number(typed)
        if valid_item:
            self.order_item_completion_selected(valid_item)
            return True
        if show_warning and typed:
            QMessageBox.warning(self, "Invalid item", "Please choose a valid item number from the list.")
        return False

    def get_item_master_row(self, item_number):
        return self.row_to_dict(
            self.db_one(
                "SELECT * FROM items WHERE UPPER(TRIM(item_number)) = UPPER(TRIM(?)) LIMIT 1",
                (item_number,),
            )
        )

    def import_orders_from_dialog(self):
        filters = "Orders files (*.csv *.xlsx *.xlsm);;CSV files (*.csv);;Excel files (*.xlsx *.xlsm);;All files (*.*)"
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Choose orders import file",
            str(self.base_dir),
            filters,
        )
        if not file_path:
            return

        try:
            imported_count = self.import_orders_file(file_path)
        except Exception as exc:
            QMessageBox.critical(self, "Orders import failed", f"Could not import orders file.\n\n{exc}")
            return

        display_text = (
            f"Last import: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
            f"File: {Path(file_path).name}\n"
            f"Rows imported: {imported_count:,}"
        )
        self.set_meta_value("orders_last_import_display", display_text)
        browser = getattr(self.ui, "lastUpdateOrders_textBrowser_3", None)
        if browser is not None:
            browser.setPlainText(display_text)

        self.refresh_order_table_on_order_column()
        if self.current_item_number:
            self.refresh_item_summary_context_boxes()
        self.rerun_order_analysis_if_ready()

        QMessageBox.information(self, "Orders imported", f"Imported {imported_count:,} rows into the orders table.")

    def import_orders_file(self, file_path):
        rows = self.read_orders_import_rows(file_path)
        cur = self.db_conn.cursor()
        cur.execute("DELETE FROM orders")
        cur.executemany(
            "INSERT INTO orders (item_number, quantity, purchase_no, order_date) VALUES (?, ?, ?, ?)",
            rows,
        )
        self.db_conn.commit()
        return len(rows)

    def read_orders_import_rows(self, file_path):
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self.read_orders_rows_from_csv(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self.read_orders_rows_from_excel(path)
        raise ValueError("Unsupported file type. Choose a CSV or Excel file.")

    def normalize_header(self, header):
        return re.sub(r"[^a-z0-9]+", "", str(header or "").strip().lower())

    def get_required_order_columns(self, headers):
        normalized = {self.normalize_header(h): h for h in headers}
        aliases = {
            "item_number": ["itemnumber", "item", "itemno"],
            "quantity": ["quantity", "qty"],
            "purchase_no": ["purchaseno", "purchasenumber", "pono", "ponumber", "po"],
            "order_date": ["shippingdate", "date", "shipdate"],
        }
        resolved = {}
        missing = []
        for field_name, options in aliases.items():
            matched = next((normalized[key] for key in options if key in normalized), None)
            if matched is None:
                missing.append(field_name)
            else:
                resolved[field_name] = matched
        if missing:
            missing_text = ", ".join(missing)
            raise ValueError(
                f"Missing required column(s): {missing_text}. Expected columns like Item Number, Quantity, Purchase No., Shipping Date."
            )
        return resolved

    def normalize_order_import_date(self, value):
        if value is None:
            return ""

        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")

        text_value = str(value).strip()
        if not text_value:
            return ""

        if text_value.isdigit():
            if len(text_value) == 8:
                for fmt in ("%d%m%Y", "%Y%m%d"):
                    try:
                        return datetime.strptime(text_value, fmt).strftime("%Y-%m-%d")
                    except ValueError:
                        pass
            try:
                serial = int(text_value)
                if 1 <= serial <= 60000:
                    excel_date = datetime(1899, 12, 30) + timedelta(days=serial)
                    return excel_date.strftime("%Y-%m-%d")
            except ValueError:
                pass

        for fmt in (
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%d%m%Y",
            "%Y%m%d",
            "%d/%m/%y",
            "%d-%m-%y",
        ):
            try:
                return datetime.strptime(text_value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass

        raise ValueError(
            f"Invalid shipping date: {text_value}. Expected a date like 7/04/2026 or 2026-04-07."
        )

    def read_orders_rows_from_csv(self, path):
        import csv

        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise ValueError("The selected CSV file has no header row.")
            column_map = self.get_required_order_columns(reader.fieldnames)
            rows = []
            for line_number, row in enumerate(reader, start=2):
                normalized = self.build_order_import_tuple(row, column_map, line_number)
                if normalized is not None:
                    rows.append(normalized)
        if not rows:
            raise ValueError("The selected file did not contain any valid order rows.")
        return rows

    def read_orders_rows_from_excel(self, path):
        if load_workbook is None:
            raise ValueError("Excel import is not available because openpyxl is not installed.")

        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not raw_headers:
                raise ValueError("The selected Excel file has no header row.")
            headers = ["" if h is None else str(h) for h in raw_headers]
            column_map = self.get_required_order_columns(headers)
            rows = []
            for line_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))}
                normalized = self.build_order_import_tuple(row_dict, column_map, line_number)
                if normalized is not None:
                    rows.append(normalized)
        finally:
            workbook.close()

        if not rows:
            raise ValueError("The selected file did not contain any valid order rows.")
        return rows

    def build_order_import_tuple(self, row, column_map, line_number):
        item_number = str(row.get(column_map["item_number"], "") or "").strip()
        quantity_value = self.parse_float(row.get(column_map["quantity"], 0))
        purchase_no = str(row.get(column_map["purchase_no"], "") or "").strip()
        shipping_date = self.normalize_order_import_date(row.get(column_map["order_date"], ""))

        if not item_number and quantity_value <= 0 and not purchase_no and not shipping_date:
            return None
        if not item_number:
            raise ValueError(f"Line {line_number}: item number is blank.")
        if item_number.startswith("\\"):
            return None
        if quantity_value <= 0:
            raise ValueError(f"Line {line_number}: quantity must be greater than 0.")

        return (item_number, quantity_value, purchase_no, shipping_date)

    def import_sales_from_dialog(self):
        filters = "Sales files (*.txt *.csv *.xlsx *.xlsm);;Text files (*.txt);;CSV files (*.csv);;Excel files (*.xlsx *.xlsm);;All files (*.*)"
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Choose sales import file",
            str(self.base_dir),
            filters,
        )
        if not file_path:
            return

        try:
            imported_count = self.import_sales_file(file_path)
        except Exception as exc:
            QMessageBox.critical(self, "Sales import failed", f"Could not import sales file.\n\n{exc}")
            return

        display_text = (
            f"Last import: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
            f"File: {Path(file_path).name}\n"
            f"Rows added: {imported_count:,}"
        )
        self.set_meta_value("sales_last_import_display", display_text)
        browser = getattr(self.ui, "lastUpdateSales_textBrowser", None)
        if browser is not None:
            browser.setPlainText(display_text)

        self.load_reference_lists()
        self.setup_customer_autocomplete()
        self.setup_item_autocomplete()
        self.setup_date_ranges()
        self.rerun_search_if_ready()
        self.rerun_item_if_ready()

        QMessageBox.information(self, "Sales imported", f"Added {imported_count:,} new rows to the sales table.")

    def import_sales_file(self, file_path):
        rows = self.read_sales_import_rows(file_path)
        if not rows:
            return 0

        date_values = [row[0] for row in rows if row and row[0]]
        min_date = min(date_values) if date_values else "0001-01-01"
        max_date = max(date_values) if date_values else "9999-12-31"

        existing_rows = self.db_all(
            """
            SELECT sale_date, customer_name, item_number, quantity, price
            FROM sales
            WHERE DATE(sale_date) BETWEEN ? AND ?
            """,
            (min_date, max_date),
        )
        existing_counts = Counter()
        for row in existing_rows:
            existing_counts[self.sales_row_signature(
                row["sale_date"],
                row["customer_name"],
                row["item_number"],
                row["quantity"],
                row["price"],
            )] += 1

        incoming_counts = Counter()
        rows_to_insert = []
        for row in rows:
            signature = self.sales_row_signature(row[0], row[1], row[2], row[5], row[6])
            if incoming_counts[signature] < existing_counts[signature]:
                incoming_counts[signature] += 1
                continue
            incoming_counts[signature] += 1
            rows_to_insert.append(row)

        if rows_to_insert:
            cur = self.db_conn.cursor()
            cur.executemany(
                """
                INSERT INTO sales (
                    sale_date, customer_name, item_number, description,
                    month_key, quantity, price, extended
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                rows_to_insert,
            )
            self.db_conn.commit()
        return len(rows_to_insert)

    def read_sales_import_rows(self, file_path):
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix in {".csv", ".txt"}:
            return self.read_sales_rows_from_csv(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self.read_sales_rows_from_excel(path)
        raise ValueError("Unsupported file type. Choose a TXT, CSV, or Excel file.")

    def get_required_sales_columns(self, headers):
        normalized = {self.normalize_header(h): h for h in headers}
        aliases = {
            "sale_date": ["date", "saledate"],
            "customer_name": ["colastname", "customer", "customername", "companylastname", "lastname", "co lastname"],
            "item_number": ["itemnumber", "item", "itemno"],
            "quantity": ["quantity", "qty"],
            "price": ["price", "unitprice", "sellprice"],
        }
        resolved = {}
        missing = []
        for field_name, options in aliases.items():
            matched = next((normalized[key] for key in options if key in normalized), None)
            if matched is None:
                missing.append(field_name)
            else:
                resolved[field_name] = matched
        if missing:
            missing_text = ", ".join(missing)
            raise ValueError(
                f"Missing required column(s): {missing_text}. Expected columns like Date, Co./Last Name, Item Number, Quantity, Price."
            )
        return resolved

    def normalize_sales_import_date(self, value):
        if value is None:
            return ""

        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")

        text_value = str(value).strip()
        if not text_value:
            return ""

        if text_value.isdigit():
            try:
                serial = int(text_value)
                if 1 <= serial <= 60000:
                    excel_date = datetime(1899, 12, 30) + timedelta(days=serial)
                    return excel_date.strftime("%Y-%m-%d")
            except ValueError:
                pass

        for fmt in (
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M",
            "%d/%m/%y",
            "%d-%m-%y",
            "%Y%m%d",
            "%d%m%Y",
        ):
            try:
                return datetime.strptime(text_value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass

        raise ValueError(
            f"Invalid sale date: {text_value}. Expected a date like 2/03/2026 or 2026-03-02."
        )

    def sales_month_key_from_date(self, iso_date_text):
        iso_date_text = (iso_date_text or "").strip()
        if not iso_date_text:
            return ""
        try:
            return datetime.strptime(iso_date_text, "%Y-%m-%d").strftime("%Y%m")
        except ValueError:
            return ""

    def sales_row_signature(self, sale_date, customer_name, item_number, quantity, price):
        return (
            self.normalize_sales_import_date(sale_date),
            str(customer_name or "").strip().upper(),
            str(item_number or "").strip().upper(),
            round(self.parse_float(quantity), 6),
            round(self.parse_float(price), 6),
        )

    def fetch_item_name_map(self, item_numbers):
        normalized_items = sorted({str(item or "").strip() for item in item_numbers if str(item or "").strip()})
        if not normalized_items:
            return {}

        result = {}
        cur = self.db_conn.cursor()
        chunk_size = 900
        for start in range(0, len(normalized_items), chunk_size):
            chunk = normalized_items[start:start + chunk_size]
            placeholders = ",".join("?" for _ in chunk)
            cur.execute(
                f"""
                SELECT TRIM(item_number) AS item_number,
                       COALESCE(NULLIF(TRIM(item_name), ''), NULLIF(TRIM(description), ''), '') AS item_name
                FROM items
                WHERE UPPER(TRIM(item_number)) IN ({placeholders})
                """,
                [item.upper() for item in chunk],
            )
            for row in cur.fetchall():
                item_key = (row[0] or "").strip().upper()
                if item_key and item_key not in result:
                    result[item_key] = (row[1] or "").strip()
        return result

    def read_sales_rows_from_csv(self, path):
        import csv
        import io

        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            raw_lines = handle.readlines()

        while raw_lines and not raw_lines[0].strip():
            raw_lines.pop(0)

        if raw_lines and raw_lines[0].strip() == '{}':
            raw_lines.pop(0)

        while raw_lines and not raw_lines[0].strip():
            raw_lines.pop(0)

        if not raw_lines:
            raise ValueError("The selected sales file has no header row.")

        reader = csv.DictReader(io.StringIO(''.join(raw_lines)))
        if not reader.fieldnames:
            raise ValueError("The selected sales file has no header row.")

        column_map = self.get_required_sales_columns(reader.fieldnames)
        raw_rows = list(reader)

        item_map = self.fetch_item_name_map(
            str(row.get(column_map["item_number"], "") or "").strip()
            for row in raw_rows
        )

        rows = []
        for line_number, row in enumerate(raw_rows, start=2):
            normalized = self.build_sales_import_tuple(row, column_map, line_number, item_map)
            if normalized is not None:
                rows.append(normalized)
        return rows

    def read_sales_rows_from_excel(self, path):
        if load_workbook is None:
            raise ValueError("Excel import is not available because openpyxl is not installed.")

        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not raw_headers:
                raise ValueError("The selected sales file has no header row.")
            headers = ["" if h is None else str(h) for h in raw_headers]
            column_map = self.get_required_sales_columns(headers)
            raw_rows = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                raw_rows.append({headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))})
        finally:
            workbook.close()

        item_map = self.fetch_item_name_map(
            str(row.get(column_map["item_number"], "") or "").strip()
            for row in raw_rows
        )

        rows = []
        for line_number, row in enumerate(raw_rows, start=2):
            normalized = self.build_sales_import_tuple(row, column_map, line_number, item_map)
            if normalized is not None:
                rows.append(normalized)
        return rows

    def build_sales_import_tuple(self, row, column_map, line_number, item_map):
        sale_date = self.normalize_sales_import_date(row.get(column_map["sale_date"], ""))
        customer_name = str(row.get(column_map["customer_name"], "") or "").strip()
        item_number = str(row.get(column_map["item_number"], "") or "").strip()
        quantity_value = self.parse_float(row.get(column_map["quantity"], 0))
        price_value = self.parse_float(row.get(column_map["price"], 0))

        if not sale_date and not customer_name and not item_number and quantity_value == 0 and price_value == 0:
            return None
        if not item_number:
            return None
        if not customer_name:
            raise ValueError(f"Line {line_number}: customer name is blank.")
        if not sale_date:
            raise ValueError(f"Line {line_number}: sale date is blank.")

        description = item_map.get(item_number.upper(), "")
        if not description:
            description = item_number
        month_key = self.sales_month_key_from_date(sale_date)
        extended = quantity_value * price_value
        return (sale_date, customer_name, item_number, description, month_key, quantity_value, price_value, extended)

    def import_stock_from_dialog(self):
        filters = "Stock files (*.xlsx *.xlsm *.csv *.txt);;Excel files (*.xlsx *.xlsm);;CSV files (*.csv *.txt);;All files (*.*)"
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Choose stock import file",
            str(self.base_dir),
            filters,
        )
        if not file_path:
            return

        try:
            imported_count = self.import_stock_file(file_path)
        except Exception as exc:
            QMessageBox.critical(self, "Stock import failed", f"Could not import stock file.\n\n{exc}")
            return

        display_text = (
            f"Last import: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
            f"File: {Path(file_path).name}\n"
            f"Rows imported: {imported_count:,}"
        )
        self.set_meta_value("stock_last_import_display", display_text)
        browser = getattr(self.ui, "lastUPdateStock_textBrowser_2", None)
        if browser is not None:
            browser.setPlainText(display_text)

        self.rerun_item_if_ready()
        self.rerun_order_analysis_if_ready()

        QMessageBox.information(self, "Stock imported", f"Imported {imported_count:,} rows into the stock table.")

    def import_stock_file(self, file_path):
        rows = self.read_stock_import_rows(file_path)
        cur = self.db_conn.cursor()
        cur.execute("DELETE FROM stock")
        cur.executemany(
            "INSERT INTO stock (item_number, on_hand, committed, on_order, available) VALUES (?, ?, ?, ?, ?)",
            [
                (
                    row.get("item_number", ""),
                    row.get("on_hand", 0),
                    row.get("committed", 0),
                    row.get("on_order", 0),
                    row.get("available", 0),
                )
                for row in rows
            ],
        )
        self.sync_items_from_stock_rows(rows)
        self.db_conn.commit()
        return len(rows)

    def read_stock_import_rows(self, file_path):
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix in {".csv", ".txt"}:
            return self.read_stock_rows_from_csv(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self.read_stock_rows_from_excel(path)
        raise ValueError("Unsupported file type. Choose an Excel or CSV stock file.")

    def get_required_stock_columns(self, headers):
        normalized = {self.normalize_header(h): h for h in headers}
        aliases = {
            "item_number": ["itemno", "itemnumber", "item"],
            "item_name": ["itemname", "description", "itemdescription", "name"],
            "on_hand": ["onhand", "qtyonhand", "quantityonhand"],
            "committed": ["committed", "qtycommitted", "quantitycommitted"],
            "on_order": ["onorder", "qtyonorder", "quantityonorder"],
            "available": ["available", "qtyavailable", "quantityavailable"],
        }
        required_fields = {"item_number", "on_hand", "committed", "on_order", "available"}
        resolved = {}
        missing = []
        for field_name, options in aliases.items():
            matched = next((normalized[key] for key in options if key in normalized), None)
            if matched is None:
                if field_name in required_fields:
                    missing.append(field_name)
            else:
                resolved[field_name] = matched
        if missing:
            missing_text = ", ".join(missing)
            raise ValueError(
                f"Missing required column(s): {missing_text}. Expected columns like Item No., On Hand, Committed, On Order, Available."
            )
        return resolved

    def build_stock_import_tuple(self, row, column_map, line_number):
        item_number = str(row.get(column_map["item_number"], "") or "").strip()
        item_name = str(row.get(column_map.get("item_name", ""), "") or "").strip() if column_map.get("item_name") else ""
        on_hand = self.parse_float(row.get(column_map["on_hand"], 0))
        committed = self.parse_float(row.get(column_map["committed"], 0))
        on_order = self.parse_float(row.get(column_map["on_order"], 0))
        available = self.parse_float(row.get(column_map["available"], 0))

        if not item_number and not item_name and on_hand == 0 and committed == 0 and on_order == 0 and available == 0:
            return None
        if not item_number:
            return None
        if item_number.startswith("\\"):
            return None

        return {
            "item_number": item_number,
            "item_name": item_name,
            "on_hand": on_hand,
            "committed": committed,
            "on_order": on_order,
            "available": available,
        }

    def locate_stock_header_row(self, rows, max_scan_rows=40):
        scan_limit = min(len(rows), max_scan_rows)
        for idx, row in enumerate(rows[:scan_limit]):
            headers = ["" if value is None else str(value) for value in row]
            try:
                column_map = self.get_required_stock_columns(headers)
                return idx, headers, column_map
            except ValueError:
                continue
        raise ValueError("Could not find the stock header row. Expected headers like Item No., On Hand, Committed, On Order, Available.")

    def read_stock_rows_from_csv(self, path):
        import csv

        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = list(csv.reader(handle))
        if not reader:
            raise ValueError("The selected stock file is empty.")

        header_index, headers, column_map = self.locate_stock_header_row(reader)
        rows = []
        for line_number, values in enumerate(reader[header_index + 1:], start=header_index + 2):
            row_dict = {headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))}
            normalized = self.build_stock_import_tuple(row_dict, column_map, line_number)
            if normalized is not None:
                rows.append(normalized)
        if not rows:
            raise ValueError("The selected stock file did not contain any valid stock rows.")
        return rows

    def read_stock_rows_from_excel(self, path):
        if load_workbook is None:
            raise ValueError("Excel import is not available because openpyxl is not installed.")

        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            raw_rows = [tuple(row) for row in sheet.iter_rows(min_row=1, max_row=60, values_only=True)]
            if not raw_rows:
                raise ValueError("The selected stock file is empty.")
            header_index, headers, column_map = self.locate_stock_header_row(raw_rows)
            rows = []
            for line_number, values in enumerate(sheet.iter_rows(min_row=header_index + 2, values_only=True), start=header_index + 2):
                row_dict = {headers[idx]: values[idx] if idx < len(headers) else "" for idx in range(len(headers))}
                normalized = self.build_stock_import_tuple(row_dict, column_map, line_number)
                if normalized is not None:
                    rows.append(normalized)
        finally:
            workbook.close()

        if not rows:
            raise ValueError("The selected stock file did not contain any valid stock rows.")
        return rows

    def sync_items_from_stock_rows(self, stock_rows):
        if self.db_conn is None or not stock_rows or not self.has_table("items"):
            return

        item_columns = {name.lower(): name for name in self.get_table_columns("items")}
        item_number_column = item_columns.get("item_number")
        if not item_number_column:
            return

        item_name_column = item_columns.get("item_name")
        description_column = item_columns.get("description")
        on_hand_column = item_columns.get("on_hand")
        committed_column = item_columns.get("committed")
        on_order_column = item_columns.get("on_order")
        available_column = item_columns.get("available")

        normalized_items = sorted({str(row.get("item_number", "") or "").strip().upper() for row in stock_rows if str(row.get("item_number", "") or "").strip()})
        existing_items = {}
        if normalized_items:
            chunk_size = 900
            for start in range(0, len(normalized_items), chunk_size):
                chunk = normalized_items[start:start + chunk_size]
                placeholders = ",".join("?" for _ in chunk)
                rows = self.db_all(
                    f"SELECT * FROM items WHERE UPPER(TRIM([{item_number_column}])) IN ({placeholders})",
                    chunk,
                )
                for existing_row in rows:
                    existing_items[str(existing_row[item_number_column] or "").strip().upper()] = self.row_to_dict(existing_row)

        cur = self.db_conn.cursor()
        for stock_row in stock_rows:
            item_number = str(stock_row.get("item_number", "") or "").strip()
            if not item_number:
                continue
            item_key = item_number.upper()
            item_name = str(stock_row.get("item_name", "") or "").strip()
            existing_row = existing_items.get(item_key)

            if existing_row is None:
                insert_columns = [item_number_column]
                insert_values = [item_number]

                if item_name_column and item_name:
                    insert_columns.append(item_name_column)
                    insert_values.append(item_name)
                if description_column and item_name and description_column not in insert_columns:
                    insert_columns.append(description_column)
                    insert_values.append(item_name)
                if on_hand_column:
                    insert_columns.append(on_hand_column)
                    insert_values.append(self.parse_float(stock_row.get("on_hand", 0)))
                if committed_column:
                    insert_columns.append(committed_column)
                    insert_values.append(self.parse_float(stock_row.get("committed", 0)))
                if on_order_column:
                    insert_columns.append(on_order_column)
                    insert_values.append(self.parse_float(stock_row.get("on_order", 0)))
                if available_column:
                    insert_columns.append(available_column)
                    insert_values.append(self.parse_float(stock_row.get("available", 0)))

                column_sql = ", ".join(f"[{name}]" for name in insert_columns)
                placeholder_sql = ", ".join("?" for _ in insert_values)
                cur.execute(
                    f"INSERT INTO items ({column_sql}) VALUES ({placeholder_sql})",
                    insert_values,
                )
                continue

            update_assignments = []
            update_values = []
            if item_name:
                if item_name_column and not str(existing_row.get(item_name_column, "") or "").strip():
                    update_assignments.append(f"[{item_name_column}] = ?")
                    update_values.append(item_name)
                if description_column and not str(existing_row.get(description_column, "") or "").strip():
                    update_assignments.append(f"[{description_column}] = ?")
                    update_values.append(item_name)

            if update_assignments:
                cur.execute(
                    f"UPDATE items SET {', '.join(update_assignments)} WHERE UPPER(TRIM([{item_number_column}])) = UPPER(TRIM(?))",
                    [*update_values, item_number],
                )


    def normalise_order_status(self, status_text="", is_urgent=False):
        status = str(status_text or "").strip().upper()
        if status == "IN CONTAINER":
            return "IN CONTAINER"
        if status == "URGENT" or is_urgent:
            return "URGENT"
        return ""

    def build_order_status_item(self, status_text="", is_urgent=False):
        status = self.normalise_order_status(status_text, is_urgent=is_urgent)
        if status == "URGENT":
            return self.make_order_table_item(
                "URGENT",
                background=QColor("#ffb3b3"),
                foreground=QColor("#7a0000"),
                bold=True,
                align=Qt.AlignCenter,
            )
        if status == "IN CONTAINER":
            return self.make_order_table_item(
                "IN CONTAINER",
                background=QColor("#d9ecff"),
                foreground=QColor("#003b73"),
                bold=True,
                align=Qt.AlignCenter,
            )
        return self.make_order_table_item("", align=Qt.AlignCenter)

    def build_order_priority_item(self, is_urgent):
        return self.build_order_status_item("URGENT" if is_urgent else "")

    def build_order_remove_item(self):
        return self.make_order_table_item(
            "Remove",
            background=QColor("#ffd6d6"),
            foreground=QColor("#8b0000"),
            bold=True,
            align=Qt.AlignCenter,
        )

    def populate_order_table_from_rows(self, rows):
        table = getattr(self.ui, "order_table", None)
        if table is None:
            return
        self._updating_order_table = True
        table.setRowCount(0)
        for line in rows:
            row = table.rowCount()
            table.insertRow(row)
            item_number = line.get("item_number", "")
            table.setItem(row, 0, self.make_order_table_item(item_number))
            table.setItem(row, 1, self.make_order_table_item(line.get("description", "")))
            table.setItem(
                row,
                self.order_on_order_column,
                self.make_order_table_item(
                    self.format_value(self.get_item_on_order_qty(item_number)),
                    align=Qt.AlignRight | Qt.AlignVCenter,
                ),
            )

            qty_value = self.parse_float(line.get("qty", 0))
            qty_item = self.make_order_table_item(
                self.format_value(qty_value),
                editable=True,
                align=Qt.AlignRight | Qt.AlignVCenter,
            )
            qty_item.setData(Qt.UserRole, qty_value)
            table.setItem(row, self.order_qty_column, qty_item)

            supplier_value = (line.get("supplier_name", "") or "").strip()
            supplier_item = self.make_order_table_item(supplier_value, editable=False)
            supplier_item.setData(Qt.UserRole, supplier_value)
            supplier_item.setToolTip("Double-click to choose a supplier.")
            table.setItem(row, self.order_supplier_column, supplier_item)

            table.setItem(row, self.order_priority_column, self.build_order_status_item(line.get("status", ""), bool(line.get("urgent"))))
            table.setItem(row, self.order_remove_column, self.build_order_remove_item())
        self._updating_order_table = False
        table.resizeRowsToContents()

    def get_order_table_rows(self):
        table = getattr(self.ui, "order_table", None)
        if table is None:
            return []
        rows = []
        for row in range(table.rowCount()):
            item_number = table.item(row, 0).text().strip() if table.item(row, 0) is not None and table.item(row, 0).text() else ""
            description = table.item(row, 1).text().strip() if table.item(row, 1) is not None and table.item(row, 1).text() else ""
            qty_item = table.item(row, self.order_qty_column)
            qty_value = qty_item.data(Qt.UserRole) if qty_item is not None else None
            if qty_value in (None, "") and qty_item is not None:
                qty_value = qty_item.text()
            supplier_item = table.item(row, self.order_supplier_column)
            supplier_value = ""
            if supplier_item is not None:
                supplier_value = str(supplier_item.data(Qt.UserRole) or supplier_item.text() or "").strip()
            priority_text = table.item(row, self.order_priority_column).text().strip().upper() if table.item(row, self.order_priority_column) is not None and table.item(row, self.order_priority_column).text() else ""
            if not item_number:
                continue
            rows.append({
                "item_number": item_number,
                "description": description,
                "qty": self.parse_float(qty_value),
                "supplier_name": supplier_value,
                "urgent": priority_text == "URGENT",
                "status": self.normalise_order_status(priority_text),
            })
        return rows

    def save_order_table_state(self):
        if self.db_conn is None or not self.has_table("to_order_lines"):
            return
        rows = self.get_order_table_rows()
        cur = self.db_conn.cursor()
        cur.execute("DELETE FROM to_order_lines")
        for line_no, row in enumerate(rows, start=1):
            cur.execute(
                """
                INSERT INTO to_order_lines (line_no, item_number, description, qty, supplier_name, urgent, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    line_no,
                    row.get("item_number", ""),
                    row.get("description", ""),
                    float(row.get("qty", 0) or 0),
                    row.get("supplier_name", ""),
                    1 if row.get("urgent") else 0,
                    row.get("status", ""),
                ),
            )
        self.db_conn.commit()

    def load_saved_order_lines(self):
        if self.db_conn is None or not self.has_table("to_order_lines"):
            return
        rows = self.db_all(
            """
            SELECT item_number, description, qty, supplier_name, urgent, status
            FROM to_order_lines
            ORDER BY line_no, id
            """
        )
        parsed_rows = [
            {
                "item_number": row["item_number"] or "",
                "description": row["description"] or "",
                "qty": self.parse_float(row["qty"]),
                "supplier_name": row["supplier_name"] or "",
                "urgent": bool(row["urgent"]),
                "status": self.normalise_order_status(row.get("status", ""), bool(row["urgent"])),
            }
            for row in rows
        ]
        self.populate_order_table_from_rows(parsed_rows)

    def prompt_for_supplier_name(self, current_supplier=""):
        dialog = QDialog(self)
        dialog.setWindowTitle("Choose Supplier")
        dialog.resize(420, 110)
        layout = QVBoxLayout(dialog)

        label = QLabel("Supplier:", dialog)
        layout.addWidget(label)

        edit = QLineEdit(dialog)
        edit.setText(current_supplier or "")
        completer = QCompleter(self.supplier_names, dialog)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchStartsWith)
        completer.setCompletionMode(QCompleter.PopupCompletion)
        completer.setModelSorting(QCompleter.CaseInsensitivelySortedModel)
        edit.setCompleter(completer)
        layout.addWidget(edit)

        buttons = QHBoxLayout()
        buttons.addStretch(1)
        ok_button = QPushButton("OK", dialog)
        cancel_button = QPushButton("Cancel", dialog)
        buttons.addWidget(ok_button)
        buttons.addWidget(cancel_button)
        layout.addLayout(buttons)

        ok_button.clicked.connect(dialog.accept)
        cancel_button.clicked.connect(dialog.reject)
        edit.returnPressed.connect(dialog.accept)
        edit.selectAll()
        edit.setFocus()

        while True:
            if dialog.exec() != QDialog.Accepted:
                return None
            resolved_supplier = self.find_supplier_name(edit.text().strip())
            if resolved_supplier:
                return resolved_supplier
            QMessageBox.warning(
                dialog,
                "Invalid supplier",
                "Supplier must match an existing supplier name or a unique starting match.",
            )
            edit.selectAll()
            edit.setFocus()

    def make_order_table_item(self, text, editable=False, align=None, background=None, foreground=None, bold=False):
        item = QTableWidgetItem("" if text is None else str(text))
        flags = Qt.ItemIsSelectable | Qt.ItemIsEnabled
        if editable:
            flags |= Qt.ItemIsEditable
        item.setFlags(flags)
        if align is not None:
            item.setTextAlignment(align)
        if background is not None:
            item.setBackground(background)
        if foreground is not None:
            item.setForeground(foreground)
        if bold:
            font = item.font()
            font.setBold(True)
            item.setFont(font)
        return item

    def add_order_line_from_inputs(self):
        item_edit = self.get_order_item_edit()
        qty_widget = self.get_order_qty_widget()
        table = getattr(self.ui, "order_table", None)
        urgent_check = getattr(self.ui, "urgent_check", None)

        if item_edit is None or qty_widget is None or table is None:
            return

        typed_item = item_edit.text().strip()
        item_number = self.find_item_number(typed_item)
        if not item_number:
            QMessageBox.warning(self, "Invalid item", "Please choose a valid item number from the list.")
            item_edit.setFocus()
            return

        qty_text = qty_widget.text().strip()
        qty_value = self.parse_float(qty_text)
        if qty_value <= 0:
            QMessageBox.warning(self, "Invalid quantity", "Please enter a quantity greater than 0.")
            qty_widget.setFocus()
            qty_widget.selectAll()
            return

        item_row = self.get_item_master_row(item_number)
        description = self.get_first(item_row, "description", "item_name", "Item Name", "Description")
        supplier = self.get_first(item_row, "supplier_name", "supplier_code", "Column1", "Supplier")
        is_urgent = bool(urgent_check and urgent_check.isChecked())
        on_order_qty = self.get_item_on_order_qty(item_number)

        row = table.rowCount()
        self._updating_order_table = True
        table.insertRow(row)
        table.setItem(row, 0, self.make_order_table_item(item_number))
        table.setItem(row, 1, self.make_order_table_item(description))
        table.setItem(row, self.order_on_order_column, self.make_order_table_item(self.format_value(on_order_qty), align=Qt.AlignRight | Qt.AlignVCenter))

        qty_item = self.make_order_table_item(
            self.format_value(qty_value),
            editable=True,
            align=Qt.AlignRight | Qt.AlignVCenter,
        )
        qty_item.setData(Qt.UserRole, qty_value)
        table.setItem(row, self.order_qty_column, qty_item)

        supplier_item = self.make_order_table_item(supplier, editable=False)
        supplier_item.setData(Qt.UserRole, supplier)
        supplier_item.setToolTip("Double-click to choose a supplier.")
        table.setItem(row, self.order_supplier_column, supplier_item)

        table.setItem(row, self.order_priority_column, self.build_order_status_item("URGENT" if is_urgent else "", is_urgent))
        table.setItem(row, self.order_remove_column, self.build_order_remove_item())
        self._updating_order_table = False

        table.resizeRowsToContents()
        self.save_order_table_state()
        self.refresh_item_summary_context_boxes()
        self.rerun_order_analysis_if_ready()
        item_edit.clear()
        order_widget.clear()
        qty_widget.clear()
        item_edit.setFocus()

    def handle_order_table_double_click(self, row, column):
        table = getattr(self.ui, "order_table", None)
        if table is None or row < 0:
            return

        if column == self.order_remove_column:
            item_number = table.item(row, 0).text() if table.item(row, 0) else "this line"
            result = QMessageBox.question(
                self,
                "Remove line",
                f"Remove {item_number} from the order sheet?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if result == QMessageBox.Yes:
                table.removeRow(row)
                self.save_order_table_state()
                self.refresh_item_summary_context_boxes()
                self.rerun_order_analysis_if_ready()
            return

        if column == self.order_qty_column:
            item = table.item(row, column)
            if item is not None:
                table.editItem(item)
            return

        if column == self.order_supplier_column:
            item = table.item(row, column)
            current_supplier = ""
            if item is not None:
                current_supplier = str(item.data(Qt.UserRole) or item.text() or "").strip()
            resolved_supplier = self.prompt_for_supplier_name(current_supplier)
            if resolved_supplier is None:
                return
            if item is None:
                item = self.make_order_table_item("", editable=False)
                table.setItem(row, column, item)
            self._updating_order_table = True
            item.setText(resolved_supplier)
            item.setData(Qt.UserRole, resolved_supplier)
            item.setToolTip("Double-click to choose a supplier.")
            self._updating_order_table = False
            item_number_item = table.item(row, 0)
            item_number = item_number_item.text().strip() if item_number_item is not None and item_number_item.text() else ""
            self.update_item_supplier_in_database(item_number, resolved_supplier)
            self.load_reference_lists()
            self.setup_supplier_autocomplete()
            self.save_order_table_state()
            return

        if column == self.order_priority_column:
            priority_item = table.item(row, column)
            current_text = (priority_item.text() or "").strip().upper() if priority_item is not None else ""
            new_status = "" if current_text == "URGENT" else "URGENT"
            self._updating_order_table = True
            table.setItem(row, column, self.build_order_status_item(new_status))
            self._updating_order_table = False
            self.save_order_table_state()
            return

    def handle_order_table_item_changed(self, item):
        if self._updating_order_table or item is None:
            return

        if item.column() == self.order_on_order_column:
            return

        if item.column() == self.order_qty_column:
            new_qty = self.parse_float(item.text())
            old_qty = self.parse_float(item.data(Qt.UserRole))
            if new_qty <= 0:
                QMessageBox.warning(self, "Invalid quantity", "Quantity must be greater than 0.")
                self._updating_order_table = True
                item.setText(self.format_value(old_qty))
                self._updating_order_table = False
                return

            self._updating_order_table = True
            item.setText(self.format_value(new_qty))
            item.setData(Qt.UserRole, new_qty)
            self._updating_order_table = False
            self.save_order_table_state()
            self.refresh_item_summary_context_boxes()
            self.rerun_order_analysis_if_ready()
            return

        if item.column() == self.order_supplier_column:
            old_supplier = str(item.data(Qt.UserRole) or "").strip()
            typed_supplier = item.text().strip()
            resolved_supplier = self.find_supplier_name(typed_supplier)

            if not resolved_supplier:
                QMessageBox.warning(
                    self,
                    "Invalid supplier",
                    "Supplier must match an existing supplier name or a unique starting match.",
                )
                self._updating_order_table = True
                item.setText(old_supplier)
                self._updating_order_table = False
                return

            self._updating_order_table = True
            item.setText(resolved_supplier)
            item.setData(Qt.UserRole, resolved_supplier)
            self._updating_order_table = False
            item_number_item = getattr(self.ui, "order_table", None).item(item.row(), 0) if getattr(self.ui, "order_table", None) is not None else None
            item_number = item_number_item.text().strip() if item_number_item is not None and item_number_item.text() else ""
            self.update_item_supplier_in_database(item_number, resolved_supplier)
            self.load_reference_lists()
            self.setup_supplier_autocomplete()
            self.save_order_table_state()
            return

    def handle_container_table_item_changed(self, _item):
        if self._updating_container_table:
            return
        self.refresh_container_totals()

    def refresh_order_table_on_order_column(self):
        table = getattr(self.ui, "order_table", None)
        if table is None:
            return
        self._updating_order_table = True
        try:
            for row in range(table.rowCount()):
                item_number = table.item(row, 0).text().strip() if table.item(row, 0) is not None and table.item(row, 0).text() else ""
                on_order_item = table.item(row, self.order_on_order_column)
                if on_order_item is None:
                    on_order_item = self.make_order_table_item("", align=Qt.AlignRight | Qt.AlignVCenter)
                    table.setItem(row, self.order_on_order_column, on_order_item)
                on_order_item.setText(self.format_value(self.get_item_on_order_qty(item_number)))
        finally:
            self._updating_order_table = False

    def get_order_sheet_item_qty(self, item_number):
        table = getattr(self.ui, "order_table", None)
        if table is None or not item_number:
            return 0.0

        total = 0.0
        target = item_number.strip().upper()
        for row in range(table.rowCount()):
            item_cell = table.item(row, 0)
            qty_cell = table.item(row, self.order_qty_column)
            item_text = item_cell.text().strip().upper() if item_cell is not None and item_cell.text() else ""
            if item_text != target:
                continue
            qty_value = qty_cell.data(Qt.UserRole) if qty_cell is not None else None
            if qty_value in (None, "") and qty_cell is not None:
                qty_value = qty_cell.text()
            total += self.parse_float(qty_value)
        return total

    def get_on_order_sheet_item_qty(self, item_number):
        table = getattr(self, "onOrder_table", None)
        if table is None or not item_number:
            return 0.0

        total = 0.0
        target = item_number.strip().upper()
        for row in range(table.rowCount()):
            item_cell = table.item(row, self.on_order_item_column)
            qty_cell = table.item(row, self.on_order_qty_column)
            item_text = item_cell.text().strip().upper() if item_cell is not None and item_cell.text() else ""
            if item_text != target:
                continue
            qty_value = qty_cell.data(Qt.UserRole) if qty_cell is not None else None
            if qty_value in (None, "") and qty_cell is not None:
                qty_value = qty_cell.text()
            total += self.parse_float(qty_value)
        return total

    def get_open_container_summary(self, item_number):
        target = (item_number or "").strip().upper()
        if not target:
            return {"qty": 0.0, "eta_text": "", "container_ref": ""}

        total = 0.0
        table = getattr(self.ui, "container_table", None)
        if table is not None:
            for row in range(table.rowCount()):
                first_item = table.item(row, 0)
                meta = first_item.data(Qt.UserRole) if first_item is not None else None
                if isinstance(meta, dict) and meta.get("note_row"):
                    continue
                item_cell = table.item(row, self.container_columns["item"])
                qty_cell = table.item(row, self.container_columns["qty"])
                item_text = item_cell.text().strip().upper() if item_cell is not None and item_cell.text() else ""
                if item_text != target:
                    continue
                qty_value = qty_cell.data(Qt.UserRole) if qty_cell is not None else None
                if qty_value in (None, "") and qty_cell is not None:
                    qty_value = qty_cell.text()
                total += self.parse_float(qty_value)

        current_item_widget = getattr(self.ui, "itemNumberContainer_line", None)
        current_qty_widget = getattr(self.ui, "qtyContainder_line", None)
        current_item = current_item_widget.text().strip().upper() if current_item_widget is not None else ""
        if current_item == target:
            total += self.parse_float(current_qty_widget.text() if current_qty_widget is not None else 0)

        eta_text = ""
        eta_date_edit = getattr(self.ui, "eta_dateEdit", None)
        if total > 0 and eta_date_edit is not None:
            qdate = eta_date_edit.date()
            if qdate.isValid():
                eta_text = qdate.toString("dd/MM/yyyy")

        container_ref = (self.current_container_ref or self.get_container_ref_text() or "").strip()
        return {"qty": total, "eta_text": eta_text, "container_ref": container_ref}

    def get_saved_next_container_summary(self, item_number):
        target = (item_number or "").strip()
        if not target or self.db_conn is None or not self.has_table("container_lines") or not self.has_table("containers"):
            return {"qty": 0.0, "eta_text": "", "container_ref": ""}

        rows = self.db_all(
            """
            SELECT
                cl.container_ref,
                SUM(COALESCE(cl.qty, 0)) AS total_qty,
                MIN(NULLIF(TRIM(c.eta_date), '')) AS eta_date,
                MIN(NULLIF(TRIM(c.updated_on), '')) AS updated_on
            FROM container_lines cl
            LEFT JOIN containers c ON UPPER(TRIM(c.container_ref)) = UPPER(TRIM(cl.container_ref))
            WHERE UPPER(TRIM(cl.item_number)) = UPPER(TRIM(?))
            GROUP BY cl.container_ref
            """,
            (target,),
        )

        best = None
        for row in rows:
            row_dict = self.row_to_dict(row)
            qty = self.parse_float(row_dict.get("total_qty", 0))
            if qty <= 0:
                continue
            eta_date = self.parse_date_value(row_dict.get("eta_date"))
            updated_date = self.parse_date_value(row_dict.get("updated_on"))
            sort_date = eta_date or updated_date or date.max
            candidate = {
                "qty": qty,
                "eta_text": eta_date.strftime("%d/%m/%Y") if eta_date is not None else "",
                "container_ref": str(row_dict.get("container_ref") or "").strip(),
                "_sort_date": sort_date,
            }
            if best is None or candidate["_sort_date"] < best["_sort_date"] or (
                candidate["_sort_date"] == best["_sort_date"] and candidate["container_ref"].upper() < best["container_ref"].upper()
            ):
                best = candidate

        if best is None:
            return {"qty": 0.0, "eta_text": "", "container_ref": ""}
        return {"qty": best["qty"], "eta_text": best["eta_text"], "container_ref": best["container_ref"]}

    def get_next_container_summary(self, item_number):
        saved = self.get_saved_next_container_summary(item_number)
        current = self.get_open_container_summary(item_number)

        if current.get("qty", 0) <= 0:
            return saved
        if saved.get("qty", 0) <= 0:
            return current

        current_ref = str(current.get("container_ref", "") or "").strip().upper()
        saved_ref = str(saved.get("container_ref", "") or "").strip().upper()
        if current_ref and saved_ref and current_ref == saved_ref:
            return current

        current_eta = self.parse_date_value(current.get("eta_text"))
        saved_eta = self.parse_date_value(saved.get("eta_text"))

        if current_eta is not None and saved_eta is not None:
            return current if current_eta <= saved_eta else saved
        if current_eta is not None:
            return current
        if saved_eta is not None:
            return saved

        return current if current_ref else saved

    def get_container_sheet_item_qty(self, item_number):
        summary = self.get_next_container_summary(item_number)
        return self.parse_float(summary.get("qty", 0))

    def get_next_container_eta_text(self, item_number):
        summary = self.get_next_container_summary(item_number)
        return str(summary.get("eta_text", "") or "")

    def get_orders_table_item_summary(self, item_number):
        row = self.db_one(
            """
            SELECT
                SUM(COALESCE(quantity, 0)) AS total_qty,
                MIN(NULLIF(TRIM(order_date), '')) AS first_order_date
            FROM orders
            WHERE UPPER(TRIM(item_number)) = UPPER(TRIM(?))
            """,
            (item_number,),
        )
        row_dict = self.row_to_dict(row)
        total_qty = self.parse_float(row_dict.get("total_qty", 0)) if row_dict else 0.0
        order_date_text = ""
        if row_dict:
            raw_date_value = row_dict.get("first_order_date")
            parsed_date = self.parse_date_value(raw_date_value)
            if parsed_date is not None:
                order_date_text = parsed_date.strftime("%d/%m/%Y")
            else:
                raw_date = "" if raw_date_value is None else str(raw_date_value).strip()
                if raw_date:
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d%m%Y"):
                        try:
                            order_date_text = datetime.strptime(raw_date, fmt).strftime("%d/%m/%Y")
                            break
                        except ValueError:
                            continue
                    if not order_date_text:
                        order_date_text = raw_date
        return total_qty, order_date_text

    def refresh_item_summary_context_boxes(self, inbound_context=None):
        if not self.current_item_number:
            return
        if not isinstance(inbound_context, dict):
            inbound_context = None
        context = inbound_context or self.get_item_inbound_context(self.current_item_number, self.get_lead_time_days())
        order_form = context.get("order_form", {})
        next_container = context.get("next_container", {})
        shipped = context.get("shipped", {})
        self.set_numeric_box("onOrderForm_box", order_form.get("qty", 0))
        self.set_numeric_box("onNextContainer_box", next_container.get("qty", 0))
        self.set_numeric_box("shippedContainer_box", shipped.get("qty", 0))
        self.set_label_text("nextContainerETA_box", next_container.get("display_date", ""))
        self.set_label_text("shippedContainerETA_box", shipped.get("display_date", ""))
        self.apply_item_summary_fallback_styles(context)

    def ensure_spin_arrow_icon(self, direction, color_hex):
        key = (direction, color_hex)
        cached = self._spin_arrow_icon_cache.get(key)
        if cached:
            return cached

        cache_dir = Path(tempfile.gettempdir()) / "windsor_widget_spin_icons"
        cache_dir.mkdir(parents=True, exist_ok=True)
        file_path = cache_dir / f"spin_{direction}_{color_hex.replace('#', '')}.png"

        if not file_path.exists():
            pixmap = QPixmap(14, 14)
            pixmap.fill(Qt.transparent)
            painter = QPainter(pixmap)
            painter.setRenderHint(QPainter.Antialiasing)
            font = QFont()
            font.setBold(True)
            font.setPixelSize(11)
            painter.setFont(font)
            painter.setPen(QColor(color_hex))
            painter.drawText(pixmap.rect(), Qt.AlignCenter, '▲' if direction == 'up' else '▼')
            painter.end()
            pixmap.save(str(file_path), 'PNG')

        url = QUrl.fromLocalFile(str(file_path)).toString()
        self._spin_arrow_icon_cache[key] = url
        return url

    def sanitize_filename_component(self, value):
        text_value = str(value or "").strip()
        if not text_value:
            return "Container"
        text_value = re.sub(r'[\/:*?"<>|]+', '_', text_value)
        text_value = re.sub(r'\s+', ' ', text_value).strip()
        text_value = text_value.rstrip('.')
        return text_value or "Container"

    def get_container_export_directory(self):
        documents_dir = Path.home() / "Documents"
        base_dir = documents_dir if documents_dir.exists() else Path.home()
        export_dir = base_dir / "Windsor Widget Exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        return export_dir

    def build_container_export_path(self):
        container_ref = self.get_container_ref_text() or (self.current_container_ref or "Container")
        eta_widget = getattr(self.ui, "eta_dateEdit", None)
        if eta_widget is not None and eta_widget.date().isValid():
            year_part = eta_widget.date().toString("yy")
        else:
            year_part = date.today().strftime("%y")
        safe_ref = self.sanitize_filename_component(container_ref)
        filename = f"YU_CONTAINER_{safe_ref}_{year_part}.xlsx"
        return self.get_container_export_directory() / filename

    def write_container_export_workbook(self, export_path):
        if Workbook is None:
            raise RuntimeError("Excel export is not available because openpyxl is not installed.")

        table = getattr(self.ui, "container_table", None)
        if table is None:
            raise RuntimeError("Container table is not available.")

        container_ref = self.get_container_ref_text() or (self.current_container_ref or "")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Container"

        header_fill = PatternFill("solid", fgColor="D9E1F2") if PatternFill is not None else None
        title_fill = PatternFill("solid", fgColor="B8CCE4") if PatternFill is not None else None
        note_fill = PatternFill("solid", fgColor="F6F1C1") if PatternFill is not None else None
        bold_font = Font(bold=True) if Font is not None else None
        title_font = Font(bold=True, size=14) if Font is not None else None
        thin_side = Side(style="thin", color="000000") if Side is not None else None
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side) if Border is not None and thin_side is not None else None

        export_headers = ["Order Number", "Item Number", "Description", "Qty", "Cartons"]
        export_column_map = [
            self.container_columns["order"],
            self.container_columns["item"],
            self.container_columns["description"],
            self.container_columns["qty"],
            self.container_columns["cartons"],
        ]

        export_date = date.today().strftime("%d/%m/%Y")
        title_text = f"{export_date} YU NEXT CONTAINER"
        if container_ref:
            title_text = f"{title_text} {container_ref}"

        worksheet.cell(row=1, column=1, value=title_text)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_headers))
        if title_font is not None:
            worksheet.cell(row=1, column=1).font = title_font
        if title_fill is not None:
            worksheet.cell(row=1, column=1).fill = title_fill
        if Alignment is not None:
            worksheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        if cell_border is not None:
            for border_col in range(1, len(export_headers) + 1):
                worksheet.cell(row=1, column=border_col).border = cell_border
        worksheet.row_dimensions[1].height = 24

        row_index = 3
        for column, header_text in enumerate(export_headers, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=header_text)
            if bold_font is not None:
                cell.font = bold_font
            if header_fill is not None:
                cell.fill = header_fill
            if Alignment is not None:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        start_table_row = row_index
        row_index += 1
        total_cartons = 0.0

        for table_row in range(table.rowCount()):
            first_item = table.item(table_row, 0)
            meta = first_item.data(Qt.UserRole) if first_item is not None else None
            is_note_row = isinstance(meta, dict) and meta.get("note_row")

            if is_note_row:
                note_text = first_item.text() if first_item is not None else ""
                worksheet.cell(row=row_index, column=1, value=note_text)
                if bold_font is not None:
                    worksheet.cell(row=row_index, column=1).font = bold_font
                if note_fill is not None:
                    worksheet.cell(row=row_index, column=1).fill = note_fill
                if Alignment is not None:
                    worksheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="left", vertical="center")
                worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(export_headers))
                if cell_border is not None:
                    for border_col in range(1, len(export_headers) + 1):
                        worksheet.cell(row=row_index, column=border_col).border = cell_border
                row_index += 1
                continue

            for export_col, table_col in enumerate(export_column_map, start=1):
                item = table.item(table_row, table_col)
                if table_col in (self.container_columns["qty"], self.container_columns["cartons"], self.container_columns["additional_cartons"]):
                    raw_value = item.data(Qt.UserRole) if item is not None else ""
                    cell_value = self.parse_float(raw_value if raw_value not in (None, "") else (item.text() if item is not None else ""))
                    if table_col == self.container_columns["cartons"]:
                        total_cartons += cell_value
                else:
                    cell_value = item.text() if item is not None else ""
                worksheet.cell(row=row_index, column=export_col, value=cell_value)
                if Alignment is not None:
                    align_horizontal = "right" if export_col in (4, 5) else "left"
                    worksheet.cell(row=row_index, column=export_col).alignment = Alignment(horizontal=align_horizontal, vertical="center")
                if cell_border is not None:
                    worksheet.cell(row=row_index, column=export_col).border = cell_border
            row_index += 1

        total_row = row_index + 1
        total_label_cell = worksheet.cell(row=total_row, column=4, value="Total Cartons")
        total_value_cell = worksheet.cell(row=total_row, column=5, value=total_cartons)
        if bold_font is not None:
            total_label_cell.font = bold_font
            total_value_cell.font = bold_font
        if header_fill is not None:
            total_label_cell.fill = header_fill
            total_value_cell.fill = header_fill
        if Alignment is not None:
            total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
            total_value_cell.alignment = Alignment(horizontal="right", vertical="center")
        if cell_border is not None:
            total_label_cell.border = cell_border
            total_value_cell.border = cell_border

        for column in range(1, len(export_headers) + 1):
            max_length = 0
            column_letter = get_column_letter(column) if get_column_letter is not None else None
            cells = worksheet[column_letter] if column_letter is not None else []
            for cell in cells:
                try:
                    cell_length = len(str(cell.value or ""))
                except Exception:
                    cell_length = 0
                max_length = max(max_length, cell_length)
            if column_letter is not None:
                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)

        worksheet.freeze_panes = f"A{start_table_row + 1}"
        workbook.save(str(export_path))

    def create_outlook_container_email(self, attachment_path, container_ref, eta_display):
        body_text = f"Container Ref: {container_ref}\r\nDate: {eta_display}\r\n"
        subject_text = f"Container {container_ref}" if container_ref else "Container Export"
        ps_script = f"""
$ErrorActionPreference = 'Stop'

function New-OutlookApp {{
    try {{
        return New-Object -ComObject Outlook.Application
    }}
    catch {{
        try {{
            Start-Process "outlook.exe" | Out-Null
            Start-Sleep -Seconds 5
            return New-Object -ComObject Outlook.Application
        }}
        catch {{
            throw "Outlook could not be opened via COM. Make sure classic Outlook desktop is installed, signed in, and not hung in the background."
        }}
    }}
}}

$outlook = New-OutlookApp
$mail = $outlook.CreateItem(0)
$mail.Subject = @"
{subject_text}
"@.Trim()
$mail.Body = @"
{body_text}
"@
$mail.Attachments.Add(@"
{str(attachment_path)}
"@.Trim()) | Out-Null
$mail.Display()
"""
        script_path = Path(tempfile.gettempdir()) / "windsor_widget_container_email.ps1"
        script_path.write_text(ps_script, encoding="utf-8")
        try:
            result = subprocess.run(
                ["powershell", "-STA", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(script_path)],
                capture_output=True,
                text=True,
                check=False,
                timeout=20,
            )
        except subprocess.TimeoutExpired:
            raise RuntimeError(
                "Outlook did not respond within 20 seconds. "
                "The Excel export succeeded, but the email draft step timed out."
            )

        if result.returncode != 0:
            stderr = (result.stderr or result.stdout or "").strip()
            message = stderr or "Could not create Outlook email draft."
            message = message.replace("\r\n", "\n").strip()
            raise RuntimeError(message)

    def export_container_to_excel_and_email(self):
        table = getattr(self.ui, "container_table", None)
        if table is None or table.rowCount() == 0:
            QMessageBox.warning(self, "Export Container", "There are no container rows to export.")
            return

        container_ref = self.get_container_ref_text() or (self.current_container_ref or "")
        if not container_ref:
            QMessageBox.warning(self, "Export Container", "Enter or load a container reference first.")
            return

        export_path = self.build_container_export_path()
        try:
            self.write_container_export_workbook(export_path)
        except Exception as exc:
            QMessageBox.critical(self, "Export Container", f"Could not create Excel export.\n\n{exc}")
            return

        eta_widget = getattr(self.ui, "eta_dateEdit", None)
        eta_display = eta_widget.date().toString("dd/MM/yyyy") if eta_widget is not None and eta_widget.date().isValid() else date.today().strftime("%d/%m/%Y")

        try:
            self.create_outlook_container_email(export_path, container_ref, eta_display)
        except Exception as exc:
            QMessageBox.warning(
                self,
                "Exported, but email not created",
                f"The Excel file was created here:\n{export_path}\n\n"
                f"But the email draft could not be opened.\n\n"
                f"{exc}\n\n"
                "The export still succeeded."
            )
            return

        QMessageBox.information(
            self,
            "Export and Email",
            f"Excel export created and attached to a new email draft.\n\nFile: {export_path}"
        )

    def setup_logo(self):
        if not hasattr(self.ui, "logo"):
            return

        logo_file = self.find_data_file("Windsor Logo.jpg", "Windsor Logo.jpeg", "Windsor Logo.png")
        if not logo_file.exists():
            return

        if self._logo_label is None:
            self._logo_label = QLabel(self.ui.logo)
            self._logo_label.setAlignment(Qt.AlignCenter)
            layout = self.ui.logo.layout()
            if layout is None:
                layout = QVBoxLayout(self.ui.logo)
                layout.setContentsMargins(6, 6, 6, 6)
            else:
                while layout.count():
                    item = layout.takeAt(0)
                    widget = item.widget()
                    if widget is not None:
                        widget.deleteLater()
            layout.addWidget(self._logo_label)

        pixmap = QPixmap(str(logo_file))
        if pixmap.isNull():
            return

        target_width = max(1, self.ui.logo.width() - 12)
        target_height = max(1, self.ui.logo.height() - 12)
        scaled = pixmap.scaled(target_width, target_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self._logo_label.setPixmap(scaled)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, "ui"):
            self.setup_logo()

    # -----------------------------
    # General helpers
    # -----------------------------
    def parse_float(self, value):
        if value is None:
            return 0.0
        text = str(value).strip().replace(",", "").replace("$", "")
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0


    def normalize_item_code(self, value):
        return re.sub(r"\s+", "", str(value or "")).upper()

    def find_item_number_by_normalized(self, normalized_code):
        target = self.normalize_item_code(normalized_code)
        if not target:
            return None
        for item_no in self.item_numbers:
            if self.normalize_item_code(item_no) == target:
                return item_no
        return None

    def get_thread_sales_item_numbers(self, item_number):
        valid_item = self.find_item_number(item_number) or self.find_item_number_by_normalized(item_number) or item_number
        valid_item = (valid_item or "").strip()
        if not valid_item:
            return []

        sales_items = [valid_item]
        if not self.is_combine_threads_enabled():
            return sales_items

        normalized = self.normalize_item_code(valid_item)

        # New Liberty thread convention:
        #   India   -> BN60 101
        #   Liberty -> BN60 101 L
        # So combine the base code with the same normalized code plus/minus trailing L.
        if normalized.endswith("L"):
            counterpart_normalized = normalized[:-1]
        else:
            counterpart_normalized = f"{normalized}L"

        counterpart = self.find_item_number_by_normalized(counterpart_normalized)
        if counterpart and self.normalize_item_code(counterpart) != normalized:
            sales_items.append(counterpart)

        unique = []
        seen = set()
        for code in sales_items:
            key = self.normalize_item_code(code)
            if key and key not in seen:
                seen.add(key)
                unique.append(code)
        return unique

    def build_sales_item_filter_clause(self, item_numbers):
        normalized_codes = []
        for code in item_numbers:
            key = self.normalize_item_code(code)
            if key:
                normalized_codes.append(key)
        clause, params = self.sql_in_clause(normalized_codes)
        return f"UPPER(REPLACE(TRIM(item_number), ' ', '')) IN {clause}", params

    def format_value(self, value):
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            if math.isfinite(float(value)) and float(value).is_integer():
                return f"{int(round(float(value))):,}"
            return f"{float(value):,.2f}"
        return str(value)

    def format_price(self, value):
        if value is None:
            return ""
        numeric = self.parse_float(value)
        return f"{numeric:,.2f}"

    def safe_text(self, value):
        if value is None:
            return ""
        text = str(value).strip()
        if text in {"0", "0.0", "0.00"}:
            return "0"
        return text

    def first_of_month(self, d):
        return date(d.year, d.month, 1)

    def next_month(self, d):
        if d.month == 12:
            return date(d.year + 1, 1, 1)
        return date(d.year, d.month + 1, 1)

    def month_list_between(self, start_date, end_date):
        months = []
        current = self.first_of_month(start_date)
        last = self.first_of_month(end_date)
        while current <= last:
            months.append(current)
            current = self.next_month(current)
        return months

    def month_start_from_picker(self, picker):
        qd = picker.date() if hasattr(picker, "date") else QDate.currentDate()
        if hasattr(picker, "monthStart"):
            qd = picker.monthStart()
        return date(qd.year(), qd.month(), 1)

    def month_end_from_picker(self, picker):
        qd = picker.date() if hasattr(picker, "date") else QDate.currentDate()
        if hasattr(picker, "monthEnd"):
            qd = picker.monthEnd()
            return date(qd.year(), qd.month(), qd.day())
        d0 = date(qd.year(), qd.month(), 1)
        return self.next_month(d0) - timedelta(days=1)

    def normalize_customer_name(self, name):
        name = (name or "").strip()
        while True:
            new_name = self.STATE_SUFFIX_RE.sub("", name).strip()
            if new_name == name:
                break
            name = new_name
        return re.sub(r"\s+", " ", name).strip()

    def find_customer_name(self, typed_text):
        typed = (typed_text or "").strip()
        if not typed:
            return None

        typed_lower = typed.lower()
        for name in self.customer_names:
            if name.lower() == typed_lower:
                return name

        prefix_matches = [name for name in self.customer_names if name.lower().startswith(typed_lower)]
        if len(prefix_matches) == 1:
            return prefix_matches[0]

        if getattr(self.ui, "combineStateAccountsCheck", None) and self.ui.combineStateAccountsCheck.isChecked():
            typed_base = self.normalize_customer_name(typed).lower()
            base_matches = [
                name for name in self.customer_names
                if self.normalize_customer_name(name).lower() == typed_base
            ]
            if len(base_matches) == 1:
                return base_matches[0]

            prefix_base_matches = [
                name for name in self.customer_names
                if self.normalize_customer_name(name).lower().startswith(typed_base)
            ]
            unique_bases = []
            seen = set()
            for name in prefix_base_matches:
                key = self.normalize_customer_name(name).lower()
                if key not in seen:
                    seen.add(key)
                    unique_bases.append(name)
            if len(unique_bases) == 1:
                return unique_bases[0]
        return None

    def find_matching_customers(self, selected_customer, combine_accounts):
        if not selected_customer:
            return []
        if not combine_accounts:
            return [selected_customer]
        selected_base = self.normalize_customer_name(selected_customer).lower()
        return [
            name for name in self.customer_names
            if self.normalize_customer_name(name).lower() == selected_base
        ]

    def find_item_number(self, typed_text):
        typed = (typed_text or "").strip()
        if not typed:
            return None

        typed_upper = typed.upper()
        for item_no in self.item_numbers:
            if item_no.upper() == typed_upper:
                return item_no

        prefix_matches = [item_no for item_no in self.item_numbers if item_no.upper().startswith(typed_upper)]
        if len(prefix_matches) == 1:
            return prefix_matches[0]
        return None

    def find_supplier_name(self, typed_text):
        typed = (typed_text or "").strip()
        if not typed:
            return None

        typed_lower = typed.lower()
        for supplier_name in self.supplier_names:
            if supplier_name.lower() == typed_lower:
                return supplier_name

        prefix_matches = [name for name in self.supplier_names if name.lower().startswith(typed_lower)]
        if len(prefix_matches) == 1:
            return prefix_matches[0]
        return None

    def rerun_order_analysis_if_ready(self, *_args):
        supplier_edit = self.get_order_analysis_supplier_edit()
        if supplier_edit is None:
            return
        typed = supplier_edit.text().strip()
        if not typed:
            return
        if self.find_supplier_name(typed):
            self.load_order_analysis(show_warning=False)

    def sql_in_clause(self, values):
        if not values:
            return "(NULL)", []
        placeholders = ",".join("?" for _ in values)
        return f"({placeholders})", list(values)

    def parse_month_key(self, month_key):
        month_key = (month_key or "").strip()
        if not month_key:
            return None
        for fmt in ("%Y%m", "%Y-%m", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(month_key, fmt)
                return date(parsed.year, parsed.month, 1)
            except ValueError:
                continue
        return None

    def previous_month(self, d):
        current = self.first_of_month(d)
        return current - timedelta(days=1)

    def month_end(self, d):
        return self.next_month(self.first_of_month(d)) - timedelta(days=1)

    def current_business_date(self):
        return date.today()

    def parse_date_value(self, value):
        if value is None:
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        if hasattr(value, "toPython"):
            try:
                py_value = value.toPython()
                if isinstance(py_value, datetime):
                    return py_value.date()
                if isinstance(py_value, date):
                    return py_value
            except Exception:
                pass
        if hasattr(value, "isValid") and hasattr(value, "year"):
            try:
                if value.isValid():
                    return date(value.year(), value.month(), value.day())
            except Exception:
                pass

        text = str(value).strip()
        if not text:
            return None

        if " " in text:
            text = text.split(" ", 1)[0].strip()
        if "T" in text:
            text = text.split("T", 1)[0].strip()

        text = text.replace(".", "/").replace("-", "/")
        text = re.sub(r"\s+", "", text)

        if re.fullmatch(r"\d{8}", text):
            day = int(text[0:2])
            month = int(text[2:4])
            year = int(text[4:8])
            try:
                return date(year, month, day)
            except ValueError:
                return None

        if re.fullmatch(r"\d{4}/\d{1,2}/\d{1,2}", text):
            year_s, month_s, day_s = text.split("/")
            try:
                return date(int(year_s), int(month_s), int(day_s))
            except ValueError:
                return None

        if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
            day_s, month_s, year_s = text.split("/")
            year = int(year_s)
            if len(year_s) == 2:
                year += 2000 if year < 70 else 1900
            try:
                return date(year, int(month_s), int(day_s))
            except ValueError:
                return None

        return None

    def format_display_date(self, value):
        parsed = self.parse_date_value(value)
        return parsed.strftime("%d/%m/%Y") if parsed else ""

    def format_short_date(self, value):
        parsed = self.parse_date_value(value)
        return parsed.strftime("%d/%m/%Y") if parsed else ""

    def format_signed_value(self, value):
        if value is None:
            return ""
        numeric = self.parse_float(value)
        prefix = "+" if numeric > 0 else ""
        return f"{prefix}{self.format_value(numeric)}"

    def set_signed_box(self, object_name, value):
        self.set_label_text(object_name, self.format_signed_value(value))

    def get_lead_time_weeks(self):
        value = int(self.lead_time_picker.value()) if self.lead_time_picker else 14
        return max(1, value)

    def get_lead_time_days(self):
        return self.get_lead_time_weeks() * 7

    def round_order_quantity(self, qty, carton_size=0, pallet_size=0):
        numeric = self.parse_float(qty)
        if abs(numeric) < 1e-9:
            return 0.0

        sign = 1 if numeric >= 0 else -1
        magnitude = abs(numeric)
        carton = self.parse_float(carton_size)
        pallet = self.parse_float(pallet_size)

        if carton > 0:
            magnitude = math.ceil(magnitude / carton) * carton
            if pallet > 0 and magnitude >= pallet:
                magnitude = math.ceil(magnitude / pallet) * pallet
        elif pallet > 0:
            magnitude = math.ceil(magnitude / pallet) * pallet
        else:
            magnitude = round(magnitude)

        return sign * magnitude

    def item_summary_accent_color_map(self):
        return {
            "suggested_min": "#E49EDD",
            "suggestedMin_box": "#E49EDD",
            "suggestedOrder_label": "#E49EDD",
            "suggestedOrder_box": "#E49EDD",
            "atRisk_label": "#E49EDD",
            "atRisk_box": "#E49EDD",
            "itemName_label_2": "#B5E6A2",
            "itemName_box": "#B5E6A2",
            "itemNumber_label_2": "#B5E6A2",
            "itemNumber_box": "#B5E6A2",
            "itemGroup_label_2": "#B5E6A2",
            "itemGroup_box": "#B5E6A2",
            "stockOnHand_label": "#B5E6A2",
            "stockOnHand_box": "#B5E6A2",
            "onOrderForm_label": "#D9E1F2",
            "onOrderForm_box": "#D9E1F2",
            "onNextContainer_label": "#D9E1F2",
            "onNextContainer_box": "#D9E1F2",
            "shippedContainer_label": "#D9E1F2",
            "shippedContainer_box": "#D9E1F2",
        }

    def build_item_summary_accent_stylesheet(self, object_name, warning=False):
        color = self.item_summary_accent_color_map().get(object_name)
        if not color:
            return ""
        border_color = "#e5b500" if warning else "#4a4a4a"
        return (
            f"background-color: {color}; "
            f"color: #111111; "
            f"border: 3px solid {border_color};"
        )

    def apply_item_summary_accent_styles(self):
        for object_name in self.item_summary_accent_color_map().keys():
            widget = getattr(self.ui, object_name, None)
            if widget is not None:
                widget.setStyleSheet(self.build_item_summary_accent_stylesheet(object_name, False))

    def set_widget_warning_state(self, object_name, enabled=False, tooltip=""):
        widget = getattr(self.ui, object_name, None)
        if widget is None:
            return
        accent_style = self.build_item_summary_accent_stylesheet(object_name, enabled)
        if accent_style:
            widget.setStyleSheet(accent_style)
        elif enabled:
            widget.setStyleSheet("border: 3px solid #e5b500; background-color: rgba(229, 181, 0, 0.18);")
        else:
            widget.setStyleSheet("")
        widget.setToolTip(tooltip or "")

    def fetch_item_monthly_qty(self, item_numbers, start_date, end_date):
        item_list = item_numbers if isinstance(item_numbers, (list, tuple, set)) else [item_numbers]
        where_clause, params = self.build_sales_item_filter_clause(item_list)
        rows = self.db_all(
            f"""
            SELECT
                COALESCE(NULLIF(TRIM(month_key), ''), STRFTIME('%Y-%m', DATE(sale_date))) AS month_key,
                SUM(COALESCE(quantity, 0)) AS total_qty
            FROM sales
            WHERE {where_clause}
              AND DATE(sale_date) BETWEEN ? AND ?
            GROUP BY COALESCE(NULLIF(TRIM(month_key), ''), STRFTIME('%Y-%m', DATE(sale_date)))
            ORDER BY month_key
            """,
            [*params, start_date.isoformat(), end_date.isoformat()],
        )

        month_totals = {}
        for row in rows:
            month_start = self.parse_month_key(row["month_key"])
            if month_start is None:
                continue
            month_totals[month_start] = self.parse_float(row["total_qty"])
        return month_totals

    def get_item_inbound_context(self, item_number, lead_days=None, today=None):
        lead_days = int(lead_days if lead_days is not None else self.get_lead_time_days())
        today = today or self.current_business_date()
        fallback_date = today + timedelta(days=lead_days)

        order_form_qty = self.get_order_sheet_item_qty(item_number) + self.get_on_order_sheet_item_qty(item_number)
        next_container_qty = self.get_container_sheet_item_qty(item_number)
        next_container_eta_text = self.get_next_container_eta_text(item_number)
        next_container_eta = self.parse_date_value(next_container_eta_text)
        next_container_used_fallback = next_container_qty > 0 and next_container_eta is None
        next_container_arrival = next_container_eta or (fallback_date if next_container_qty > 0 else None)

        shipped_qty, shipped_eta_text = self.get_orders_table_item_summary(item_number)
        shipped_eta = self.parse_date_value(shipped_eta_text)
        shipped_used_fallback = shipped_qty > 0 and shipped_eta is None
        shipped_arrival = shipped_eta or (fallback_date if shipped_qty > 0 else None)

        def entry(qty, arrival_date, used_fallback, source_name, include_in_calcs, tooltip_override=None):
            tooltip = tooltip_override or ""
            if qty > 0 and arrival_date:
                if used_fallback:
                    detail = f"using lead-time fallback ETA {self.format_display_date(arrival_date)}"
                else:
                    detail = f"ETA {self.format_display_date(arrival_date)}"
                if include_in_calcs:
                    tooltip = tooltip or f"{source_name}: {detail}. Counted as inbound."
                else:
                    tooltip = tooltip or f"{source_name}: {detail}. Display only and not counted as hard inbound."
            elif qty > 0 and not tooltip:
                tooltip = f"{source_name}: qty shown for reference only."
            return {
                "qty": qty,
                "arrival_date": arrival_date,
                "display_date": self.format_display_date(arrival_date),
                "used_fallback": bool(used_fallback and qty > 0),
                "tooltip": tooltip,
                "include_in_calcs": bool(include_in_calcs and qty > 0 and arrival_date is not None),
                "source_name": source_name,
            }

        order_form_tooltip = (
            "To order + on order sheets: qty shown for reference only. Not counted as inbound until it is physically assigned "
            "to a real shipment or container."
            if order_form_qty > 0 else ""
        )

        return {
            "today": today,
            "lead_days": lead_days,
            "horizon_date": fallback_date,
            "order_form": entry(order_form_qty, None, False, "To order / on order", False, order_form_tooltip),
            "next_container": entry(next_container_qty, next_container_arrival, next_container_used_fallback, "Next container", True),
            "shipped": entry(shipped_qty, shipped_arrival, shipped_used_fallback, "Shipped", True),
        }

    def inbound_qty_by_date(self, inbound_context, cutoff_date):
        if cutoff_date is None:
            return 0.0
        total = 0.0
        for key in ("shipped", "next_container", "order_form"):
            entry = inbound_context.get(key, {})
            if not entry.get("include_in_calcs", False):
                continue
            arrival_date = entry.get("arrival_date")
            if entry.get("qty", 0) > 0 and arrival_date is not None and arrival_date <= cutoff_date:
                total += self.parse_float(entry.get("qty", 0))
        return total

    def earliest_inbound_date(self, inbound_context):
        dates = []
        for key in ("shipped", "next_container", "order_form"):
            entry = inbound_context.get(key, {})
            if not entry.get("include_in_calcs", False):
                continue
            qty = self.parse_float(entry.get("qty", 0))
            arrival_date = entry.get("arrival_date")
            if qty > 0 and arrival_date is not None:
                dates.append(arrival_date)
        return min(dates) if dates else None

    def calculate_suggested_order(self, avg_monthly_qty, on_hand, inbound_context, lead_days, carton_size=0, pallet_size=0):
        lead_demand = self.parse_float(avg_monthly_qty) * (lead_days / 30.4375)
        hard_inbound = self.inbound_qty_by_date(inbound_context, inbound_context.get("horizon_date"))
        inventory_position = self.parse_float(on_hand) + hard_inbound
        suggested_raw = max(0.0, lead_demand - inventory_position)
        suggested_rounded = self.round_order_quantity(suggested_raw, carton_size, pallet_size)
        return {
            "lead_demand": lead_demand,
            "inventory_position": inventory_position,
            "hard_inbound": hard_inbound,
            "raw": suggested_raw,
            "rounded": suggested_rounded,
        }

    def calculate_at_risk(self, avg_monthly_qty, on_hand, inbound_context, lead_days, carton_size=0, pallet_size=0):
        today = inbound_context.get("today") or self.current_business_date()
        cutoff_date = self.earliest_inbound_date(inbound_context) or (today + timedelta(days=lead_days))
        days_until_cutoff = max(0, (cutoff_date - today).days)
        demand_until_cutoff = self.parse_float(avg_monthly_qty) * (days_until_cutoff / 30.4375)
        supply_cutoff = cutoff_date - timedelta(days=1) if cutoff_date else None
        supply_before_cutoff = self.parse_float(on_hand) + self.inbound_qty_by_date(inbound_context, supply_cutoff)
        at_risk_raw = max(0.0, demand_until_cutoff - supply_before_cutoff)
        at_risk_rounded = self.round_order_quantity(at_risk_raw, carton_size, pallet_size)
        return {
            "cutoff_date": cutoff_date,
            "days_until_cutoff": days_until_cutoff,
            "demand_until_cutoff": demand_until_cutoff,
            "supply_by_cutoff": supply_before_cutoff,
            "raw": at_risk_raw,
            "rounded": at_risk_rounded,
        }

    def calculate_trending_adjustment(self, months, monthly_qty_map, lead_days, carton_size=0, pallet_size=0, today=None):
        today = today or self.current_business_date()
        current_month_start = date(today.year, today.month, 1)
        completed_months = [m for m in months if m < current_month_start]
        if len(completed_months) < 2:
            return {
                "recent_count": len(completed_months),
                "previous_count": 0,
                "recent_avg": 0.0,
                "previous_avg": 0.0,
                "raw": 0.0,
                "rounded": 0.0,
            }

        total_completed = len(completed_months)
        if total_completed >= 4:
            recent_count = 3
        else:
            recent_count = max(1, total_completed - 1)
        previous_count = min(3, total_completed - recent_count)

        recent_months = completed_months[-recent_count:]
        previous_months = completed_months[-(recent_count + previous_count):-recent_count] if previous_count > 0 else []

        if not previous_months:
            return {
                "recent_count": len(recent_months),
                "previous_count": 0,
                "recent_avg": 0.0,
                "previous_avg": 0.0,
                "raw": 0.0,
                "rounded": 0.0,
            }

        recent_avg = sum(self.parse_float(monthly_qty_map.get(m, 0.0)) for m in recent_months) / len(recent_months)
        previous_avg = sum(self.parse_float(monthly_qty_map.get(m, 0.0)) for m in previous_months) / len(previous_months)
        monthly_delta = recent_avg - previous_avg
        trend_raw = monthly_delta * (lead_days / 30.4375)
        trend_rounded = self.round_order_quantity(trend_raw, carton_size, pallet_size)
        return {
            "recent_count": len(recent_months),
            "previous_count": len(previous_months),
            "recent_avg": recent_avg,
            "previous_avg": previous_avg,
            "raw": trend_raw,
            "rounded": trend_rounded,
        }

    def calculate_seasonal_adjustment(self, item_numbers, avg_monthly_qty, arrival_date, lead_days=0, carton_size=0, pallet_size=0, today=None):
        today = today or self.current_business_date()
        last_full_year = today.year - 1
        start_year = last_full_year - 2
        if start_year < 1 or arrival_date is None:
            return {
                "years_with_sales": 0,
                "factor": 1.0,
                "month_factor": 1.0,
                "december_factor": 1.0,
                "prebuild_weight": 0.0,
                "january_unwind_weight": 0.0,
                "raw": 0.0,
                "rounded": 0.0,
            }

        start_date = date(start_year, 1, 1)
        end_date = date(last_full_year, 12, 31)
        month_totals = self.fetch_item_monthly_qty(item_numbers, start_date, end_date)

        all_months = []
        current = start_date
        while current <= end_date:
            all_months.append(current)
            current = self.next_month(current)

        totals_with_gaps = {month_start: self.parse_float(month_totals.get(month_start, 0.0)) for month_start in all_months}
        years_with_sales = {
            year for year in range(start_year, last_full_year + 1)
            if any(totals_with_gaps.get(date(year, month, 1), 0.0) > 0 for month in range(1, 13))
        }

        if len(years_with_sales) < 3:
            return {
                "years_with_sales": len(years_with_sales),
                "factor": 1.0,
                "month_factor": 1.0,
                "december_factor": 1.0,
                "prebuild_weight": 0.0,
                "january_unwind_weight": 0.0,
                "raw": 0.0,
                "rounded": 0.0,
            }

        overall_avg = sum(totals_with_gaps.values()) / len(totals_with_gaps) if totals_with_gaps else 0.0
        if overall_avg <= 0:
            return {
                "years_with_sales": len(years_with_sales),
                "factor": 1.0,
                "month_factor": 1.0,
                "december_factor": 1.0,
                "prebuild_weight": 0.0,
                "january_unwind_weight": 0.0,
                "raw": 0.0,
                "rounded": 0.0,
            }

        month_factor_map = {}
        for month in range(1, 13):
            month_values = [totals_with_gaps[date(year, month, 1)] for year in range(start_year, last_full_year + 1)]
            month_avg = sum(month_values) / len(month_values) if month_values else 0.0
            month_factor_map[month] = month_avg / overall_avg if overall_avg > 0 else 1.0

        target_month = arrival_date.month
        month_factor = month_factor_map.get(target_month, 1.0)
        december_factor = month_factor_map.get(12, 1.0)
        december_lift = max(0.0, december_factor - 1.0)

        # Seasonal action window:
        # - Sep arrivals get a small prebuild lift
        # - Oct arrivals get a medium prebuild lift
        # - Nov arrivals get the full prebuild lift
        # - Jan arrivals can unwind part of the December build
        # - All other arrival months stay neutral at 0
        prebuild_weight = {
            9: 0.25,
            10: 0.50,
            11: 1.00,
        }.get(target_month, 0.0)
        january_unwind_weight = 0.25 if target_month == 1 else 0.0

        if prebuild_weight > 0.0:
            effective_lift = december_lift * prebuild_weight
        elif january_unwind_weight > 0.0:
            effective_lift = -(december_lift * january_unwind_weight)
        else:
            effective_lift = 0.0

        seasonal_factor = max(0.0, 1.0 + effective_lift)

        coverage_months = max(1.0, self.parse_float(lead_days) / 30.4375) if self.parse_float(avg_monthly_qty) > 0 else 1.0
        seasonal_raw = effective_lift * self.parse_float(avg_monthly_qty) * coverage_months
        seasonal_rounded = self.round_order_quantity(seasonal_raw, carton_size, pallet_size)
        return {
            "years_with_sales": len(years_with_sales),
            "factor": seasonal_factor,
            "month_factor": month_factor,
            "december_factor": december_factor,
            "prebuild_weight": prebuild_weight,
            "january_unwind_weight": january_unwind_weight,
            "active_window": target_month in (1, 9, 10, 11),
            "raw": seasonal_raw,
            "rounded": seasonal_rounded,
        }

    def apply_item_summary_fallback_styles(self, inbound_context=None):
        context = inbound_context or {}
        order_form = context.get("order_form", {})
        next_container = context.get("next_container", {})
        shipped = context.get("shipped", {})

        self.set_widget_warning_state(
            "onOrderForm_box",
            order_form.get("used_fallback", False),
            order_form.get("tooltip", ""),
        )
        self.set_widget_warning_state(
            "onNextContainer_box",
            next_container.get("used_fallback", False),
            next_container.get("tooltip", ""),
        )
        self.set_widget_warning_state(
            "nextContainerETA_box",
            next_container.get("used_fallback", False),
            next_container.get("tooltip", ""),
        )
        self.set_widget_warning_state(
            "shippedContainer_box",
            shipped.get("used_fallback", False),
            shipped.get("tooltip", ""),
        )
        self.set_widget_warning_state(
            "shippedContainerETA_box",
            shipped.get("used_fallback", False),
            shipped.get("tooltip", ""),
        )

    def set_label_text(self, object_name, text):
        widget = getattr(self.ui, object_name, None)
        if widget is not None:
            widget.setText("" if text is None else str(text))

    def clear_item_summary_fields(self):
        self.current_item_number = None
        label_names = [
            "itemNumber_box", "itemName_box", "itemGroup_box", "rollSpool_box", "mtUnit_box",
            "box_box", "palletCarton_box", "totalQtySold_box", "avrMonthlySales_box",
            "suggestedMin_box", "stockOnHand_box", "stockCommited_box", "stockOnOrder_box",
            "stockAvailable_box", "onOrderForm_box", "onNextContainer_box", "shippedContainer_box",
            "nextContainerETA_box", "shippedContainerETA_box", "suggestedOrder_box", "atRisk_box",
            "seasonalOrder_box", "trendingOrder_box", "adjustedOrder_box",
        ]
        for label_name in label_names:
            self.set_label_text(label_name, "")
        for widget_name in (
            "onOrderForm_box", "onNextContainer_box", "nextContainerETA_box",
            "shippedContainer_box", "shippedContainerETA_box",
        ):
            self.set_widget_warning_state(widget_name, False, "")
        self.populate_customer_purchase_table([])

    def set_numeric_box(self, object_name, value):
        self.set_label_text(object_name, self.format_value(value))

    # -----------------------------
    # Theme handling
    # -----------------------------
    def restore_theme(self):
        theme = self.settings.value("theme", "dark")
        if theme == "light" and hasattr(self.ui, "radioLight"):
            self.ui.radioLight.setChecked(True)
        elif theme == "high" and hasattr(self.ui, "radioHighContrast"):
            self.ui.radioHighContrast.setChecked(True)
        elif hasattr(self.ui, "radioDark"):
            self.ui.radioDark.setChecked(True)
        self.apply_theme(theme)

    def apply_theme(self, theme_name):
        self.settings.setValue("theme", theme_name)
        app = QApplication.instance()
        if app is None:
            return

        if theme_name == "light":
            arrow_color = "#1b1f23"
        elif theme_name == "high":
            arrow_color = "#ffff00"
        else:
            arrow_color = "#e8eaed"

        up_arrow_url = self.ensure_spin_arrow_icon("up", arrow_color)
        down_arrow_url = self.ensure_spin_arrow_icon("down", arrow_color)

        if theme_name == "light":
            app.setStyleSheet(f"""
                QWidget {{ background: #e3e8ee; color: #1b1f23; }}
                QMainWindow, QStackedWidget, QScrollArea, QScrollArea > QWidget > QWidget {{
                    background: #e3e8ee;
                    color: #1b1f23;
                }}
                QFrame#frame_3, QFrame#frame_11, QFrame#frame_12, QFrame#frame_18,
                QFrame#frame_23, QFrame#frame_28, QFrame#frame_30, QFrame#frame_31,
                QFrame#frame_32, QFrame#frame_33, QFrame#frame_37, QFrame#frame_38,
                QFrame#frame_39, QFrame#frame_40, QFrame#frame_41, QFrame#frame_42,
                QFrame#frame_43, QFrame#frame_44, QFrame#frame_45 {{
                    background: #eef2f5;
                    border: 1px solid #c3cbd3;
                    border-radius: 8px;
                }}
                QLineEdit, QComboBox, QSpinBox, QTableWidget, QTableView, QTextEdit, QTextBrowser, QDateEdit {{
                    background: #f8fafb;
                    color: #111111;
                    border: 1px solid #aab4be;
                    border-radius: 4px;
                    selection-background-color: #c9def5;
                    selection-color: #111111;
                }}
                QSpinBox {{
                    padding-right: 24px;
                }}
                QSpinBox::up-button, QSpinBox::down-button {{
                    subcontrol-origin: border;
                    width: 22px;
                    background: #dde3e9;
                    border-left: 1px solid #aab4be;
                }}
                QSpinBox::up-button {{
                    subcontrol-position: top right;
                    border-bottom: 1px solid #aab4be;
                }}
                QSpinBox::down-button {{
                    subcontrol-position: bottom right;
                }}
                QSpinBox::up-button:hover, QSpinBox::down-button:hover {{
                    background: #d1d9e1;
                }}
                QSpinBox::up-arrow {{
                    image: url("{up_arrow_url}");
                    width: 12px;
                    height: 12px;
                }}
                QSpinBox::down-arrow {{
                    image: url("{down_arrow_url}");
                    width: 12px;
                    height: 12px;
                }}
                QPushButton {{
                    background: #d6dce3;
                    color: #1b1f23;
                    border: 1px solid #97a4b0;
                    border-radius: 5px;
                    padding: 6px;
                }}
                QPushButton:hover {{ background: #c9d1d9; }}
                QHeaderView::section {{
                    background: #d7dde4;
                    color: #111111;
                    border: 1px solid #aab4be;
                    padding: 4px;
                }}
                QRadioButton {{
                    color: #1b1f23; spacing: 6px; padding: 3px 6px; border-radius: 4px;
                }}
                QRadioButton:checked {{
                    color: white; background: #5c6b7a; font-weight: 700;
                }}
            """)
        elif theme_name == "high":
            app.setStyleSheet(f"""
                QWidget {{ background: black; color: yellow; }}
                QLineEdit, QComboBox, QSpinBox, QTableWidget, QTableView, QTextEdit, QTextBrowser, QDateEdit {{
                    background: black; color: yellow; border: 2px solid yellow;
                }}
                QSpinBox {{
                    padding-right: 24px;
                }}
                QSpinBox::up-button, QSpinBox::down-button {{
                    subcontrol-origin: border;
                    width: 22px;
                    background: #111111;
                    border-left: 2px solid yellow;
                }}
                QSpinBox::up-button {{
                    subcontrol-position: top right;
                    border-bottom: 2px solid yellow;
                }}
                QSpinBox::down-button {{
                    subcontrol-position: bottom right;
                }}
                QSpinBox::up-button:hover, QSpinBox::down-button:hover {{
                    background: #1c1c1c;
                }}
                QSpinBox::up-arrow {{
                    image: url("{up_arrow_url}");
                    width: 12px;
                    height: 12px;
                }}
                QSpinBox::down-arrow {{
                    image: url("{down_arrow_url}");
                    width: 12px;
                    height: 12px;
                }}
                QPushButton {{
                    background: black; color: cyan; border: 2px solid cyan; padding: 6px;
                }}
                QHeaderView::section {{
                    background: black; color: yellow; border: 1px solid yellow;
                }}
                QRadioButton {{
                    color: yellow; spacing: 6px; padding: 3px 6px; border-radius: 4px;
                }}
                QRadioButton:checked {{
                    color: white; background: #005a9e; font-weight: 700;
                }}
            """)
        else:
            app.setStyleSheet(f"""
                QWidget {{ background: #202124; color: #e8eaed; }}
                QLineEdit, QComboBox, QSpinBox, QTableWidget, QTableView, QTextEdit, QTextBrowser, QDateEdit {{
                    background: #2b2d30; color: #e8eaed; border: 1px solid #5f6368;
                }}
                QSpinBox {{
                    padding-right: 24px;
                }}
                QSpinBox::up-button, QSpinBox::down-button {{
                    subcontrol-origin: border;
                    width: 22px;
                    background: #37393d;
                    border-left: 1px solid #5f6368;
                }}
                QSpinBox::up-button {{
                    subcontrol-position: top right;
                    border-bottom: 1px solid #5f6368;
                }}
                QSpinBox::down-button {{
                    subcontrol-position: bottom right;
                }}
                QSpinBox::up-button:hover, QSpinBox::down-button:hover {{
                    background: #43464b;
                }}
                QSpinBox::up-arrow {{
                    image: url("{up_arrow_url}");
                    width: 12px;
                    height: 12px;
                }}
                QSpinBox::down-arrow {{
                    image: url("{down_arrow_url}");
                    width: 12px;
                    height: 12px;
                }}
                QPushButton {{
                    background: #303134; border: 1px solid #5f6368; padding: 6px;
                }}
                QPushButton:hover {{ background: #3c4043; }}
                QHeaderView::section {{
                    background: #303134; color: #e8eaed; border: 1px solid #5f6368;
                }}
                QRadioButton {{
                    color: #e8eaed; spacing: 6px; padding: 3px 6px; border-radius: 4px;
                }}
                QRadioButton:checked {{
                    color: white; background: #4a4d52; font-weight: 700;
                }}
            """)

        for spin in (self.lead_time_picker, getattr(self.ui, 'additional_spinner', None)):
            if spin is not None:
                spin.setButtonSymbols(QAbstractSpinBox.UpDownArrows)
                spin.update()

        self.apply_item_summary_accent_styles()
        self.update_navigation_button_highlight()
        self.redraw_charts()

    def redraw_charts(self):
        if self.current_customer_pivot and self.current_customer_months:
            totals = [sum(data["months"].get(m, 0.0) for data in self.current_customer_pivot.values()) for m in self.current_customer_months]
            self.draw_monthly_line_chart(self.customer_chart_view, self.current_customer_months, totals, "Monthly Units")
        current_item = self.find_item_number(getattr(self.ui, "enterItem", None).text().strip()) if hasattr(self.ui, "enterItem") else None
        if current_item:
            self.load_item_summary()

    def chart_text_color(self):
        if hasattr(self.ui, "radioHighContrast") and self.ui.radioHighContrast.isChecked():
            return QColor("yellow")
        if hasattr(self.ui, "radioLight") and self.ui.radioLight.isChecked():
            return QColor("#202124")
        return QColor("#e8eaed")

    def chart_background_color(self):
        if hasattr(self.ui, "radioHighContrast") and self.ui.radioHighContrast.isChecked():
            return QColor("black")
        if hasattr(self.ui, "radioLight") and self.ui.radioLight.isChecked():
            return QColor("#eef2f5")
        return QColor("#202124")

    # -----------------------------
    # Customer page
    # -----------------------------
    def rerun_search_if_ready(self, *_args):
        if hasattr(self.ui, "customerEdit") and (self.ui.customerEdit.text() or "").strip():
            self.search_customer_sales()

    def search_customer_sales(self):
        typed_customer = self.ui.customerEdit.text() if hasattr(self.ui, "customerEdit") else ""
        valid_customer = self.find_customer_name(typed_customer)

        if not valid_customer:
            QMessageBox.warning(self, "Invalid customer", "Please choose a valid customer from the list.")
            return

        start_date = self.month_start_from_picker(self.customer_start_picker)
        end_date = self.month_end_from_picker(self.customer_end_picker or self.customer_start_picker)
        combine_accounts = bool(getattr(self.ui, "combineStateAccountsCheck", None) and self.ui.combineStateAccountsCheck.isChecked())

        if start_date > end_date:
            QMessageBox.warning(self, "Invalid date range", "Start date cannot be after end date.")
            return

        matched_customers = self.find_matching_customers(valid_customer, combine_accounts)
        in_clause, in_params = self.sql_in_clause(matched_customers)

        rows = self.db_all(
            f"""
            SELECT
                TRIM(item_number) AS item_number,
                COALESCE(NULLIF(TRIM(month_key), ''), STRFTIME('%Y-%m', DATE(sale_date))) AS month_key,
                SUM(COALESCE(quantity, 0)) AS total_qty
            FROM sales
            WHERE DATE(sale_date) BETWEEN ? AND ?
              AND customer_name IN {in_clause}
            GROUP BY TRIM(item_number), COALESCE(NULLIF(TRIM(month_key), ''), STRFTIME('%Y-%m', DATE(sale_date)))
            ORDER BY item_number COLLATE NOCASE, month_key
            """,
            [start_date.isoformat(), end_date.isoformat(), *in_params],
        )

        last_price_rows = self.db_all(
            f"""
            SELECT
                TRIM(item_number) AS item_number,
                DATE(sale_date) AS sale_date,
                COALESCE(price, 0) AS price
            FROM sales
            WHERE DATE(sale_date) BETWEEN ? AND ?
              AND customer_name IN {in_clause}
              AND TRIM(COALESCE(item_number, '')) <> ''
            ORDER BY item_number COLLATE NOCASE, DATE(sale_date) DESC
            """,
            [start_date.isoformat(), end_date.isoformat(), *in_params],
        )

        months = self.month_list_between(start_date, end_date)
        pivot = {}
        monthly_totals = {m: 0.0 for m in months}

        item_name_map = self.fetch_item_name_map((row["item_number"] for row in rows))
        latest_sale_info_by_item = {}
        for row in last_price_rows:
            item_number = (row["item_number"] or "").strip()
            if not item_number or item_number in latest_sale_info_by_item:
                continue
            latest_sale_info_by_item[item_number] = {
                "last_price": self.parse_float(row["price"]),
                "last_sale_date": self.parse_date_value(row["sale_date"]),
            }

        for row in rows:
            item_number = (row["item_number"] or "").strip()
            if not item_number:
                continue
            month_start = self.parse_month_key(row["month_key"])
            qty = self.parse_float(row["total_qty"])

            if month_start is None or month_start not in monthly_totals:
                continue

            if item_number not in pivot:
                pivot[item_number] = {
                    "description": item_name_map.get(item_number.upper(), ""),
                    "months": {m: 0.0 for m in months},
                    "last_price": latest_sale_info_by_item.get(item_number, {}).get("last_price"),
                    "last_sale_date": latest_sale_info_by_item.get(item_number, {}).get("last_sale_date"),
                    "total_qty": 0.0,
                }

            pivot[item_number]["months"][month_start] += qty
            pivot[item_number]["total_qty"] += qty
            monthly_totals[month_start] += qty

        self.current_customer_months = months
        self.current_customer_pivot = pivot
        self.current_customer_name = valid_customer
        self.current_customer_combine_accounts = combine_accounts

        self.populate_monthly_item_table(pivot, months, valid_customer, combine_accounts)
        self.populate_customer_info(valid_customer)
        self.draw_monthly_line_chart(
            self.customer_chart_view,
            months,
            [monthly_totals[m] for m in months],
            "Monthly Units",
        )

    def populate_customer_info(self, customer_name):
        table = getattr(self.ui, "customer_Info", None)
        if table is None:
            return

        row = self.db_one(
            "SELECT * FROM customers WHERE LOWER(TRIM(customer_name)) = LOWER(TRIM(?)) LIMIT 1",
            (customer_name,),
        )
        row = self.row_to_dict(row)
        self.current_customer_file_path = self.get_first(row, "matched_file", "Matched File", default="")
        freight_value = self.get_customer_freight_value(row)
        self.update_charge_freight_box(self.is_yes_like(freight_value), freight_value)

        if row:
            values = [
                self.get_first(row, "customer_name", "Co./Last Name", default=customer_name),
                self.get_first(row, "addr1_line1", "Addr 1 - Line 1"),
                self.get_first(row, "addr1_city", "Addr 1 - City"),
                self.get_first(row, "addr1_state", "Addr 1 - State"),
                self.get_first(row, "addr1_postcode", "Addr 1 - Postcode"),
                self.get_first(row, "phone", "Addr 1 - Phone No. 1"),
            ]
        else:
            values = [customer_name]
            self.current_customer_file_path = ""

        model = QStandardItemModel(0, 1, self)
        for value in values:
            item = QStandardItem("" if value is None else str(value))
            item.setEditable(False)
            model.appendRow(item)

        table.setModel(model)
        table.resizeRowsToContents()

    def get_customer_files_root(self):
        root_text = str(self.settings.value("customerFilesRoot", "") or "").strip()
        if root_text:
            root_path = Path(root_text)
            if root_path.exists() and root_path.is_dir():
                return root_path

        default_candidates = [
            self.base_dir / "CustomerFiles",
            self.base_dir / "data" / "CustomerFiles",
        ]
        for candidate in default_candidates:
            if candidate.exists() and candidate.is_dir():
                return candidate
        return None

    def set_customer_files_root(self, folder_path):
        folder = Path(folder_path)
        try:
            folder = folder.resolve()
        except Exception:
            pass
        self.settings.setValue("customerFilesRoot", str(folder))

    def save_current_customer_file_name(self, selected_path):
        selected = Path(selected_path)
        file_name_only = selected.name.strip()
        customer_name = (self.current_customer_name or "").strip()
        if not customer_name:
            QMessageBox.warning(self, "Customer File", "No current customer is loaded, so the selected file could not be saved.")
            return False
        if not file_name_only:
            return False

        self.set_customer_files_root(selected.parent)
        cur = self.db_conn.cursor()
        cur.execute(
            "UPDATE customers SET matched_file = ? WHERE LOWER(TRIM(customer_name)) = LOWER(TRIM(?))",
            (file_name_only, customer_name),
        )
        self.db_conn.commit()
        self.current_customer_file_path = file_name_only
        self.populate_customer_info(customer_name)
        return True

    def select_customer_file_for_current_customer(self, parent_dialog=None):
        current_path_text = (self.current_customer_file_path or "").strip()
        current_resolved = self.resolve_customer_file_path(current_path_text) if current_path_text else None
        start_dir = None
        if current_resolved is not None:
            start_dir = current_resolved.parent
        else:
            start_dir = self.get_customer_files_root()
        if start_dir is None:
            start_dir = self.base_dir

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select customer file",
            str(start_dir),
            "Customer files (*.xlsx *.xlsm *.xltx *.xltm *.csv *.txt *.log *.md);;All files (*.*)",
        )
        if not file_path:
            return False

        selected = Path(file_path)
        if not self.save_current_customer_file_name(selected):
            return False

        if parent_dialog is not None and hasattr(parent_dialog, "load_new_path"):
            try:
                parent_dialog.load_new_path(selected)
            except Exception:
                pass
            return True

        if parent_dialog is not None:
            try:
                parent_dialog.accept()
            except Exception:
                pass

        self.show_customer_file_popup(selected)
        return True

    def resolve_customer_file_path(self, path_text):
        path_text = (path_text or "").strip()
        if not path_text:
            return None

        raw_path = Path(path_text)
        candidates = []
        customer_files_root = self.get_customer_files_root()

        if raw_path.is_absolute():
            candidates.append(raw_path)
        else:
            if customer_files_root is not None:
                candidates.extend([
                    customer_files_root / path_text,
                    customer_files_root / raw_path.name,
                ])
            candidates.extend([
                self.base_dir / path_text,
                self.base_dir / "CustomerFiles" / path_text,
                self.base_dir / "CustomerFiles" / raw_path.name,
                self.base_dir / "data" / "CustomerFiles" / path_text,
                self.base_dir / "data" / "CustomerFiles" / raw_path.name,
            ])

        seen = set()
        for candidate in candidates:
            try:
                resolved = candidate.resolve()
            except Exception:
                resolved = candidate
            key = str(resolved).lower()
            if key in seen:
                continue
            seen.add(key)
            if resolved.exists():
                return resolved

        search_dirs = []
        if customer_files_root is not None:
            search_dirs.append(customer_files_root)
        search_dirs.extend([
            self.base_dir / "CustomerFiles",
            self.base_dir / "data" / "CustomerFiles",
        ])
        target_name = raw_path.name.lower()
        seen_dirs = set()
        for directory in search_dirs:
            try:
                resolved_dir = directory.resolve()
            except Exception:
                resolved_dir = directory
            dir_key = str(resolved_dir).lower()
            if dir_key in seen_dirs or not resolved_dir.exists() or not resolved_dir.is_dir():
                continue
            seen_dirs.add(dir_key)
            for candidate in resolved_dir.iterdir():
                if candidate.is_file() and candidate.name.lower() == target_name:
                    return candidate.resolve()

        return None

    def open_customer_file(self):
        path_text = (self.current_customer_file_path or "").strip()
        path = self.resolve_customer_file_path(path_text) if path_text else None
        if path is not None:
            try:
                dialog = CustomerFileViewerDialog(path, self)
                dialog.is_customer_file_context = True
                self.customer_file_preview_dialog = dialog
                dialog.exec()
            except Exception as exc:
                self.customer_file_preview_dialog = None
                QMessageBox.warning(
                    self,
                    "Load Customer File",
                    f"Could not open customer file viewer.\n\n{exc}",
                )
                return
            self.customer_file_preview_dialog = None
        else:
            self.show_customer_file_popup(path, path_text)

    def open_file_with_default_app(self, path):
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(path))
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(str(path)))
        except Exception as exc:
            QMessageBox.warning(self, "Open File", f"Could not open file:\n{path}\n\n{exc}")

    def show_customer_file_popup(self, path=None, stored_value=""):
        dialog = QDialog(self)
        title_text = path.name if path is not None else (Path(stored_value).name if stored_value else "No customer file selected")
        dialog.setWindowTitle(f"Customer File - {title_text}")
        dialog.resize(1100, 720)

        layout = QVBoxLayout(dialog)

        title = QLabel(title_text)
        title_font = title.font()
        title_font.setPointSize(title_font.pointSize() + 1)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        if path is not None:
            path_display = str(path)
        elif stored_value:
            path_display = f"Stored file name: {stored_value}"
        else:
            path_display = "No matched customer file is currently stored."

        path_label = QLabel(path_display)
        path_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        path_label.setWordWrap(True)
        layout.addWidget(path_label)

        if path is not None:
            preview_widget = self.build_customer_file_preview(path)
        else:
            expected_root = self.get_customer_files_root()
            if expected_root is not None and stored_value:
                expected_text = str(expected_root / Path(stored_value).name)
            elif stored_value:
                expected_text = str(self.base_dir / "CustomerFiles" / Path(stored_value).name)
            else:
                expected_text = ""
            message = "No preview available because the stored customer file could not be found."
            if stored_value:
                message += f"\n\nStored value: {stored_value}"
            if expected_text:
                message += f"\nExpected location: {expected_text}"
            message += "\n\nUse Select File to choose the correct file. Only the file name will be stored in the database."
            preview_widget = self.build_text_message_widget(message)
        layout.addWidget(preview_widget, 1)

        button_row = QHBoxLayout()
        button_row.addStretch(1)

        select_button = QPushButton("Select File")
        select_button.clicked.connect(lambda: self.select_customer_file_for_current_customer(dialog))
        button_row.addWidget(select_button)

        open_button = QPushButton("Open in Excel")
        open_button.setEnabled(path is not None)
        if path is not None:
            open_button.clicked.connect(lambda: self.open_file_with_default_app(path))
        button_row.addWidget(open_button)

        close_button = QPushButton("Close")
        close_button.clicked.connect(dialog.accept)
        button_row.addWidget(close_button)

        layout.addLayout(button_row)

        self.customer_file_preview_dialog = dialog
        dialog.exec()
        self.customer_file_preview_dialog = None

    def build_customer_file_preview(self, path):
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"} and load_workbook is not None:
            return self.build_excel_preview_widget(path)
        if suffix == ".csv":
            return self.build_csv_preview_widget(path)
        if suffix in {".txt", ".log", ".md"}:
            return self.build_text_preview_widget(path)

        message = QLabel(
            "Preview is not available for this file type in the app. "
            "Use the button below to open the full file in Excel."
        )
        message.setWordWrap(True)
        message.setAlignment(Qt.AlignCenter)
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.addStretch(1)
        layout.addWidget(message)
        layout.addStretch(1)
        return container

    def build_excel_preview_widget(self, path, max_rows=200, max_cols=30, max_sheets=10):
        tabs = QTabWidget()
        workbook = None
        try:
            workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
            worksheet_names = workbook.sheetnames[:max_sheets]
            if not worksheet_names:
                return self.build_text_message_widget("This workbook does not contain any worksheets.")

            for sheet_name in worksheet_names:
                sheet = workbook[sheet_name]
                table = QTableWidget()
                table.setEditTriggers(QAbstractItemView.NoEditTriggers)
                table.setSelectionBehavior(QAbstractItemView.SelectItems)
                table.setAlternatingRowColors(True)
                table.verticalHeader().setVisible(False)

                rows = []
                for row in sheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    while values and values[-1] == "":
                        values.pop()
                    rows.append(values)

                column_count = max((len(row) for row in rows), default=1)
                table.setColumnCount(column_count)
                if get_column_letter is not None:
                    table.setHorizontalHeaderLabels([get_column_letter(i + 1) for i in range(column_count)])

                table.setRowCount(len(rows))
                for row_index, row_values in enumerate(rows):
                    for column_index, value in enumerate(row_values):
                        table.setItem(row_index, column_index, QTableWidgetItem(value))

                table.resizeColumnsToContents()
                table.horizontalHeader().setStretchLastSection(True)

                tab_page = QWidget()
                tab_layout = QVBoxLayout(tab_page)
                info = QLabel(f"Showing up to {max_rows} rows and {max_cols} columns.")
                info.setWordWrap(True)
                tab_layout.addWidget(info)
                tab_layout.addWidget(table, 1)
                tabs.addTab(tab_page, sheet_name)
        except Exception as exc:
            return self.build_text_message_widget(f"Could not preview workbook.\n\n{exc}")
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass

        return tabs

    def build_csv_preview_widget(self, path, max_rows=200, max_cols=30):
        import csv

        table = QTableWidget()
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectItems)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)

        rows = []
        try:
            with open(path, "r", newline="", encoding="utf-8-sig", errors="replace") as handle:
                reader = csv.reader(handle)
                for row_index, row in enumerate(reader):
                    if row_index >= max_rows:
                        break
                    rows.append([cell for cell in row[:max_cols]])
        except Exception as exc:
            return self.build_text_message_widget(f"Could not preview CSV file.\n\n{exc}")

        column_count = max((len(row) for row in rows), default=1)
        table.setColumnCount(column_count)
        table.setRowCount(len(rows))
        for row_index, row_values in enumerate(rows):
            for column_index, value in enumerate(row_values):
                table.setItem(row_index, column_index, QTableWidgetItem(value))

        table.resizeColumnsToContents()
        table.horizontalHeader().setStretchLastSection(True)

        container = QWidget()
        layout = QVBoxLayout(container)
        info = QLabel(f"Showing up to {max_rows} rows and {max_cols} columns.")
        info.setWordWrap(True)
        layout.addWidget(info)
        layout.addWidget(table, 1)
        return container

    def build_text_preview_widget(self, path, max_chars=50000):
        editor = QTextEdit()
        editor.setReadOnly(True)
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            return self.build_text_message_widget(f"Could not preview text file.\n\n{exc}")

        if len(text) > max_chars:
            text = text[:max_chars] + "\n\n[Preview truncated]"
        editor.setPlainText(text)
        return editor

    def build_text_message_widget(self, message):
        label = QLabel(message)
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignCenter)
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.addStretch(1)
        layout.addWidget(label)
        layout.addStretch(1)
        return container

    def populate_monthly_item_table(self, pivot, months, customer_name, combine_accounts):
        table = getattr(self.ui, "salesTable", None)
        if table is None:
            return

        item_numbers = sorted(pivot.keys(), key=str.lower)
        headers = ["Item Number", "Description", "Last Sale Date", "Last Price", "Total Units"] + [m.strftime("%b %Y") for m in months]

        table.clear()
        table.setRowCount(len(item_numbers))
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)

        for row_index, item_number in enumerate(item_numbers):
            item_data = pivot[item_number]
            table.setItem(row_index, 0, QTableWidgetItem(item_number))
            table.setItem(row_index, 1, QTableWidgetItem(item_data.get("description", "")))
            table.setItem(row_index, 2, QTableWidgetItem(self.format_short_date(item_data.get("last_sale_date"))))
            last_price = item_data.get("last_price")
            table.setItem(row_index, 3, QTableWidgetItem(self.format_price(last_price) if last_price is not None else ""))
            table.setItem(row_index, 4, QTableWidgetItem(self.format_value(item_data.get("total_qty", 0.0))))

            for month_index, month_start in enumerate(months, start=5):
                qty = item_data["months"].get(month_start, 0.0)
                month_item = QTableWidgetItem(self.format_value(qty))
                month_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                month_item.setData(Qt.UserRole, {
                    "item_number": item_number,
                    "month_start": month_start.isoformat(),
                    "qty": qty,
                })
                table.setItem(row_index, month_index, month_item)

            for numeric_col in (3, 4):
                num_item = table.item(row_index, numeric_col)
                if num_item is not None:
                    num_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

        header = table.horizontalHeader()
        header.setStretchLastSection(False)
        for col in range(table.columnCount()):
            header.setSectionResizeMode(col, QHeaderView.Interactive)
            table.resizeColumnToContents(col)

        if table.columnCount() > 0:
            table.setColumnWidth(0, max(table.columnWidth(0), 130))
        if table.columnCount() > 1:
            table.setColumnWidth(1, max(table.columnWidth(1), 280))
        if table.columnCount() > 2:
            table.setColumnWidth(2, max(table.columnWidth(2), 110))
        if table.columnCount() > 3:
            table.setColumnWidth(3, max(table.columnWidth(3), 100))
        if table.columnCount() > 4:
            table.setColumnWidth(4, max(table.columnWidth(4), 100))
        table.resizeRowsToContents()

        if combine_accounts:
            base_name = self.normalize_customer_name(customer_name)
            self.setWindowTitle(f"Monthly unit sales by item - {base_name} (combined state accounts)")
        else:
            self.setWindowTitle(f"Monthly unit sales by item - {customer_name}")

    def show_customer_row_chart(self, row, _column):
        table = getattr(self.ui, "salesTable", None)
        if table is None or not self.current_customer_pivot or not self.current_customer_months:
            return
        item_number_item = table.item(row, 0)
        if item_number_item is None:
            return
        item_number = item_number_item.text().strip()
        item_data = self.current_customer_pivot.get(item_number)
        if not item_data:
            return

        totals = [item_data["months"].get(month, 0.0) for month in self.current_customer_months]
        self.draw_monthly_line_chart(self.customer_chart_view, self.current_customer_months, totals, item_number)

    def handle_customer_table_double_click(self, row, column):
        table = getattr(self.ui, "salesTable", None)
        if table is None or row < 0 or column < 0:
            return

        if column == 0:
            item_number_item = table.item(row, 0)
            if item_number_item is None:
                return

            item_number = self.find_item_number(item_number_item.text())
            if not item_number:
                return

            self.open_item_summary_from_customer(item_number)
            return

        if column >= 5:
            self.show_customer_month_invoice_lines(row, column)
            return


    def fetch_customer_month_invoice_lines(self, customer_name, item_number, month_start, combine_accounts=False):
        month_start = self.parse_date_value(month_start)
        if month_start is None:
            return []

        matched_customers = self.find_matching_customers(customer_name, combine_accounts)
        in_clause, in_params = self.sql_in_clause(matched_customers)
        month_end = self.next_month(month_start) - timedelta(days=1)

        rows = self.db_all(
            f"""
            SELECT
                DATE(sale_date) AS sale_date,
                TRIM(customer_name) AS customer_name,
                TRIM(item_number) AS item_number,
                COALESCE(NULLIF(TRIM(description), ''), TRIM(item_number)) AS description,
                COALESCE(quantity, 0) AS quantity,
                COALESCE(price, 0) AS price,
                COALESCE(extended, COALESCE(quantity, 0) * COALESCE(price, 0)) AS extended
            FROM sales
            WHERE DATE(sale_date) BETWEEN ? AND ?
              AND customer_name IN {in_clause}
              AND UPPER(TRIM(COALESCE(item_number, ''))) = UPPER(TRIM(?))
            ORDER BY DATE(sale_date), customer_name COLLATE NOCASE, description COLLATE NOCASE
            """,
            [month_start.isoformat(), month_end.isoformat(), *in_params, item_number],
        )
        return [self.row_to_dict(row) for row in rows]

    def show_customer_month_invoice_lines(self, row, column):
        table = getattr(self.ui, "salesTable", None)
        if table is None or not self.current_customer_months:
            return
        if column < 5 or (column - 5) >= len(self.current_customer_months):
            return

        item_number_item = table.item(row, 0)
        if item_number_item is None:
            return
        item_number = (item_number_item.text() or "").strip()
        if not item_number:
            return

        month_start = self.current_customer_months[column - 5]
        combine_accounts = bool(getattr(self, "current_customer_combine_accounts", False))
        detail_rows = self.fetch_customer_month_invoice_lines(
            self.current_customer_name,
            item_number,
            month_start,
            combine_accounts=combine_accounts,
        )

        title_month = month_start.strftime("%b %Y")
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Invoice Lines - {item_number} - {title_month}")
        dialog.resize(980, 620)
        layout = QVBoxLayout(dialog)

        summary = QLabel(
            f"Customer: {self.current_customer_name}\n"
            f"Item: {item_number}\n"
            f"Month: {title_month}"
            + ("\n(Combined state accounts)" if combine_accounts else "")
        )
        summary.setWordWrap(True)
        layout.addWidget(summary)

        detail_table = QTableWidget(dialog)
        detail_table.setColumnCount(6)
        detail_table.setHorizontalHeaderLabels([
            "Sale Date",
            "Customer",
            "Description / Invoice Line",
            "Qty",
            "Price",
            "Extended",
        ])
        detail_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        detail_table.setSelectionMode(QAbstractItemView.SingleSelection)
        detail_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        detail_table.verticalHeader().setVisible(False)
        layout.addWidget(detail_table, 1)

        total_qty = 0.0
        total_value = 0.0
        detail_table.setRowCount(len(detail_rows))
        for row_index, line in enumerate(detail_rows):
            sale_date_text = self.format_short_date(line.get("sale_date"))
            customer_text = (line.get("customer_name") or "").strip()
            description_text = (line.get("description") or "").strip()
            qty_value = self.parse_float(line.get("quantity", 0))
            price_value = self.parse_float(line.get("price", 0))
            extended_value = self.parse_float(line.get("extended", qty_value * price_value))

            total_qty += qty_value
            total_value += extended_value

            values = [
                sale_date_text,
                customer_text,
                description_text,
                self.format_value(qty_value),
                self.format_price(price_value),
                self.format_price(extended_value),
            ]
            for col_index, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col_index in (3, 4, 5):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                detail_table.setItem(row_index, col_index, item)

        header = detail_table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        detail_table.resizeRowsToContents()

        footer = QLabel(
            f"Lines: {len(detail_rows)}    "
            f"Qty Total: {self.format_value(total_qty)}    "
            f"Value Total: {self.format_price(total_value)}"
        )
        footer.setWordWrap(True)
        layout.addWidget(footer)

        button_row = QHBoxLayout()
        button_row.addStretch(1)
        close_button = QPushButton("Close", dialog)
        close_button.clicked.connect(dialog.accept)
        button_row.addWidget(close_button)
        layout.addLayout(button_row)

        dialog.exec()

    def open_item_summary_from_customer(self, item_number):
        item_page = getattr(self.ui, "itemSummary_page", None)
        stacked_widget = getattr(self.ui, "stackedWidget", None)
        enter_item = getattr(self.ui, "enterItem", None)

        if enter_item is None or stacked_widget is None or item_page is None:
            return

        start_qdate = self.customer_start_picker.date() if self.customer_start_picker else QDate.currentDate()
        end_qdate = self.customer_end_picker.date() if self.customer_end_picker else start_qdate

        blockers = []
        if self.item_start_picker is not None:
            blockers.append(QSignalBlocker(self.item_start_picker))
        if self.item_end_picker is not None:
            blockers.append(QSignalBlocker(self.item_end_picker))

        try:
            if self.item_start_picker is not None:
                self.item_start_picker.setDate(start_qdate)
            if self.item_end_picker is not None:
                self.item_end_picker.setDate(end_qdate)
        finally:
            del blockers

        enter_item.setText(item_number)
        stacked_widget.setCurrentWidget(item_page)
        self.load_item_summary()

    def handle_customer_purchase_double_click(self, index):
        if not index.isValid():
            return

        model = getattr(self.ui, "customerPurchase_table", None)
        if model is None:
            return

        customer_index = index.siblingAtColumn(0)
        customer_name = customer_index.data(Qt.DisplayRole) if customer_index.isValid() else ""
        customer_name = (customer_name or "").strip()
        if not customer_name:
            return

        self.open_customer_summary_from_item(customer_name)

    def open_customer_summary_from_item(self, customer_name):
        customer_page = getattr(self.ui, "customerSummary_page", None)
        stacked_widget = getattr(self.ui, "stackedWidget", None)
        customer_edit = getattr(self.ui, "customerEdit", None)

        if customer_edit is None or stacked_widget is None or customer_page is None:
            return

        valid_customer = self.find_customer_name(customer_name)
        if not valid_customer:
            valid_customer = customer_name

        start_qdate = self.item_start_picker.date() if self.item_start_picker else QDate.currentDate()
        end_qdate = self.item_end_picker.date() if self.item_end_picker else start_qdate

        blockers = []
        if self.customer_start_picker is not None:
            blockers.append(QSignalBlocker(self.customer_start_picker))
        if self.customer_end_picker is not None:
            blockers.append(QSignalBlocker(self.customer_end_picker))

        try:
            if self.customer_start_picker is not None:
                self.customer_start_picker.setDate(start_qdate)
            if self.customer_end_picker is not None:
                self.customer_end_picker.setDate(end_qdate)
        finally:
            del blockers

        customer_edit.setText(valid_customer)
        stacked_widget.setCurrentWidget(customer_page)
        self.search_customer_sales()

    # -----------------------------
    # Item page
    # -----------------------------
    def rerun_item_if_ready(self, *_args):
        if hasattr(self.ui, "enterItem") and (self.ui.enterItem.text() or "").strip():
            valid_item = self.find_item_number(self.ui.enterItem.text())
            if valid_item:
                self.load_item_summary()


    def load_item_summary(self):
        typed_item = self.ui.enterItem.text() if hasattr(self.ui, "enterItem") else ""
        valid_item = self.find_item_number(typed_item)

        if not valid_item:
            QMessageBox.warning(self, "Invalid item", "Please choose a valid item number from the list.")
            self.clear_item_summary_fields()
            return

        start_date = self.month_start_from_picker(self.item_start_picker or self.item_end_picker)
        end_date = self.month_end_from_picker(self.item_end_picker or self.item_start_picker)
        if start_date > end_date:
            QMessageBox.warning(self, "Invalid date range", "Start month cannot be after end month.")
            return

        months = self.month_list_between(start_date, end_date)
        months_count = max(1, len(months))
        monthly_qty = {m: 0.0 for m in months}
        sales_item_numbers = self.get_thread_sales_item_numbers(valid_item)
        sales_where_clause, sales_params = self.build_sales_item_filter_clause(sales_item_numbers)

        sales_rows = self.db_all(
            f"""
            SELECT
                COALESCE(NULLIF(TRIM(month_key), ''), STRFTIME('%Y-%m', DATE(sale_date))) AS month_key,
                SUM(COALESCE(quantity, 0)) AS total_qty,
                SUM(COALESCE(extended, COALESCE(quantity, 0) * COALESCE(price, 0))) AS total_value
            FROM sales
            WHERE {sales_where_clause}
              AND DATE(sale_date) BETWEEN ? AND ?
            GROUP BY COALESCE(NULLIF(TRIM(month_key), ''), STRFTIME('%Y-%m', DATE(sale_date)))
            ORDER BY month_key
            """,
            [*sales_params, start_date.isoformat(), end_date.isoformat()],
        )

        total_qty = 0.0
        total_value = 0.0
        for row in sales_rows:
            month_start = self.parse_month_key(row["month_key"])
            if month_start is None or month_start not in monthly_qty:
                continue
            qty = self.parse_float(row["total_qty"])
            value = self.parse_float(row["total_value"])
            monthly_qty[month_start] = qty
            total_qty += qty
            total_value += value

        avg_monthly_qty = total_qty / months_count if months_count else 0.0
        lead_weeks = self.get_lead_time_weeks()
        lead_days = self.get_lead_time_days()
        suggested_min = avg_monthly_qty * (lead_days / 30.4375) if months_count else 0.0

        item_row = self.row_to_dict(
            self.db_one(
                "SELECT * FROM items WHERE UPPER(TRIM(item_number)) = UPPER(TRIM(?)) LIMIT 1",
                (valid_item,),
            )
        )
        stock_row = self.row_to_dict(
            self.db_one(
                "SELECT * FROM stock WHERE UPPER(TRIM(item_number)) = UPPER(TRIM(?)) LIMIT 1",
                (valid_item,),
            )
        )

        on_hand = self.parse_float(self.get_first(stock_row, "on_hand", "On Hand", default=0))
        committed = self.parse_float(self.get_first(stock_row, "committed", "Committed", default=0))
        on_order = self.parse_float(self.get_first(stock_row, "on_order", "On Order", default=0))
        available = self.parse_float(self.get_first(stock_row, "available", "Available", default=0))
        carton_size = self.parse_float(self.get_first(item_row, "carton", "CARTON", default=0))
        pallet_size = self.parse_float(self.get_first(item_row, "pallet", "PALLET", default=0))

        inbound_context = self.get_item_inbound_context(valid_item, lead_days)
        suggested_result = self.calculate_suggested_order(
            avg_monthly_qty, on_hand, inbound_context, lead_days, carton_size, pallet_size
        )
        at_risk_result = self.calculate_at_risk(
            avg_monthly_qty, on_hand, inbound_context, lead_days, carton_size, pallet_size
        )
        trend_result = self.calculate_trending_adjustment(
            months, monthly_qty, lead_days, carton_size, pallet_size
        )
        arrival_date = inbound_context.get("horizon_date")
        seasonal_result = self.calculate_seasonal_adjustment(
            sales_item_numbers, avg_monthly_qty, arrival_date, lead_days, carton_size, pallet_size
        )
        adjusted_order = max(
            0.0,
            self.parse_float(suggested_result["rounded"])
            + self.parse_float(trend_result["rounded"])
            + self.parse_float(seasonal_result["rounded"])
        )
        adjusted_order = self.round_order_quantity(adjusted_order, carton_size, pallet_size)

        self.current_item_number = valid_item
        self.set_label_text("itemNumber_box", valid_item)
        self.set_label_text("itemName_box", self.get_first(item_row, "item_name", "Item Name", "description", "Description"))
        self.set_label_text("itemGroup_box", self.get_first(item_row, "Custom List 1", default=""))
        self.set_label_text("rollSpool_box", self.safe_text(self.get_first(item_row, "roll", "ROLL")))
        self.set_label_text("mtUnit_box", self.safe_text(self.get_first(item_row, "per_roll", "PER ROLL")))
        self.set_label_text("box_box", self.safe_text(self.get_first(item_row, "carton", "CARTON")))
        self.set_label_text("palletCarton_box", self.safe_text(self.get_first(item_row, "pallet", "PALLET")))

        self.set_numeric_box("totalQtySold_box", total_qty)
        self.set_numeric_box("avrMonthlySales_box", avg_monthly_qty)
        self.set_numeric_box("suggestedMin_box", suggested_min)
        self.set_numeric_box("stockOnHand_box", on_hand)
        self.set_numeric_box("stockCommited_box", committed)
        self.set_numeric_box("stockOnOrder_box", on_order)
        self.set_numeric_box("stockAvailable_box", available)

        self.set_numeric_box("suggestedOrder_box", suggested_result["rounded"])
        self.set_numeric_box("atRisk_box", at_risk_result["rounded"])
        self.set_signed_box("trendingOrder_box", trend_result["rounded"])
        self.set_signed_box("seasonalOrder_box", seasonal_result["rounded"])
        self.set_numeric_box("adjustedOrder_box", adjusted_order)

        trend_tooltip = (
            f"Trend compares {trend_result['recent_count']} recent completed month(s) "
            f"vs {trend_result['previous_count']} previous month(s)."
        )
        seasonal_tooltip = (
            (
                f"Seasonal adjustment for arrival month {arrival_date.strftime('%b %Y')}: "
                f"September = small lift, October = medium lift, November = full lift, January = unwind. "
                f"December factor {seasonal_result['december_factor']:.2f}x, "
                f"prebuild weight {seasonal_result['prebuild_weight']:.2f}, "
                f"January unwind {seasonal_result['january_unwind_weight']:.2f}, "
                f"effective factor {seasonal_result['factor']:.2f}x."
                if seasonal_result.get('active_window') else
                f"Seasonal adjustment is 0 for {arrival_date.strftime('%b %Y')} because only Sep/Oct/Nov prebuild and Jan unwind are active."
            )
            if arrival_date else
            "Seasonal adjustment unavailable."
        )
        self.set_widget_warning_state("trendingOrder_box", False, trend_tooltip)
        self.set_widget_warning_state("seasonalOrder_box", False, seasonal_tooltip)
        self.set_widget_warning_state(
            "adjustedOrder_box",
            False,
            "Adjusted Order = Suggested Order + Trending Adjustment + Seasonal Adjustment",
        )
        self.set_widget_warning_state(
            "suggestedOrder_box",
            False,
            (
                f"Lead time {lead_weeks} week(s). Hard inbound counted through "
                f"{self.format_display_date(inbound_context.get('horizon_date'))}. "
                f"To-order qty is display only until it is assigned to a real shipment or container."
            ),
        )
        self.set_widget_warning_state(
            "atRisk_box",
            False,
            (
                f"At-risk check runs to {self.format_display_date(at_risk_result.get('cutoff_date'))}. "
                f"Supply arriving on that cutoff date is not counted as protection before the gap."
            ),
        )

        self.draw_monthly_line_chart(
            self.item_chart_view,
            months,
            [monthly_qty[m] for m in months],
            f"{valid_item} monthly units",
        )

        self.populate_customer_purchase_table(self.fetch_item_customer_summary(sales_item_numbers, start_date, end_date))
        self._last_item_total_value = total_value
        self.refresh_item_summary_context_boxes(inbound_context)

    def fetch_item_customer_summary(self, item_numbers, start_date, end_date):
        item_list = item_numbers if isinstance(item_numbers, (list, tuple, set)) else [item_numbers]
        where_clause, params = self.build_sales_item_filter_clause(item_list)
        months = self.month_list_between(start_date, end_date)
        month_rows = self.db_all(
            f"""
            SELECT
                TRIM(customer_name) AS customer_name,
                COALESCE(NULLIF(TRIM(month_key), ''), STRFTIME('%Y-%m', DATE(sale_date))) AS month_key,
                SUM(COALESCE(quantity, 0)) AS total_qty
            FROM sales
            WHERE {where_clause}
              AND DATE(sale_date) BETWEEN ? AND ?
              AND TRIM(COALESCE(customer_name, '')) <> ''
            GROUP BY TRIM(customer_name), COALESCE(NULLIF(TRIM(month_key), ''), STRFTIME('%Y-%m', DATE(sale_date)))
            ORDER BY customer_name COLLATE NOCASE, month_key
            """,
            [*params, start_date.isoformat(), end_date.isoformat()],
        )

        latest_price_rows = self.db_all(
            f"""
            SELECT
                TRIM(customer_name) AS customer_name,
                sale_date,
                COALESCE(price, 0) AS price
            FROM sales
            WHERE {where_clause}
              AND TRIM(COALESCE(customer_name, '')) <> ''
            ORDER BY customer_name COLLATE NOCASE,
                     DATE(sale_date) DESC,
                     sale_date DESC
            """,
            params,
        )

        pivot = {}
        for row in month_rows:
            customer_name = (row["customer_name"] or "").strip()
            if not customer_name:
                continue
            month_start = self.parse_month_key(row["month_key"])
            if month_start is None or month_start not in months:
                continue
            qty = self.parse_float(row["total_qty"])
            if customer_name not in pivot:
                pivot[customer_name] = {
                    "months": {month: 0.0 for month in months},
                    "last_price": None,
                    "last_sale_date": None,
                }
            pivot[customer_name]["months"][month_start] += qty

        latest_sale_info_by_customer = {}
        for row in latest_price_rows:
            customer_name = (row["customer_name"] or "").strip()
            if not customer_name or customer_name in latest_sale_info_by_customer:
                continue
            last_sale_date = self.parse_date_value(row["sale_date"])
            latest_sale_info_by_customer[customer_name] = {
                "last_price": self.parse_float(row["price"]),
                "last_sale_date": last_sale_date,
            }

        for customer_name, info in latest_sale_info_by_customer.items():
            if customer_name in pivot:
                pivot[customer_name]["last_price"] = info.get("last_price")
                pivot[customer_name]["last_sale_date"] = info.get("last_sale_date")

        customer_rows = []
        for customer_name, data in pivot.items():
            month_map = data.get("months", {}) or {}
            total_qty = sum(month_map.values())
            customer_rows.append({
                "customer_name": customer_name,
                "last_sale_date": data.get("last_sale_date"),
                "last_price": data.get("last_price"),
                "months": month_map,
                "total_qty": total_qty,
            })

        customer_rows.sort(key=lambda row: (-row["total_qty"], row["customer_name"].lower()))
        return {"months": months, "rows": customer_rows}

    def populate_customer_purchase_table(self, summary):
        table = getattr(self.ui, "customerPurchase_table", None)
        if table is None:
            return

        months = []
        rows = []
        if isinstance(summary, dict):
            months = summary.get("months") or []
            rows = summary.get("rows") or []

        headers = ["Customer", "Last Sale Date", "Last Price"] + [month.strftime("%b %Y") for month in months] + ["Total"]
        model = QStandardItemModel(0, len(headers), self)
        model.setHorizontalHeaderLabels(headers)

        for row in rows:
            last_sale_date = row.get("last_sale_date")
            if isinstance(last_sale_date, datetime):
                last_sale_date = last_sale_date.date()
            last_sale_date_text = last_sale_date.strftime("%d/%m/%Y") if isinstance(last_sale_date, date) else ""

            values = [row.get("customer_name", ""), last_sale_date_text]
            values.append(self.format_price(row.get("last_price")) if row.get("last_price") is not None else "")
            month_map = row.get("months", {}) or {}
            for month in months:
                values.append(self.format_value(self.parse_float(month_map.get(month, 0.0))))
            values.append(self.format_value(self.parse_float(row.get("total_qty", 0.0))))

            model_row = []
            for i, value in enumerate(values):
                item = QStandardItem(str(value))
                item.setEditable(False)
                if i > 0:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                model_row.append(item)
            model.appendRow(model_row)

        table.setModel(model)
        header = table.horizontalHeader()
        header.setStretchLastSection(False)
        for col in range(len(headers)):
            header.setSectionResizeMode(col, QHeaderView.Interactive)
            table.resizeColumnToContents(col)
        if len(headers) > 0:
            table.setColumnWidth(0, max(table.columnWidth(0), 220))
        if len(headers) > 1:
            table.setColumnWidth(1, max(table.columnWidth(1), 110))
        if len(headers) > 2:
            table.setColumnWidth(2, max(table.columnWidth(2), 95))
        table.resizeRowsToContents()


    def create_order_analysis_progress_dialog(self, total_items, supplier_name):
        total_steps = max(1, int(total_items or 0))
        dialog = QProgressDialog(
            f"Building order analysis for {supplier_name}...",
            "",
            0,
            total_steps,
            self,
        )
        dialog.setWindowTitle("Order Analysis")
        dialog.setWindowModality(Qt.WindowModal)
        dialog.setMinimumDuration(0)
        dialog.setAutoClose(True)
        dialog.setAutoReset(True)
        dialog.setCancelButton(None)
        dialog.setValue(0)
        QApplication.processEvents()
        return dialog

    def update_order_analysis_progress(self, dialog, value, total_items, supplier_name):
        if dialog is None:
            return
        total_steps = max(1, int(total_items or 0))
        current = min(max(0, int(value or 0)), total_steps)
        dialog.setMaximum(total_steps)
        dialog.setLabelText(
            f"Building order analysis for {supplier_name}... {current}/{total_steps}"
        )
        dialog.setValue(current)
        QApplication.processEvents()


    def load_order_analysis(self, show_warning=True):
        supplier_edit = self.get_order_analysis_supplier_edit()
        table = self.get_order_analysis_table()
        if supplier_edit is None or table is None:
            return

        typed_supplier = supplier_edit.text().strip()
        supplier_name = self.find_supplier_name(typed_supplier)
        if not supplier_name:
            if show_warning:
                QMessageBox.warning(self, "Invalid supplier", "Please enter a valid supplier name or a unique supplier prefix.")
            return

        supplier_edit.setText(supplier_name)
        self.current_order_analysis_supplier = supplier_name

        start_date = self.month_start_from_picker(self.item_start_picker or self.item_end_picker)
        end_date = self.month_end_from_picker(self.item_end_picker or self.item_start_picker)
        if start_date > end_date:
            if show_warning:
                QMessageBox.warning(self, "Invalid date range", "Start month cannot be after end month.")
            return

        supplier_expr = self.build_items_supplier_expression(alias=False)
        supplier_match_clause, supplier_columns = self.build_items_supplier_match_clause()
        if not supplier_expr or not supplier_match_clause:
            table.setRowCount(0)
            return

        item_rows = self.db_all(
            f"""
            SELECT
                TRIM(item_number) AS item_number,
                COALESCE(NULLIF(TRIM(item_name), ''), NULLIF(TRIM(description), ''), '') AS item_name,
                COALESCE({supplier_expr}, '') AS supplier_name,
                COALESCE(carton, 0) AS carton,
                COALESCE(pallet, 0) AS pallet
            FROM items
            WHERE TRIM(COALESCE(item_number, '')) <> ''
              AND ({supplier_match_clause})
            ORDER BY item_number COLLATE NOCASE
            """,
            tuple([supplier_name] * len(supplier_columns)),
        )

        months = self.month_list_between(start_date, end_date)
        months_count = max(1, len(months))
        lead_days = self.get_lead_time_days()
        rows_to_show = []
        progress = self.create_order_analysis_progress_dialog(len(item_rows), supplier_name)

        try:
            for index, item_row in enumerate(item_rows, start=1):
                item_number = (item_row["item_number"] or "").strip()
                if not item_number:
                    self.update_order_analysis_progress(progress, index, len(item_rows), supplier_name)
                    continue

                monthly_qty = self.fetch_item_monthly_qty(item_number, start_date, end_date)
                total_qty = sum(self.parse_float(monthly_qty.get(month, 0.0)) for month in months)
                avg_monthly_qty = total_qty / months_count if months_count else 0.0

                stock_row = self.row_to_dict(
                    self.db_one(
                        "SELECT * FROM stock WHERE UPPER(TRIM(item_number)) = UPPER(TRIM(?)) LIMIT 1",
                        (item_number,),
                    )
                )
                on_hand = self.parse_float(self.get_first(stock_row, "on_hand", "On Hand", default=0))
                stock_on_order = self.parse_float(self.get_first(stock_row, "on_order", "On Order", default=0))
                carton_size = self.parse_float(item_row["carton"])
                pallet_size = self.parse_float(item_row["pallet"])

                inbound_context = self.get_item_inbound_context(item_number, lead_days)
                suggested_result = self.calculate_suggested_order(
                    avg_monthly_qty, on_hand, inbound_context, lead_days, carton_size, pallet_size
                )
                at_risk_result = self.calculate_at_risk(
                    avg_monthly_qty, on_hand, inbound_context, lead_days, carton_size, pallet_size
                )

                suggested_qty = self.parse_float(suggested_result["rounded"])
                at_risk_qty = self.parse_float(at_risk_result["rounded"])
                if suggested_qty > 0 or at_risk_qty > 0:
                    rows_to_show.append({
                        "item_number": item_number,
                        "item_name": (item_row["item_name"] or "").strip(),
                        "sales_for_period": total_qty,
                        "avg_monthly_sales": avg_monthly_qty,
                        "soh": on_hand,
                        "stock_on_order": stock_on_order,
                        "on_order_form": self.parse_float(inbound_context.get("order_form", {}).get("qty", 0)),
                        "on_next_container": self.parse_float(inbound_context.get("next_container", {}).get("qty", 0)),
                        "shipped_container": self.parse_float(inbound_context.get("shipped", {}).get("qty", 0)),
                        "suggested_order": suggested_qty,
                        "at_risk": at_risk_qty,
                    })

                self.update_order_analysis_progress(progress, index, len(item_rows), supplier_name)
        finally:
            self.update_order_analysis_progress(progress, len(item_rows), len(item_rows), supplier_name)
            progress.close()

        rows_to_show.sort(
            key=lambda row: (
                -self.parse_float(row["suggested_order"]),
                -self.parse_float(row["at_risk"]),
                row["item_number"].lower(),
            )
        )

        self.populate_order_analysis_table(rows_to_show)

    def populate_order_analysis_table(self, rows):
        table = self.get_order_analysis_table()
        if table is None:
            return

        table.setRowCount(0)
        numeric_columns = {
            self.order_analysis_columns["sales_for_period"],
            self.order_analysis_columns["avg_monthly_sales"],
            self.order_analysis_columns["soh"],
            self.order_analysis_columns["stock_on_order"],
            self.order_analysis_columns["on_order_form"],
            self.order_analysis_columns["on_next_container"],
            self.order_analysis_columns["shipped_container"],
            self.order_analysis_columns["suggested_order"],
            self.order_analysis_columns["at_risk"],
        }

        for row_data in rows:
            row = table.rowCount()
            table.insertRow(row)
            values = [
                row_data.get("item_number", ""),
                row_data.get("item_name", ""),
                self.format_value(row_data.get("sales_for_period", 0)),
                self.format_value(row_data.get("avg_monthly_sales", 0)),
                self.format_value(row_data.get("soh", 0)),
                self.format_value(row_data.get("stock_on_order", 0)),
                self.format_value(row_data.get("on_order_form", 0)),
                self.format_value(row_data.get("on_next_container", 0)),
                self.format_value(row_data.get("shipped_container", 0)),
                self.format_value(row_data.get("suggested_order", 0)),
                self.format_value(row_data.get("at_risk", 0)),
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                if col in numeric_columns:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                table.setItem(row, col, item)

        table.resizeRowsToContents()

    def handle_order_analysis_table_double_click(self, row, column):
        table = self.get_order_analysis_table()
        if table is None or row < 0:
            return

        item_cell = table.item(row, self.order_analysis_columns["item_number"])
        item_number = item_cell.text().strip() if item_cell is not None and item_cell.text() else ""
        if not item_number:
            return

        if column == self.order_analysis_columns["item_number"]:
            self.open_item_summary_from_order_analysis(item_number)
        elif column == self.order_analysis_columns["on_order_form"]:
            self.open_to_order_sheet_for_item(item_number)

    def open_item_summary_from_order_analysis(self, item_number):
        item_page = getattr(self.ui, "itemSummary_page", None)
        stacked_widget = getattr(self.ui, "stackedWidget", None)
        enter_item = getattr(self.ui, "enterItem", None)
        if item_page is None or stacked_widget is None or enter_item is None:
            return

        enter_item.setText(item_number)
        stacked_widget.setCurrentWidget(item_page)
        self.load_item_summary()

    def open_to_order_sheet_for_item(self, item_number):
        stacked_widget = getattr(self.ui, "stackedWidget", None)
        order_page = getattr(self.ui, "toOrderSheet_page", None)
        order_table = getattr(self.ui, "order_table", None)
        if stacked_widget is None or order_page is None or order_table is None:
            return

        stacked_widget.setCurrentWidget(order_page)
        target = (item_number or "").strip().upper()
        for row in range(order_table.rowCount()):
            item_cell = order_table.item(row, 0)
            item_text = item_cell.text().strip().upper() if item_cell is not None and item_cell.text() else ""
            if item_text != target:
                continue
            order_table.selectRow(row)
            order_table.setCurrentCell(row, 0)
            if item_cell is not None:
                order_table.scrollToItem(item_cell, QAbstractItemView.PositionAtCenter)
            return

        QMessageBox.information(self, "To Order Sheet", f"{item_number} is not currently on the to-order sheet.")

    def find_saba_customer_name(self, typed_text):
        typed = (typed_text or "").strip()
        if not typed or typed.lower() == "all":
            return "All"

        typed_lower = typed.lower()
        exact_matches = [name for name in self.customer_names if name.lower() == typed_lower]
        if exact_matches:
            return exact_matches[0]

        prefix_matches = [name for name in self.customer_names if name.lower().startswith(typed_lower)]
        if len(prefix_matches) == 1:
            return prefix_matches[0]
        return None

    def handle_saba_all_customers_changed(self, state):
        edit = self.get_saba_customer_edit()
        checked = bool(state)
        if edit is not None:
            if checked:
                edit.setText("All")
            elif (edit.text() or "").strip().lower() == "all":
                edit.clear()
        self.load_saba_review(show_warning=False)

    def parse_saba_pack_info(self, item_number):
        text = re.sub(r"\s+", " ", (item_number or "").upper()).strip()
        color = None
        pack_key = "OTHER"
        pack_label = "Other"
        size_kg = None

        if "IBC" in text:
            pack_key = "IBC"
            pack_label = "IBC (1100 kg)"
            size_kg = 1100
            match = re.search(r"\bIBC([BN])\b", text)
            if match:
                color = "Blue" if match.group(1) == "B" else "Natural"
        elif "BIB" in text:
            pack_key = "BIB"
            pack_label = "Bag in Box (550 kg)"
            size_kg = 550
            match = re.search(r"\bBIB([BN])\b", text)
            if match:
                color = "Blue" if match.group(1) == "B" else "Natural"
        elif re.search(r"\b21([BN])?\b", text):
            pack_key = "KEG"
            pack_label = "Keg (21 kg)"
            size_kg = 21
            match = re.search(r"\b21([BN])\b", text)
            if match:
                color = "Blue" if match.group(1) == "B" else "Natural"

        if color:
            pack_label = f"{pack_label} - {color}"

        return {
            "pack_key": pack_key,
            "pack_label": pack_label,
            "size_kg": size_kg,
            "color": color,
        }

    def get_saba_item_rows(self):
        normalized_item_expr = "UPPER(REPLACE(TRIM(COALESCE(item_number, '')), ' ', ''))"

        # Narrow the review to the SA3392 family the user explicitly called out.
        # This avoids false positives from broader SA*/SABA matching and is tolerant
        # of spacing differences such as `SA3392 21B` vs `SA339221B`.
        where_parts = [
            "TRIM(COALESCE(item_number, '')) <> ''",
            f"{normalized_item_expr} LIKE 'SA3392%'",
            "("
            f"{normalized_item_expr} LIKE 'SA339221%' OR "
            f"{normalized_item_expr} LIKE 'SA3392BIB%' OR "
            f"{normalized_item_expr} LIKE 'SA3392IBC%'"
            ")",
        ]

        return self.db_all(
            f"""
            SELECT
                TRIM(item_number) AS item_number,
                COALESCE(NULLIF(TRIM(item_name), ''), NULLIF(TRIM(description), ''), TRIM(item_number)) AS item_name
            FROM items
            WHERE {' AND '.join(where_parts)}
            ORDER BY item_number COLLATE NOCASE
            """
        )

    def compute_saba_review_rows(self, selected_customer=None):
        item_rows = self.get_saba_item_rows()
        item_map = {}
        for row in item_rows:
            item_number = (row["item_number"] or "").strip()
            if not item_number:
                continue
            normalized_item_number = re.sub(r"\s+", "", item_number.upper())
            item_map[normalized_item_number] = {
                "item_number": item_number,
                "item_name": (row["item_name"] or item_number).strip(),
            }

        if not item_map:
            return []

        item_numbers = sorted(item_map.keys())
        item_clause, item_params = self.sql_in_clause(item_numbers)
        sql = f"""
            SELECT
                TRIM(customer_name) AS customer_name,
                TRIM(item_number) AS item_number,
                DATE(sale_date) AS sale_date
            FROM sales
            WHERE TRIM(COALESCE(customer_name, '')) <> ''
              AND TRIM(COALESCE(item_number, '')) <> ''
              AND UPPER(REPLACE(TRIM(item_number), ' ', '')) IN {item_clause}
              AND sale_date IS NOT NULL
        """
        params = list(item_params)
        if selected_customer and selected_customer != "All":
            sql += " AND UPPER(TRIM(customer_name)) = UPPER(TRIM(?))"
            params.append(selected_customer)
        sql += " ORDER BY customer_name COLLATE NOCASE, item_number COLLATE NOCASE, DATE(sale_date)"

        sales_rows = self.db_all(sql, tuple(params))
        grouped_dates = {}
        for row in sales_rows:
            customer_name = (row["customer_name"] or "").strip()
            item_number = (row["item_number"] or "").strip()
            sale_date = self.parse_date_value(row["sale_date"])
            if not customer_name or not item_number or sale_date is None:
                continue
            normalized_item_number = re.sub(r"\s+", "", item_number.upper())
            key = (customer_name, normalized_item_number)
            grouped_dates.setdefault(key, set()).add(sale_date)

        today = self.current_business_date()
        result_rows = []
        for (customer_name, item_key), sale_dates in grouped_dates.items():
            ordered_dates = sorted(sale_dates)
            meta = item_map.get(item_key, {"item_number": item_key, "item_name": item_key})
            pack_info = self.parse_saba_pack_info(meta["item_number"])
            last_purchase = ordered_dates[-1] if ordered_dates else None
            weeks_since_last = 0.0
            if last_purchase is not None:
                weeks_since_last = max(0.0, (today - last_purchase).days / 7.0)

            avg_weeks = None
            purchase_dates_count = len(ordered_dates)
            if purchase_dates_count >= 2:
                gaps = [
                    (ordered_dates[idx] - ordered_dates[idx - 1]).days / 7.0
                    for idx in range(1, purchase_dates_count)
                ]
                if gaps:
                    avg_weeks = sum(gaps) / len(gaps)

            if avg_weeks is None:
                status = "Grey"
                status_rank = 0
            else:
                green_at = avg_weeks - 3
                yellow_at = avg_weeks - 2
                red_at = avg_weeks - 1
                if weeks_since_last >= red_at:
                    status = "Red"
                    status_rank = 4
                elif weeks_since_last >= yellow_at:
                    status = "Yellow"
                    status_rank = 3
                elif weeks_since_last >= green_at:
                    status = "Green"
                    status_rank = 2
                else:
                    status = "OK"
                    status_rank = 1

            result_rows.append({
                "customer_name": customer_name,
                "item_number": meta["item_number"],
                "item_name": meta["item_name"],
                "pack_key": pack_info["pack_key"],
                "pack_label": pack_info["pack_label"],
                "avg_weeks": avg_weeks,
                "weeks_since_last": weeks_since_last,
                "last_purchase_date": last_purchase,
                "purchase_dates_count": purchase_dates_count,
                "status": status,
                "status_rank": status_rank,
            })

        result_rows.sort(
            key=lambda row: (
                -int(row.get("status_rank", 0)),
                -self.parse_float(row.get("weeks_since_last", 0)),
                (row.get("customer_name") or "").lower(),
                (row.get("item_number") or "").lower(),
            )
        )
        return result_rows

    def get_saba_status_brush(self, status):
        status = (status or "").strip().lower()
        if status == "red":
            return QBrush(QColor(255, 199, 206))
        if status == "yellow":
            return QBrush(QColor(255, 235, 156))
        if status == "green":
            return QBrush(QColor(198, 239, 206))
        if status == "grey":
            return QBrush(QColor(217, 217, 217))
        return None

    def get_saba_status_foreground_brush(self, status):
        brush = self.get_saba_status_brush(status)
        if brush is None:
            return None
        return QBrush(QColor(17, 17, 17))

    def populate_saba_review_table(self, rows):
        table = self.get_saba_table()
        if table is None:
            return

        headers = [
            "Customer",
            "Item Number",
            "Item Name",
            "Pack",
            "Avg Weeks Between Purchases",
            "Weeks Since Last Purchase",
            "Last Purchase Date",
            "Purchase Dates",
            "Status",
        ]
        model = QStandardItemModel(0, len(headers), self)
        model.setHorizontalHeaderLabels(headers)

        for row in rows:
            values = [
                row.get("customer_name", ""),
                row.get("item_number", ""),
                row.get("item_name", ""),
                row.get("pack_label", ""),
                "Insufficient history" if row.get("avg_weeks") is None else f"{self.parse_float(row.get('avg_weeks')):.1f}",
                f"{self.parse_float(row.get('weeks_since_last')):.1f}",
                self.format_display_date(row.get("last_purchase_date")),
                self.format_value(row.get("purchase_dates_count", 0)),
                row.get("status", ""),
            ]
            model_row = []
            brush = self.get_saba_status_brush(row.get("status"))
            for index, value in enumerate(values):
                item = QStandardItem(str(value))
                item.setEditable(False)
                if index >= 4 and index != 8:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if brush is not None:
                    item.setBackground(brush)
                    foreground_brush = self.get_saba_status_foreground_brush(row.get("status"))
                    if foreground_brush is not None:
                        item.setForeground(foreground_brush)
                model_row.append(item)
            model.appendRow(model_row)

        table.setModel(model)
        header = table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        for col in range(3, len(headers)):
            header.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        table.resizeRowsToContents()

    def summarize_saba_pack_rows(self, rows, pack_key):
        pack_rows = [row for row in rows if row.get("pack_key") == pack_key]
        if not pack_rows:
            return "No matching rows", "No matching rows", ""

        pack_rows.sort(
            key=lambda row: (
                -int(row.get("status_rank", 0)),
                -self.parse_float(row.get("weeks_since_last", 0)),
                (row.get("customer_name") or "").lower(),
            )
        )
        top_row = pack_rows[0]
        avg_text = [top_row.get("customer_name", ""), top_row.get("item_number", "")]
        if top_row.get("avg_weeks") is None:
            avg_text.append("Insufficient history")
        else:
            avg_text.append(f"Avg: {self.parse_float(top_row.get('avg_weeks')):.1f} weeks")
        avg_text.append(f"Rows: {len(pack_rows)}")

        since_text = [
            f"Since last: {self.parse_float(top_row.get('weeks_since_last')):.1f} weeks",
            f"Status: {top_row.get('status', '')}",
            f"Last: {self.format_display_date(top_row.get('last_purchase_date'))}",
        ]
        return "\n".join(avg_text), "\n".join(since_text), top_row.get("status", "")

    def update_saba_summary_boxes(self, rows):
        mapping = [
            ("BIB", "bagInBox_textBrowser", "bagInBoxWeeksLast_textBrowser"),
            ("KEG", "keg_textBrowser", "kegWeeksLast_textBrowser"),
            ("IBC", "ibc_textBrowser", "ibcWeeksLast_textBrowser"),
        ]
        for pack_key, avg_name, since_name in mapping:
            avg_text, since_text, status = self.summarize_saba_pack_rows(rows, pack_key)
            avg_widget = getattr(self.ui, avg_name, None)
            since_widget = getattr(self.ui, since_name, None)
            brush = self.get_saba_status_brush(status)
            foreground_brush = self.get_saba_status_foreground_brush(status)
            style = ""
            if brush is not None:
                color = brush.color()
                text_color = foreground_brush.color() if foreground_brush is not None else QColor(17, 17, 17)
                style = (
                    f"background-color: rgb({color.red()}, {color.green()}, {color.blue()});"
                    f"color: rgb({text_color.red()}, {text_color.green()}, {text_color.blue()});"
                )
            if avg_widget is not None:
                avg_widget.setPlainText(avg_text)
                avg_widget.setStyleSheet(style)
            if since_widget is not None:
                since_widget.setPlainText(since_text)
                since_widget.setStyleSheet(style)

    def load_saba_review(self, show_warning=True):
        edit = self.get_saba_customer_edit()
        table = self.get_saba_table()
        if edit is None or table is None:
            return

        if self.saba_show_all_checkbox is not None and self.saba_show_all_checkbox.isChecked():
            selected_customer = "All"
            edit.setText("All")
        else:
            typed = (edit.text() or "").strip()
            selected_customer = self.find_saba_customer_name(typed)
            if selected_customer is None:
                if show_warning:
                    QMessageBox.warning(self, "Invalid customer", "Please enter a valid customer name, a unique customer prefix, or All.")
                return
            if selected_customer != "All":
                edit.setText(selected_customer)

        self.current_saba_customer = selected_customer
        rows = self.compute_saba_review_rows(selected_customer=None if selected_customer == "All" else selected_customer)
        self.populate_saba_review_table(rows)
        self.update_saba_summary_boxes(rows)

    # -----------------------------
    # Charting
    # -----------------------------
    def build_trend_series(self, totals, color=None):
        if not totals or len(totals) < 2:
            return None

        n = len(totals)
        xs = list(range(n))
        sum_x = sum(xs)
        sum_y = sum(totals)
        sum_xy = sum(x * y for x, y in zip(xs, totals))
        sum_x2 = sum(x * x for x in xs)

        denominator = n * sum_x2 - sum_x * sum_x
        if denominator == 0:
            return None

        slope = (n * sum_xy - sum_x * sum_y) / denominator
        intercept = (sum_y - slope * sum_x) / n

        trend_series = QLineSeries()
        trend_series.setName("Trend")
        pen = QPen(color or QColor("#FFB74D"))
        pen.setStyle(Qt.DashLine)
        pen.setWidth(2)
        trend_series.setPen(pen)

        for x in xs:
            trend_series.append(x, intercept + slope * x)

        return trend_series

    def draw_monthly_line_chart(self, chart_view, months, totals, series_name):
        if chart_view is None:
            return

        month_labels = [m.strftime("%b\n%Y") for m in months]

        chart = QChart()
        chart.setTitle("")
        chart.setMargins(QMargins(12, 8, 12, 18))
        chart.legend().setVisible(True)
        chart.setBackgroundVisible(True)
        chart.setBackgroundBrush(self.chart_background_color())

        is_item_chart = chart_view is self.item_chart_view

        if hasattr(self.ui, "radioHighContrast") and self.ui.radioHighContrast.isChecked():
            main_color = QColor("#00FFFF") if is_item_chart else QColor("#00FF00")
            trend_color = QColor("#FF8C00")
            grid_color = QColor(255, 255, 0, 70)
        elif hasattr(self.ui, "radioLight") and self.ui.radioLight.isChecked():
            main_color = QColor("#005BBB") if is_item_chart else QColor("#006400")
            trend_color = QColor("#C45500")
            grid_color = QColor(32, 33, 36, 45)
        else:
            main_color = QColor("#64B5F6") if is_item_chart else QColor("#81C784")
            trend_color = QColor("#FFB74D")
            grid_color = QColor(232, 234, 237, 45)

        series = QLineSeries()
        series.setName(series_name if series_name else "Monthly Units")
        series.setPointsVisible(True)
        series.setPointLabelsVisible(True)
        series.setPointLabelsFormat("@yPoint")

        pen = QPen(main_color)
        pen.setWidth(3 if is_item_chart else 2)
        series.setPen(pen)

        label_font = QFont()
        label_font.setPointSize(9)
        label_font.setBold(True)
        try:
            series.setPointLabelsColor(main_color)
            series.setPointLabelsFont(label_font)
        except Exception:
            pass

        for idx, total in enumerate(totals):
            series.append(idx, float(total))

        chart.addSeries(series)

        trend_series = self.build_trend_series(totals, trend_color)
        if trend_series is not None:
            chart.addSeries(trend_series)

        x_axis = QBarCategoryAxis()
        x_axis.append(month_labels)
        x_axis.setLabelsAngle(0)

        axis_font = QFont()
        axis_font.setPointSize(10)
        axis_font.setBold(True)
        x_axis.setLabelsFont(axis_font)

        y_axis = QValueAxis()
        max_total = max(totals) if totals else 0.0
        y_axis.setMin(0)
        y_axis.setMax(max(1.0, max_total * 1.15))
        y_axis.setLabelFormat("%.0f")
        y_axis.setTitleText("Units")
        y_axis.setLabelsFont(axis_font)

        chart.addAxis(x_axis, Qt.AlignBottom)
        chart.addAxis(y_axis, Qt.AlignLeft)

        series.attachAxis(x_axis)
        series.attachAxis(y_axis)
        if trend_series is not None:
            trend_series.attachAxis(x_axis)
            trend_series.attachAxis(y_axis)

        text_color = self.chart_text_color()
        axis_pen = QPen(text_color)
        for axis in chart.axes():
            axis.setLabelsColor(text_color)
            axis.setLinePen(axis_pen)
            axis.setGridLinePen(QPen(grid_color))

        chart.legend().setLabelColor(text_color)
        chart_view.setChart(chart)


    def get_yu_order_drafts_dir(self):
        base_dir = Path(__file__).resolve().parent / "data" / "yu_order_drafts"
        base_dir.mkdir(parents=True, exist_ok=True)
        return str(base_dir)

    def get_yu_order_temp_dir(self):
        base_dir = Path(__file__).resolve().parent / "data" / "yu_order_temp"
        base_dir.mkdir(parents=True, exist_ok=True)
        return str(base_dir)

    def get_yu_order_output_dir(self):
        base_dir = Path(__file__).resolve().parent / "data" / "yu_order_exports"
        base_dir.mkdir(parents=True, exist_ok=True)
        return str(base_dir)

    def ensure_yu_template_path(self):
        expected_name = "yuchang_order_form_Widget.xlsx"

        candidate_paths = []
        saved_path = str(self.settings.value("yu/template_path", "") or "").strip()
        if saved_path:
            candidate_paths.append(Path(saved_path))

        # Common local fallback locations
        app_dir = Path(__file__).resolve().parent
        candidate_paths.extend([
            app_dir / expected_name,
            app_dir / "data" / expected_name,
            self.base_dir / expected_name,
            self.base_dir / "data" / expected_name,
        ])

        seen = set()
        for candidate in candidate_paths:
            try:
                resolved = str(Path(candidate).resolve())
            except Exception:
                resolved = str(candidate)
            if resolved in seen:
                continue
            seen.add(resolved)
            if Path(candidate).exists():
                final_path = str(Path(candidate))
                self.settings.setValue("yu/template_path", final_path)
                return final_path

        start_dir = str(app_dir)
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            f"Choose {expected_name}",
            start_dir,
            "Excel Files (*.xlsx *.xlsm)"
        )
        if not file_path:
            return ""
        self.settings.setValue("yu/template_path", file_path)
        return file_path

    def open_yu_order_entry_dialog(self, initial_order_number="", initial_lines=None):
        dialog = YUOrderEntryDialog(
            self,
            self,
            initial_order_number=initial_order_number,
            initial_lines=initial_lines or [],
        )
        self.yu_order_entry_dialog = dialog
        dialog.exec()

    def open_yu_order_review_window(self, csv_path):
        template_path = self.ensure_yu_template_path()
        if not template_path:
            QMessageBox.warning(self, "Missing YU template", "Choose the YU workbook template before continuing.")
            return False
        try:
            module = load_yu_review_module(self)
            db_helper = module.SQLHelper()
            window = module.YUOrderReviewWindow(
                db=db_helper,
                prefix="yu_test",
                template_path=template_path,
                order_csv_path=csv_path,
                output_dir=self.get_yu_order_output_dir(),
            )
        except Exception as exc:
            QMessageBox.critical(self, "YU Order Review", f"Could not open YU order review window.\n\n{exc}")
            return False

        self.yu_order_review_db_helpers.append(db_helper)
        self.yu_order_review_windows.append(window)
        try:
            window.destroyed.connect(lambda *_args, w=window, db=db_helper: self._cleanup_yu_order_review_window(w, db))
        except Exception:
            pass
        window.show()
        if window.isMinimized():
            window.showNormal()
        window.raise_()
        window.activateWindow()
        return True

    def _cleanup_yu_order_review_window(self, window, db_helper):
        try:
            if window in self.yu_order_review_windows:
                self.yu_order_review_windows.remove(window)
        except Exception:
            pass
        try:
            if db_helper in self.yu_order_review_db_helpers:
                self.yu_order_review_db_helpers.remove(db_helper)
        except Exception:
            pass
        try:
            db_helper.close()
        except Exception:
            pass


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
