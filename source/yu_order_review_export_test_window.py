
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from collections import defaultdict
from copy import copy, deepcopy
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

try:
    import pyodbc
except Exception:
    pyodbc = None

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter

from PySide6.QtCore import Qt, QSettings, QSize, QUrl
from PySide6.QtGui import QColor, QBrush, QDesktopServices
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QStatusBar,
    QTableWidget,
    QTableWidgetItem,
    QTextBrowser,
    QVBoxLayout,
    QWidget,
)


TABLE_FONT_SIZE_OPTIONS = (8, 9, 10, 11, 12, 14, 16, 18, 20)
TABLE_FONT_SETTINGS_PREFIX = "table_font_sizes"

DEFAULT_PREFIX = "yu_test"
DEFAULT_ORDER_CSV = ""
DEFAULT_TEMPLATE = ""
DEFAULT_OUTPUT_DIR = "yu_exports"


# ------------------------------------------------------------
# Table font helpers copied to match main app behaviour
# ------------------------------------------------------------
def _safe_table_settings_token(table, fallback_scope="table"):
    token = str(table.property("_table_font_settings_token") or "").strip()
    if token:
        return token

    parts = []
    current = table
    while current is not None:
        name = ""
        try:
            name = str(current.objectName() or "").strip()
        except Exception:
            name = ""
        if name and not name.startswith("qt_"):
            parts.append(name)
        try:
            current = current.parentWidget()
        except Exception:
            current = None
    parts.reverse()

    token = "/".join(parts) if parts else fallback_scope
    token = re.sub(r"[^A-Za-z0-9_./-]+", "_", token)
    return token or fallback_scope


def table_font_settings_key(scope_key, table):
    token = _safe_table_settings_token(table, scope_key)
    return f"{TABLE_FONT_SETTINGS_PREFIX}/{scope_key}/{token}"


def stored_table_font_size(settings, scope_key, table):
    if settings is None or table is None:
        return None
    try:
        value = settings.value(table_font_settings_key(scope_key, table), None)
    except Exception:
        return None
    if value in (None, ""):
        return None
    try:
        size = int(value)
    except Exception:
        return None
    return size if size > 0 else None


def current_table_font_size(table):
    if table is None:
        return 10
    try:
        size = int(round(float(table.font().pointSizeF())))
        if size > 0:
            return size
    except Exception:
        pass
    try:
        size = int(table.font().pointSize())
        if size > 0:
            return size
    except Exception:
        pass
    return 10


def sync_table_item_font_sizes(table, size):
    if table is None or not isinstance(table, QTableWidget):
        return
    for row in range(table.rowCount()):
        for column in range(table.columnCount()):
            item = table.item(row, column)
            if item is None:
                continue
            item_font = item.font()
            item_font.setPointSize(size)
            item.setFont(item_font)


def apply_table_font_size(table, size, settings=None, scope_key=None, persist=True):
    if table is None:
        return
    try:
        size = int(size)
    except Exception:
        return
    if size < 6:
        size = 6

    font = table.font()
    font.setPointSize(size)
    table.setFont(font)
    sync_table_item_font_sizes(table, size)

    vertical_header = getattr(table, "verticalHeader", lambda: None)()
    if vertical_header is not None:
        try:
            vertical_header.setDefaultSectionSize(max(24, size * 2 + 8))
        except Exception:
            pass
    try:
        table.resizeRowsToContents()
    except Exception:
        pass
    try:
        table.viewport().update()
    except Exception:
        pass

    if persist and settings is not None and scope_key:
        try:
            settings.setValue(table_font_settings_key(scope_key, table), size)
        except Exception:
            pass


def reset_table_font_size(table, settings=None, scope_key=None):
    if table is None:
        return
    default_size = None
    try:
        default_size = int(table.property("_table_font_default_size") or 0)
    except Exception:
        default_size = None
    if not default_size:
        default_size = 10

    if settings is not None and scope_key:
        try:
            settings.remove(table_font_settings_key(scope_key, table))
        except Exception:
            pass
    apply_table_font_size(table, default_size, settings=settings, scope_key=scope_key, persist=False)


def install_table_font_context_menu(owner, table, settings, scope_key, settings_token=None):
    if table is None:
        return
    if table.property("_table_font_menu_installed"):
        saved_size = stored_table_font_size(settings, scope_key, table)
        if saved_size:
            apply_table_font_size(table, saved_size, settings=settings, scope_key=scope_key, persist=False)
        return

    if settings_token:
        table.setProperty("_table_font_settings_token", settings_token)
    elif not str(table.objectName() or "").strip():
        table.setObjectName(_safe_table_settings_token(table, scope_key))

    table.setProperty("_table_font_menu_installed", True)
    table.setProperty("_table_font_scope_key", scope_key)
    table.setProperty("_table_font_default_size", current_table_font_size(table))
    table.setContextMenuPolicy(Qt.CustomContextMenu)
    table.customContextMenuRequested.connect(
        lambda pos, tbl=table, own=owner, s=settings, scope=scope_key: show_table_font_context_menu(own, tbl, pos, s, scope)
    )

    saved_size = stored_table_font_size(settings, scope_key, table)
    if saved_size:
        apply_table_font_size(table, saved_size, settings=settings, scope_key=scope_key, persist=False)


def show_table_font_context_menu(owner, table, pos, settings, scope_key):
    if table is None:
        return

    menu = QMenu(table)
    font_menu = menu.addMenu("Font Size")
    current_size = current_table_font_size(table)

    for size in TABLE_FONT_SIZE_OPTIONS:
        action = font_menu.addAction(str(size))
        action.setCheckable(True)
        action.setChecked(size == current_size)
        action.triggered.connect(
            lambda _checked=False, tbl=table, selected_size=size, s=settings, scope=scope_key: apply_table_font_size(
                tbl, selected_size, settings=s, scope_key=scope, persist=True
            )
        )

    font_menu.addSeparator()
    reset_action = font_menu.addAction("Reset")
    reset_action.triggered.connect(
        lambda _checked=False, tbl=table, s=settings, scope=scope_key: reset_table_font_size(tbl, settings=s, scope_key=scope)
    )

    viewport = getattr(table, "viewport", lambda: None)()
    if viewport is not None:
        global_pos = viewport.mapToGlobal(pos)
    else:
        global_pos = table.mapToGlobal(pos)
    menu.exec(global_pos)


# ------------------------------------------------------------
# DB
# ------------------------------------------------------------
def get_database_config_candidate_paths(base_dir: Path | None = None) -> list[Path]:
    base_dir = Path(base_dir or Path.cwd())
    app_dir = Path(__file__).resolve().parent

    candidate_paths: list[Path] = []
    env_path = os.environ.get("WINDSOR_WIDGET_CONFIG", "").strip()
    if env_path:
        candidate_paths.append(Path(env_path))

    candidate_paths.extend([
        app_dir / "client_config.json",
        app_dir / "data" / "client_config.json",
        Path(os.environ.get("ProgramData", r"C:\ProgramData")) / "WindsorWidget" / "client_config.json",
        base_dir / "client_config.json",
        base_dir / "data" / "client_config.json",
    ])

    deduped: list[Path] = []
    seen: set[str] = set()
    for path in candidate_paths:
        try:
            resolved = str(path.resolve())
        except Exception:
            resolved = str(path)
        if resolved in seen:
            continue
        seen.add(resolved)
        deduped.append(path)
    return deduped


def load_database_config(base_dir: Path | None = None) -> dict:
    last_error = None
    checked_paths = []
    for path in get_database_config_candidate_paths(base_dir=base_dir):
        checked_paths.append(str(path))
        if not path.exists():
            continue
        for encoding_name in ("utf-8-sig", "utf-8"):
            try:
                data = json.loads(path.read_text(encoding=encoding_name))
                if isinstance(data, dict):
                    data["_loaded_from_path"] = str(path)
                    return data
            except Exception as exc:
                last_error = f"{path}: {exc}"
                continue
    return {"_checked_paths": checked_paths, "_last_error": last_error}


class SQLHelper:
    def __init__(self, base_dir: Path | None = None) -> None:
        if pyodbc is None:
            raise RuntimeError("pyodbc is required for SQL Server mode.")

        db_config = load_database_config(base_dir=base_dir)
        provider = str(db_config.get("provider", "")).strip().lower()
        server = str(db_config.get("server", "")).strip()
        database = str(db_config.get("database", "")).strip()

        if provider != "sqlserver" or not server or not database:
            checked_paths = db_config.get("_checked_paths", [])
            last_error = db_config.get("_last_error", "")
            checked_text = "\n".join(checked_paths) if checked_paths else "(none)"
            extra = f"\n\nLast config parse error:\n{last_error}" if last_error else ""
            raise RuntimeError(
                "A valid SQL Server client_config.json was not found "
                "or does not declare provider='sqlserver'.\n\nChecked paths:\n"
                f"{checked_text}{extra}"
            )

        driver = db_config.get("driver") or "ODBC Driver 18 for SQL Server"
        port = str(db_config.get("port", "")).strip()
        username = str(db_config.get("username", "")).strip()
        password = str(db_config.get("password", "")).strip()
        trusted = bool(db_config.get("trusted_connection", False))
        encrypt = str(db_config.get("encrypt", "no")).strip().lower()
        trust_cert = str(db_config.get("trust_server_certificate", "yes")).strip().lower()
        timeout = int(db_config.get("timeout", 5) or 5)
        server_spec = f"{server},{port}" if port else server

        parts = [
            f"DRIVER={{{driver}}}",
            f"SERVER={server_spec}",
            f"DATABASE={database}",
            f"Encrypt={'yes' if encrypt in {'1', 'true', 'yes'} else 'no'}",
            f"TrustServerCertificate={'yes' if trust_cert in {'1', 'true', 'yes'} else 'no'}",
            f"Connection Timeout={timeout}",
        ]
        if trusted:
            parts.append("Trusted_Connection=yes")
        else:
            parts.append(f"UID={username}")
            parts.append(f"PWD={password}")

        self.loaded_from_path = db_config.get("_loaded_from_path", "")
        self.conn = pyodbc.connect(";".join(parts))

    def close(self):
        self.conn.close()

    def all(self, sql: str, params: tuple | list = ()) -> list[dict[str, Any]]:
        cur = self.conn.cursor()
        cur.execute(sql, params)
        columns = [desc[0] for desc in cur.description] if cur.description else []
        rows = []
        for row in cur.fetchall():
            rows.append(dict(zip(columns, row)))
        return rows

    def one(self, sql: str, params: tuple | list = ()) -> dict[str, Any] | None:
        cur = self.conn.cursor()
        cur.execute(sql, params)
        row = cur.fetchone()
        if row is None:
            return None
        columns = [desc[0] for desc in cur.description] if cur.description else []
        return dict(zip(columns, row))

    def execute(self, sql: str, params: tuple | list = ()) -> None:
        cur = self.conn.cursor()
        cur.execute(sql, params)
        self.conn.commit()

    def scalar(self, sql: str, params: tuple | list = ()) -> Any:
        cur = self.conn.cursor()
        cur.execute(sql, params)
        row = cur.fetchone()
        return row[0] if row else None


# ------------------------------------------------------------
# Order CSV + resolution + export
# ------------------------------------------------------------
@dataclass
class ReviewHit:
    source_row: int
    match_type: str
    confidence: float | None
    note: str


@dataclass
class OrderResolveResult:
    item_number: str
    quantity: float
    date: str
    order_number: str
    status: str
    source_row: int | None
    source: str
    note: str
    review_hits: list[ReviewHit]


def read_order_csv(csv_path: str) -> list[dict]:
    rows: list[dict] = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        expected = {"Date", "Order Number", "Item Number", "QTY"}
        missing = expected.difference(reader.fieldnames or [])
        if missing:
            raise ValueError(f"CSV is missing required columns: {sorted(missing)}")
        for line_no, row in enumerate(reader, start=2):
            item_number = str(row.get("Item Number", "")).strip()
            if not item_number:
                continue
            try:
                qty = float(row.get("QTY", 0) or 0)
            except Exception as exc:
                raise ValueError(f"Invalid QTY on CSV line {line_no}: {row.get('QTY')!r}") from exc
            rows.append(
                {
                    "Date": str(row.get("Date", "")).strip(),
                    "Order Number": str(row.get("Order Number", "")).strip(),
                    "Item Number": item_number,
                    "QTY": qty,
                }
            )
    return rows


def group_order_rows(rows: Iterable[dict]) -> list[dict]:
    grouped: dict[tuple[str, str, str], float] = defaultdict(float)
    for row in rows:
        key = (row["Date"], row["Order Number"], row["Item Number"])
        grouped[key] += float(row["QTY"])
    out: list[dict] = []
    for (date_text, order_no, item_number), qty in sorted(grouped.items(), key=lambda x: (x[0][1], x[0][0], x[0][2])):
        out.append({
            "Date": date_text,
            "Order Number": order_no,
            "Item Number": item_number,
            "QTY": qty,
        })
    return out


def write_audit_csv(audit_path: str, results: list[OrderResolveResult]) -> None:
    out_path = Path(audit_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "date",
                "order_number",
                "item_number",
                "quantity",
                "status",
                "resolved_source_row",
                "resolution_source",
                "note",
                "review_hit_1",
                "review_hit_2",
                "review_hit_3",
            ]
        )
        for result in results:
            hit_text = []
            for hit in result.review_hits[:3]:
                conf = "" if hit.confidence is None else f" {hit.confidence:.1%}"
                hit_text.append(f"row {hit.source_row} | {hit.match_type}{conf} | {hit.note}")
            while len(hit_text) < 3:
                hit_text.append("")
            writer.writerow(
                [
                    result.date,
                    result.order_number,
                    result.item_number,
                    result.quantity,
                    result.status,
                    result.source_row or "",
                    result.source,
                    result.note,
                    *hit_text,
                ]
            )


def export_yuchang_po_compact_by_rows(
    template_path: str,
    output_path: str,
    order_date: str,
    order_number: str,
    source_rows_with_qty: list[tuple[int, float | int]],
    *,
    sheet_name: str = "Sheet1",
    qty_col: str = "L",
    date_cell: str = "C10",
    order_no_cell: str = "H10",
    export_min_col: str = "B",
    export_max_col: str = "N",
    header_start_row: int = 1,
    header_end_row: int = 14,
    footer_start_row: int = 3045,
    footer_end_row: int = 3050,
    max_blank_scan: int = 3,
) -> dict:
    template_path = str(template_path)
    output_path = str(output_path)

    wb = load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

    src = wb[sheet_name]

    qty_idx = column_index_from_string(qty_col)
    export_min_idx = column_index_from_string(export_min_col)
    export_max_idx = column_index_from_string(export_max_col)

    if export_min_idx > export_max_idx:
        raise ValueError("export_min_col cannot be after export_max_col")

    src[date_cell] = order_date
    src[order_no_cell] = order_number

    qty_by_row: dict[int, float] = defaultdict(float)
    for source_row, qty in source_rows_with_qty:
        qty_by_row[int(source_row)] += float(qty)

    matched_rows: set[int] = set(qty_by_row.keys())
    for row, qty in qty_by_row.items():
        src.cell(row, qty_idx).value = qty

    def row_has_item(r: int) -> bool:
        if src.cell(r, 2).value in (None, ""):
            return False
        detail_cols = (3, 4, 5, 6, 7, 8)
        return any(src.cell(r, c).value not in (None, "") for c in detail_cols)

    def row_has_export_content(r: int) -> bool:
        for c in range(export_min_idx, export_max_idx + 1):
            if src.cell(r, c).value not in (None, ""):
                return True
        return False

    rows_to_keep: set[int] = set(range(header_start_row, header_end_row + 1))
    rows_to_keep.update(range(footer_start_row, footer_end_row + 1))
    rows_to_keep.update(matched_rows)

    for row in sorted(matched_rows):
        scan = row - 1
        blank_count = 0
        while scan > header_end_row:
            if row_has_item(scan):
                break
            if row_has_export_content(scan):
                rows_to_keep.add(scan)
                blank_count = 0
            else:
                blank_count += 1
                if blank_count > max_blank_scan:
                    break
            scan -= 1

    kept_rows = sorted(rows_to_keep)
    row_map = {old_row: new_row for new_row, old_row in enumerate(kept_rows, start=1)}

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name

    out_ws.sheet_view.showGridLines = src.sheet_view.showGridLines
    out_ws.sheet_properties = copy(src.sheet_properties)
    out_ws.page_margins = copy(src.page_margins)
    out_ws.page_setup = copy(src.page_setup)
    out_ws.print_options = copy(src.print_options)
    out_ws.sheet_format = copy(src.sheet_format)

    if src.freeze_panes:
        try:
            fp_col = src.freeze_panes.column
            fp_row = src.freeze_panes.row
            if fp_row in row_map and export_min_idx <= fp_col <= export_max_idx:
                new_fp = f"{get_column_letter(fp_col - export_min_idx + 1)}{row_map[fp_row]}"
                out_ws.freeze_panes = new_fp
        except Exception:
            pass

    for src_col in range(export_min_idx, export_max_idx + 1):
        out_col = src_col - export_min_idx + 1
        src_letter = get_column_letter(src_col)
        out_letter = get_column_letter(out_col)
        src_dim = src.column_dimensions[src_letter]
        out_dim = out_ws.column_dimensions[out_letter]
        out_dim.width = src_dim.width
        out_dim.hidden = src_dim.hidden
        out_dim.bestFit = src_dim.bestFit
        out_dim.collapsed = src_dim.collapsed
        out_dim.outlineLevel = src_dim.outlineLevel

    for old_row in kept_rows:
        new_row = row_map[old_row]
        src_row_dim = src.row_dimensions[old_row]
        out_row_dim = out_ws.row_dimensions[new_row]
        out_row_dim.height = src_row_dim.height
        out_row_dim.hidden = src_row_dim.hidden
        out_row_dim.outlineLevel = src_row_dim.outlineLevel
        out_row_dim.collapsed = src_row_dim.collapsed

        for src_col in range(export_min_idx, export_max_idx + 1):
            out_col = src_col - export_min_idx + 1
            s = src.cell(old_row, src_col)
            d = out_ws.cell(new_row, out_col)
            if isinstance(s.value, str) and s.value.startswith("="):
                old_coord = f"{get_column_letter(src_col)}{old_row}"
                new_coord = f"{get_column_letter(out_col)}{new_row}"
                try:
                    d.value = Translator(s.value, origin=old_coord).translate_formula(new_coord)
                except Exception:
                    d.value = s.value
            else:
                d.value = s.value

            if s.has_style:
                d.font = copy(s.font)
                d.fill = copy(s.fill)
                d.border = copy(s.border)
                d.alignment = copy(s.alignment)
                d.number_format = s.number_format
                d.protection = copy(s.protection)
            if s.hyperlink:
                d._hyperlink = copy(s.hyperlink)
            if s.comment:
                d.comment = copy(s.comment)

    for merged_range in src.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col < export_min_idx or max_col > export_max_idx:
            continue
        if not all(r in row_map for r in range(min_row, max_row + 1)):
            continue
        out_ws.merge_cells(
            start_row=row_map[min_row],
            start_column=min_col - export_min_idx + 1,
            end_row=row_map[max_row],
            end_column=max_col - export_min_idx + 1,
        )

    for img in getattr(src, "_images", []):
        try:
            anc = img.anchor
            if not hasattr(anc, "_from"):
                continue

            old_from_row = anc._from.row + 1
            old_from_col = anc._from.col + 1

            # Keep header images like the logo even when they sit outside the export column window.
            if old_from_row not in row_map:
                if old_from_row <= header_end_row:
                    target_row = old_from_row
                else:
                    continue
            else:
                target_row = row_map[old_from_row]

            img_bytes = img._data()
            new_img = XLImage(BytesIO(img_bytes))
            new_anchor = deepcopy(anc)
            new_anchor._from.row = max(0, target_row - 1)

            if old_from_col < export_min_idx:
                new_anchor._from.col = 0
            elif old_from_col > export_max_idx:
                new_anchor._from.col = export_max_idx - export_min_idx
            else:
                new_anchor._from.col = old_from_col - export_min_idx

            if hasattr(new_anchor, "to") and hasattr(new_anchor.to, "row") and hasattr(new_anchor.to, "col"):
                old_to_row = anc.to.row + 1
                old_to_col = anc.to.col + 1

                if old_to_row not in row_map:
                    if old_to_row <= header_end_row:
                        target_to_row = old_to_row
                    else:
                        target_to_row = target_row
                else:
                    target_to_row = row_map[old_to_row]

                new_anchor.to.row = max(0, target_to_row - 1)

                if old_to_col < export_min_idx:
                    new_anchor.to.col = 0
                elif old_to_col > export_max_idx:
                    new_anchor.to.col = export_max_idx - export_min_idx
                else:
                    new_anchor.to.col = old_to_col - export_min_idx

            new_img.anchor = new_anchor
            out_ws.add_image(new_img)
        except Exception:
            pass

    amount_src_col = column_index_from_string("M")
    if export_min_idx <= amount_src_col <= export_max_idx and footer_end_row in row_map:
        amount_out_col = get_column_letter(amount_src_col - export_min_idx + 1)
        total_row = row_map[footer_end_row]
        detail_rows = [row_map[r] for r in kept_rows if header_end_row < r < footer_start_row]
        if detail_rows:
            first_detail = min(detail_rows)
            last_detail = max(detail_rows)
            out_ws[f"{amount_out_col}{total_row}"] = f"=SUM({amount_out_col}{first_detail}:{amount_out_col}{last_detail})"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    return {
        "output_path": output_path,
        "matched_rows": sorted(matched_rows),
        "kept_source_rows": kept_rows,
    }


# ------------------------------------------------------------
# Window
# ------------------------------------------------------------
class YUOrderReviewWindow(QMainWindow):
    def __init__(
        self,
        db: SQLHelper,
        prefix: str,
        template_path: str,
        order_csv_path: str,
        output_dir: str,
    ):
        super().__init__()
        self.db = db
        self.prefix = prefix
        self.tables = {
            "supplier_lines": f"{prefix}_supplier_lines",
            "match_review": f"{prefix}_match_review",
            "match_candidates": f"{prefix}_match_candidates",
            "import_runs": f"{prefix}_import_runs",
        }
        self.template_path = str(template_path or "").strip()
        self.order_csv_path = str(order_csv_path or "").strip()
        self.output_dir = str(output_dir or "").strip()
        self.settings = QSettings("Windsor", "WidgetApp")

        self.raw_order_rows: list[dict] = []
        self.grouped_order_rows: list[dict] = []
        self.current_rows: list[dict] = []
        self.current_selected_detail: dict[str, Any] | None = None

        self.setObjectName("YUOrderReviewWindow")
        self.setWindowTitle("YU Order Review")
        self.resize(1480, 920)
        self.setMinimumSize(QSize(1280, 760))

        self.build_ui()
        self.apply_theme_from_settings()
        self.check_required_tables()
        self.load_order_csv()
        self.refresh_all()

    # ---------------- UI
    def build_ui(self):
        central = QWidget(self)
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(8)

        self.top_frame = QFrame(central)
        self.top_frame.setObjectName("top_frame")
        self.top_frame.setMinimumHeight(58)
        top_layout = QHBoxLayout(self.top_frame)

        self.last_import_browser = QTextBrowser(self.top_frame)
        self.last_import_browser.setObjectName("last_import_browser")
        self.last_import_browser.setMaximumWidth(180)
        self.last_import_browser.setMaximumHeight(42)
        self.last_import_browser.setReadOnly(True)
        top_layout.addWidget(self.last_import_browser)

        self.title_label = QLabel("YU Order Review", self.top_frame)
        self.title_label.setObjectName("title_label")
        top_layout.addWidget(self.title_label)

        self.refresh_button = QPushButton("Refresh", self.top_frame)
        self.refresh_button.clicked.connect(self.refresh_all)
        top_layout.addWidget(self.refresh_button)

        self.open_workbook_button = QPushButton("Open Workbook", self.top_frame)
        self.open_workbook_button.clicked.connect(self.open_workbook)
        self.open_workbook_button.setEnabled(bool(self.template_path))
        top_layout.addWidget(self.open_workbook_button)

        self.export_visible_button = QPushButton("Export Visible", self.top_frame)
        self.export_visible_button.clicked.connect(self.export_visible_orders)
        top_layout.addWidget(self.export_visible_button)

        self.open_output_button = QPushButton("Open Output Folder", self.top_frame)
        self.open_output_button.clicked.connect(self.open_output_dir)
        top_layout.addWidget(self.open_output_button)

        top_layout.addStretch(1)

        top_layout.addWidget(QLabel("Order", self.top_frame))
        self.order_filter_combo = QComboBox(self.top_frame)
        self.order_filter_combo.currentTextChanged.connect(self.load_main_table)
        top_layout.addWidget(self.order_filter_combo)

        top_layout.addWidget(QLabel("Filter", self.top_frame))
        self.filter_combo = QComboBox(self.top_frame)
        self.filter_combo.addItems(["All", "Resolved", "Needs Review", "Unmatched", "Error"])
        self.filter_combo.currentTextChanged.connect(self.load_main_table)
        top_layout.addWidget(self.filter_combo)

        self.search_edit = QLineEdit(self.top_frame)
        self.search_edit.setPlaceholderText("Search item or description...")
        self.search_edit.returnPressed.connect(self.load_main_table)
        self.search_edit.textChanged.connect(self.on_search_text_changed)
        self.search_edit.setMinimumWidth(240)
        top_layout.addWidget(self.search_edit)

        self.search_button = QPushButton("Show Data", self.top_frame)
        self.search_button.clicked.connect(self.load_main_table)
        top_layout.addWidget(self.search_button)

        root.addWidget(self.top_frame)

        stats_frame = QFrame(central)
        stats_frame.setObjectName("stats_frame")
        stats_layout = QHBoxLayout(stats_frame)
        stats_layout.setSpacing(8)

        self.total_box = self.make_stat_box("Total Lines")
        self.resolved_box = self.make_stat_box("Resolved")
        self.needs_review_box = self.make_stat_box("Needs Review")
        self.unmatched_box = self.make_stat_box("Unmatched / Error")

        stats_layout.addWidget(self.total_box["frame"])
        stats_layout.addWidget(self.resolved_box["frame"])
        stats_layout.addWidget(self.needs_review_box["frame"])
        stats_layout.addWidget(self.unmatched_box["frame"])
        stats_layout.addStretch(1)

        root.addWidget(stats_frame)

        splitter = QSplitter(Qt.Horizontal, central)

        left_panel = QWidget(splitter)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)

        self.main_table = QTableWidget(left_panel)
        self.main_table.setObjectName("main_table")
        self.main_table.setColumnCount(8)
        self.main_table.setHorizontalHeaderLabels([
            "Date", "Order No", "Item Number", "Qty", "Status", "Source Row", "Source", "Best Hit"
        ])
        self.main_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.main_table.setSelectionMode(QTableWidget.SingleSelection)
        self.main_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.main_table.verticalHeader().setVisible(False)
        self.main_table.itemSelectionChanged.connect(self.on_main_selection_changed)
        left_layout.addWidget(self.main_table)

        right_panel = QWidget(splitter)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(8)

        self.detail_frame = QFrame(right_panel)
        self.detail_frame.setObjectName("detail_frame")
        detail_layout = QVBoxLayout(self.detail_frame)
        detail_layout.setSpacing(6)

        self.date_box = self.make_value_row(detail_layout, "Date")
        self.order_no_box = self.make_value_row(detail_layout, "Order No")
        self.item_box = self.make_value_row(detail_layout, "Item Number")
        self.qty_box = self.make_value_row(detail_layout, "Qty")
        self.status_box = self.make_value_row(detail_layout, "Status")
        self.resolved_row_box = self.make_value_row(detail_layout, "Resolved Row")

        right_layout.addWidget(self.detail_frame)

        self.preview_frame = QFrame(right_panel)
        self.preview_frame.setObjectName("preview_frame")
        preview_layout = QVBoxLayout(self.preview_frame)
        preview_layout.addWidget(QLabel("Supplier Preview / Notes", self.preview_frame))
        self.preview_browser = QTextBrowser(self.preview_frame)
        self.preview_browser.setObjectName("preview_browser")
        preview_layout.addWidget(self.preview_browser)
        right_layout.addWidget(self.preview_frame, 1)

        self.candidate_frame = QFrame(right_panel)
        self.candidate_frame.setObjectName("candidate_frame")
        candidate_layout = QVBoxLayout(self.candidate_frame)
        candidate_layout.addWidget(QLabel("Candidate Supplier Rows", self.candidate_frame))

        self.candidate_table = QTableWidget(self.candidate_frame)
        self.candidate_table.setObjectName("candidate_table")
        self.candidate_table.setColumnCount(5)
        self.candidate_table.setHorizontalHeaderLabels(["Row", "Hit Type", "Conf %", "Description", "Labelled As"])
        self.candidate_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.candidate_table.setSelectionMode(QTableWidget.SingleSelection)
        self.candidate_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.candidate_table.verticalHeader().setVisible(False)
        self.candidate_table.itemSelectionChanged.connect(self.on_candidate_selection_changed)
        self.candidate_table.itemDoubleClicked.connect(lambda *_args: self.confirm_selected_candidate())
        candidate_layout.addWidget(self.candidate_table)

        action_bar = QHBoxLayout()
        self.confirm_button = QPushButton("Confirm Candidate", self.candidate_frame)
        self.confirm_button.clicked.connect(self.confirm_selected_candidate)
        action_bar.addWidget(self.confirm_button)

        self.clear_button = QPushButton("Clear Confirm", self.candidate_frame)
        self.clear_button.clicked.connect(self.clear_confirmation)
        action_bar.addWidget(self.clear_button)

        self.manual_row_edit = QLineEdit(self.candidate_frame)
        self.manual_row_edit.setPlaceholderText("Type source row...")
        self.manual_row_edit.returnPressed.connect(self.confirm_manual_row)
        action_bar.addWidget(self.manual_row_edit)

        self.manual_row_button = QPushButton("Use Source Row", self.candidate_frame)
        self.manual_row_button.clicked.connect(self.confirm_manual_row)
        action_bar.addWidget(self.manual_row_button)

        candidate_layout.addLayout(action_bar)
        right_layout.addWidget(self.candidate_frame, 1)

        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([920, 520])
        root.addWidget(splitter, 1)

        status_bar = QStatusBar(self)
        self.setStatusBar(status_bar)

        install_table_font_context_menu(self, self.main_table, self.settings, "yu_order_review", "yu_order_review/main_table")
        install_table_font_context_menu(self, self.candidate_table, self.settings, "yu_order_review", "yu_order_review/candidate_table")

    def make_stat_box(self, label_text: str) -> dict[str, QWidget]:
        frame = QFrame(self)
        frame.setObjectName("stat_box_frame")
        frame.setMinimumSize(QSize(160, 78))
        frame.setMaximumHeight(88)
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(4)
        label = QLabel(label_text, frame)
        label.setObjectName("stat_label")
        value = QLabel("", frame)
        value.setObjectName("stat_value")
        value.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        layout.addWidget(value)
        return {"frame": frame, "label": label, "value": value}

    def make_value_row(self, parent_layout: QVBoxLayout, label_text: str) -> QLabel:
        row = QHBoxLayout()
        label = QLabel(label_text, self.detail_frame)
        label.setObjectName("field_label")
        label.setMinimumWidth(120)
        value = QLabel("", self.detail_frame)
        value.setObjectName("field_value")
        value.setMinimumHeight(34)
        value.setAlignment(Qt.AlignCenter)
        row.addWidget(label)
        row.addWidget(value, 1)
        parent_layout.addLayout(row)
        return value

    def apply_theme_from_settings(self):
        theme_name = str(self.settings.value("theme", "dark"))
        app = QApplication.instance()
        if app is None:
            return
        if theme_name == "light":
            app.setStyleSheet("""
                QWidget { background: #e3e8ee; color: #1b1f23; }
                QMainWindow, QSplitter, QScrollArea, QScrollArea > QWidget > QWidget {
                    background: #e3e8ee; color: #1b1f23;
                }
                QFrame#top_frame, QFrame#stats_frame, QFrame#detail_frame,
                QFrame#preview_frame, QFrame#candidate_frame, QFrame#stat_box_frame {
                    background: #eef2f5;
                    border: 1px solid #c3cbd3;
                    border-radius: 8px;
                }
                QLabel#title_label {
                    font-size: 16px;
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_label, QLabel#stat_label {
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_value, QLabel#stat_value {
                    background: #f8fafb;
                    color: #111111;
                    border: 1px solid #aab4be;
                    border-radius: 4px;
                    padding: 6px;
                    font-weight: 700;
                }
                QLineEdit, QComboBox, QTableWidget, QTextBrowser {
                    background: #f8fafb;
                    color: #111111;
                    border: 1px solid #aab4be;
                    border-radius: 4px;
                    selection-background-color: #c9def5;
                    selection-color: #111111;
                }
                QHeaderView::section {
                    background: #d9e0e7;
                    color: #1b1f23;
                    border: 1px solid #bac4ce;
                    padding: 4px;
                    font-weight: 600;
                }
                QPushButton {
                    background: #d6dde5;
                    color: #1b1f23;
                    border: 1px solid #aab4be;
                    border-radius: 6px;
                    padding: 7px 12px;
                    min-height: 28px;
                    font-weight: 600;
                }
                QPushButton:hover {
                    background: #c9d3dc;
                }
                QPushButton:disabled {
                    background: #e3e8ee;
                    color: #808890;
                }
            """)
        elif theme_name == "high":
            app.setStyleSheet("""
                QWidget { background: #000000; color: #ffff00; }
                QMainWindow, QSplitter { background: #000000; color: #ffff00; }
                QFrame#top_frame, QFrame#stats_frame, QFrame#detail_frame,
                QFrame#preview_frame, QFrame#candidate_frame, QFrame#stat_box_frame {
                    background: #000000;
                    border: 2px solid #ffff00;
                    border-radius: 8px;
                }
                QLabel#title_label {
                    font-size: 16px;
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_label, QLabel#stat_label {
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_value, QLabel#stat_value {
                    background: #000000;
                    color: #ffff00;
                    border: 2px solid #ffff00;
                    border-radius: 4px;
                    padding: 6px;
                    font-weight: 700;
                }
                QLineEdit, QComboBox, QTableWidget, QTextBrowser {
                    background: #000000;
                    color: #ffff00;
                    border: 2px solid #ffff00;
                    border-radius: 4px;
                    selection-background-color: #ffff00;
                    selection-color: #000000;
                }
                QHeaderView::section {
                    background: #000000;
                    color: #ffff00;
                    border: 2px solid #ffff00;
                    padding: 4px;
                    font-weight: 700;
                }
                QPushButton {
                    background: #000000;
                    color: #ffff00;
                    border: 2px solid #ffff00;
                    border-radius: 6px;
                    padding: 7px 12px;
                    min-height: 28px;
                    font-weight: 700;
                }
                QPushButton:hover {
                    background: #333300;
                }
            """)
        else:
            app.setStyleSheet("""
                QWidget { background: #23272e; color: #e8eaed; }
                QMainWindow, QSplitter, QScrollArea, QScrollArea > QWidget > QWidget {
                    background: #23272e; color: #e8eaed;
                }
                QFrame#top_frame, QFrame#stats_frame, QFrame#detail_frame,
                QFrame#preview_frame, QFrame#candidate_frame, QFrame#stat_box_frame {
                    background: #2d333b;
                    border: 1px solid #48515b;
                    border-radius: 8px;
                }
                QLabel#title_label {
                    font-size: 16px;
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_label, QLabel#stat_label {
                    font-weight: 700;
                    background: transparent;
                    border: none;
                }
                QLabel#field_value, QLabel#stat_value {
                    background: #3b424d;
                    color: #f1f3f4;
                    border: 1px solid #59626d;
                    border-radius: 4px;
                    padding: 6px;
                    font-weight: 700;
                }
                QLineEdit, QComboBox, QTableWidget, QTextBrowser {
                    background: #3b424d;
                    color: #f1f3f4;
                    border: 1px solid #59626d;
                    border-radius: 4px;
                    selection-background-color: #5b7394;
                    selection-color: #ffffff;
                }
                QHeaderView::section {
                    background: #2f353d;
                    color: #e8eaed;
                    border: 1px solid #59626d;
                    padding: 4px;
                    font-weight: 600;
                }
                QPushButton {
                    background: #3c444f;
                    color: #e8eaed;
                    border: 1px solid #59626d;
                    border-radius: 6px;
                    padding: 7px 12px;
                    min-height: 28px;
                    font-weight: 600;
                }
                QPushButton:hover {
                    background: #4a5461;
                }
                QPushButton:disabled {
                    background: #323840;
                    color: #7b838c;
                }
            """)

    # ---------------- data load
    def check_required_tables(self):
        missing = []
        for table_name in self.tables.values():
            exists = self.db.scalar(
                "SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE' AND TABLE_NAME = ?",
                (table_name,),
            )
            if not exists:
                missing.append(table_name)
        if missing:
            raise RuntimeError(
                "Required test tables were not found.\n\n"
                "Run the importer first, for example:\n"
                "python yu_sqlserver_test_import_v3.py rebuild --workbook \"yuchang_order_form_matched_third_pass.xlsx\"\n\n"
                f"Missing tables: {', '.join(missing)}"
            )

    def load_order_csv(self):
        if not self.order_csv_path:
            raise RuntimeError("An order CSV path is required.")
        self.raw_order_rows = read_order_csv(self.order_csv_path)
        self.grouped_order_rows = group_order_rows(self.raw_order_rows)
        order_numbers = sorted({str(row["Order Number"]) for row in self.grouped_order_rows}, key=str.lower)

        self.order_filter_combo.blockSignals(True)
        self.order_filter_combo.clear()
        self.order_filter_combo.addItem("All")
        for order_no in order_numbers:
            self.order_filter_combo.addItem(order_no)
        self.order_filter_combo.blockSignals(False)

    def refresh_all(self):
        self.load_last_import_date()
        self.load_counts()
        self.load_main_table()

    def load_last_import_date(self):
        row = self.db.one(
            f"SELECT TOP 1 imported_at, workbook_path FROM dbo.{self.tables['import_runs']} ORDER BY imported_at DESC"
        )
        if row is None:
            self.last_import_browser.setPlainText("")
            return
        text = str(row.get("imported_at") or "")
        self.last_import_browser.setPlainText(text)

    def resolve_item(self, item_number: str, quantity: float, date_text: str, order_number: str) -> OrderResolveResult:
        item_number = str(item_number or "").strip()

        direct_rows_raw = self.db.all(
            f"""
            SELECT DISTINCT source_row
            FROM dbo.{self.tables["supplier_lines"]}
            WHERE row_kind = 'detail'
              AND (
                    ISNULL(current_item_number, '') = ?
                 OR ISNULL(literal_item_number, '') = ?
              )
            ORDER BY source_row
            """,
            (item_number, item_number),
        )
        direct_rows = [int(row["source_row"]) for row in direct_rows_raw]

        if len(direct_rows) == 1:
            return OrderResolveResult(
                item_number=item_number,
                quantity=quantity,
                date=date_text,
                order_number=order_number,
                status="resolved",
                source_row=direct_rows[0],
                source="sheet1_direct",
                note="Resolved by exact item number already present on supplier rows.",
                review_hits=[],
            )
        if len(direct_rows) > 1:
            return OrderResolveResult(
                item_number=item_number,
                quantity=quantity,
                date=date_text,
                order_number=order_number,
                status="error",
                source_row=None,
                source="sheet1_direct_duplicate",
                note=f"Item number exists more than once on supplier rows: {direct_rows}.",
                review_hits=[],
            )

        final_rows_raw = self.db.all(
            f"""
            SELECT DISTINCT source_row
            FROM dbo.{self.tables["match_review"]}
            WHERE ISNULL(final_selection, '') = ?
            ORDER BY source_row
            """,
            (item_number,),
        )
        final_rows = [int(row["source_row"]) for row in final_rows_raw]
        if len(final_rows) == 1:
            return OrderResolveResult(
                item_number=item_number,
                quantity=quantity,
                date=date_text,
                order_number=order_number,
                status="resolved",
                source_row=final_rows[0],
                source="match_review_final",
                note="Resolved by approved Final Selection.",
                review_hits=[],
            )
        if len(final_rows) > 1:
            return OrderResolveResult(
                item_number=item_number,
                quantity=quantity,
                date=date_text,
                order_number=order_number,
                status="error",
                source_row=None,
                source="match_review_final_duplicate",
                note=f"Item number exists more than once in Final Selection: {final_rows}.",
                review_hits=[],
            )

        hits: list[ReviewHit] = []

        suggested_hits = self.db.all(
            f"""
            SELECT source_row, confidence_pct, review_row
            FROM dbo.{self.tables["match_review"]}
            WHERE ISNULL(suggested_match, '') = ?
            ORDER BY source_row
            """,
            (item_number,),
        )
        for row in suggested_hits:
            conf = row.get("confidence_pct")
            hits.append(
                ReviewHit(
                    source_row=int(row["source_row"]),
                    match_type="suggested_match",
                    confidence=float(conf) if conf is not None else None,
                    note=f"suggested in Match_Review row {row.get('review_row')}",
                )
            )

        candidate_hits = self.db.all(
            f"""
            SELECT source_row, confidence_pct, review_row, candidate_rank
            FROM dbo.{self.tables["match_candidates"]}
            WHERE candidate_item_number = ?
            ORDER BY source_row, candidate_rank
            """,
            (item_number,),
        )
        for row in candidate_hits:
            conf = row.get("confidence_pct")
            hits.append(
                ReviewHit(
                    source_row=int(row["source_row"]),
                    match_type=f"candidate_{row.get('candidate_rank')}",
                    confidence=float(conf) if conf is not None else None,
                    note=f"candidate in Match_Review row {row.get('review_row')}",
                )
            )

        if hits:
            hits_sorted = sorted(
                hits,
                key=lambda x: (x.confidence if x.confidence is not None else -1),
                reverse=True,
            )
            best = hits_sorted[0]
            confidence_txt = "" if best.confidence is None else f" Best confidence: {best.confidence:.1%}."
            note = (
                f"No approved mapping yet. Best review hit is source row {best.source_row} via {best.match_type}.{confidence_txt} "
                "Confirm a candidate row first."
            )
            return OrderResolveResult(
                item_number=item_number,
                quantity=quantity,
                date=date_text,
                order_number=order_number,
                status="needs_review",
                source_row=None,
                source="review_hit_only",
                note=note,
                review_hits=hits_sorted,
            )

        return OrderResolveResult(
            item_number=item_number,
            quantity=quantity,
            date=date_text,
            order_number=order_number,
            status="unmatched",
            source_row=None,
            source="not_found",
            note="No direct match, no approved final selection, and no candidate review hits were found.",
            review_hits=[],
        )

    def load_counts(self):
        resolved = 0
        needs_review = 0
        unmatched_error = 0

        for row in self.grouped_order_rows:
            result = self.resolve_item(
                item_number=row["Item Number"],
                quantity=float(row["QTY"]),
                date_text=row["Date"],
                order_number=row["Order Number"],
            )
            if result.status == "resolved":
                resolved += 1
            elif result.status == "needs_review":
                needs_review += 1
            else:
                unmatched_error += 1

        self.total_box["value"].setText(str(len(self.grouped_order_rows)))
        self.resolved_box["value"].setText(str(resolved))
        self.needs_review_box["value"].setText(str(needs_review))
        self.unmatched_box["value"].setText(str(unmatched_error))

    def passes_filters(self, row: dict, result: OrderResolveResult) -> bool:
        order_filter = (self.order_filter_combo.currentText() or "").strip()
        if order_filter and order_filter.lower() != "all":
            if str(row["Order Number"]) != order_filter:
                return False

        status_filter = (self.filter_combo.currentText() or "").strip().lower()
        if status_filter == "resolved" and result.status != "resolved":
            return False
        if status_filter == "needs review" and result.status != "needs_review":
            return False
        if status_filter == "unmatched" and result.status != "unmatched":
            return False
        if status_filter == "error" and result.status != "error":
            return False

        search = (self.search_edit.text() or "").strip().lower()
        if search:
            haystack = " ".join([
                str(row["Date"]),
                str(row["Order Number"]),
                str(row["Item Number"]),
                str(result.note),
                str(result.source),
            ]).lower()
            if search not in haystack:
                return False

        return True

    def load_main_table(self):
        self.current_rows = []
        table = self.main_table
        table.setRowCount(0)

        for row in self.grouped_order_rows:
            result = self.resolve_item(
                item_number=row["Item Number"],
                quantity=float(row["QTY"]),
                date_text=row["Date"],
                order_number=row["Order Number"],
            )
            if not self.passes_filters(row, result):
                continue

            detail = dict(row)
            detail["resolve_result"] = result
            self.current_rows.append(detail)

            row_index = table.rowCount()
            table.insertRow(row_index)
            values = [
                row["Date"],
                row["Order Number"],
                row["Item Number"],
                self.format_qty(row["QTY"]),
                result.status,
                "" if result.source_row is None else str(result.source_row),
                result.source,
                self.best_hit_text(result),
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem("" if value is None else str(value))
                if col in {3, 5}:
                    item.setTextAlignment(Qt.AlignCenter)
                if col == 2:
                    item.setData(Qt.UserRole, row["Item Number"])
                table.setItem(row_index, col, item)

            self.apply_main_row_styles(row_index, result)

        self.auto_size_all_columns(table)
        table.resizeRowsToContents()

        if table.rowCount() > 0:
            table.selectRow(0)
            self.on_main_selection_changed()
        else:
            self.current_selected_detail = None
            self.clear_detail_panel()

        self.statusBar().showMessage(f"Loaded {table.rowCount()} order lines.", 4000)

    def apply_main_row_styles(self, row_index: int, result: OrderResolveResult):
        if result.status == "resolved":
            bg = QColor(0, 255, 0)
            fg = QColor(Qt.black)
        elif result.status == "needs_review":
            bg = QColor(255, 199, 206)
            fg = QColor(156, 0, 6)
        elif result.status == "unmatched":
            bg = QColor(214, 214, 214) if row_index % 2 == 0 else QColor(190, 190, 190)
            fg = QColor(Qt.black)
        else:
            bg = QColor(255, 235, 156)
            fg = QColor(156, 101, 0)

        for col in range(self.main_table.columnCount()):
            item = self.main_table.item(row_index, col)
            if item is not None:
                item.setBackground(QBrush(bg))
                item.setForeground(QBrush(fg))

    def best_hit_text(self, result: OrderResolveResult) -> str:
        if result.status == "resolved":
            return ""
        if result.review_hits:
            hit = result.review_hits[0]
            conf = "" if hit.confidence is None else f" {hit.confidence:.1%}"
            return f"row {hit.source_row} {hit.match_type}{conf}"
        return ""

    def selected_row_detail(self) -> dict[str, Any] | None:
        row_index = self.main_table.currentRow()
        if row_index < 0 or row_index >= len(self.current_rows):
            return None
        return self.current_rows[row_index]

    def load_candidates_for_item(self, item_number: str, current_result: OrderResolveResult) -> list[dict]:
        rows: list[dict] = []

        direct_rows = self.db.all(
            f"""
            SELECT source_row, description, labelled_as
            FROM dbo.{self.tables["supplier_lines"]}
            WHERE row_kind = 'detail'
              AND (ISNULL(current_item_number, '') = ? OR ISNULL(literal_item_number, '') = ?)
            ORDER BY source_row
            """,
            (item_number, item_number),
        )
        for row in direct_rows:
            rows.append({
                "source_row": int(row["source_row"]),
                "hit_type": "direct",
                "confidence": None,
                "description": str(row.get("description") or ""),
                "labelled_as": str(row.get("labelled_as") or ""),
            })

        final_rows = self.db.all(
            f"""
            SELECT r.source_row, s.description, s.labelled_as
            FROM dbo.{self.tables["match_review"]} r
            LEFT JOIN dbo.{self.tables["supplier_lines"]} s ON s.source_row = r.source_row
            WHERE ISNULL(r.final_selection, '') = ?
            ORDER BY r.source_row
            """,
            (item_number,),
        )
        for row in final_rows:
            rows.append({
                "source_row": int(row["source_row"]),
                "hit_type": "final_selection",
                "confidence": None,
                "description": str(row.get("description") or ""),
                "labelled_as": str(row.get("labelled_as") or ""),
            })

        suggested_rows = self.db.all(
            f"""
            SELECT r.source_row, r.confidence_pct, s.description, s.labelled_as
            FROM dbo.{self.tables["match_review"]} r
            LEFT JOIN dbo.{self.tables["supplier_lines"]} s ON s.source_row = r.source_row
            WHERE ISNULL(r.suggested_match, '') = ?
            ORDER BY r.source_row
            """,
            (item_number,),
        )
        for row in suggested_rows:
            rows.append({
                "source_row": int(row["source_row"]),
                "hit_type": "suggested_match",
                "confidence": float(row["confidence_pct"]) if row.get("confidence_pct") is not None else None,
                "description": str(row.get("description") or ""),
                "labelled_as": str(row.get("labelled_as") or ""),
            })

        candidate_rows = self.db.all(
            f"""
            SELECT c.source_row, c.confidence_pct, c.candidate_rank, s.description, s.labelled_as
            FROM dbo.{self.tables["match_candidates"]} c
            LEFT JOIN dbo.{self.tables["supplier_lines"]} s ON s.source_row = c.source_row
            WHERE c.candidate_item_number = ?
            ORDER BY c.source_row, c.candidate_rank
            """,
            (item_number,),
        )
        for row in candidate_rows:
            rows.append({
                "source_row": int(row["source_row"]),
                "hit_type": f"candidate_{row.get('candidate_rank')}",
                "confidence": float(row["confidence_pct"]) if row.get("confidence_pct") is not None else None,
                "description": str(row.get("description") or ""),
                "labelled_as": str(row.get("labelled_as") or ""),
            })

        if current_result.source_row is not None:
            source_row = int(current_result.source_row)
            if not any(r["source_row"] == source_row and r["hit_type"] == "current" for r in rows):
                src = self.db.one(
                    f"""
                    SELECT source_row, description, labelled_as
                    FROM dbo.{self.tables["supplier_lines"]}
                    WHERE source_row = ?
                    """,
                    (source_row,),
                )
                if src is not None:
                    rows.insert(0, {
                        "source_row": source_row,
                        "hit_type": "current",
                        "confidence": None,
                        "description": str(src.get("description") or ""),
                        "labelled_as": str(src.get("labelled_as") or ""),
                    })

        deduped = []
        seen = set()
        for row in rows:
            key = (row["source_row"], row["hit_type"])
            if key in seen:
                continue
            seen.add(key)
            deduped.append(row)
        return deduped

    def on_main_selection_changed(self):
        detail = self.selected_row_detail()
        self.current_selected_detail = detail
        if detail is None:
            self.clear_detail_panel()
            return

        result: OrderResolveResult = detail["resolve_result"]

        self.date_box.setText(str(detail["Date"]))
        self.order_no_box.setText(str(detail["Order Number"]))
        self.item_box.setText(str(detail["Item Number"]))
        self.qty_box.setText(self.format_qty(detail["QTY"]))
        self.status_box.setText(result.status)
        self.resolved_row_box.setText("" if result.source_row is None else str(result.source_row))

        preview_lines = [
            f"<b>Resolution Source:</b> {self.html_text(result.source)}",
            f"<b>Note:</b> {self.html_text(result.note)}",
        ]
        self.preview_browser.setHtml("<br>".join(preview_lines))

        candidates = self.load_candidates_for_item(str(detail["Item Number"]), result)
        self.candidate_table.setRowCount(0)
        for cand in candidates:
            row_index = self.candidate_table.rowCount()
            self.candidate_table.insertRow(row_index)
            values = [
                str(cand["source_row"]),
                cand["hit_type"],
                self.format_confidence(cand["confidence"]),
                cand["description"],
                cand["labelled_as"],
            ]
            for col, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col in {0, 2}:
                    item.setTextAlignment(Qt.AlignCenter)
                if col == 0:
                    item.setData(Qt.UserRole, int(cand["source_row"]))
                self.candidate_table.setItem(row_index, col, item)

        self.auto_size_all_columns(self.candidate_table)
        self.candidate_table.resizeRowsToContents()
        if self.candidate_table.rowCount() > 0:
            self.candidate_table.selectRow(0)
            self.on_candidate_selection_changed()

    def on_candidate_selection_changed(self):
        row_index = self.candidate_table.currentRow()
        if row_index < 0:
            return
        item = self.candidate_table.item(row_index, 0)
        if item is None:
            return
        source_row = item.data(Qt.UserRole)
        try:
            source_row = int(source_row)
        except Exception:
            return

        preview = self.db.one(
            f"""
            SELECT
                s.source_row,
                s.current_item_number,
                s.literal_item_number,
                s.description,
                s.size_text,
                s.colour,
                s.pack_type,
                s.labelled_as,
                r.final_selection,
                r.suggested_match,
                r.review_reasons
            FROM dbo.{self.tables["supplier_lines"]} s
            LEFT JOIN dbo.{self.tables["match_review"]} r ON r.source_row = s.source_row
            WHERE s.source_row = ?
            """,
            (source_row,),
        )
        if preview is None:
            return

        lines = [
            f"<b>Source Row:</b> {self.html_text(preview.get('source_row'))}",
            f"<b>Current Item:</b> {self.html_text(preview.get('current_item_number') or preview.get('literal_item_number'))}",
            f"<b>Description:</b> {self.html_text(preview.get('description'))}",
            f"<b>Size:</b> {self.html_text(preview.get('size_text'))}",
            f"<b>Colour:</b> {self.html_text(preview.get('colour'))}",
            f"<b>Pack:</b> {self.html_text(preview.get('pack_type'))}",
            f"<b>Labelled As:</b> {self.html_text(preview.get('labelled_as'))}",
            f"<b>Final Selection:</b> {self.html_text(preview.get('final_selection'))}",
            f"<b>Suggested Match:</b> {self.html_text(preview.get('suggested_match'))}",
            "",
            f"<b>Review Reasons:</b><br>{self.html_text(preview.get('review_reasons'))}",
        ]
        self.preview_browser.setHtml("<br>".join(lines))

    def clear_detail_panel(self):
        for box in (self.date_box, self.order_no_box, self.item_box, self.qty_box, self.status_box, self.resolved_row_box):
            box.setText("")
        self.preview_browser.clear()
        self.candidate_table.setRowCount(0)
        self.manual_row_edit.clear()

    # ---------------- actions
    def selected_candidate_source_row(self) -> int | None:
        row_index = self.candidate_table.currentRow()
        if row_index < 0:
            return None
        item = self.candidate_table.item(row_index, 0)
        if item is None:
            return None
        try:
            return int(item.data(Qt.UserRole))
        except Exception:
            return None

    def next_review_row(self) -> int:
        value = self.db.scalar(f"SELECT ISNULL(MAX(review_row), 4) + 1 FROM dbo.{self.tables['match_review']}")
        return int(value or 5)

    def ensure_match_review_row(self, source_row: int) -> None:
        exists = self.db.scalar(
            f"SELECT COUNT(*) FROM dbo.{self.tables['match_review']} WHERE source_row = ?",
            (source_row,),
        )
        if exists:
            return
        src = self.db.one(
            f"""
            SELECT source_row, current_item_number, size_text, colour, pack_type, labelled_as
            FROM dbo.{self.tables['supplier_lines']}
            WHERE source_row = ?
            """,
            (source_row,),
        ) or {}
        self.db.execute(
            f"""
            INSERT INTO dbo.{self.tables["match_review"]} (
                source_row,
                review_row,
                final_selection,
                suggested_match,
                confidence_pct,
                gap_score,
                review_reasons,
                duplicate_state,
                source_item,
                size_text,
                colour,
                pack_type,
                unit_size,
                labelled_as,
                review_status
            )
            VALUES (?, ?, NULL, NULL, NULL, NULL, NULL, NULL, ?, ?, ?, ?, NULL, ?, 'unmatched')
            """,
            (
                source_row,
                self.next_review_row(),
                src.get("current_item_number"),
                src.get("size_text"),
                src.get("colour"),
                src.get("pack_type"),
                src.get("labelled_as"),
            ),
        )

    def reset_row_status_for_item(self, item_number: str, exclude_source_row: int | None = None):
        rows = self.db.all(
            f"""
            SELECT r.source_row,
                   CASE
                     WHEN ISNULL(r.suggested_match, '') <> ''
                          OR EXISTS (
                              SELECT 1 FROM dbo.{self.tables["match_candidates"]} c
                              WHERE c.source_row = r.source_row
                          )
                     THEN 'needs_review'
                     ELSE 'unmatched'
                   END AS new_status
            FROM dbo.{self.tables["match_review"]} r
            WHERE ISNULL(r.final_selection, '') = ?
            """,
            (item_number,),
        )
        for row in rows:
            source_row = int(row["source_row"])
            if exclude_source_row is not None and source_row == int(exclude_source_row):
                continue
            self.db.execute(
                f"""
                UPDATE dbo.{self.tables["match_review"]}
                SET final_selection = NULL, review_status = ?
                WHERE source_row = ?
                """,
                (str(row["new_status"]), source_row),
            )

    def confirm_item_to_source_row(self, item_number: str, source_row: int):
        self.ensure_match_review_row(source_row)
        self.reset_row_status_for_item(item_number, exclude_source_row=source_row)
        self.db.execute(
            f"""
            UPDATE dbo.{self.tables["match_review"]}
            SET final_selection = ?, review_status = 'confirmed'
            WHERE source_row = ?
            """,
            (item_number, source_row),
        )
        self.statusBar().showMessage(f"Confirmed {item_number} to source row {source_row}.", 5000)
        self.refresh_all()
        self.reselect_item(item_number)

    def confirm_selected_candidate(self):
        detail = self.current_selected_detail
        if detail is None:
            QMessageBox.warning(self, "YU Order Review", "Select an order line first.")
            return
        source_row = self.selected_candidate_source_row()
        if source_row is None:
            QMessageBox.warning(self, "YU Order Review", "Select a candidate supplier row first.")
            return
        self.confirm_item_to_source_row(str(detail["Item Number"]), int(source_row))

    def confirm_manual_row(self):
        detail = self.current_selected_detail
        if detail is None:
            QMessageBox.warning(self, "YU Order Review", "Select an order line first.")
            return
        text = (self.manual_row_edit.text() or "").strip()
        if not text:
            QMessageBox.warning(self, "YU Order Review", "Type a source row first.")
            self.manual_row_edit.setFocus()
            return
        try:
            source_row = int(text)
        except Exception:
            QMessageBox.warning(self, "YU Order Review", "Source row must be a whole number.")
            return

        exists = self.db.scalar(
            f"SELECT COUNT(*) FROM dbo.{self.tables['supplier_lines']} WHERE source_row = ?",
            (source_row,),
        )
        if not exists:
            QMessageBox.warning(self, "YU Order Review", f"Source row {source_row} was not found.")
            return

        self.confirm_item_to_source_row(str(detail["Item Number"]), source_row)

    def clear_confirmation(self):
        detail = self.current_selected_detail
        if detail is None:
            QMessageBox.warning(self, "YU Order Review", "Select an order line first.")
            return
        item_number = str(detail["Item Number"])

        rows = self.db.all(
            f"""
            SELECT r.source_row,
                   CASE
                     WHEN ISNULL(r.suggested_match, '') <> ''
                          OR EXISTS (
                              SELECT 1 FROM dbo.{self.tables["match_candidates"]} c
                              WHERE c.source_row = r.source_row
                          )
                     THEN 'needs_review'
                     ELSE 'unmatched'
                   END AS new_status
            FROM dbo.{self.tables["match_review"]} r
            WHERE ISNULL(r.final_selection, '') = ?
            """,
            (item_number,),
        )
        if not rows:
            QMessageBox.information(self, "YU Order Review", f"There is no confirmed mapping for {item_number}.")
            return

        for row in rows:
            self.db.execute(
                f"""
                UPDATE dbo.{self.tables["match_review"]}
                SET final_selection = NULL, review_status = ?
                WHERE source_row = ?
                """,
                (str(row["new_status"]), int(row["source_row"])),
            )

        self.statusBar().showMessage(f"Cleared confirmation for {item_number}.", 5000)
        self.refresh_all()
        self.reselect_item(item_number)

    def reselect_item(self, item_number: str):
        for row_index in range(self.main_table.rowCount()):
            item = self.main_table.item(row_index, 2)
            if item and (item.text() or "") == item_number:
                self.main_table.selectRow(row_index)
                self.main_table.scrollToItem(item)
                self.on_main_selection_changed()
                break

    def export_visible_orders(self):
        if not self.template_path:
            QMessageBox.warning(self, "YU Order Review", "No workbook template path is set.")
            return

        if not self.current_rows:
            QMessageBox.warning(self, "YU Order Review", "There are no visible order rows to export.")
            return

        output_dir = QFileDialog.getExistingDirectory(
            self,
            "Choose export output folder",
            self.output_dir or str(Path.cwd() / DEFAULT_OUTPUT_DIR),
        )
        if not output_dir:
            return

        visible_rows = [dict(row) for row in self.current_rows]
        unresolved = [row for row in visible_rows if row["resolve_result"].status != "resolved"]
        audit_path = str(Path(output_dir) / "yu_order_audit.csv")
        write_audit_csv(audit_path, [row["resolve_result"] for row in visible_rows])

        if unresolved:
            item_list = ", ".join(sorted({row["Item Number"] for row in unresolved}))
            QMessageBox.warning(
                self,
                "YU Order Review",
                "Export stopped because some visible items are still unresolved.\n\n"
                f"Items: {item_list}\n\nAudit written to:\n{audit_path}",
            )
            return

        grouped: dict[tuple[str, str], list[tuple[int, float]]] = defaultdict(list)
        for row in visible_rows:
            result: OrderResolveResult = row["resolve_result"]
            grouped[(row["Date"], row["Order Number"])].append((int(result.source_row), float(row["QTY"])))

        exports = []
        for (date_text, order_number), resolved_rows in sorted(grouped.items(), key=lambda x: (x[0][1], x[0][0])):
            filename = f"yuchang_order_{order_number}.xlsx"
            output_path = str(Path(output_dir) / filename)
            export_yuchang_po_compact_by_rows(
                template_path=self.template_path,
                output_path=output_path,
                order_date=date_text,
                order_number=order_number,
                source_rows_with_qty=resolved_rows,
            )
            exports.append(output_path)

        self.output_dir = output_dir
        self.statusBar().showMessage(f"Exported {len(exports)} workbook(s).", 5000)
        QMessageBox.information(
            self,
            "YU Order Review",
            "Export complete.\n\n"
            + "\n".join(exports)
            + f"\n\nAudit:\n{audit_path}",
        )

    def open_workbook(self):
        if not self.template_path:
            QMessageBox.information(self, "YU Order Review", "No workbook path is set.")
            return
        try:
            resolved = str(Path(self.template_path).resolve())
            if sys.platform.startswith("win"):
                os.startfile(resolved)
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(resolved))
        except Exception as exc:
            QMessageBox.warning(self, "YU Order Review", f"Could not open workbook:\n{exc}")

    def open_output_dir(self):
        target = Path(self.output_dir or (Path.cwd() / DEFAULT_OUTPUT_DIR))
        target.mkdir(parents=True, exist_ok=True)
        try:
            resolved = str(target.resolve())
            if sys.platform.startswith("win"):
                os.startfile(resolved)
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(resolved))
        except Exception as exc:
            QMessageBox.warning(self, "YU Order Review", f"Could not open output folder:\n{exc}")

    def on_search_text_changed(self, text: str):
        if not text.strip():
            self.load_main_table()

    # helpers
    def format_confidence(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        try:
            return f"{float(value):.1%}"
        except Exception:
            return str(value)

    def format_qty(self, value: Any) -> str:
        try:
            number = float(value)
        except Exception:
            return str(value)
        if number.is_integer():
            return f"{int(number)}"
        return f"{number:,.2f}"

    def html_text(self, value: Any) -> str:
        text = "" if value is None else str(value)
        return (
            text.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace("\n", "<br>")
        )

    def auto_size_all_columns(self, table: QTableWidget):
        for column_index in range(table.columnCount()):
            try:
                table.resizeColumnToContents(column_index)
            except Exception:
                pass


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Standalone YU order review/export test window.")
    parser.add_argument("--prefix", default=DEFAULT_PREFIX, help=f"Table prefix. Default: {DEFAULT_PREFIX}")
    parser.add_argument("--template", default=DEFAULT_TEMPLATE, help="Path to the matched workbook/template")
    parser.add_argument("--order-csv", default=DEFAULT_ORDER_CSV, help="Path to the order CSV")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help=f"Default output folder. Default: {DEFAULT_OUTPUT_DIR}")
    parser.add_argument("--base-dir", default=None, help="Optional base directory to search for client_config.json")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    app = QApplication(sys.argv)
    db = None
    try:
        db = SQLHelper(base_dir=Path(args.base_dir).resolve() if args.base_dir else None)
        window = YUOrderReviewWindow(
            db=db,
            prefix=args.prefix,
            template_path=args.template,
            order_csv_path=args.order_csv,
            output_dir=args.output_dir,
        )
        window.show()
        return app.exec()
    except Exception as exc:
        QMessageBox.critical(None, "YU Order Review", str(exc))
        return 1
    finally:
        if db is not None:
            try:
                db.close()
            except Exception:
                pass


if __name__ == "__main__":
    raise SystemExit(main())
